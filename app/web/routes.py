# pyright: reportGeneralTypeIssues=false
# pyright: reportArgumentType=false
# pyright: reportAttributeAccessIssue=false
# pyright: reportOptionalMemberAccess=false
# pyright: reportOperatorIssue=false
# pyright: reportCallIssue=false
# pyright: reportOptionalOperand=false
# pyright: reportMissingImports=false
# pyright: reportAssignmentType=false
# pyright: reportOptionalSubscript=false
# The above directives suppress SQLAlchemy ORM type checker false positives
# SQLAlchemy uses metaprogramming that type checkers don't understand
from __future__ import annotations

from typing import Optional
from pathlib import Path
from datetime import date, datetime, timedelta, time, timezone
import calendar as pycalendar
import os
import uuid
import asyncio
import logging

# Set up logger for this module
logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, Form, HTTPException, Request, File, UploadFile, Query, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_session
from app.core.security import verify_password, get_password_hash
from app.core.email import send_email
from app.core.email_to_ticket_v2 import get_local_time
from app.models.project import Project
from app.models.task import Task
from app.models.user import User
from app.models.enums import TaskStatus, TaskPriority, MeetingPlatform
from app.models.workspace import Workspace
from app.models.assignment import Assignment
from app.models.comment import Comment
from app.models.comment_attachment import CommentAttachment
from app.models.task_history import TaskHistory
from app.models.notification import Notification
from app.models.chat import Chat, ChatMember, Message
from app.models.meeting import Meeting, MeetingAttendee
from app.models.company import Company
from app.models.contact import Contact
from app.models.lead import Lead
from app.models.deal import Deal
from app.models.activity import Activity
from app.core.bubbles_personality import get_conversational_response

BASE_DIR = Path(__file__).resolve().parents[1]
templates = Jinja2Templates(directory=str(BASE_DIR / 'templates'))

# Store original TemplateResponse
_original_template_response = templates.TemplateResponse

def enhanced_template_response(name: str, context: dict, *args, **kwargs):
    """Enhanced TemplateResponse that adds workspace from request state"""
    # ALWAYS add workspace to context if request is present
    if 'request' in context:
        request = context['request']
        # Check if workspace exists in request.state
        if hasattr(request, 'state') and hasattr(request.state, 'workspace'):
            context['workspace'] = request.state.workspace
        # If not in state, try to get it from context (already passed)
        elif 'workspace' not in context:
            context['workspace'] = None
    
    return _original_template_response(name, context, *args, **kwargs)

# Replace the TemplateResponse method - this affects ALL template calls
templates.TemplateResponse = enhanced_template_response

def format_datetime_tz(dt, tz_name=None, format_str="%Y-%m-%d %H:%M"):
    """Convert UTC datetime to specified timezone and format it.
    
    The tz_name should come from the workspace timezone setting.
    If tz_name is None or empty, defaults to UTC (no conversion).
    """
    if dt is None:
        return ""
    
    # If no timezone specified, default to UTC (no offset applied)
    if not tz_name:
        tz_name = "UTC"
    
    from datetime import timezone as dt_timezone
    
    # Handle UTC directly without needing pytz/zoneinfo
    if tz_name == "UTC":
        if isinstance(dt, datetime):
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=dt_timezone.utc)
            return dt.strftime(format_str)
        return dt.strftime(format_str) if isinstance(dt, datetime) else str(dt)
    
    try:
        from zoneinfo import ZoneInfo
        has_zoneinfo = True
    except ImportError:
        has_zoneinfo = False
    
    try:
        # Try pytz first as it's more reliable on Windows
        import pytz
        if isinstance(dt, datetime):
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=pytz.UTC)
            target_tz = pytz.timezone(tz_name)
            local_dt = dt.astimezone(target_tz)
            return local_dt.strftime(format_str)
    except ImportError:
        pass
    
    # Fallback to zoneinfo
    if has_zoneinfo:
        try:
            if isinstance(dt, datetime):
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=dt_timezone.utc)
                target_tz = ZoneInfo(tz_name)  # type: ignore[possibly-undefined]
                local_dt = dt.astimezone(target_tz)
                return local_dt.strftime(format_str)
        except Exception:
            pass
    
    # Ultimate fallback - treat as UTC (no offset) to avoid wrong times
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=dt_timezone.utc)
        return dt.strftime(format_str)
    
    # Last resort - just format as-is
    return dt.strftime(format_str) if isinstance(dt, datetime) else str(dt)

async def get_workspace_for_user(user_id: int, db: AsyncSession) -> Optional[Workspace]:
    """Get workspace with branding for a user"""
    try:
        user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
        if not user:
            return None
        workspace = (await db.execute(select(Workspace).where(Workspace.id == user.workspace_id))).scalar_one_or_none()
        return workspace
    except Exception:
        return None

# Add helper functions to Jinja2 globals for use in templates
templates.env.globals['now'] = datetime.utcnow

# Add timezone formatting filter
templates.env.filters['format_datetime_tz'] = format_datetime_tz

router = APIRouter(tags=['web'])


# --------------------------
# Authentication Dependencies
# --------------------------
async def get_current_user(request: Request, db: AsyncSession = Depends(get_session)) -> User:
    """Get current authenticated user or raise HTTPException"""
    user_id = request.session.get('user_id')
    if not user_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        raise HTTPException(status_code=401, detail="User not found or inactive")
    
    return user


async def get_current_admin(request: Request, db: AsyncSession = Depends(get_session)) -> User:
    """Get current authenticated admin user or raise HTTPException"""
    user = await get_current_user(request, db)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


# --------------------------
# App Download Page
# --------------------------
@router.get('/download', response_class=HTMLResponse)
async def download_app(request: Request):
    """Download page for mobile app - shows Android APK download or iOS PWA instructions"""
    return templates.TemplateResponse('download.html', {'request': request})


# --------------------------
# Auth (session-based for web)
# --------------------------
@router.get('/login', response_class=HTMLResponse)
async def web_login(request: Request):
    return templates.TemplateResponse('auth/login.html', {'request': request, 'error': None})


@router.post('/login')
async def web_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_session),
):
    result = await db.execute(select(User).where(User.username == username))
    user = result.scalar_one_or_none()
    if not user or not verify_password(password, user.hashed_password):
        return templates.TemplateResponse('auth/login.html', {'request': request, 'error': 'Invalid username or password'}, status_code=400)
    
    # Check if user is active
    if not user.is_active:
        return templates.TemplateResponse('auth/login.html', {'request': request, 'error': 'Your account has been deactivated. Please contact your administrator.'}, status_code=403)
    
    request.session['user_id'] = user.id
    request.session['workspace_id'] = user.workspace_id
    # Redirect to profile completion if not completed
    if not user.profile_completed:
        return RedirectResponse('/web/profile/complete', status_code=303)
    return RedirectResponse('/web/dashboard', status_code=303)


@router.get('/signup', response_class=HTMLResponse)
async def web_signup(request: Request):
    return templates.TemplateResponse('auth/signup.html', {'request': request, 'error': None})


@router.post('/signup')
async def web_signup_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    company_name: str = Form(...),
    db: AsyncSession = Depends(get_session),
):
    # Import validation function
    from app.core.security import validate_password
    
    # Validate password
    is_valid, error_msg = validate_password(password)
    if not is_valid:
        return templates.TemplateResponse('auth/signup.html', {'request': request, 'error': error_msg}, status_code=400)
    
    exists = await db.execute(select(User).where(User.username == username))
    if exists.scalar_one_or_none():
        return templates.TemplateResponse('auth/signup.html', {'request': request, 'error': 'Username already taken'}, status_code=400)
    # Create workspace and user
    # Self-registered users become admin of their own workspace
    ws = Workspace(
        name=f"{username}'s Workspace",
        site_title=company_name  # Use company name as the site title
    )
    db.add(ws)
    await db.flush()
    user = User(
        username=username, 
        hashed_password=get_password_hash(password), 
        workspace_id=ws.id,
        profile_completed=False,
        email_verified=True,
        is_admin=True  # Self-registered users are admins of their workspace
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    request.session['user_id'] = user.id
    return RedirectResponse('/web/profile/complete', status_code=303)


@router.post('/logout')
async def web_logout(request: Request):
    request.session.clear()
    return RedirectResponse('/', status_code=303)


# --------------------------
# Profile Completion
# --------------------------
@router.get('/profile/complete', response_class=HTMLResponse)
async def web_profile_complete(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    if user.profile_completed:
        return RedirectResponse('/web/projects', status_code=303)
    return templates.TemplateResponse('auth/profile_complete.html', {'request': request, 'user': user, 'error': None})


@router.post('/profile/complete')
async def web_profile_complete_post(
    request: Request,
    full_name: str = Form(...),
    email: str = Form(...),
    preferred_meeting_platform: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    # Check if email is already used by another user
    if email:
        exists = await db.execute(select(User).where(User.email == email, User.id != user_id))
        if exists.scalar_one_or_none():
            return templates.TemplateResponse('auth/profile_complete.html', {'request': request, 'user': user, 'error': 'Email already in use'}, status_code=400)
    user.full_name = full_name
    user.email = email
    if preferred_meeting_platform:
        try:
            user.preferred_meeting_platform = MeetingPlatform(preferred_meeting_platform)
        except ValueError:
            pass
    user.profile_completed = True
    await db.commit()
    return RedirectResponse('/web/projects', status_code=303)


@router.get('/profile', response_class=HTMLResponse)
async def web_profile(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    return templates.TemplateResponse('auth/profile.html', {'request': request, 'user': user})


@router.post('/profile')
async def web_profile_post(
    request: Request,
    full_name: str = Form(...),
    email: str = Form(...),
    preferred_meeting_platform: Optional[str] = Form(None),
    calendar_color: Optional[str] = Form(None),
    mute_ticket_notifications: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Check if email is already used by another user
    if email:
        exists = await db.execute(select(User).where(User.email == email, User.id != user_id))
        if exists.scalar_one_or_none():
            return templates.TemplateResponse('auth/profile.html', {
                'request': request, 
                'user': user, 
                'error': 'Email already in use'
            }, status_code=400)
    
    user.full_name = full_name
    user.email = email
    if preferred_meeting_platform:
        try:
            user.preferred_meeting_platform = MeetingPlatform(preferred_meeting_platform)
        except ValueError:
            user.preferred_meeting_platform = None
    else:
        user.preferred_meeting_platform = None
    if calendar_color:
        user.calendar_color = calendar_color
    
    # Update mute_ticket_notifications (for all users)
    user.mute_ticket_notifications = mute_ticket_notifications == 'true'
    
    await db.commit()
    return templates.TemplateResponse('auth/profile.html', {
        'request': request, 
        'user': user, 
        'success': 'Profile updated successfully'
    })


@router.post('/profile/picture')
async def web_profile_picture_upload(
    request: Request,
    profile_picture: UploadFile = File(...),
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Validate file type
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
    if profile_picture.content_type not in allowed_types:
        return templates.TemplateResponse('auth/profile.html', {
            'request': request,
            'user': user,
            'error': 'Invalid file type. Please upload a JPEG, PNG, GIF, or WebP image.'
        }, status_code=400)
    
    # Validate file size (max 5MB)
    content = await profile_picture.read()
    if len(content) > 5 * 1024 * 1024:
        return templates.TemplateResponse('auth/profile.html', {
            'request': request,
            'user': user,
            'error': 'File too large. Maximum size is 5MB.'
        }, status_code=400)
    
    # Create uploads directory if it doesn't exist
    upload_dir = BASE_DIR / 'uploads' / 'profile_pictures'
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Delete old profile picture if it exists
    if user.profile_picture:
        old_file = BASE_DIR / user.profile_picture.lstrip('/')
        if old_file.exists():
            old_file.unlink()
    
    # Generate unique filename
    filename_str = profile_picture.filename or 'upload.jpg'
    file_extension = filename_str.split('.')[-1] if '.' in filename_str else 'jpg'
    filename = f"{user_id}_{uuid.uuid4().hex[:8]}.{file_extension}"
    file_path = upload_dir / filename
    
    # Save file
    with open(file_path, 'wb') as f:
        f.write(content)
    
    # Update user profile picture path (relative to BASE_DIR)
    user.profile_picture = f"/uploads/profile_pictures/{filename}"
    await db.commit()
    
    return RedirectResponse('/web/profile?success=picture', status_code=303)


@router.get('/uploads/profile_pictures/{filename}')
async def serve_profile_picture(filename: str):
    """Serve profile picture files"""
    # Prevent path traversal attacks
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    file_path = BASE_DIR / 'uploads' / 'profile_pictures' / filename
    
    # Ensure the resolved path is within the uploads directory
    try:
        file_path = file_path.resolve()
        upload_base = (BASE_DIR / 'uploads' / 'profile_pictures').resolve()
        if not str(file_path).startswith(str(upload_base)):
            raise HTTPException(status_code=403, detail="Access denied")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file path")
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Profile picture not found")
    return FileResponse(file_path)


# --------------------------
# Google OAuth Integration
# --------------------------
@router.get('/auth/google/link')
async def web_google_oauth_link(request: Request, db: AsyncSession = Depends(get_session)):
    """Initiate Google OAuth flow to link account"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    try:
        from app.core.config import get_settings
        from app.core.google_oauth import get_authorization_url
        
        settings = get_settings()
        
        # Check if Google OAuth is configured
        if not settings.google_client_id or not settings.google_client_secret:
            logger.warning("Google OAuth not configured")
            return RedirectResponse('/web/profile?error=google_config', status_code=303)
        
        logger.info(f"Starting Google OAuth for user {user_id}")
        
        auth_url, state = get_authorization_url(user_id)
        # Store state in session for verification
        request.session['google_oauth_state'] = state
        return RedirectResponse(auth_url, status_code=303)
    except Exception as e:
        logger.error(f"Error initiating Google OAuth: {e}", exc_info=True)
        return RedirectResponse('/web/profile?error=google_config', status_code=303)


@router.get('/auth/google/callback')
async def web_google_oauth_callback(
    request: Request,
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    db: AsyncSession = Depends(get_session)
):
    """Handle Google OAuth callback"""
    if error:
        return RedirectResponse(f'/web/profile?error=google_denied', status_code=303)
    
    if not code or not state:
        return RedirectResponse(f'/web/profile?error=google_invalid', status_code=303)
    
    # Verify state matches session
    session_state = request.session.get('google_oauth_state')
    if not session_state or session_state != state:
        return RedirectResponse(f'/web/profile?error=google_state_mismatch', status_code=303)
    
    try:
        user_id = int(state)
        user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
        
        if not user or not user.is_active:
            return RedirectResponse('/web/login', status_code=303)
        
        from app.core.google_oauth import exchange_code_for_tokens, get_google_user_info
        
        # Exchange code for tokens
        token_info = exchange_code_for_tokens(code, state)
        
        # Get user info from Google
        google_user_info = get_google_user_info(token_info['access_token'])
        
        # Update user with Google credentials
        user.google_id = google_user_info.get('id')
        user.google_access_token = token_info['access_token']
        user.google_refresh_token = token_info['refresh_token']
        user.google_token_expiry = token_info['token_expiry']
        
        await db.commit()
        
        # Clear OAuth state from session
        request.session.pop('google_oauth_state', None)
        
        return RedirectResponse('/web/profile?success=google_linked', status_code=303)
        
    except Exception as e:
        logger.error(f"Error in Google OAuth callback: {e}", exc_info=True)
        return RedirectResponse(f'/web/profile?error=google_failed', status_code=303)


@router.post('/auth/google/unlink')
async def web_google_oauth_unlink(request: Request, db: AsyncSession = Depends(get_session)):
    """Unlink Google account"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Clear Google OAuth fields
    user.google_id = None
    user.google_access_token = None
    user.google_refresh_token = None
    user.google_token_expiry = None
    
    await db.commit()
    
    return RedirectResponse('/web/profile?success=google_unlinked', status_code=303)


# --------------------------
# Dashboard
# --------------------------
@router.get('/dashboard', response_class=HTMLResponse)
async def web_dashboard(request: Request, view: Optional[str] = None, user_id: Optional[int] = None, db: AsyncSession = Depends(get_session)):
    """Main dashboard with stats and overview"""
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can see user view
    if view == 'user' and not user.is_admin:
        view = 'personal'
    
    from app.models.ticket import Ticket
    from app.models.project_member import ProjectMember
    from datetime import timedelta
    from sqlalchemy import func
    
    today = date.today()
    week_ago = today - timedelta(days=7)
    
    # Get user's projects (admin sees all, user sees assigned)
    if user.is_admin:
        projects_result = await db.execute(
            select(Project)
            .where(Project.workspace_id == user.workspace_id, Project.is_archived == False)
        )
        projects = projects_result.scalars().all()
        project_ids = [p.id for p in projects]
    else:
        member_result = await db.execute(
            select(ProjectMember.project_id)
            .where(ProjectMember.user_id == current_user_id)
        )
        project_ids = [r[0] for r in member_result.fetchall()]
        projects_result = await db.execute(
            select(Project)
            .where(Project.id.in_(project_ids), Project.is_archived == False)
        )
        projects = projects_result.scalars().all()
    
    # My Tasks stats
    my_tasks_result = await db.execute(
        select(Task)
        .join(Assignment, Task.id == Assignment.task_id)
        .where(Assignment.assignee_id == current_user_id, Task.status != TaskStatus.done)
    )
    my_tasks = my_tasks_result.scalars().all()
    
    # Tasks done this week
    my_done_result = await db.execute(
        select(func.count())
        .select_from(Task)
        .join(Assignment, Task.id == Assignment.task_id)
        .where(
            Assignment.assignee_id == current_user_id,
            Task.status == TaskStatus.done,
            Task.updated_at >= datetime.combine(week_ago, time.min)
        )
    )
    my_tasks_done = my_done_result.scalar() or 0
    
    # Tasks due soon (next 7 days)
    tasks_due_result = await db.execute(
        select(Task, Project.name.label('project_name'))
        .join(Project, Task.project_id == Project.id)
        .join(Assignment, Task.id == Assignment.task_id)
        .where(
            Assignment.assignee_id == current_user_id,
            Task.status != TaskStatus.done,
            Task.due_date.isnot(None),
            Task.due_date <= today + timedelta(days=7)
        )
        .order_by(Task.due_date)
    )
    tasks_due_rows = tasks_due_result.fetchall()
    tasks_due_soon = []
    for row in tasks_due_rows:
        task = row[0]
        # Create a dict-like object to add project_name
        task_dict = {
            "id": task.id,
            "title": task.title,
            "description": task.description,
            "status": task.status,
            "priority": task.priority,
            "due_date": task.due_date,
            "project_id": task.project_id,
            "project_name": row[1]
        }
        tasks_due_soon.append(task_dict)
    
    # Open tickets (exclude archived)
    tickets_result = await db.execute(
        select(Ticket)
        .where(
            Ticket.workspace_id == user.workspace_id,
            Ticket.status.in_(['open', 'in_progress', 'waiting']),
            Ticket.is_archived == False
        )
    )
    open_tickets = tickets_result.scalars().all()
    urgent_tickets = len([t for t in open_tickets if t.priority == 'urgent'])
    high_tickets = len([t for t in open_tickets if t.priority == 'high'])
    
    # Meetings today
    meetings_result = await db.execute(
        select(func.count())
        .select_from(Meeting)
        .join(MeetingAttendee, Meeting.id == MeetingAttendee.meeting_id)
        .where(
            MeetingAttendee.user_id == current_user_id,
            Meeting.date == today,
            Meeting.is_cancelled == False
        )
    )
    meetings_today = meetings_result.scalar() or 0
    
    # Upcoming meetings (next 7 days)
    upcoming_result = await db.execute(
        select(Meeting)
        .join(MeetingAttendee, Meeting.id == MeetingAttendee.meeting_id)
        .where(
            MeetingAttendee.user_id == current_user_id,
            Meeting.date >= today,
            Meeting.date <= today + timedelta(days=7),
            Meeting.is_cancelled == False
        )
        .order_by(Meeting.date, Meeting.start_time)
        .limit(5)
    )
    upcoming_meetings = upcoming_result.scalars().all()
    
    # Team members count
    team_result = await db.execute(
        select(func.count())
        .select_from(User)
        .where(User.workspace_id == user.workspace_id, User.is_active == True)
    )
    total_team_members = team_result.scalar() or 0
    
    # ========== ENHANCED RECENT ACTIVITY ==========
    from app.models.task_history import TaskHistory
    
    # 1. User's recent activities (what the user did)
    my_activities_result = await db.execute(
        select(TaskHistory, Task.title.label('task_title'), Project.name.label('project_name'))
        .join(Task, TaskHistory.task_id == Task.id)
        .join(Project, Task.project_id == Project.id)
        .where(
            TaskHistory.editor_id == current_user_id,
            Task.project_id.in_(project_ids) if project_ids else True
        )
        .order_by(TaskHistory.created_at.desc())
        .limit(10)
    )
    my_activities_raw = my_activities_result.fetchall()
    
    recent_activities = []
    for activity, task_title, project_name in my_activities_raw:
        activity_type = 'task_created' if activity.field == 'created' else 'task_updated'
        if activity.field == 'status' and activity.new_value == 'done':
            activity_type = 'task_completed'
        
        # Create a readable description
        if activity.field == 'created':
            desc = f"You created task: {task_title}"
        elif activity.field == 'status':
            desc = f"You changed status to {activity.new_value} on: {task_title}"
        elif activity.field == 'assigned':
            desc = f"You assigned task: {task_title}"
        else:
            desc = f"You updated {activity.field.replace('_', ' ')} on: {task_title}"
        
        recent_activities.append({
            'type': activity_type,
            'description': desc,
            'task_id': activity.task_id,
            'task_title': task_title,
            'project_name': project_name,
            'created_at': activity.created_at
        })
    
    # 2. Tasks allocated TO the user (by others)
    tasks_allocated_to_me_result = await db.execute(
        select(Task, Project.name.label('project_name'), User.full_name.label('creator_name'), User.username.label('creator_username'))
        .join(Project, Task.project_id == Project.id)
        .join(Assignment, Task.id == Assignment.task_id)
        .join(User, Task.creator_id == User.id)
        .where(
            Assignment.assignee_id == current_user_id,
            Task.creator_id != current_user_id,  # Not created by me
            Task.status != TaskStatus.done,
            Project.workspace_id == user.workspace_id
        )
        .order_by(Task.created_at.desc())
        .limit(5)
    )
    tasks_allocated_to_me = []
    for task, project_name, creator_name, creator_username in tasks_allocated_to_me_result.fetchall():
        tasks_allocated_to_me.append({
            'id': task.id,
            'title': task.title,
            'project_name': project_name,
            'creator': creator_name or creator_username,
            'due_date': task.due_date,
            'priority': task.priority.value if hasattr(task.priority, 'value') else task.priority,
            'status': task.status.value if hasattr(task.status, 'value') else task.status,
            'created_at': task.created_at
        })
    
    # 3. Tasks the user allocated to OTHERS
    tasks_allocated_by_me_result = await db.execute(
        select(Task, Project.name.label('project_name'), User.full_name.label('assignee_name'), User.username.label('assignee_username'))
        .join(Project, Task.project_id == Project.id)
        .join(Assignment, Task.id == Assignment.task_id)
        .join(User, Assignment.assignee_id == User.id)
        .where(
            Task.creator_id == current_user_id,
            Assignment.assignee_id != current_user_id,  # Assigned to someone else
            Task.status != TaskStatus.done,
            Project.workspace_id == user.workspace_id
        )
        .order_by(Task.created_at.desc())
        .limit(5)
    )
    tasks_allocated_by_me = []
    for task, project_name, assignee_name, assignee_username in tasks_allocated_by_me_result.fetchall():
        tasks_allocated_by_me.append({
            'id': task.id,
            'title': task.title,
            'project_name': project_name,
            'assignee': assignee_name or assignee_username,
            'due_date': task.due_date,
            'priority': task.priority.value if hasattr(task.priority, 'value') else task.priority,
            'status': task.status.value if hasattr(task.status, 'value') else task.status,
            'created_at': task.created_at
        })
    
    # 4. Overdue tasks assigned to this user
    overdue_tasks_result = await db.execute(
        select(Task, Project.name.label('project_name'))
        .join(Project, Task.project_id == Project.id)
        .join(Assignment, Task.id == Assignment.task_id)
        .where(
            Assignment.assignee_id == current_user_id,
            Task.status != TaskStatus.done,
            Task.due_date < today,
            Project.workspace_id == user.workspace_id
        )
        .order_by(Task.due_date.asc())
        .limit(20)
    )
    overdue_tasks = []
    for task, project_name in overdue_tasks_result.fetchall():
        days_overdue = (today - task.due_date).days
        overdue_tasks.append({
            'id': task.id,
            'title': task.title,
            'project_name': project_name,
            'due_date': task.due_date,
            'days_overdue': days_overdue,
            'priority': task.priority.value if hasattr(task.priority, 'value') else task.priority,
            'status': task.status.value if hasattr(task.status, 'value') else task.status
        })
    
    # 5. Task progress updates (what others did on tasks related to user)
    # Get tasks assigned to user or created by user
    my_related_tasks = await db.execute(
        select(Task.id)
        .join(Assignment, Task.id == Assignment.task_id, isouter=True)
        .where(
            or_(
                Assignment.assignee_id == current_user_id,
                Task.creator_id == current_user_id
            )
        )
    )
    my_task_ids = [r[0] for r in my_related_tasks.fetchall()]
    
    task_progress_updates = []
    if my_task_ids:
        progress_result = await db.execute(
            select(TaskHistory, Task.title.label('task_title'), User.full_name.label('user_name'), User.username.label('user_username'))
            .join(Task, TaskHistory.task_id == Task.id)
            .join(User, TaskHistory.editor_id == User.id)
            .where(
                TaskHistory.task_id.in_(my_task_ids),
                TaskHistory.editor_id != current_user_id,  # Done by others
                TaskHistory.created_at >= datetime.combine(week_ago, time.min)
            )
            .order_by(TaskHistory.created_at.desc())
            .limit(10)
        )
        
        for history, task_title, user_name, user_username in progress_result.fetchall():
            user_display = user_name or user_username
            if history.field == 'status':
                desc = f"{user_display} changed status to {history.new_value}"
            elif history.field == 'comment':
                desc = f"{user_display} added a comment"
            else:
                desc = f"{user_display} updated {history.field.replace('_', ' ')}"
            
            task_progress_updates.append({
                'task_id': history.task_id,
                'task_title': task_title,
                'description': desc,
                'user_name': user_display,
                'action': history.field,
                'new_value': history.new_value,
                'created_at': history.created_at
            })
    
    # Admin stats - team performance
    team_tasks_completed = 0
    team_tickets_resolved = 0
    avg_response_time = None
    team_members = []
    selected_user = None
    selected_user_stats = {}
    selected_user_tasks = []
    selected_user_projects = []
    
    if user.is_admin:
        # Tasks completed by team this week
        team_done_result = await db.execute(
            select(func.count())
            .select_from(Task)
            .where(
                Task.project_id.in_(project_ids) if project_ids else True,
                Task.status == TaskStatus.done,
                Task.updated_at >= datetime.combine(week_ago, time.min)
            )
        )
        team_tasks_completed = team_done_result.scalar() or 0
        
        # Tickets resolved this week (exclude archived)
        resolved_result = await db.execute(
            select(func.count())
            .select_from(Ticket)
            .where(
                Ticket.workspace_id == user.workspace_id,
                Ticket.status.in_(['resolved', 'closed']),
                Ticket.resolved_at >= datetime.combine(week_ago, time.min),
                Ticket.is_archived == False
            )
        )
        team_tickets_resolved = resolved_result.scalar() or 0
        
        # User view - get all team members for selector
        if view == 'user':
            # Get all team members for the selector
            team_members_result = await db.execute(
                select(User)
                .where(User.workspace_id == user.workspace_id, User.is_active == True)
                .order_by(User.full_name)
            )
            team_members = team_members_result.scalars().all()
            
            # If a specific user_id is selected, get their data
            if user_id:
                selected_user_result = await db.execute(
                    select(User).where(User.id == user_id, User.workspace_id == user.workspace_id)
                )
                selected_user = selected_user_result.scalar_one_or_none()
                
                if selected_user:
                    # Get selected user's open tasks
                    su_tasks_result = await db.execute(
                        select(func.count())
                        .select_from(Task)
                        .join(Assignment, Task.id == Assignment.task_id)
                        .where(Assignment.assignee_id == user_id, Task.status != TaskStatus.done)
                    )
                    su_open_tasks = su_tasks_result.scalar() or 0
                    
                    # Get selected user's done tasks this week
                    su_done_result = await db.execute(
                        select(func.count())
                        .select_from(Task)
                        .join(Assignment, Task.id == Assignment.task_id)
                        .where(
                            Assignment.assignee_id == user_id,
                            Task.status == TaskStatus.done,
                            Task.updated_at >= datetime.combine(week_ago, time.min)
                        )
                    )
                    su_done_tasks = su_done_result.scalar() or 0
                    
                    # Get selected user's open tickets (exclude archived)
                    su_tickets_result = await db.execute(
                        select(func.count())
                        .select_from(Ticket)
                        .where(
                            Ticket.assigned_to_id == user_id,
                            Ticket.status.in_(['open', 'in_progress', 'waiting']),
                            Ticket.is_archived == False
                        )
                    )
                    su_open_tickets = su_tickets_result.scalar() or 0
                    
                    # Get selected user's overdue tasks
                    su_overdue_result = await db.execute(
                        select(func.count())
                        .select_from(Task)
                        .join(Assignment, Task.id == Assignment.task_id)
                        .where(
                            Assignment.assignee_id == user_id,
                            Task.status != TaskStatus.done,
                            Task.due_date < today
                        )
                    )
                    su_overdue_tasks = su_overdue_result.scalar() or 0
                    
                    # Get selected user's projects
                    su_projects_result = await db.execute(
                        select(Project)
                        .join(ProjectMember, Project.id == ProjectMember.project_id)
                        .where(ProjectMember.user_id == user_id, Project.is_archived == False)
                    )
                    selected_user_projects = su_projects_result.scalars().all()
                    
                    selected_user_stats = {
                        'open_tasks': su_open_tasks,
                        'done_this_week': su_done_tasks,
                        'open_tickets': su_open_tickets,
                        'overdue_tasks': su_overdue_tasks,
                        'projects': len(selected_user_projects)
                    }
                    
                    # Get selected user's tasks due soon
                    su_tasks_due_result = await db.execute(
                        select(Task, Project.name.label('project_name'))
                        .join(Project, Task.project_id == Project.id)
                        .join(Assignment, Task.id == Assignment.task_id)
                        .where(
                            Assignment.assignee_id == user_id,
                            Task.status != TaskStatus.done,
                            Task.due_date.isnot(None),
                            Task.due_date <= today + timedelta(days=14)
                        )
                        .order_by(Task.due_date)
                        .limit(10)
                    )
                    su_tasks_rows = su_tasks_due_result.fetchall()
                    for row in su_tasks_rows:
                        task = row[0]
                        selected_user_tasks.append({
                            "id": task.id,
                            "title": task.title,
                            "status": task.status,
                            "priority": task.priority,
                            "due_date": task.due_date,
                            "project_name": row[1]
                        })
    
    stats = {
        'my_tasks': len(my_tasks),
        'my_tasks_done': my_tasks_done,
        'open_tickets': len(open_tickets),
        'urgent_tickets': urgent_tickets,
        'high_tickets': high_tickets,
        'active_projects': len(projects),
        'total_team_members': total_team_members,
        'meetings_today': meetings_today,
        'team_tasks_completed': team_tasks_completed,
        'team_tickets_resolved': team_tickets_resolved,
        'avg_response_time': avg_response_time
    }
    
    workspace = await get_workspace_for_user(current_user_id, db)
    
    return templates.TemplateResponse('dashboard/index.html', {
        'request': request,
        'user': user,
        'stats': stats,
        'tasks_due_soon': tasks_due_soon,
        'upcoming_meetings': upcoming_meetings,
        'recent_activities': recent_activities,
        'tasks_allocated_to_me': tasks_allocated_to_me,
        'tasks_allocated_by_me': tasks_allocated_by_me,
        'overdue_tasks': overdue_tasks,
        'task_progress_updates': task_progress_updates,
        'projects': projects,
        'today': today,
        'workspace': workspace,
        'view': view or 'personal',
        'team_members': team_members,
        'selected_user': selected_user,
        'selected_user_stats': selected_user_stats,
        'selected_user_tasks': selected_user_tasks,
        'selected_user_projects': selected_user_projects
    })


@router.post('/dashboard/quick-task')
async def web_dashboard_quick_task(
    request: Request,
    title: str = Form(...),
    project_id: int = Form(...),
    priority: str = Form('medium'),
    due_date: Optional[date] = Form(None),
    customer_name: Optional[str] = Form(None),
    customer_surname: Optional[str] = Form(None),
    customer_email: Optional[str] = Form(None),
    customer_phone: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Quick add task from dashboard"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify project exists and user has access
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not project:
        return RedirectResponse('/web/dashboard?error=invalid_project', status_code=303)
    
    # Create task
    task = Task(
        title=title,
        project_id=project_id,
        creator_id=user_id,
        priority=TaskPriority(priority),
        due_date=due_date,
        customer_name=customer_name or None,
        customer_surname=customer_surname or None,
        customer_email=customer_email or None,
        customer_phone=customer_phone or None
    )
    db.add(task)
    await db.flush()
    
    # Auto-assign to creator
    assignment = Assignment(task_id=task.id, assignee_id=user_id)
    db.add(assignment)
    
    # Add task history
    history = TaskHistory(
        task_id=task.id,
        editor_id=user_id,
        field='created',
        new_value=title
    )
    db.add(history)
    
    # Track user behavior for learning
    from app.core.smart_suggestions import track_user_action
    await track_user_action(
        db, user_id, user.workspace_id, 'task_created', 'task',
        entity_id=task.id, project_id=project_id,
        field_name='priority', field_value=priority
    )
    
    await db.commit()
    
    return RedirectResponse('/web/dashboard', status_code=303)


# --------------------------
# What Changed While You Were Away - AI Summary Feature
# --------------------------
@router.get('/dashboard/what-changed')
async def get_what_changed(request: Request, db: AsyncSession = Depends(get_session)):
    """Get summary of changes since user's last visit"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'error': 'User not found'}, status_code=404)
    
    from app.models.ticket import Ticket
    from app.models.task_history import TaskHistory
    from app.models.project_member import ProjectMember
    from datetime import timedelta
    
    # Determine the time range to check
    now = datetime.utcnow()
    last_seen = user.last_seen_at or (now - timedelta(hours=24))  # Default to 24 hours if never tracked
    time_away = now - last_seen
    
    # Only show if away for more than 30 minutes
    if time_away < timedelta(minutes=30):
        return JSONResponse({
            'show_summary': False,
            'reason': 'Not away long enough',
            'minutes_away': int(time_away.total_seconds() / 60)
        })
    
    # Get user's projects
    if user.is_admin:
        projects_result = await db.execute(
            select(Project.id)
            .where(Project.workspace_id == user.workspace_id, Project.is_archived == False)
        )
        project_ids = [r[0] for r in projects_result.fetchall()]
    else:
        member_result = await db.execute(
            select(ProjectMember.project_id)
            .where(ProjectMember.user_id == user_id)
        )
        project_ids = [r[0] for r in member_result.fetchall()]
    
    changes = {
        'time_away_hours': round(time_away.total_seconds() / 3600, 1),
        'time_away_display': format_time_away(time_away),
        'last_seen': last_seen.isoformat() if last_seen else None,
        'preference': user.away_summary_preference or 'ask',
        'new_tasks_assigned': [],
        'task_updates': [],
        'new_tickets': [],
        'ticket_updates': [],
        'completed_tasks': [],
        'meetings_scheduled': [],
        'summary': ''
    }
    
    # 1. New tasks assigned to user since last seen
    new_tasks_result = await db.execute(
        select(Task, Project.name.label('project_name'), User.full_name.label('creator_name'))
        .join(Project, Task.project_id == Project.id)
        .join(Assignment, Task.id == Assignment.task_id)
        .join(User, Task.creator_id == User.id)
        .where(
            Assignment.assignee_id == user_id,
            Task.created_at > last_seen,
            Task.creator_id != user_id  # Not self-assigned
        )
        .order_by(Task.created_at.desc())
    )
    for task, project_name, creator_name in new_tasks_result.fetchall():
        changes['new_tasks_assigned'].append({
            'id': task.id,
            'title': task.title,
            'project': project_name,
            'creator': creator_name or 'Unknown',
            'priority': task.priority.value if hasattr(task.priority, 'value') else str(task.priority),
            'due_date': task.due_date.isoformat() if task.due_date else None
        })
    
    # 2. Updates on tasks assigned to user or created by user
    my_task_ids_result = await db.execute(
        select(Task.id)
        .join(Assignment, Task.id == Assignment.task_id, isouter=True)
        .where(
            or_(Assignment.assignee_id == user_id, Task.creator_id == user_id)
        )
    )
    my_task_ids = [r[0] for r in my_task_ids_result.fetchall()]
    
    if my_task_ids:
        task_updates_result = await db.execute(
            select(TaskHistory, Task.title.label('task_title'), User.full_name.label('editor_name'))
            .join(Task, TaskHistory.task_id == Task.id)
            .join(User, TaskHistory.editor_id == User.id)
            .where(
                TaskHistory.task_id.in_(my_task_ids),
                TaskHistory.created_at > last_seen,
                TaskHistory.editor_id != user_id  # Changes by others
            )
            .order_by(TaskHistory.created_at.desc())
            .limit(20)
        )
        for history, task_title, editor_name in task_updates_result.fetchall():
            if history.field == 'status' and history.new_value == 'done':
                changes['completed_tasks'].append({
                    'task_id': history.task_id,
                    'title': task_title,
                    'completed_by': editor_name or 'Unknown'
                })
            else:
                changes['task_updates'].append({
                    'task_id': history.task_id,
                    'title': task_title,
                    'field': history.field,
                    'new_value': history.new_value,
                    'editor': editor_name or 'Unknown'
                })
    
    # 3. New tickets in workspace (for admins/agents)
    if user.is_admin or user.can_see_all_tickets:
        new_tickets_result = await db.execute(
            select(Ticket)
            .where(
                Ticket.workspace_id == user.workspace_id,
                Ticket.created_at > last_seen
            )
            .order_by(Ticket.created_at.desc())
            .limit(10)
        )
        for ticket in new_tickets_result.scalars().all():
            changes['new_tickets'].append({
                'id': ticket.id,
                'subject': ticket.subject,
                'priority': ticket.priority,
                'status': ticket.status
            })
    
    # 4. Ticket updates on assigned tickets
    assigned_tickets_result = await db.execute(
        select(Ticket)
        .where(
            Ticket.assigned_to_id == user_id,
            Ticket.updated_at > last_seen,
            Ticket.updated_at != Ticket.created_at  # Actually updated, not just created
        )
    )
    for ticket in assigned_tickets_result.scalars().all():
        changes['ticket_updates'].append({
            'id': ticket.id,
            'subject': ticket.subject,
            'status': ticket.status
        })
    
    # 5. Meetings scheduled during absence
    meetings_result = await db.execute(
        select(Meeting)
        .join(MeetingAttendee, Meeting.id == MeetingAttendee.meeting_id)
        .where(
            MeetingAttendee.user_id == user_id,
            Meeting.created_at > last_seen,
            Meeting.is_cancelled == False
        )
        .order_by(Meeting.date, Meeting.start_time)
    )
    for meeting in meetings_result.scalars().all():
        changes['meetings_scheduled'].append({
            'id': meeting.id,
            'title': meeting.title,
            'date': meeting.date.isoformat() if meeting.date else None,
            'time': meeting.start_time.strftime('%H:%M') if meeting.start_time else None
        })
    
    # Generate AI-like summary
    changes['summary'] = generate_away_summary(changes)
    
    # Determine if we should show the summary
    total_changes = (
        len(changes['new_tasks_assigned']) + 
        len(changes['task_updates']) + 
        len(changes['new_tickets']) + 
        len(changes['completed_tasks']) +
        len(changes['meetings_scheduled'])
    )
    
    changes['show_summary'] = total_changes > 0
    changes['total_changes'] = total_changes
    
    return JSONResponse(changes)


@router.post('/dashboard/update-last-seen')
async def update_last_seen(request: Request, db: AsyncSession = Depends(get_session)):
    """Update user's last seen timestamp"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if user:
        user.last_seen_at = datetime.utcnow()
        await db.commit()
    
    return JSONResponse({'success': True})


@router.post('/dashboard/away-preference')
async def update_away_preference(request: Request, db: AsyncSession = Depends(get_session)):
    """Update user's preference for away summary"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    data = await request.json()
    preference = data.get('preference', 'ask')
    
    if preference not in ['always', 'ask', 'never']:
        return JSONResponse({'error': 'Invalid preference'}, status_code=400)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if user:
        user.away_summary_preference = preference
        await db.commit()
    
    return JSONResponse({'success': True, 'preference': preference})


def format_time_away(delta: timedelta) -> str:
    """Format time away into human-readable string"""
    total_seconds = int(delta.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    
    if hours >= 24:
        days = hours // 24
        remaining_hours = hours % 24
        if days == 1:
            return f"1 day and {remaining_hours} hours" if remaining_hours else "1 day"
        return f"{days} days and {remaining_hours} hours" if remaining_hours else f"{days} days"
    elif hours > 0:
        if hours == 1:
            return f"1 hour and {minutes} minutes" if minutes else "1 hour"
        return f"{hours} hours and {minutes} minutes" if minutes else f"{hours} hours"
    else:
        return f"{minutes} minutes"


def generate_away_summary(changes: dict) -> str:
    """Generate a friendly AI-style summary of changes"""
    parts = []
    
    time_away = changes.get('time_away_display', 'a while')
    
    # New tasks assigned
    new_tasks = len(changes.get('new_tasks_assigned', []))
    if new_tasks > 0:
        if new_tasks == 1:
            task = changes['new_tasks_assigned'][0]
            parts.append(f"📋 **{task['creator']}** assigned you a new task: \"{task['title']}\"")
        else:
            parts.append(f"📋 You have **{new_tasks} new tasks** assigned to you")
    
    # Completed tasks
    completed = len(changes.get('completed_tasks', []))
    if completed > 0:
        if completed == 1:
            task = changes['completed_tasks'][0]
            parts.append(f"✅ **{task['completed_by']}** completed the task: \"{task['title']}\"")
        else:
            parts.append(f"✅ **{completed} tasks** were completed by your team")
    
    # Task updates
    updates = len(changes.get('task_updates', []))
    if updates > 0:
        parts.append(f"📝 **{updates} updates** were made on tasks you're involved with")
    
    # New tickets
    tickets = len(changes.get('new_tickets', []))
    if tickets > 0:
        if tickets == 1:
            ticket = changes['new_tickets'][0]
            parts.append(f"🎫 **1 new ticket**: \"{ticket['subject']}\" ({ticket['priority']} priority)")
        else:
            parts.append(f"🎫 **{tickets} new tickets** were created")
    
    # Ticket updates
    ticket_updates = len(changes.get('ticket_updates', []))
    if ticket_updates > 0:
        parts.append(f"🔄 **{ticket_updates} tickets** assigned to you were updated")
    
    # Meetings
    meetings = len(changes.get('meetings_scheduled', []))
    if meetings > 0:
        if meetings == 1:
            meeting = changes['meetings_scheduled'][0]
            parts.append(f"📅 **1 meeting** was scheduled: \"{meeting['title']}\"")
        else:
            parts.append(f"📅 **{meetings} meetings** were scheduled for you")
    
    if not parts:
        return f"🎉 All caught up! Nothing significant happened while you were away."
    
    intro = f"Welcome back! Here's what happened in the last **{time_away}**:\n\n"
    return intro + "\n\n".join(parts)


# --------------------------
# Email Verification (kept for later, not enforced)
# --------------------------
@router.get('/verify-email', response_class=HTMLResponse)
async def web_verify_email(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    return templates.TemplateResponse('auth/verify_email.html', {'request': request, 'user': user, 'error': None, 'sent': False})


@router.post('/verify-email/request')
async def web_verify_email_request(request: Request, db: AsyncSession = Depends(get_session)):
    # No-op while OTP disabled
    return templates.TemplateResponse('auth/verify_email.html', {'request': request, 'user': None, 'error': None, 'sent': True})


@router.post('/verify-email/confirm')
async def web_verify_email_confirm(request: Request, code: str = Form(...), db: AsyncSession = Depends(get_session)):
    # No-op while OTP disabled
    return RedirectResponse('/web/projects', status_code=303)


# --------------------------
# Smart Suggestions API
# --------------------------
@router.get('/api/suggestions/assignees')
async def api_get_suggested_assignees(
    request: Request,
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_session)
):
    """Get suggested assignees based on user's history"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'suggestions': []})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'suggestions': []})
    
    from app.core.smart_suggestions import get_suggested_assignees
    suggestions = await get_suggested_assignees(db, user_id, user.workspace_id, project_id)
    return JSONResponse({'suggestions': suggestions})


@router.get('/api/suggestions/priority')
async def api_get_suggested_priority(
    request: Request,
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_session)
):
    """Get suggested priority based on user's history"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'suggestion': None})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'suggestion': None})
    
    from app.core.smart_suggestions import get_suggested_priority
    suggestion = await get_suggested_priority(db, user_id, user.workspace_id, project_id)
    return JSONResponse({'suggestion': suggestion})


@router.get('/api/suggestions/similar-tasks')
async def api_get_similar_tasks(
    request: Request,
    title: str,
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_session)
):
    """Find similar tasks based on title keywords"""
    user_id = request.session.get('user_id')
    if not user_id or not title:
        return JSONResponse({'similar_tasks': []})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'similar_tasks': []})
    
    from app.core.smart_suggestions import get_similar_tasks
    similar = await get_similar_tasks(db, user_id, user.workspace_id, title, project_id)
    return JSONResponse({'similar_tasks': similar})


@router.get('/api/suggestions/insights')
async def api_get_work_insights(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Get work pattern insights for current user"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'insights': {}})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'insights': {}})
    
    from app.core.smart_suggestions import get_work_pattern_insights
    insights = await get_work_pattern_insights(db, user_id, user.workspace_id)
    return JSONResponse({'insights': insights})


# --------------------------
# Global Search
# --------------------------
# TODO: FUTURE FEATURE - Add Attachment Search
# Currently searches: Tasks, Tickets, Projects
# Need to add search for attachments:
#   - TaskAttachment (search by filename, and future 'label' field)
#   - TicketAttachment (search by filename, and future 'label' field)
#   - CommentAttachment (search by filename, and future 'label' field)
# Implementation:
#   1. Add 'attachments' to results dict
#   2. Query all attachment tables where filename.ilike(search_term) OR label.ilike(search_term)
#   3. Join with parent entity (Task/Ticket/Comment) to get context
#   4. Return: filename, label, parent_type (task/ticket/comment), parent_id, parent_title
#   5. Update search/results.html template to display attachment results
#   6. Link attachments to their parent entity detail page
@router.get('/search')
async def web_global_search(
    request: Request,
    q: str = '',
    db: AsyncSession = Depends(get_session)
):
    """Global search across tasks, tickets, projects, and comments"""
    from app.models.ticket import Ticket
    
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    results = {
        'tasks': [],
        'tickets': [],
        'projects': [],
        'comments': [],
        'query': q
    }
    
    if q and len(q) >= 2:
        search_term = f'%{q}%'
        
        # Search tasks
        task_results = await db.execute(
            select(Task, Project.name.label('project_name'))
            .join(Project, Task.project_id == Project.id)
            .where(
                Project.workspace_id == user.workspace_id,
                Task.is_archived == False,
                (Task.title.ilike(search_term) | Task.description.ilike(search_term))
            )
            .limit(10)
        )
        for row in task_results.fetchall():
            task = row[0]
            results['tasks'].append({
                'id': task.id,
                'title': task.title,
                'project_name': row[1],
                'status': task.status.value if hasattr(task.status, 'value') else task.status,
                'priority': task.priority.value if hasattr(task.priority, 'value') else task.priority
            })
        
        # Search tickets (includes guest info and comments)
        from sqlalchemy import exists, case, literal
        from app.models.ticket import TicketComment
        
        # Subquery to find tickets with matching comments
        comment_match_subq = exists().where(
            TicketComment.ticket_id == Ticket.id,
            TicketComment.content.ilike(search_term)
        )
        
        # Relevance ordering: name/email matches first, then subject, then description/comments
        relevance = case(
            (Ticket.guest_name.ilike(search_term), literal(1)),
            (Ticket.guest_surname.ilike(search_term), literal(1)),
            (Ticket.guest_email.ilike(search_term), literal(1)),
            (Ticket.guest_company.ilike(search_term), literal(1)),
            (Ticket.ticket_number.ilike(search_term), literal(1)),
            (Ticket.subject.ilike(search_term), literal(2)),
            else_=literal(3)
        )
        
        ticket_results = await db.execute(
            select(Ticket)
            .where(
                Ticket.workspace_id == user.workspace_id,
                or_(
                    Ticket.subject.ilike(search_term),
                    Ticket.description.ilike(search_term),
                    Ticket.ticket_number.ilike(search_term),
                    Ticket.guest_email.ilike(search_term),
                    Ticket.guest_name.ilike(search_term),
                    Ticket.guest_surname.ilike(search_term),
                    Ticket.guest_company.ilike(search_term),
                    comment_match_subq
                )
            )
            .order_by(relevance, Ticket.created_at.desc())
            .limit(10)
        )
        for ticket in ticket_results.scalars().all():
            results['tickets'].append({
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'title': ticket.subject,
                'status': ticket.status,
                'priority': ticket.priority,
                'guest_name': f"{ticket.guest_name or ''} {ticket.guest_surname or ''}".strip(),
                'guest_email': ticket.guest_email or '',
                'guest_company': ticket.guest_company or ''
            })
        
        # Search projects
        project_results = await db.execute(
            select(Project)
            .where(
                Project.workspace_id == user.workspace_id,
                Project.is_archived == False,
                (Project.name.ilike(search_term) | Project.description.ilike(search_term))
            )
            .limit(10)
        )
        for project in project_results.scalars().all():
            results['projects'].append({
                'id': project.id,
                'name': project.name,
                'description': project.description[:100] if project.description else None
            })
        
        # Search comments
        comment_results = await db.execute(
            select(Comment, Task.id.label('task_id'), Task.title.label('task_title'), 
                   Project.name.label('project_name'), User.full_name.label('author_name'))
            .join(Task, Comment.task_id == Task.id)
            .join(Project, Task.project_id == Project.id)
            .join(User, Comment.author_id == User.id)
            .where(
                Project.workspace_id == user.workspace_id,
                Comment.content.ilike(search_term)
            )
            .order_by(Comment.created_at.desc())
            .limit(10)
        )
        for row in comment_results.fetchall():
            comment = row[0]
            results['comments'].append({
                'id': comment.id,
                'content': comment.content[:200] + '...' if len(comment.content) > 200 else comment.content,
                'task_id': row[1],
                'task_title': row[2],
                'project_name': row[3],
                'author_name': row[4],
                'created_at': comment.created_at
            })
    
    # Check if AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JSONResponse(results)
    
    return templates.TemplateResponse('search/results.html', {
        'request': request,
        'user': user,
        'results': results,
        'query': q
    })


@router.get('/api/search')
async def api_global_search(
    request: Request,
    q: str = '',
    db: AsyncSession = Depends(get_session)
):
    """API endpoint for global search (used by search modal)"""
    from app.models.ticket import Ticket
    
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'tasks': [], 'tickets': [], 'projects': [], 'comments': []})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'tasks': [], 'tickets': [], 'projects': [], 'comments': []})
    
    results = {'tasks': [], 'tickets': [], 'projects': [], 'comments': []}
    
    if q and len(q) >= 2:
        search_term = f'%{q}%'
        
        # Search tasks
        task_results = await db.execute(
            select(Task, Project.name.label('project_name'))
            .join(Project, Task.project_id == Project.id)
            .where(
                Project.workspace_id == user.workspace_id,
                Task.is_archived == False,
                (Task.title.ilike(search_term) | Task.description.ilike(search_term))
            )
            .limit(5)
        )
        for row in task_results.fetchall():
            task = row[0]
            results['tasks'].append({
                'id': task.id,
                'title': task.title,
                'project_name': row[1],
                'url': f'/web/tasks/{task.id}'
            })
        
        # Search tickets (includes guest info and comments)
        from sqlalchemy import exists, or_, case, literal
        from app.models.ticket import TicketComment
        
        # Subquery to find tickets with matching comments
        comment_match_subq2 = exists().where(
            TicketComment.ticket_id == Ticket.id,
            TicketComment.content.ilike(search_term)
        )
        
        # Relevance ordering: name/email matches first, then subject, then description/comments
        relevance2 = case(
            (Ticket.guest_name.ilike(search_term), literal(1)),
            (Ticket.guest_surname.ilike(search_term), literal(1)),
            (Ticket.guest_email.ilike(search_term), literal(1)),
            (Ticket.guest_company.ilike(search_term), literal(1)),
            (Ticket.ticket_number.ilike(search_term), literal(1)),
            (Ticket.subject.ilike(search_term), literal(2)),
            else_=literal(3)
        )
        
        ticket_results = await db.execute(
            select(Ticket)
            .where(
                Ticket.workspace_id == user.workspace_id,
                or_(
                    Ticket.subject.ilike(search_term),
                    Ticket.description.ilike(search_term),
                    Ticket.ticket_number.ilike(search_term),
                    Ticket.guest_email.ilike(search_term),
                    Ticket.guest_name.ilike(search_term),
                    Ticket.guest_surname.ilike(search_term),
                    Ticket.guest_company.ilike(search_term),
                    comment_match_subq2
                )
            )
            .order_by(relevance2, Ticket.created_at.desc())
            .limit(5)
        )
        for ticket in ticket_results.scalars().all():
            results['tickets'].append({
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'title': ticket.subject,
                'url': f'/web/tickets/{ticket.id}',
                'guest_name': f"{ticket.guest_name or ''} {ticket.guest_surname or ''}".strip(),
                'guest_email': ticket.guest_email or ''
            })
        
        # Search projects
        project_results = await db.execute(
            select(Project)
            .where(
                Project.workspace_id == user.workspace_id,
                Project.is_archived == False,
                Project.name.ilike(search_term)
            )
            .limit(5)
        )
        for project in project_results.scalars().all():
            results['projects'].append({
                'id': project.id,
                'name': project.name,
                'url': f'/web/projects/{project.id}'
            })
        
        # Search comments
        comment_results = await db.execute(
            select(Comment, Task.id.label('task_id'), Task.title.label('task_title'))
            .join(Task, Comment.task_id == Task.id)
            .join(Project, Task.project_id == Project.id)
            .where(
                Project.workspace_id == user.workspace_id,
                Comment.content.ilike(search_term)
            )
            .order_by(Comment.created_at.desc())
            .limit(5)
        )
        for row in comment_results.fetchall():
            comment = row[0]
            results['comments'].append({
                'id': comment.id,
                'content': comment.content[:100] + '...' if len(comment.content) > 100 else comment.content,
                'task_id': row[1],
                'task_title': row[2],
                'url': f'/web/tasks/{row[1]}'
            })
    
    return JSONResponse(results)


# --------------------------
# Task Duplication
# --------------------------
@router.post('/tasks/{task_id}/duplicate')
async def web_task_duplicate(
    request: Request,
    task_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Duplicate a task with all its details"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get original task
    original = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not original:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Create duplicate
    new_task = Task(
        title=f"Copy of {original.title}",
        description=original.description,
        project_id=original.project_id,
        creator_id=user_id,
        status=TaskStatus.todo,
        priority=original.priority,
        due_date=original.due_date,
        due_time=original.due_time,
        start_date=original.start_date,
        start_time=original.start_time,
        estimated_hours=original.estimated_hours,
        working_days=original.working_days
    )
    db.add(new_task)
    await db.flush()
    
    # Copy assignments
    assignments = (await db.execute(
        select(Assignment).where(Assignment.task_id == task_id)
    )).scalars().all()
    
    for orig_assign in assignments:
        new_assign = Assignment(
            task_id=new_task.id,
            assignee_id=orig_assign.assignee_id
        )
        db.add(new_assign)
    
    # Copy subtasks
    from app.models.subtask import Subtask
    subtasks = (await db.execute(
        select(Subtask).where(Subtask.task_id == task_id)
    )).scalars().all()
    
    for orig_subtask in subtasks:
        new_subtask = Subtask(
            task_id=new_task.id,
            title=orig_subtask.title,
            is_completed=False,
            order=orig_subtask.order
        )
        db.add(new_subtask)
    
    # Add history
    history = TaskHistory(
        task_id=new_task.id,
        editor_id=user_id,
        field='created',
        new_value=f"Duplicated from task #{task_id}"
    )
    db.add(history)
    
    # Track behavior
    from app.core.smart_suggestions import track_user_action
    await track_user_action(
        db, user_id, user.workspace_id, 'task_duplicated', 'task',
        entity_id=new_task.id, project_id=new_task.project_id
    )
    
    await db.commit()
    
    return RedirectResponse(f'/web/tasks/{new_task.id}', status_code=303)


# --------------------------
# Time Tracking
# --------------------------
@router.post('/tasks/{task_id}/time-log')
async def web_task_add_time_log(
    request: Request,
    task_id: int,
    hours: float = Form(...),
    description: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Log time spent on a task"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify task exists
    task = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Create time log
    from app.models.task_extensions import TimeLog
    time_log = TimeLog(
        task_id=task_id,
        user_id=user_id,
        hours=hours,
        description=description
    )
    db.add(time_log)
    
    # Track behavior
    from app.core.smart_suggestions import track_user_action
    await track_user_action(
        db, user_id, user.workspace_id, 'time_logged', 'task',
        entity_id=task_id, project_id=task.project_id,
        field_name='hours', field_value=str(hours)
    )
    
    await db.commit()
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.get('/tasks/{task_id}/time-logs')
async def web_task_get_time_logs(
    request: Request,
    task_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Get time logs for a task"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'logs': [], 'total_hours': 0})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'logs': [], 'total_hours': 0})
    
    # Verify task exists
    task = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not task:
        return JSONResponse({'logs': [], 'total_hours': 0})
    
    # Get time logs
    from app.models.task_extensions import TimeLog
    logs_result = await db.execute(
        select(TimeLog, User.full_name, User.username)
        .join(User, TimeLog.user_id == User.id)
        .where(TimeLog.task_id == task_id)
        .order_by(TimeLog.logged_at.desc())
    )
    
    logs = []
    total_hours = 0
    for row in logs_result.fetchall():
        log = row[0]
        logs.append({
            'id': log.id,
            'hours': log.hours,
            'description': log.description,
            'user_name': row[1] or row[2],
            'logged_at': log.logged_at.isoformat() if log.logged_at else None
        })
        total_hours += log.hours
    
    return JSONResponse({'logs': logs, 'total_hours': total_hours})


# --------------------------
# Notifications (minimal for layout)
# --------------------------
@router.get('/notifications/count', response_class=HTMLResponse)
async def web_notifications_count(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return HTMLResponse('0')
    result = await db.execute(select(Notification).where(Notification.user_id == user_id, Notification.read_at.is_(None)))
    count = len(result.scalars().all())
    return HTMLResponse(str(count))


@router.get('/notifications/peek', response_class=HTMLResponse)
async def web_notifications_peek(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return HTMLResponse('')
    # Show nothing while OTP is disabled (simplify)
    return HTMLResponse('')


@router.get('/notifications', response_class=HTMLResponse)
async def web_notifications(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get all notifications for the user, ordered by newest first
    notifications = (await db.execute(
        select(Notification)
        .where(Notification.user_id == user_id)
        .order_by(Notification.created_at.desc())
        .limit(50)
    )).scalars().all()
    
    return templates.TemplateResponse('notifications/list.html', {
        'request': request,
        'user': user,
        'notifications': notifications
    })


@router.post('/notifications/{notification_id}/read')
async def web_notification_mark_read(
    request: Request,
    notification_id: int,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    notification = (await db.execute(
        select(Notification)
        .where(Notification.id == notification_id, Notification.user_id == user_id)
    )).scalar_one_or_none()
    
    if notification:
        from datetime import datetime
        notification.read_at = datetime.utcnow()
        await db.commit()
        
        # Smart navigation based on notification type
        if notification.url:
            return RedirectResponse(notification.url, status_code=303)
        elif notification.type == 'meeting' and notification.related_id:
            return RedirectResponse(f'/web/meetings?highlight={notification.related_id}', status_code=303)
        elif notification.type in ['task', 'assignment'] and notification.related_id:
            return RedirectResponse(f'/web/tasks/my?highlight={notification.related_id}', status_code=303)
        elif notification.type == 'message':
            # For messages, navigate to the chat
            return RedirectResponse(notification.url or '/web/chats', status_code=303)
    
    return RedirectResponse('/web/notifications', status_code=303)


@router.post('/notifications/{notification_id}/delete')
async def web_notification_delete(
    request: Request,
    notification_id: int,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    notification = (await db.execute(
        select(Notification)
        .where(Notification.id == notification_id, Notification.user_id == user_id)
    )).scalar_one_or_none()
    
    if notification:
        await db.delete(notification)
        await db.commit()
    
    return RedirectResponse('/web/notifications', status_code=303)


@router.post('/notifications/{notification_id}/dismiss')
async def web_notification_dismiss(
    request: Request,
    notification_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Mark notification popup as dismissed (auto-dismiss after 1 minute)"""
    user_id = request.session.get('user_id')
    if not user_id:
        return HTMLResponse('', status_code=401)
    
    notification = (await db.execute(
        select(Notification)
        .where(Notification.id == notification_id, Notification.user_id == user_id)
    )).scalar_one_or_none()
    
    if notification:
        from datetime import datetime
        notification.dismissed_at = datetime.utcnow()
        await db.commit()
    
    return HTMLResponse('OK')


@router.get('/notifications/unread', response_class=HTMLResponse)
async def web_notifications_unread(request: Request, db: AsyncSession = Depends(get_session)):
    """Get unread and undismissed notifications for popup display"""
    user_id = request.session.get('user_id')
    if not user_id:
        return HTMLResponse('[]')
    
    notifications = (await db.execute(
        select(Notification)
        .where(
            Notification.user_id == user_id,
            Notification.read_at.is_(None),
            Notification.dismissed_at.is_(None)
        )
        .order_by(Notification.created_at.desc())
        .limit(5)
    )).scalars().all()
    
    import json
    notification_data = [{
        'id': n.id,
        'type': n.type,
        'message': n.message,
        'url': n.url,
        'related_id': n.related_id,
        'created_at': n.created_at.isoformat() if n.created_at else None
    } for n in notifications]
    
    return HTMLResponse(json.dumps(notification_data), media_type='application/json')


@router.post('/notifications/mark-all-read')
async def web_notifications_mark_all_read(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    from datetime import datetime
    await db.execute(
        select(Notification)
        .where(Notification.user_id == user_id, Notification.read_at.is_(None))
    )
    
    # Update all unread notifications
    result = await db.execute(
        select(Notification)
        .where(Notification.user_id == user_id, Notification.read_at.is_(None))
    )
    notifications = result.scalars().all()
    
    for notification in notifications:
        notification.read_at = datetime.utcnow()
    
    await db.commit()
    return RedirectResponse('/web/notifications', status_code=303)


# --------------------------
# User Management (all users can view, only admins can create)
# --------------------------
@router.get('/admin/users', response_class=HTMLResponse)
async def web_admin_users_list(
    request: Request,
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # All users can view the user list (not just admins)
    # Get all users in the workspace
    users = (
        await db.execute(
            select(User)
            .where(User.workspace_id == user.workspace_id)
            .order_by(User.username)
        )
    ).scalars().all()
    
    return templates.TemplateResponse(
        'admin/users_list.html',
        {
            'request': request,
            'user': user,
            'users': users,
        },
    )


@router.get('/admin/users/create', response_class=HTMLResponse)
async def web_admin_create_user(
    request: Request,
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    return templates.TemplateResponse(
        'admin/create_user.html',
        {
            'request': request,
            'user': user,
            'error': None,
            'success': None,
        },
    )


@router.post('/admin/users/create')
async def web_admin_create_user_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(...),
    email: str = Form(...),
    is_admin: bool = Form(False),
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Import validation function
    from app.core.security import validate_password
    
    # Validate password
    is_valid, error_msg = validate_password(password)
    if not is_valid:
        return templates.TemplateResponse(
            'admin/create_user.html',
            {
                'request': request,
                'user': user,
                'error': error_msg,
                'success': None,
            },
            status_code=400
        )
    
    # Check if username already exists
    exists = await db.execute(select(User).where(User.username == username))
    if exists.scalar_one_or_none():
        return templates.TemplateResponse(
            'admin/create_user.html',
            {
                'request': request,
                'user': user,
                'error': 'Username already taken',
                'success': None,
            },
            status_code=400
        )
    
    # Check if email is already used
    if email:
        exists = await db.execute(select(User).where(User.email == email))
        if exists.scalar_one_or_none():
            return templates.TemplateResponse(
                'admin/create_user.html',
                {
                    'request': request,
                    'user': user,
                    'error': 'Email already in use',
                    'success': None,
                },
                status_code=400
            )
    
    # Create new user
    new_user = User(
        username=username,
        hashed_password=get_password_hash(password),
        full_name=full_name,
        email=email,
        workspace_id=user.workspace_id,
        is_admin=is_admin,
        profile_completed=True,  # Admin sets all details
        email_verified=True,
        is_active=True,
    )
    db.add(new_user)
    await db.commit()
    await db.refresh(new_user)
    
    return templates.TemplateResponse(
        'admin/create_user.html',
        {
            'request': request,
            'user': user,
            'error': None,
            'success': f'User "{username}" created successfully!',
        },
    )


@router.post('/admin/users/{user_id}/deactivate')
async def web_admin_deactivate_user(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session),
):
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Can't deactivate yourself
    if user_id == current_user_id:
        raise HTTPException(status_code=400, detail="Cannot deactivate yourself")
    
    # Get the user to deactivate
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Deactivate the user
    target_user.is_active = False
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


# User Activity Reports
# --------------------------
@router.get('/admin/reports/user-activity', response_class=HTMLResponse)
async def web_admin_user_activity_report(
    request: Request,
    db: AsyncSession = Depends(get_session),
):
    """Admin page to generate user activity reports"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get all users in workspace
    users = (
        await db.execute(
            select(User)
            .where(User.workspace_id == user.workspace_id)
            .order_by(User.full_name, User.username)
        )
    ).scalars().all()
    
    return templates.TemplateResponse(
        'admin/user_activity_report.html',
        {
            'request': request,
            'user': user,
            'users': users,
        },
    )


@router.get('/admin/reports/user-activity/{target_user_id}/pdf')
async def web_admin_generate_user_activity_pdf(
    request: Request,
    target_user_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session),
):
    """Generate PDF report of user activity"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get target user
    target_user = (await db.execute(select(User).where(User.id == target_user_id))).scalar_one_or_none()
    if not target_user or target_user.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    # Gather activity data
    from sqlalchemy import text
    
    # 1. Tasks created - filter by workspace through project
    tasks_created = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Task.creator_id == target_user_id)
        .where(Task.created_at >= start_dt)
        .where(Task.created_at < end_dt)
        .order_by(Task.created_at.desc())
    )).scalars().all()
    
    # 2. Task assignments - filter by workspace through project
    task_assignments = (await db.execute(
        select(Task, Assignment)
        .join(Assignment, Task.id == Assignment.task_id)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Assignment.assignee_id == target_user_id)
        .where(Task.created_at >= start_dt)
        .where(Task.created_at < end_dt)
        .order_by(Task.created_at.desc())
    )).all()
    
    # 3. Task edits - filter by workspace through task->project
    try:
        task_edits_result = await db.execute(
            text("""
                SELECT th.id, th.task_id, th.editor_id, th.field, th.old_value, th.new_value, th.created_at
                FROM taskhistory th
                JOIN task t ON th.task_id = t.id
                JOIN project p ON t.project_id = p.id
                WHERE th.editor_id = :user_id
                AND p.workspace_id = :workspace_id
                AND th.created_at >= :start_dt
                AND th.created_at < :end_dt
                ORDER BY th.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        task_edits_raw = task_edits_result.fetchall()
        # Convert to objects with attributes for compatibility
        class TaskEditRow:
            def __init__(self, row):
                self.id, self.task_id, self.editor_id, self.field, self.old_value, self.new_value, created_at_val = row
                # Parse created_at if it's a string
                if isinstance(created_at_val, str):
                    try:
                        self.created_at = datetime.fromisoformat(created_at_val.replace('Z', '+00:00'))
                    except Exception:
                        self.created_at = datetime.utcnow()
                else:
                    self.created_at = created_at_val
        task_edits = [TaskEditRow(row) for row in task_edits_raw]
    except Exception as e:
        logger.error(f"Error fetching task edits: {e}")
        task_edits = []
    
    # 4. Comments - filter by workspace through task->project
    try:
        comments_result = await db.execute(
            text("""
                SELECT c.id, c.task_id, c.author_id, c.content, c.created_at 
                FROM comment c
                JOIN task t ON c.task_id = t.id
                JOIN project p ON t.project_id = p.id
                WHERE c.author_id = :user_id 
                AND p.workspace_id = :workspace_id
                AND c.created_at >= :start_dt 
                AND c.created_at < :end_dt 
                ORDER BY c.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        comments = comments_result.fetchall()
    except Exception as e:
        logger.error(f"Error fetching comments: {e}")
        comments = []
    
    # 5. Projects created - filter by workspace
    projects_created = (await db.execute(
        select(Project)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Project.owner_id == target_user_id)
        .where(Project.created_at >= start_dt)
        .where(Project.created_at < end_dt)
        .order_by(Project.created_at.desc())
    )).scalars().all()
    
    # 6. Activities logged - filter by workspace
    activities = (await db.execute(
        select(Activity)
        .where(Activity.workspace_id == target_user.workspace_id)
        .where(Activity.created_by == target_user_id)
        .where(Activity.created_at >= start_dt)
        .where(Activity.created_at < end_dt)
        .order_by(Activity.created_at.desc())
    )).scalars().all()
    
    # 7. Tickets closed - Include archived tickets as well
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    tickets_closed = (await db.execute(
        select(Ticket)
        .where(Ticket.workspace_id == target_user.workspace_id)
        .where(Ticket.closed_by_id == target_user_id)
        .where(Ticket.closed_at >= start_dt)
        .where(Ticket.closed_at < end_dt)
        .order_by(Ticket.closed_at.desc())
    )).scalars().all()
    
    # 8. Ticket comments - filter by workspace through ticket
    try:
        ticket_comments_result = await db.execute(
            text("""
                SELECT tc.id, tc.ticket_id, tc.user_id, tc.content, tc.is_internal, tc.created_at 
                FROM ticketcomment tc
                JOIN ticket t ON tc.ticket_id = t.id
                WHERE tc.user_id = :user_id
                AND t.workspace_id = :workspace_id
                AND tc.created_at >= :start_dt 
                AND tc.created_at < :end_dt 
                ORDER BY tc.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        ticket_comments = ticket_comments_result.fetchall()
    except Exception as e:
        logger.error(f"Error fetching ticket comments: {e}")
        ticket_comments = []
    
    # 9. Tickets assigned to user AND where user made changes
    # Double verification: ticket is assigned to user + user has activity in tickethistory
    try:
        # Get tickets that are assigned to user AND have history entries by this user
        tickets_assigned_result = await db.execute(
            text("""
                SELECT DISTINCT t.id
                FROM ticket t
                INNER JOIN tickethistory th ON th.ticket_id = t.id AND th.user_id = :user_id
                WHERE t.assigned_to_id = :user_id
                AND t.workspace_id = :workspace_id
                AND t.created_at >= :start_dt
                AND t.created_at < :end_dt
            """),
            {"user_id": target_user_id, "start_dt": start_dt, "end_dt": end_dt, "workspace_id": target_user.workspace_id}
        )
        assigned_ticket_ids = [row[0] for row in tickets_assigned_result.fetchall()]
        
        if assigned_ticket_ids:
            tickets_assigned = (await db.execute(
                select(Ticket)
                .where(Ticket.id.in_(assigned_ticket_ids))
                .order_by(Ticket.created_at.desc())
            )).scalars().all()
        else:
            tickets_assigned = []
    except Exception as e:
        logger.error(f"Error fetching tickets assigned with activity: {e}")
        # Fallback to just assigned query
        tickets_assigned = (await db.execute(
            select(Ticket)
            .where(Ticket.workspace_id == target_user.workspace_id)
            .where(Ticket.assigned_to_id == target_user_id)
            .where(Ticket.created_at >= start_dt)
            .where(Ticket.created_at < end_dt)
            .order_by(Ticket.created_at.desc())
        )).scalars().all()
    
    # Generate PDF
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    # Create PDF in memory
    import io
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20,
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubheading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=8,
    )
    
    # Title
    elements.append(Paragraph(f"User Activity Report", title_style))
    elements.append(Paragraph(f"{target_user.full_name or target_user.username}", heading_style))
    elements.append(Paragraph(
        f"Period: {start_dt.strftime('%B %d, %Y')} - {end_dt.strftime('%B %d, %Y')}",
        subheading_style
    ))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        subheading_style
    ))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary section with enhanced metrics
    elements.append(Paragraph("Activity Summary", heading_style))
    summary_data = [
        ['Activity Type', 'Count'],
        ['Tasks Created', str(len(tasks_created))],
        ['Task Assignments Received', str(len(task_assignments))],
        ['Task Edits Made', str(len(task_edits))],
        ['Task Comments Posted', str(len(comments))],
        ['Projects Created', str(len(projects_created))],
        ['Activities Logged (Calls/Emails/Meetings)', str(len(activities))],
        ['Tickets Assigned', str(len(tickets_assigned))],
        ['Ticket Comments Posted', str(len(ticket_comments))],
        ['Tickets Closed', str(len(tickets_closed))],
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # OVERDUE TASKS SECTION (Critical!)
    from datetime import date as date_type
    now = datetime.now()
    now_date = now.date()
    
    # Helper function to handle both date and datetime
    def is_overdue(due_date):
        if not due_date:
            return False
        # Convert to date if it's a datetime
        if isinstance(due_date, datetime):
            due_date = due_date.date()
        return due_date < now_date
    
    overdue_tasks = []
    for task in tasks_created:
        if task.due_date and is_overdue(task.due_date) and task.status.value not in ['done', 'completed', 'archived']:
            overdue_tasks.append(task)
    for task, assignment in task_assignments:
        if task.due_date and is_overdue(task.due_date) and task.status.value not in ['done', 'completed', 'archived']:
            if task not in overdue_tasks:
                overdue_tasks.append(task)
    
    if overdue_tasks:
        elements.append(Paragraph("⚠️ OVERDUE TASKS", heading_style))
        overdue_data = [['Task Title', 'Due Date', 'Days Overdue', 'Priority', 'Status']]
        for task in sorted(overdue_tasks, key=lambda t: t.due_date if isinstance(t.due_date, date_type) else t.due_date.date()):
            task_due_date = task.due_date if isinstance(task.due_date, date_type) else task.due_date.date()
            days_overdue = (now_date - task_due_date).days
            overdue_data.append([
                Paragraph(task.title, styles['Normal']),
                task_due_date.strftime('%Y-%m-%d'),
                str(days_overdue),
                task.priority.value.title(),
                task.status.value.replace('_', ' ').title(),
            ])
        
        overdue_table = Table(overdue_data, colWidths=[2.3*inch, 1*inch, 1.2*inch, 0.9*inch, 1*inch])
        overdue_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightpink, colors.mistyrose]),
        ]))
        elements.append(overdue_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Tasks Created with enhanced details
    elements.append(Paragraph("Tasks Created", heading_style))
    if tasks_created:
        task_data = [['Date Created', 'Title', 'Due Date', 'Priority', 'Status']]
        for task in tasks_created:
            due_str = task.due_date.strftime('%Y-%m-%d') if task.due_date else 'No due date'
            if task.due_date and is_overdue(task.due_date) and task.status.value not in ['done', 'completed', 'archived']:
                due_str += ' (OVERDUE)'
            task_data.append([
                task.created_at.strftime('%Y-%m-%d'),
                task.title[:35],
                due_str,
                task.priority.value.title(),
                task.status.value.replace('_', ' ').title(),
            ])
        
        task_table = Table(task_data, colWidths=[1.1*inch, 2*inch, 1.3*inch, 0.9*inch, 1.1*inch])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(task_table)
    else:
        elements.append(Paragraph("No tasks created during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Task Assignments Received
    elements.append(Paragraph("Task Assignments Received", heading_style))
    if task_assignments:
        assignment_data = [['Date', 'Title', 'Assigned By', 'Due Date', 'Status']]
        for task, assignment in task_assignments:
            # Task creator is the one who assigned it
            assigner = (await db.execute(select(User).where(User.id == task.creator_id))).scalar_one_or_none()
            assigner_name = assigner.full_name or assigner.username if assigner else 'Unknown'
            due_str = task.due_date.strftime('%Y-%m-%d') if task.due_date else 'None'
            if task.due_date and is_overdue(task.due_date) and task.status.value not in ['done', 'completed', 'archived']:
                due_str += ' (LATE)'
            assignment_data.append([
                task.created_at.strftime('%Y-%m-%d'),
                Paragraph(task.title, styles['Normal']),
                Paragraph(assigner_name, styles['Normal']),
                due_str,
                task.status.value.replace('_', ' ').title(),
            ])
        
        assignment_table = Table(assignment_data, colWidths=[1*inch, 1.8*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        assignment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(assignment_table)
    else:
        elements.append(Paragraph("No task assignments received during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Task Edits
    elements.append(Paragraph("Recent Task Edits", heading_style))
    if task_edits:
        edit_data = [['Date', 'Task ID', 'Field Changed', 'Old Value', 'New Value']]
        for edit in task_edits:
            edit_data.append([
                edit.created_at.strftime('%Y-%m-%d %H:%M'),
                str(edit.task_id),
                edit.field.replace('_', ' ').title(),
                Paragraph(edit.old_value or 'None', styles['Normal']),
                Paragraph(edit.new_value or 'None', styles['Normal']),
            ])
        
        edit_table = Table(edit_data, colWidths=[1.3*inch, 0.7*inch, 1.2*inch, 1.5*inch, 1.5*inch])
        edit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(edit_table)
    else:
        elements.append(Paragraph("No task edits made during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Comments
    elements.append(Paragraph("Recent Comments", heading_style))
    if comments:
        comment_data = [['Date', 'Task ID', 'Comment']]
        for comment in comments:
            # comment is a tuple: (id, task_id, author_id, content, created_at)
            comment_id, task_id, author_id, content, created_at = comment
            comment_data.append([
                created_at.strftime('%Y-%m-%d %H:%M') if isinstance(created_at, datetime) else str(created_at)[:16],
                str(task_id),
                Paragraph(content or '', styles['Normal']),
            ])
        
        comment_table = Table(comment_data, colWidths=[1.5*inch, 0.8*inch, 4*inch])
        comment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(comment_table)
    else:
        elements.append(Paragraph("No comments posted during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Projects Created
    elements.append(Paragraph("Projects Created", heading_style))
    if projects_created:
        from app.models.project_member import ProjectMember
        project_data = [['Date', 'Project Name', 'Status', 'Members']]
        for project in projects_created:
            member_count = (await db.execute(
                select(ProjectMember).where(ProjectMember.project_id == project.id)
            )).scalars().all()
            project_data.append([
                project.created_at.strftime('%Y-%m-%d'),
                Paragraph(project.name, styles['Normal']),
                project.status.value.title(),
                str(len(member_count)),
            ])
        
        project_table = Table(project_data, colWidths=[1.1*inch, 3*inch, 1.1*inch, 1.1*inch])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lavender]),
        ]))
        elements.append(project_table)
    else:
        elements.append(Paragraph("No projects created during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Activities Logged (Calls, Emails, Meetings, Notes)
    elements.append(Paragraph("Activities Logged (Calls, Emails, Meetings, Notes)", heading_style))
    if activities:
        activity_data = [['Date', 'Type', 'Subject', 'Related To']]
        for activity in activities:
            related = ''
            if activity.project_id:
                proj = (await db.execute(select(Project).where(Project.id == activity.project_id))).scalar_one_or_none()
                related = f"Project: {proj.name[:20]}" if proj else 'Project'
            elif activity.contact_id:
                related = f'Contact ID: {activity.contact_id}'
            
            activity_data.append([
                activity.created_at.strftime('%Y-%m-%d'),
                activity.activity_type.replace('_', ' ').title(),
                Paragraph(activity.subject or '', styles['Normal']),
                Paragraph(related, styles['Normal']),
            ])
        
        activity_table = Table(activity_data, colWidths=[1*inch, 1.1*inch, 2.5*inch, 1.7*inch])
        activity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgoldenrodyellow]),
        ]))
        elements.append(activity_table)
    else:
        elements.append(Paragraph("No activities logged during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Ticket Comments Posted
    elements.append(Paragraph("Ticket Comments Posted", heading_style))
    if ticket_comments:
        tcomment_data = [['Date', 'Ticket ID', 'Comment Preview', 'Internal']]
        for tc in ticket_comments:
            # tc is a tuple: (id, ticket_id, user_id, content, is_internal, created_at)
            tc_id, ticket_id, user_id, content, is_internal, created_at = tc
            tcomment_data.append([
                created_at.strftime('%Y-%m-%d') if isinstance(created_at, datetime) else str(created_at)[:10],
                str(ticket_id),
                Paragraph(content or '', styles['Normal']),
                'Yes' if is_internal else 'No',
            ])
        
        tcomment_table = Table(tcomment_data, colWidths=[1.1*inch, 1*inch, 3*inch, 1.2*inch])
        tcomment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06B6D4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightblue]),
        ]))
        elements.append(tcomment_table)
    else:
        elements.append(Paragraph("No ticket comments posted during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tickets Assigned (verified via TicketHistory that user actually worked on them)
    elements.append(Paragraph("Tickets Assigned", heading_style))
    if tickets_assigned:
        tassigned_data = [['Date Assigned', 'Ticket #', 'Subject', 'Priority', 'Status']]
        for ticket in tickets_assigned:
            tassigned_data.append([
                ticket.created_at.strftime('%Y-%m-%d'),
                ticket.ticket_number,
                Paragraph(ticket.subject, styles['Normal']),
                ticket.priority.title(),
                ticket.status.title(),
            ])
        
        tassigned_table = Table(tassigned_data, colWidths=[1.2*inch, 1.1*inch, 2.2*inch, 0.9*inch, 0.9*inch])
        tassigned_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14B8A6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(tassigned_table)
    else:
        elements.append(Paragraph("No tickets assigned to this user during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tickets Closed
    elements.append(Paragraph("Tickets Closed", heading_style))
    if tickets_closed:
        ticket_data = [['Date Closed', 'Ticket #', 'Subject', 'Priority']]
        for ticket in tickets_closed:
            ticket_data.append([
                ticket.closed_at.strftime('%Y-%m-%d %H:%M'),
                ticket.ticket_number,
                Paragraph(ticket.subject, styles['Normal']),
                ticket.priority.title(),
            ])
        
        ticket_table = Table(ticket_data, colWidths=[1.5*inch, 1.3*inch, 2.5*inch, 1*inch])
        ticket_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(ticket_table)
    else:
        elements.append(Paragraph("No tickets closed during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return PDF file
    from fastapi.responses import StreamingResponse
    filename = f"user_activity_{target_user.username}_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.get('/admin/reports/user-activity/{target_user_id}/excel')
async def web_admin_generate_user_activity_excel(
    request: Request,
    target_user_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session),
):
    """Generate Excel report of user activity"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get target user
    target_user = (await db.execute(select(User).where(User.id == target_user_id))).scalar_one_or_none()
    if not target_user or target_user.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    from sqlalchemy import text
    
    # Gather all the same data as PDF
    # 1. Tasks created
    tasks_created = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Task.creator_id == target_user_id)
        .where(Task.created_at >= start_dt)
        .where(Task.created_at < end_dt)
        .order_by(Task.created_at.desc())
    )).scalars().all()
    
    # 2. Task assignments
    task_assignments = (await db.execute(
        select(Task, Assignment)
        .join(Assignment, Task.id == Assignment.task_id)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Assignment.assignee_id == target_user_id)
        .where(Task.created_at >= start_dt)
        .where(Task.created_at < end_dt)
        .order_by(Task.created_at.desc())
    )).all()
    
    # 3. Task edits
    try:
        task_edits_result = await db.execute(
            text("""
                SELECT th.id, th.task_id, th.editor_id, th.field, th.old_value, th.new_value, th.created_at
                FROM taskhistory th
                JOIN task t ON th.task_id = t.id
                JOIN project p ON t.project_id = p.id
                WHERE th.editor_id = :user_id
                AND p.workspace_id = :workspace_id
                AND th.created_at >= :start_dt
                AND th.created_at < :end_dt
                ORDER BY th.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        task_edits = task_edits_result.fetchall()
    except Exception:
        task_edits = []
    
    # 4. Comments
    try:
        comments_result = await db.execute(
            text("""
                SELECT c.id, c.task_id, c.author_id, c.content, c.created_at 
                FROM comment c
                JOIN task t ON c.task_id = t.id
                JOIN project p ON t.project_id = p.id
                WHERE c.author_id = :user_id 
                AND p.workspace_id = :workspace_id
                AND c.created_at >= :start_dt 
                AND c.created_at < :end_dt 
                ORDER BY c.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        comments = comments_result.fetchall()
    except Exception:
        comments = []
    
    # 5. Projects created
    projects_created = (await db.execute(
        select(Project)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Project.owner_id == target_user_id)
        .where(Project.created_at >= start_dt)
        .where(Project.created_at < end_dt)
        .order_by(Project.created_at.desc())
    )).scalars().all()
    
    # 6. Activities
    activities = (await db.execute(
        select(Activity)
        .where(Activity.workspace_id == target_user.workspace_id)
        .where(Activity.created_by == target_user_id)
        .where(Activity.created_at >= start_dt)
        .where(Activity.created_at < end_dt)
        .order_by(Activity.created_at.desc())
    )).scalars().all()
    
    # 7. Tickets
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    tickets_closed = (await db.execute(
        select(Ticket)
        .where(Ticket.workspace_id == target_user.workspace_id)
        .where(Ticket.closed_by_id == target_user_id)
        .where(Ticket.closed_at >= start_dt)
        .where(Ticket.closed_at < end_dt)
        .order_by(Ticket.closed_at.desc())
    )).scalars().all()
    
    # 8. Ticket comments
    try:
        ticket_comments_result = await db.execute(
            text("""
                SELECT tc.id, tc.ticket_id, tc.user_id, tc.content, tc.is_internal, tc.created_at 
                FROM ticketcomment tc
                JOIN ticket t ON tc.ticket_id = t.id
                WHERE tc.user_id = :user_id
                AND t.workspace_id = :workspace_id
                AND tc.created_at >= :start_dt 
                AND tc.created_at < :end_dt 
                ORDER BY tc.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        ticket_comments = ticket_comments_result.fetchall()
    except Exception:
        ticket_comments = []
    
    # 9. Tickets assigned
    try:
        tickets_assigned_result = await db.execute(
            text("""
                SELECT DISTINCT t.id
                FROM ticket t
                INNER JOIN tickethistory th ON th.ticket_id = t.id AND th.user_id = :user_id
                WHERE t.assigned_to_id = :user_id
                AND t.workspace_id = :workspace_id
                AND t.created_at >= :start_dt
                AND t.created_at < :end_dt
            """),
            {"user_id": target_user_id, "start_dt": start_dt, "end_dt": end_dt, "workspace_id": target_user.workspace_id}
        )
        assigned_ticket_ids = [row[0] for row in tickets_assigned_result.fetchall()]
        
        if assigned_ticket_ids:
            tickets_assigned = (await db.execute(
                select(Ticket)
                .where(Ticket.id.in_(assigned_ticket_ids))
                .order_by(Ticket.created_at.desc())
            )).scalars().all()
        else:
            tickets_assigned = []
    except Exception:
        tickets_assigned = []
    
    # Generate Excel file
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    import io
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_width(ws):
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["User Activity Report"])
    ws_summary.append([f"User: {target_user.full_name or target_user.username}"])
    ws_summary.append([f"Period: {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])
    ws_summary.append(["Activity Type", "Count"])
    style_header_row(ws_summary, 6, 2)
    ws_summary.append(["Tasks Created", len(tasks_created)])
    ws_summary.append(["Task Assignments Received", len(task_assignments)])
    ws_summary.append(["Task Edits Made", len(task_edits)])
    ws_summary.append(["Task Comments Posted", len(comments)])
    ws_summary.append(["Projects Created", len(projects_created)])
    ws_summary.append(["Activities Logged", len(activities)])
    ws_summary.append(["Tickets Assigned", len(tickets_assigned)])
    ws_summary.append(["Ticket Comments Posted", len(ticket_comments)])
    ws_summary.append(["Tickets Closed", len(tickets_closed)])
    auto_width(ws_summary)
    
    # Sheet 2: Tasks Created
    ws_tasks = wb.create_sheet("Tasks Created")
    ws_tasks.append(["Date Created", "Title", "Due Date", "Priority", "Status"])
    style_header_row(ws_tasks, 1, 5)
    for task in tasks_created:
        ws_tasks.append([
            task.created_at.strftime('%Y-%m-%d'),
            task.title,
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '',
            task.priority.value,
            task.status.value,
        ])
    auto_width(ws_tasks)
    
    # Sheet 3: Task Assignments
    ws_assignments = wb.create_sheet("Task Assignments")
    ws_assignments.append(["Date", "Title", "Due Date", "Status"])
    style_header_row(ws_assignments, 1, 4)
    for task, assignment in task_assignments:
        ws_assignments.append([
            task.created_at.strftime('%Y-%m-%d'),
            task.title,
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '',
            task.status.value,
        ])
    auto_width(ws_assignments)
    
    # Sheet 4: Task Edits
    ws_edits = wb.create_sheet("Task Edits")
    ws_edits.append(["Date", "Task ID", "Field Changed", "Old Value", "New Value"])
    style_header_row(ws_edits, 1, 5)
    for edit in task_edits:
        created_at = edit[6] if len(edit) > 6 else ''
        if isinstance(created_at, str):
            date_str = created_at[:19] if created_at else ''
        else:
            date_str = created_at.strftime('%Y-%m-%d %H:%M') if created_at else ''
        ws_edits.append([
            date_str,
            edit[1],  # task_id
            edit[3],  # field
            edit[4] or '',  # old_value
            edit[5] or '',  # new_value
        ])
    auto_width(ws_edits)
    
    # Sheet 5: Task Comments
    ws_comments = wb.create_sheet("Task Comments")
    ws_comments.append(["Date", "Task ID", "Comment"])
    style_header_row(ws_comments, 1, 3)
    for comment in comments:
        created_at = comment[4] if len(comment) > 4 else ''
        if isinstance(created_at, str):
            date_str = created_at[:19] if created_at else ''
        else:
            date_str = created_at.strftime('%Y-%m-%d %H:%M') if created_at else ''
        ws_comments.append([
            date_str,
            comment[1],  # task_id
            comment[3] or '',  # content
        ])
    auto_width(ws_comments)
    
    # Sheet 6: Projects
    ws_projects = wb.create_sheet("Projects Created")
    ws_projects.append(["Date", "Project Name", "Status"])
    style_header_row(ws_projects, 1, 3)
    for project in projects_created:
        ws_projects.append([
            project.created_at.strftime('%Y-%m-%d'),
            project.name,
            project.status.value,
        ])
    auto_width(ws_projects)
    
    # Sheet 7: Activities
    ws_activities = wb.create_sheet("Activities")
    ws_activities.append(["Date", "Type", "Subject"])
    style_header_row(ws_activities, 1, 3)
    for activity in activities:
        ws_activities.append([
            activity.created_at.strftime('%Y-%m-%d'),
            activity.activity_type,
            activity.subject or '',
        ])
    auto_width(ws_activities)
    
    # Sheet 8: Tickets Assigned
    ws_tickets_assigned = wb.create_sheet("Tickets Assigned")
    ws_tickets_assigned.append(["Date", "Ticket #", "Subject", "Priority", "Status"])
    style_header_row(ws_tickets_assigned, 1, 5)
    for ticket in tickets_assigned:
        ws_tickets_assigned.append([
            ticket.created_at.strftime('%Y-%m-%d'),
            ticket.ticket_number,
            ticket.subject,
            ticket.priority,
            ticket.status,
        ])
    auto_width(ws_tickets_assigned)
    
    # Sheet 9: Ticket Comments
    ws_ticket_comments = wb.create_sheet("Ticket Comments")
    ws_ticket_comments.append(["Date", "Ticket ID", "Comment", "Internal"])
    style_header_row(ws_ticket_comments, 1, 4)
    for tc in ticket_comments:
        created_at = tc[5] if len(tc) > 5 else ''
        if isinstance(created_at, str):
            date_str = created_at[:10] if created_at else ''
        else:
            date_str = created_at.strftime('%Y-%m-%d') if created_at else ''
        ws_ticket_comments.append([
            date_str,
            tc[1],  # ticket_id
            tc[3] or '',  # content
            'Yes' if tc[4] else 'No',  # is_internal
        ])
    auto_width(ws_ticket_comments)
    
    # Sheet 10: Tickets Closed
    ws_tickets_closed = wb.create_sheet("Tickets Closed")
    ws_tickets_closed.append(["Date Closed", "Ticket #", "Subject", "Priority"])
    style_header_row(ws_tickets_closed, 1, 4)
    for ticket in tickets_closed:
        ws_tickets_closed.append([
            ticket.closed_at.strftime('%Y-%m-%d %H:%M') if ticket.closed_at else '',
            ticket.ticket_number,
            ticket.subject,
            ticket.priority,
        ])
    auto_width(ws_tickets_closed)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return Excel file
    from fastapi.responses import StreamingResponse
    filename = f"user_activity_{target_user.username}_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.get('/admin/reports/user-activity/{target_user_id}/view', response_class=HTMLResponse)
async def web_admin_user_activity_view(
    request: Request,
    target_user_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session),
):
    """View HTML report of user activity with comprehensive metrics"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get target user
    target_user = (await db.execute(select(User).where(User.id == target_user_id))).scalar_one_or_none()
    if not target_user or target_user.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get workspace for timezone
    workspace = (await db.execute(
        select(Workspace).where(Workspace.id == user.workspace_id)
    )).scalar_one_or_none()
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    from sqlalchemy import text
    from app.models.ticket import Ticket, TicketComment
    from datetime import date as date_type
    
    now = datetime.now()
    now_date = now.date()
    
    def is_overdue(due_date):
        if not due_date:
            return False
        if isinstance(due_date, datetime):
            due_date = due_date.date()
        return due_date < now_date
    
    # Gather activity data
    # 1. Tasks created - filter by workspace through project
    tasks_created = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Task.creator_id == target_user_id)
        .where(Task.created_at >= start_dt)
        .where(Task.created_at < end_dt)
        .order_by(Task.created_at.desc())
    )).scalars().all()
    
    # Get project names for tasks
    tasks_with_project = []
    for task in tasks_created:
        project = None
        if task.project_id:
            project = (await db.execute(select(Project).where(Project.id == task.project_id))).scalar_one_or_none()
        tasks_with_project.append({
            'id': task.id,
            'title': task.title,
            'project_name': project.name if project else None,
            'created_at': task.created_at,
            'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
            'priority': task.priority.value,
            'status': task.status.value,
        })
    
    # 2. Task assignments with details - filter by workspace through project
    task_assignments_raw = (await db.execute(
        select(Task, Assignment)
        .join(Assignment, Task.id == Assignment.task_id)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Assignment.assignee_id == target_user_id)
        .where(Task.created_at >= start_dt)
        .where(Task.created_at < end_dt)
        .order_by(Task.created_at.desc())
    )).all()
    
    task_assignments = []
    for task, assignment in task_assignments_raw:
        assigner = (await db.execute(select(User).where(User.id == task.creator_id))).scalar_one_or_none()
        # Calculate time spent on task (timeentry table may not exist)
        time_spent = 0
        try:
            time_entries_result = await db.execute(
                text("""
                    SELECT SUM(duration_hours) as total FROM timeentry 
                    WHERE task_id = :task_id AND user_id = :user_id
                """),
                {"task_id": task.id, "user_id": target_user_id}
            )
            time_spent = time_entries_result.scalar() or 0
        except Exception:
            pass  # timeentry table doesn't exist
        
        task_assignments.append({
            'id': task.id,
            'title': task.title,
            'assigned_by': assigner.full_name or assigner.username if assigner else 'Unknown',
            'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
            'status': task.status.value,
            'time_spent_hours': round(time_spent, 1) if time_spent else None,
        })
    
    # 3. Task edits - filter by workspace through task->project
    try:
        task_edits_result = await db.execute(
            text("""
                SELECT th.id, th.task_id, th.editor_id, th.field, th.old_value, th.new_value, th.created_at
                FROM taskhistory th
                JOIN task t ON th.task_id = t.id
                JOIN project p ON t.project_id = p.id
                WHERE th.editor_id = :user_id
                AND p.workspace_id = :workspace_id
                AND th.created_at >= :start_dt
                AND th.created_at < :end_dt
                ORDER BY th.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        task_edits_raw = task_edits_result.fetchall()
        class TaskEditRow:
            def __init__(self, row):
                self.id, self.task_id, self.editor_id, self.field, self.old_value, self.new_value, created_at_val = row
                # Parse created_at if it's a string
                if isinstance(created_at_val, str):
                    try:
                        self.created_at = datetime.fromisoformat(created_at_val.replace('Z', '+00:00'))
                    except Exception:
                        self.created_at = datetime.utcnow()
                else:
                    self.created_at = created_at_val
        task_edits = [TaskEditRow(row) for row in task_edits_raw]
    except Exception as e:
        logger.error(f"Error fetching task edits: {e}")
        task_edits = []
    
    # 4. Comments - filter by workspace through task->project
    try:
        comments_result = await db.execute(
            text("""
                SELECT c.id, c.task_id, c.author_id, c.content, c.created_at 
                FROM comment c
                JOIN task t ON c.task_id = t.id
                JOIN project p ON t.project_id = p.id
                WHERE c.author_id = :user_id 
                AND p.workspace_id = :workspace_id
                AND c.created_at >= :start_dt 
                AND c.created_at < :end_dt 
                ORDER BY c.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        comments = comments_result.fetchall()
    except Exception as e:
        logger.error(f"Error fetching comments: {e}")
        comments = []
    
    # 5. Projects created - filter by workspace
    projects_created = (await db.execute(
        select(Project)
        .where(Project.workspace_id == target_user.workspace_id)
        .where(Project.owner_id == target_user_id)
        .where(Project.created_at >= start_dt)
        .where(Project.created_at < end_dt)
        .order_by(Project.created_at.desc())
    )).scalars().all()
    
    # 6. Activities logged - filter by workspace
    activities = (await db.execute(
        select(Activity)
        .where(Activity.workspace_id == target_user.workspace_id)
        .where(Activity.created_by == target_user_id)
        .where(Activity.created_at >= start_dt)
        .where(Activity.created_at < end_dt)
        .order_by(Activity.created_at.desc())
    )).scalars().all()
    
    # 7. Tickets closed by this user - Include archived tickets as well
    from app.models.ticket import TicketHistory
    tickets_closed = (await db.execute(
        select(Ticket)
        .where(Ticket.workspace_id == target_user.workspace_id)
        .where(Ticket.closed_by_id == target_user_id)
        .where(Ticket.closed_at >= start_dt)
        .where(Ticket.closed_at < end_dt)
        .order_by(Ticket.closed_at.desc())
    )).scalars().all()

    # Tickets assigned to user AND where user made changes (double verification)
    try:
        # Get tickets that are assigned to user AND have history entries by this user
        tickets_assigned_result = await db.execute(
            text("""
                SELECT DISTINCT t.id
                FROM ticket t
                INNER JOIN tickethistory th ON th.ticket_id = t.id AND th.user_id = :user_id
                WHERE t.assigned_to_id = :user_id
                AND t.workspace_id = :workspace_id
                AND t.created_at >= :start_dt
                AND t.created_at < :end_dt
            """),
            {"user_id": target_user_id, "start_dt": start_dt, "end_dt": end_dt, "workspace_id": target_user.workspace_id}
        )
        assigned_ticket_ids = [row[0] for row in tickets_assigned_result.fetchall()]
        
        if assigned_ticket_ids:
            tickets_assigned = (await db.execute(
                select(Ticket)
                .where(Ticket.id.in_(assigned_ticket_ids))
                .order_by(Ticket.created_at.desc())
            )).scalars().all()
        else:
            tickets_assigned = []
    except Exception as e:
        logger.error(f"Error fetching tickets assigned with activity: {e}")
        # Fallback to just assigned query
        tickets_assigned = (await db.execute(
            select(Ticket)
            .where(Ticket.workspace_id == target_user.workspace_id)
            .where(Ticket.assigned_to_id == target_user_id)
            .where(Ticket.created_at >= start_dt)
            .where(Ticket.created_at < end_dt)
            .order_by(Ticket.created_at.desc())
        )).scalars().all()
    
    # 8. Ticket comments - filter by workspace through ticket
    try:
        ticket_comments_result = await db.execute(
            text("""
                SELECT tc.id, tc.ticket_id, tc.user_id, tc.content, tc.is_internal, tc.created_at 
                FROM ticketcomment tc
                JOIN ticket t ON tc.ticket_id = t.id
                WHERE tc.user_id = :user_id
                AND t.workspace_id = :workspace_id
                AND tc.created_at >= :start_dt 
                AND tc.created_at < :end_dt 
                ORDER BY tc.created_at DESC
            """),
            {"user_id": target_user_id, "workspace_id": target_user.workspace_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        ticket_comments = ticket_comments_result.fetchall()
    except Exception as e:
        logger.error(f"Error fetching ticket comments: {e}")
        ticket_comments = []
    
    # Calculate overdue tasks
    overdue_tasks = []
    all_assigned_task_ids = set()
    for task, _ in task_assignments_raw:
        all_assigned_task_ids.add(task.id)
        if task.due_date and is_overdue(task.due_date) and task.status.value not in ['done', 'completed', 'archived']:
            task_due = task.due_date if isinstance(task.due_date, date_type) else task.due_date.date()
            overdue_tasks.append({
                'id': task.id,
                'title': task.title,
                'due_date': task_due.strftime('%Y-%m-%d'),
                'days_overdue': (now_date - task_due).days,
                'priority': task.priority.value,
                'status': task.status.value,
            })
    
    for task in tasks_created:
        if task.id not in all_assigned_task_ids:
            if task.due_date and is_overdue(task.due_date) and task.status.value not in ['done', 'completed', 'archived']:
                task_due = task.due_date if isinstance(task.due_date, date_type) else task.due_date.date()
                overdue_tasks.append({
                    'id': task.id,
                    'title': task.title,
                    'due_date': task_due.strftime('%Y-%m-%d'),
                    'days_overdue': (now_date - task_due).days,
                    'priority': task.priority.value,
                    'status': task.status.value,
                })
    
    # Calculate total time logged (timeentry table may not exist)
    total_hours = 0
    try:
        total_time_result = await db.execute(
            text("""
                SELECT SUM(duration_hours) as total FROM timeentry 
                WHERE user_id = :user_id 
                AND start_time >= :start_dt 
                AND start_time < :end_dt
            """),
            {"user_id": target_user_id, "start_dt": start_dt, "end_dt": end_dt}
        )
        total_hours = total_time_result.scalar() or 0
    except Exception:
        pass  # timeentry table doesn't exist
    
    # Calculate completed tasks
    completed_count = sum(1 for t, _ in task_assignments_raw if t.status.value in ['done', 'completed'])
    completion_rate = round((completed_count / len(task_assignments_raw) * 100) if task_assignments_raw else 0, 1)
    
    # Calculate average time per task
    tasks_with_time = [t for t in task_assignments if t['time_spent_hours']]
    avg_time_per_task = round(sum(t['time_spent_hours'] for t in tasks_with_time) / len(tasks_with_time), 1) if tasks_with_time else 0
    
    # Count active projects (projectmember table may not exist)
    active_projects = 0
    try:
        active_projects_result = await db.execute(
            text("""
                SELECT COUNT(DISTINCT project_id) FROM projectmember 
                WHERE user_id = :user_id
            """),
            {"user_id": target_user_id}
        )
        active_projects = active_projects_result.scalar() or 0
    except Exception:
        pass  # projectmember table doesn't exist
    
    # Build recent activity timeline
    recent_activity = []
    
    for task in tasks_created[:10]:
        recent_activity.append({
            'type': 'task_created',
            'description': f'Created task: {task.title[:50]}',
            'detail': None,
            'created_at': task.created_at,
        })
    
    for task, _ in task_assignments_raw[:10]:
        if task.status.value in ['done', 'completed']:
            recent_activity.append({
                'type': 'task_completed',
                'description': f'Completed task: {task.title[:50]}',
                'detail': None,
                'created_at': task.updated_at or task.created_at,
            })
    
    for ticket in tickets_closed[:10]:
        recent_activity.append({
            'type': 'ticket_closed',
            'description': f'Closed ticket: {ticket.ticket_number}',
            'detail': ticket.subject[:60],
            'created_at': ticket.closed_at,
        })
    
    for comment in comments[:10]:
        # comment[4] may be string or datetime from raw SQL
        comment_date = comment[4]
        if isinstance(comment_date, str):
            try:
                comment_date = datetime.fromisoformat(comment_date.replace('Z', '+00:00'))
            except Exception:
                comment_date = None
        recent_activity.append({
            'type': 'comment',
            'description': f'Commented on task #{comment[1]}',
            'detail': (comment[3] or '')[:60],
            'created_at': comment_date,
        })
    
    # Sort by date - ensure all values are datetime
    def get_sort_date(x):
        dt = x['created_at']
        if dt is None:
            return datetime.min
        if isinstance(dt, str):
            try:
                return datetime.fromisoformat(dt.replace('Z', '+00:00'))
            except Exception:
                return datetime.min
        return dt
    
    recent_activity.sort(key=get_sort_date, reverse=True)
    
    # Prepare summary data
    summary = {
        'tasks_created': len(tasks_created),
        'tasks_assigned': len(task_assignments),
        'task_edits': len(task_edits),
        'comments': len(comments),
        'projects_created': len(projects_created),
        'activities': len(activities),
        'tickets_assigned': len(tickets_assigned),
        'ticket_comments': len(ticket_comments),
        'tickets_closed': len(tickets_closed),
    }
    
    # Prepare metrics
    metrics = {
        'tasks_assigned': len(task_assignments),
        'tasks_completed': completed_count,
        'completion_rate': completion_rate,
        'overdue_tasks': len(overdue_tasks),
        'total_hours': round(total_hours, 1),
        'tickets_closed': len(tickets_closed),
        'avg_time_per_task': avg_time_per_task,
        'on_time_completion': round(100 - (len(overdue_tasks) / max(len(task_assignments), 1) * 100), 1),
        'avg_days_to_complete': 3,  # Would need more complex calculation
        'active_projects': active_projects,
    }
    
    return templates.TemplateResponse(
        'admin/user_activity_view.html',
        {
            'request': request,
            'user': user,
            'target_user': target_user,
            'workspace': workspace,
            'start_date': start_dt.strftime('%Y-%m-%d'),
            'end_date': (end_dt - timedelta(days=1)).strftime('%Y-%m-%d'),
            'now': datetime.utcnow(),
            'summary': summary,
            'metrics': metrics,
            'overdue_tasks': sorted(overdue_tasks, key=lambda x: x['days_overdue'], reverse=True),
            'tasks_created': tasks_with_project,
            'task_assignments': task_assignments,
            'tickets_assigned': tickets_assigned,
            'tickets_closed': tickets_closed,
            'recent_activity': recent_activity[:25],
        },
    )


@router.post('/admin/users/{user_id}/activate')
async def web_admin_activate_user(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session),
):
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get the user to activate
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Activate the user
    target_user.is_active = True
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


@router.post('/admin/users/{user_id}/toggle-admin')
async def web_admin_toggle_admin(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session),
):
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Can't modify your own admin rights
    if user_id == current_user_id:
        raise HTTPException(status_code=400, detail="Cannot modify your own admin rights")
    
    # Get the target user
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Toggle admin status
    target_user.is_admin = not target_user.is_admin
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


@router.post('/admin/users/{user_id}/toggle-ticket-visibility')
async def web_admin_toggle_ticket_visibility(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session),
):
    """Toggle whether a user can see all tickets or only their project tickets"""
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get the target user
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Toggle ticket visibility
    target_user.can_see_all_tickets = not target_user.can_see_all_tickets
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


@router.post('/admin/users/{user_id}/toggle-bubbles-analytics')
async def web_admin_toggle_bubbles_analytics(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session),
):
    """Toggle whether a user can view Bubbles analytics dashboard"""
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get the target user
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Toggle bubbles analytics permission
    target_user.show_bubbles_analytics = not target_user.show_bubbles_analytics
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


@router.post('/admin/users/{user_id}/delete')
async def web_admin_delete_user(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session),
):
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Can't delete yourself
    if user_id == current_user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    # Get the user to delete
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Hard delete: Remove user from database
    await db.delete(target_user)
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


@router.post('/admin/users/{user_id}/change-password')
async def web_admin_change_user_password(
    request: Request,
    user_id: int,
    new_password: str = Form(...),
    db: AsyncSession = Depends(get_session),
):
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get the target user
    target_user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Must be in same workspace
    if target_user.workspace_id != current_user.workspace_id:
        raise HTTPException(status_code=403, detail="User not in your workspace")
    
    # Validate password length
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    # Hash and update password
    from app.core.security import get_password_hash
    target_user.hashed_password = get_password_hash(new_password)
    
    await db.commit()
    
    return RedirectResponse('/web/admin/users', status_code=303)


# --------------------------
# Admin - Database Backup Management
# --------------------------
@router.get('/admin/backups', response_class=HTMLResponse)
async def web_admin_backups(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.backup import backup_manager
    stats = backup_manager.get_backup_stats()
    
    # Get list of all backups (both .db and .zip)
    backups = []
    for backup_file in sorted(
        [f for f in backup_manager.backup_dir.glob("backup_*.*") 
         if f.suffix in ['.db', '.zip'] and 'latest' not in f.name and 'corrupted' not in f.name],
        key=lambda x: x.stat().st_mtime, 
        reverse=True
    ):
        backup_type = "MANUAL" if "_MANUAL_" in backup_file.name else ("AUTO" if "_AUTO_" in backup_file.name else "UPLOADED")
        includes_attachments = backup_file.suffix == '.zip'
        
        backups.append({
            'filename': backup_file.name,
            'type': backup_type,
            'includes_attachments': includes_attachments,
            'size': backup_file.stat().st_size,
            'size_mb': round(backup_file.stat().st_size / (1024 * 1024), 2),
            'created': datetime.fromtimestamp(backup_file.stat().st_mtime).strftime('%d/%m/%Y %H:%M:%S'),
            'created_timestamp': backup_file.stat().st_mtime
        })
    
    # Get recent system logs for display
    from app.models.system_log import SystemLog
    from sqlalchemy import func as sa_func
    try:
        log_count_result = await db.execute(select(sa_func.count()).select_from(SystemLog))
        log_count = log_count_result.scalar() or 0
        recent_logs_result = await db.execute(
            select(SystemLog).order_by(SystemLog.timestamp.desc()).limit(20)
        )
        recent_logs = recent_logs_result.scalars().all()
    except Exception:
        log_count = 0
        recent_logs = []
    
    return templates.TemplateResponse('admin/backups.html', {
        'request': request,
        'user': user,
        'stats': stats,
        'backups': backups,
        'log_count': log_count,
        'recent_logs': recent_logs
    })


@router.post('/admin/backups/create')
async def web_admin_backup_create(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    from app.core.backup import backup_manager
    import asyncio
    
    if backup_manager.backup_status == 'running':
        return JSONResponse({'success': False, 'error': 'A backup is already in progress'})
    
    # Fire-and-forget: start backup in background thread
    async def _run_backup():
        try:
            backup_manager.backup_status = 'running'
            backup_manager.backup_progress = 'Starting backup...'
            backup_manager.backup_result_file = None
            backup_file = await asyncio.to_thread(
                backup_manager.create_backup, is_manual=True, include_attachments=True
            )
            if backup_file:
                backup_manager.backup_status = 'done'
                backup_manager.backup_progress = 'Backup created successfully'
                backup_manager.backup_result_file = backup_file.name
            else:
                backup_manager.backup_status = 'error'
                backup_manager.backup_progress = 'Backup creation failed'
        except Exception as e:
            backup_manager.backup_status = 'error'
            backup_manager.backup_progress = f'Error: {str(e)[:200]}'
    
    asyncio.create_task(_run_backup())
    return JSONResponse({'success': True, 'message': 'Backup started'})


@router.get('/admin/backups/status')
async def web_admin_backup_status(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Poll endpoint for backup progress"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    from app.core.backup import backup_manager
    return JSONResponse({
        'status': backup_manager.backup_status,
        'progress': backup_manager.backup_progress,
        'filename': backup_manager.backup_result_file
    })


@router.get('/admin/backups/download/{filename}')
async def web_admin_backup_download(
    request: Request,
    filename: str,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.backup import backup_manager
    from fastapi.responses import FileResponse
    
    # Security: prevent path traversal
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    backup_path = (backup_manager.backup_dir / filename).resolve()
    # Ensure the resolved path is still inside the backup directory
    if not str(backup_path).startswith(str(backup_manager.backup_dir.resolve())):
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")
    
    return FileResponse(
        path=str(backup_path),
        filename=filename,
        media_type='application/octet-stream'
    )


@router.post('/admin/backups/upload')
async def web_admin_backup_upload(
    request: Request,
    backup_file: UploadFile = File(...),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.backup import backup_manager
    
    # Validate file extension
    if not backup_file.filename.endswith(('.db', '.zip')):
        return RedirectResponse('/web/admin/backups?error=invalid_file', status_code=303)
    
    # Read file content
    content = await backup_file.read()
    
    # Save the uploaded backup
    saved_path = backup_manager.save_uploaded_backup(content, backup_file.filename)
    
    if saved_path:
        return RedirectResponse('/web/admin/backups?success=backup_uploaded', status_code=303)
    else:
        return RedirectResponse('/web/admin/backups?error=upload_failed', status_code=303)


@router.post('/admin/backups/restore')
async def web_admin_backup_restore(
    request: Request,
    backup_file: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.backup import backup_manager
    from pathlib import Path
    
    backup_path = backup_manager.backup_dir / backup_file
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")
    
    success = backup_manager.restore_from_backup(backup_path)
    
    if success:
        return RedirectResponse('/web/admin/backups?success=restore_complete', status_code=303)
    else:
        return RedirectResponse('/web/admin/backups?error=restore_failed', status_code=303)


@router.post('/admin/backups/delete')
async def web_admin_backup_delete(
    request: Request,
    backup_file: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.backup import backup_manager
    
    # Security: prevent path traversal
    if '..' in backup_file or '/' in backup_file or '\\' in backup_file:
        return RedirectResponse('/web/admin/backups?error=invalid_filename', status_code=303)
    
    success = backup_manager.delete_backup(backup_file)
    
    if success:
        return RedirectResponse('/web/admin/backups?success=backup_deleted', status_code=303)
    else:
        return RedirectResponse('/web/admin/backups?error=delete_failed', status_code=303)


@router.get('/admin/backups/logs/download')
async def web_admin_logs_download(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Download system diagnostic logs as a text file for a given time range."""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.system_log import SystemLog
    from fastapi.responses import Response
    
    # Parse date range from query params
    from_str = request.query_params.get('from', '')
    to_str = request.query_params.get('to', '')
    level_filter = request.query_params.get('level', '')
    
    try:
        from_dt = datetime.fromisoformat(from_str) if from_str else datetime.utcnow() - timedelta(days=1)
        to_dt = datetime.fromisoformat(to_str) if to_str else datetime.utcnow()
    except ValueError:
        from_dt = datetime.utcnow() - timedelta(days=1)
        to_dt = datetime.utcnow()
    
    # Build query
    query = select(SystemLog).where(
        SystemLog.timestamp >= from_dt,
        SystemLog.timestamp <= to_dt
    )
    
    if level_filter == 'ERROR':
        query = query.where(SystemLog.level == 'ERROR')
    elif level_filter == 'INFO':
        query = query.where(SystemLog.level.in_(['INFO', 'WARN', 'ERROR']))
    # DEBUG = all levels, empty = all levels
    
    query = query.order_by(SystemLog.timestamp.asc())
    
    result = await db.execute(query)
    logs = result.scalars().all()
    
    if not logs:
        return RedirectResponse('/web/admin/backups?error=no_logs', status_code=303)
    
    # Build compact text report
    lines = [
        f"=== System Diagnostic Report ===",
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
        f"Range: {from_dt.strftime('%Y-%m-%d %H:%M')} to {to_dt.strftime('%Y-%m-%d %H:%M')}",
        f"Filter: {level_filter or 'ALL'}",
        f"Entries: {len(logs)}",
        f"{'='*60}",
        "",
    ]
    
    for log in logs:
        ts = log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        line = f"[{ts}] [{log.level}] [{log.source}] {log.message}"
        if log.details:
            line += f" | {log.details}"
        lines.append(line)
    
    report_text = "\n".join(lines)
    filename = f"system_log_{from_dt.strftime('%Y%m%d')}_{to_dt.strftime('%Y%m%d')}.txt"
    
    return Response(
        content=report_text,
        media_type="text/plain",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# --------------------------
# Admin - System Updates
# --------------------------

@router.get('/admin/updates')
async def web_admin_updates(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.update_manager import update_manager
    
    try:
        # Get current version
        current_version = await update_manager.get_current_version()
        
        # Get commit history
        commit_history = await update_manager.get_commit_history(limit=30)
    except Exception as e:
        # If update manager fails, provide defaults
        current_version = {
            'hash': 'unknown',
            'message': 'Unable to determine version',
            'date': 'unknown',
            'branch': 'unknown'
        }
        commit_history = []
    
    # Get SSH public key if it exists
    ssh_public_key = None
    from pathlib import Path
    pub_key_path = Path.home() / ".ssh" / "crm_deploy_key.pub"
    if pub_key_path.exists():
        ssh_public_key = pub_key_path.read_text().strip()
    
    return templates.TemplateResponse('admin/updates.html', {
        'request': request,
        'user': user,
        'current_version': current_version,
        'commit_history': commit_history,
        'ssh_public_key': ssh_public_key,
        'success': request.query_params.get('success'),
        'error': request.query_params.get('error')
    })


@router.get('/admin/setup-ssh')
async def web_admin_setup_ssh(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Generate SSH key and display public key for GitHub"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    import subprocess
    from pathlib import Path
    
    home = Path.home()
    ssh_dir = home / ".ssh"
    key_path = ssh_dir / "crm_deploy_key"
    pub_key_path = ssh_dir / "crm_deploy_key.pub"
    
    message = ""
    
    # Create .ssh directory
    ssh_dir.mkdir(mode=0o700, exist_ok=True)
    
    # Generate key if it doesn't exist
    if not key_path.exists():
        try:
            subprocess.run([
                "ssh-keygen", "-t", "ed25519",
                "-C", "crm-server-deploy-key",
                "-f", str(key_path),
                "-N", ""
            ], check=True, capture_output=True)
            message = "SSH key generated successfully!"
        except Exception as e:
            message = f"Failed to generate key: {e}"
    else:
        message = "SSH key already exists"
    
    # Add GitHub to known hosts
    try:
        known_hosts = ssh_dir / "known_hosts"
        result = subprocess.run(["ssh-keyscan", "github.com"], capture_output=True, text=True)
        if result.stdout:
            with open(known_hosts, "a") as f:
                f.write(result.stdout)
    except Exception:
        pass
    
    # Update git remote to SSH
    try:
        subprocess.run([
            "git", "remote", "set-url", "origin",
            "git@github.com:dadad132/cem-backend.git"
        ], capture_output=True)
    except Exception:
        pass
    
    # Read public key
    ssh_public_key = None
    if pub_key_path.exists():
        ssh_public_key = pub_key_path.read_text().strip()
    
    return templates.TemplateResponse('admin/ssh_setup.html', {
        'request': request,
        'user': user,
        'ssh_public_key': ssh_public_key,
        'message': message
    })


@router.post('/admin/updates/latest')
async def web_admin_update_latest(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.update_manager import update_manager
    
    result = await update_manager.update_to_latest()
    
    if result["success"]:
        # Restart service after update
        await update_manager.restart_service()
        return RedirectResponse('/web/admin/updates?success=update_complete', status_code=303)
    else:
        error_msg = result.get("error", "Unknown error")
        return RedirectResponse(f'/web/admin/updates?error={error_msg}', status_code=303)


@router.post('/admin/updates/rollback')
async def web_admin_update_rollback(
    request: Request,
    commit_hash: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.core.update_manager import update_manager
    
    result = await update_manager.rollback_to_commit(commit_hash)
    
    if result["success"]:
        # Restart service after rollback
        await update_manager.restart_service()
        return RedirectResponse('/web/admin/updates?success=rollback_complete', status_code=303)
    else:
        error_msg = result.get("error", "Unknown error")
        return RedirectResponse(f'/web/admin/updates?error={error_msg}', status_code=303)


# --------------------------
# Task/Ticket Completion Email Helper
# --------------------------
async def send_completion_notification_email(
    db: AsyncSession,
    workspace_id: int,
    notification_type: str,  # 'task' or 'ticket'
    item_id: str,  # Task ID or Ticket Number
    title: str,
    status: str,
    priority: str,
    completed_by_name: str,
    created_at: datetime,
    completed_at: datetime,
    additional_details: str = ""
):
    """Send completion notification email for tasks or tickets"""
    from app.models.email_settings import EmailSettings
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    try:
        # Get email settings
        settings = (await db.execute(
            select(EmailSettings).where(EmailSettings.workspace_id == workspace_id)
        )).scalar_one_or_none()
        
        if not settings:
            logger.warning(f"No email settings for workspace {workspace_id}, skipping completion notification")
            return False
        
        if not settings.completion_notify_enabled:
            logger.debug(f"Completion notifications disabled for workspace {workspace_id}")
            return False
        
        if not settings.completion_notify_email:
            logger.warning(f"No completion notify email set for workspace {workspace_id}")
            return False
        
        # Check if we should notify for this type
        if notification_type == 'task' and not settings.completion_notify_task:
            return False
        if notification_type == 'ticket' and not settings.completion_notify_ticket:
            return False
        
        # Calculate time to complete
        time_diff = completed_at - created_at
        days = time_diff.days
        hours, remainder = divmod(time_diff.seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        
        if days > 0:
            time_to_complete = f"{days} day(s), {hours} hour(s), {minutes} minute(s)"
        elif hours > 0:
            time_to_complete = f"{hours} hour(s), {minutes} minute(s)"
        else:
            time_to_complete = f"{minutes} minute(s)"
        
        # Prepare variables
        type_display = "Task" if notification_type == 'task' else "Ticket"
        type_id_label = "Task ID" if notification_type == 'task' else "Ticket Number"
        
        # Build email subject
        subject = (settings.completion_email_subject or "{type} Completed - {title}").format(
            type=type_display,
            title=title,
            type_id=item_id
        )
        
        # Build email body
        default_body = """Good day,

Please see the {type} that has been completed:

{type} Details:
--------------
{type_id_label}: {type_id}
Title/Subject: {title}
Status: {status}
Priority: {priority}
Completed By: {completed_by}
Completed At: {completed_at}

Time to Complete: {time_to_complete}
Created At: {created_at}

{additional_details}

Best regards,
{company_name}

---
This is an automated notification from your CRM system."""
        
        body_template = settings.completion_email_body or default_body
        body = body_template.format(
            type=type_display,
            type_id=item_id,
            type_id_label=type_id_label,
            title=title,
            status=status,
            priority=priority,
            completed_by=completed_by_name,
            completed_at=completed_at.strftime('%Y-%m-%d %H:%M'),
            time_to_complete=time_to_complete,
            created_at=created_at.strftime('%Y-%m-%d %H:%M'),
            additional_details=additional_details,
            company_name=settings.company_name or "Support Team"
        )
        
        # Send email
        msg = MIMEMultipart()
        msg['From'] = f"{settings.smtp_from_name} <{settings.smtp_from_email}>"
        msg['To'] = settings.completion_notify_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        if settings.smtp_use_tls:
            server = smtplib.SMTP(settings.smtp_host, settings.smtp_port)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(settings.smtp_host, settings.smtp_port)
        
        server.login(settings.smtp_username, settings.smtp_password)
        server.sendmail(settings.smtp_from_email, settings.completion_notify_email, msg.as_string())
        server.quit()
        
        logger.info(f"Sent {notification_type} completion notification for {item_id} to {settings.completion_notify_email}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send completion notification: {e}")
        return False


# --------------------------
# Admin - Email Settings
# --------------------------
@router.get('/admin/email-settings', response_class=HTMLResponse)
async def web_admin_email_settings(request: Request, db: AsyncSession = Depends(get_session)):
    """Admin page to configure email settings"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.email_settings import EmailSettings
    
    # Get existing settings
    settings = (await db.execute(
        select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    default_body = """Dear {guest_name} {guest_surname},

Thank you for contacting us. Your support ticket has been successfully created.

Ticket Details:
--------------
Ticket Number: {ticket_number}
Subject: {subject}
Status: Open
Priority: {priority}

Our team will review your request and someone will assist you as soon as possible.

You can reference your ticket number {ticket_number} in any future communication.

Best regards,
{company_name} Support Team

---
This is an automated message. Please do not reply to this email."""

    default_completion_body = """Good day,

Please see the {type} that has been completed:

{type} Details:
--------------
{type_id_label}: {type_id}
Title/Subject: {title}
Status: {status}
Priority: {priority}
Completed By: {completed_by}
Completed At: {completed_at}

Time to Complete: {time_to_complete}
Created At: {created_at}

{additional_details}

Best regards,
{company_name}

---
This is an automated notification from your CRM system."""
    
    return templates.TemplateResponse('admin/email_settings.html', {
        'request': request,
        'user': user,
        'settings': settings,
        'default_body': default_body,
        'default_completion_body': default_completion_body
    })


@router.post('/admin/email-settings/save')
async def web_admin_email_settings_save(
    request: Request,
    smtp_host: str = Form(...),
    smtp_port: int = Form(...),
    smtp_username: str = Form(...),
    smtp_password: str = Form(...),
    smtp_from_email: str = Form(...),
    smtp_from_name: str = Form(...),
    smtp_use_tls: Optional[str] = Form(None),
    incoming_mail_type: Optional[str] = Form("IMAP"),
    incoming_mail_host: Optional[str] = Form(None),
    incoming_mail_port: Optional[int] = Form(993),
    incoming_mail_username: Optional[str] = Form(None),
    incoming_mail_password: Optional[str] = Form(None),
    incoming_mail_use_ssl: Optional[str] = Form(None),
    webmail_url: Optional[str] = Form(None),
    confirmation_subject: str = Form(...),
    confirmation_body: str = Form(...),
    company_name: str = Form(...),
    auto_reply_enabled: Optional[str] = Form(None),
    completion_notify_enabled: Optional[str] = Form(None),
    completion_notify_email: Optional[str] = Form(None),
    completion_notify_task: Optional[str] = Form(None),
    completion_notify_ticket: Optional[str] = Form(None),
    completion_email_subject: Optional[str] = Form("{type} Completed - {title}"),
    completion_email_body: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Save email settings"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.email_settings import EmailSettings
    from datetime import datetime
    
    try:
        # Get or create settings
        settings = (await db.execute(
            select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
        )).scalar_one_or_none()
        
        if settings:
            # Update existing
            settings.smtp_host = smtp_host
            settings.smtp_port = smtp_port
            settings.smtp_username = smtp_username
            settings.smtp_password = smtp_password
            settings.smtp_from_email = smtp_from_email
            settings.smtp_from_name = smtp_from_name
            settings.smtp_use_tls = smtp_use_tls == 'true'
            settings.incoming_mail_type = incoming_mail_type
            settings.incoming_mail_host = incoming_mail_host
            settings.incoming_mail_port = incoming_mail_port
            settings.incoming_mail_username = incoming_mail_username
            settings.incoming_mail_password = incoming_mail_password
            settings.incoming_mail_use_ssl = incoming_mail_use_ssl == 'true'
            settings.webmail_url = webmail_url
            settings.confirmation_subject = confirmation_subject
            settings.confirmation_body = confirmation_body
            settings.company_name = company_name
            settings.auto_reply_enabled = auto_reply_enabled == 'true'
            # Completion notification settings
            settings.completion_notify_enabled = completion_notify_enabled == 'true'
            settings.completion_notify_email = completion_notify_email
            settings.completion_notify_task = completion_notify_task == 'true'
            settings.completion_notify_ticket = completion_notify_ticket == 'true'
            settings.completion_email_subject = completion_email_subject or "{type} Completed - {title}"
            settings.completion_email_body = completion_email_body
            settings.updated_at = datetime.utcnow()
        else:
            # Create new
            settings = EmailSettings(
                workspace_id=user.workspace_id,
                smtp_host=smtp_host,
                smtp_port=smtp_port,
                smtp_username=smtp_username,
                smtp_password=smtp_password,
                smtp_from_email=smtp_from_email,
                smtp_from_name=smtp_from_name,
                smtp_use_tls=smtp_use_tls == 'true',
                incoming_mail_type=incoming_mail_type,
                incoming_mail_host=incoming_mail_host,
                incoming_mail_port=incoming_mail_port,
                incoming_mail_username=incoming_mail_username,
                incoming_mail_password=incoming_mail_password,
                incoming_mail_use_ssl=incoming_mail_use_ssl == 'true',
                webmail_url=webmail_url,
                confirmation_subject=confirmation_subject,
                confirmation_body=confirmation_body,
                company_name=company_name,
                auto_reply_enabled=auto_reply_enabled == 'true',
                completion_notify_enabled=completion_notify_enabled == 'true',
                completion_notify_email=completion_notify_email,
                completion_notify_task=completion_notify_task == 'true',
                completion_notify_ticket=completion_notify_ticket == 'true',
                completion_email_subject=completion_email_subject or "{type} Completed - {title}",
                completion_email_body=completion_email_body
            )
            db.add(settings)
        
        await db.commit()
        
        request.session['flash_message'] = "✓ Email settings saved successfully"
        request.session['flash_type'] = 'success'
        
    except Exception as e:
        request.session['flash_message'] = f"✗ Failed to save settings: {str(e)}"
        request.session['flash_type'] = 'error'
    
    return RedirectResponse('/web/admin/email-settings', status_code=303)


@router.post('/admin/email-settings/test')
async def web_admin_email_settings_test(request: Request, db: AsyncSession = Depends(get_session)):
    """Send test email"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    from app.models.email_settings import EmailSettings
    import smtplib
    from email.mime.text import MIMEText
    
    try:
        settings = (await db.execute(
            select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
        )).scalar_one_or_none()
        
        if not settings:
            return JSONResponse({'success': False, 'error': 'Email settings not configured'})
        
        # Send test email
        msg = MIMEText("This is a test email from your CRM system. Email settings are working correctly!")
        msg['Subject'] = "Test Email - CRM System"
        msg['From'] = f"{settings.smtp_from_name} <{settings.smtp_from_email}>"
        msg['To'] = settings.smtp_username
        
        server = smtplib.SMTP(settings.smtp_host, settings.smtp_port)
        if settings.smtp_use_tls:
            server.starttls()
        server.login(settings.smtp_username, settings.smtp_password)
        server.send_message(msg)
        server.quit()
        
        return JSONResponse({'success': True})
        
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)})


@router.post('/admin/email-settings/check-emails')
async def web_admin_check_emails(request: Request, db: AsyncSession = Depends(get_session)):
    """Manually trigger email check - wakes the background scheduler immediately."""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'success': False, 'error': 'User not found'})
    
    try:
        from app.core.email_scheduler_v2 import email_scheduler
        
        if not email_scheduler.running:
            return JSONResponse({'success': False, 'error': 'Email scheduler is not running. Please restart the server.'})
        
        # Wake the scheduler to run immediately (fire-and-forget)
        await email_scheduler.check_now()
        
        return JSONResponse({
            'success': True,
            'message': 'Email check triggered'
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"Email check error: {error_details}")
        return JSONResponse({'success': False, 'error': str(e), 'details': error_details})


@router.get('/admin/email-settings/check-emails/status')
async def web_admin_check_emails_status(request: Request):
    """Poll endpoint: returns whether the email scheduler is currently checking or has completed."""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    from app.core.email_scheduler_v2 import email_scheduler
    
    completed_at = None
    if email_scheduler._last_check_completed_at:
        completed_at = email_scheduler._last_check_completed_at.isoformat()
    
    return JSONResponse({
        'checking': email_scheduler._checking,
        'completed_at': completed_at,
    })


@router.get('/admin/email-settings/diagnose')
async def web_admin_email_diagnose(request: Request, db: AsyncSession = Depends(get_session)):
    """Diagnose email account connections - tests each account and reports status"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    import imaplib as _imaplib
    import poplib as _poplib
    import socket as _socket
    
    from app.models.email_settings import EmailSettings
    from app.models.incoming_email_account import IncomingEmailAccount
    from app.models.processed_mail import ProcessedMail
    
    results = []
    
    # 1. Check legacy EmailSettings
    settings = (await db.execute(
        select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if settings and settings.incoming_mail_host:
        def test_legacy():
            step = "initializing"
            try:
                _socket.setdefaulttimeout(15)
                host = settings.incoming_mail_host
                port = settings.incoming_mail_port or 993
                use_ssl = settings.incoming_mail_use_ssl
                username = settings.incoming_mail_username
                
                # Auto-detect Gmail and force correct IMAP settings
                is_gmail = 'gmail' in host.lower() or 'google' in host.lower()
                if is_gmail and (not use_ssl or port in (110, 143, 0)):
                    print(f"[Diagnose] Gmail detected ({host}) - forcing SSL on port 993")
                    use_ssl = True
                    port = 993
                
                step = f"connecting to {host}:{port} (SSL={use_ssl})"
                if use_ssl:
                    m = _imaplib.IMAP4_SSL(host, port)
                else:
                    m = _imaplib.IMAP4(host, port or 143)
                
                step = f"authenticating as {username}"
                m.login(username, settings.incoming_mail_password)
                
                step = "selecting INBOX"
                m.select('INBOX')
                
                from datetime import datetime as dt2, timedelta as td2
                date_since = (dt2.now() - td2(days=7)).strftime("%d-%b-%Y")
                status2, messages2 = m.search(None, f'SINCE {date_since}')
                email_count = len(messages2[0].split()) if messages2[0] else 0
                
                gmail_folders = {}
                for folder in ['[Gmail]/Spam', '[Gmail]/All Mail', '[Gmail]/Trash']:
                    try:
                        s, _ = m.select(folder)
                        if s == 'OK':
                            s2, m2 = m.search(None, f'SINCE {date_since}')
                            gmail_folders[folder] = len(m2[0].split()) if m2[0] else 0
                    except Exception:
                        pass
                
                m.close()
                m.logout()
                _socket.setdefaulttimeout(None)
                return {
                    'status': 'success',
                    'inbox_emails_7days': email_count,
                    'gmail_folders': gmail_folders if gmail_folders else None
                }
            except Exception as e:
                _socket.setdefaulttimeout(None)
                return {'status': 'error', 'error': f"Failed at {step}: {str(e)}"}
        
        import asyncio as _asyncio
        try:
            result = await _asyncio.wait_for(_asyncio.to_thread(test_legacy), timeout=30.0)
        except _asyncio.TimeoutError:
            result = {'status': 'error', 'error': 'Connection timed out after 30 seconds'}
        
        results.append({
            'type': 'Legacy EmailSettings',
            'name': 'Main Workspace Email',
            'host': settings.incoming_mail_host,
            'port': settings.incoming_mail_port,
            'username': settings.incoming_mail_username,
            'use_ssl': settings.incoming_mail_use_ssl,
            'mail_type': settings.incoming_mail_type,
            **result
        })
    else:
        results.append({
            'type': 'Legacy EmailSettings',
            'name': 'Main Workspace Email',
            'status': 'not_configured',
            'error': 'No incoming mail settings configured'
        })
    
    # 2. Check IncomingEmailAccount entries
    accounts = (await db.execute(
        select(IncomingEmailAccount).where(
            IncomingEmailAccount.workspace_id == user.workspace_id
        )
    )).scalars().all()
    
    for account in accounts:
        acct_host = account.imap_host
        acct_port = account.imap_port or 993
        acct_ssl = account.imap_use_ssl
        acct_user = account.imap_username
        acct_pass = account.imap_password
        acct_protocol = getattr(account, 'protocol', 'imap')
        
        def test_account(h=acct_host, p=acct_port, ssl_on=acct_ssl, u=acct_user, pw=acct_pass, proto=acct_protocol):
            step = "initializing"
            try:
                _socket.setdefaulttimeout(15)
                if proto == 'pop3':
                    step = f"POP3 connecting to {h}:{p}"
                    if ssl_on:
                        conn = _poplib.POP3_SSL(h, p)
                    else:
                        conn = _poplib.POP3(h, p or 110)
                    step = f"authenticating as {u}"
                    conn.user(u)
                    conn.pass_(pw)
                    count, size = conn.stat()
                    conn.quit()
                    _socket.setdefaulttimeout(None)
                    return {'status': 'success', 'total_emails': count, 'protocol': 'pop3'}
                else:
                    step = f"IMAP connecting to {h}:{p}"
                    if ssl_on:
                        m = _imaplib.IMAP4_SSL(h, p)
                    else:
                        m = _imaplib.IMAP4(h, p or 143)
                        try:
                            m.starttls()
                        except Exception:
                            pass
                    
                    step = f"authenticating as {u}"
                    m.login(u, pw)
                    m.select('INBOX')
                    
                    from datetime import datetime as dt3, timedelta as td3
                    date_since = (dt3.now() - td3(days=7)).strftime("%d-%b-%Y")
                    status3, messages3 = m.search(None, f'SINCE {date_since}')
                    email_count = len(messages3[0].split()) if messages3[0] else 0
                    
                    gmail_folders = {}
                    if 'gmail' in h.lower() or 'google' in h.lower():
                        for folder in ['[Gmail]/Spam', '[Gmail]/All Mail', '[Gmail]/Trash']:
                            try:
                                s, _ = m.select(folder)
                                if s == 'OK':
                                    s2, m2 = m.search(None, f'SINCE {date_since}')
                                    gmail_folders[folder] = len(m2[0].split()) if m2[0] else 0
                            except Exception:
                                pass
                    
                    m.close()
                    m.logout()
                    _socket.setdefaulttimeout(None)
                    return {
                        'status': 'success',
                        'inbox_emails_7days': email_count,
                        'gmail_folders': gmail_folders if gmail_folders else None,
                        'protocol': 'imap'
                    }
            except Exception as e:
                _socket.setdefaulttimeout(None)
                return {'status': 'error', 'error': f"Failed at {step}: {str(e)}"}
        
        import asyncio as _asyncio2
        try:
            result = await _asyncio2.wait_for(_asyncio2.to_thread(test_account), timeout=30.0)
        except _asyncio2.TimeoutError:
            result = {'status': 'error', 'error': 'Connection timed out after 30 seconds'}
        
        results.append({
            'type': 'IncomingEmailAccount',
            'name': account.name,
            'email': account.email_address,
            'host': account.imap_host,
            'port': account.imap_port,
            'username': account.imap_username,
            'use_ssl': account.imap_use_ssl,
            'is_active': account.is_active,
            'protocol': acct_protocol,
            **result
        })
    
    # 3. Get processed email count
    processed_result = await db.execute(
        select(ProcessedMail).where(ProcessedMail.workspace_id == user.workspace_id)
    )
    processed_count = len(processed_result.scalars().all())
    
    return JSONResponse({
        'success': True,
        'workspace_id': user.workspace_id,
        'total_email_sources': len(results),
        'total_processed_emails': processed_count,
        'accounts': results
    })


@router.get('/admin/email-settings/debug')
async def web_admin_debug_settings(request: Request, db: AsyncSession = Depends(get_session)):
    """Debug: Show current email settings from database"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    from app.models.email_settings import EmailSettings
    from app.models.processed_mail import ProcessedMail
    
    settings = (await db.execute(
        select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not settings:
        return JSONResponse({'success': False, 'error': 'No email settings found'})
    
    # Get recent processedmail entries
    processed_emails = (await db.execute(
        select(ProcessedMail)
        .where(ProcessedMail.workspace_id == user.workspace_id)
        .order_by(ProcessedMail.processed_at.desc())
        .limit(10)
    )).scalars().all()
    
    return JSONResponse({
        'success': True,
        'settings': {
            'smtp_host': settings.smtp_host,
            'smtp_port': settings.smtp_port,
            'smtp_username': settings.smtp_username,
            'incoming_mail_host': settings.incoming_mail_host,
            'incoming_mail_port': settings.incoming_mail_port,
            'incoming_mail_username': settings.incoming_mail_username,
            'incoming_mail_use_ssl': settings.incoming_mail_use_ssl,
            'mail_type': settings.incoming_mail_type,
            'workspace_id': settings.workspace_id
        },
        'processed_emails': [{
            'message_id': p.message_id[:50] + '...' if len(p.message_id) > 50 else p.message_id,
            'email_from': p.email_from,
            'subject': p.subject,
            'ticket_id': p.ticket_id,
            'processed_at': p.processed_at.isoformat() if p.processed_at else None
        } for p in processed_emails]
    })


@router.get('/admin/email-settings/comment-logs', response_class=HTMLResponse)
async def web_admin_comment_logs(request: Request, db: AsyncSession = Depends(get_session)):
    """List all comment email log files"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return RedirectResponse('/web/dashboard', status_code=303)
    
    from pathlib import Path
    from datetime import datetime
    log_dir = Path("logs/comment_emails")
    
    logs = []
    if log_dir.exists():
        for log_file in sorted(log_dir.glob("*.log"), reverse=True):
            stat = log_file.stat()
            logs.append({
                'name': log_file.name,
                'size': stat.st_size,
                'modified': stat.st_mtime
            })
    
    return enhanced_template_response('admin/comment_logs.html', {
        'request': request,
        'logs': logs,
        'user': user
    })


@router.get('/admin/email-settings/comment-logs/{filename}')
async def web_admin_download_comment_log(filename: str, request: Request, db: AsyncSession = Depends(get_session)):
    """Download or view a specific comment email log file"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    from pathlib import Path
    from fastapi.responses import FileResponse, PlainTextResponse
    
    log_file = Path("logs/comment_emails") / filename
    
    # Security: prevent path traversal
    try:
        log_file.resolve().relative_to(Path("logs/comment_emails").resolve())
    except ValueError:
        return JSONResponse({'success': False, 'error': 'Invalid file path'})
    
    if not log_file.exists():
        return JSONResponse({'success': False, 'error': 'File not found'})
    
    # If viewing (not downloading), return as plain text
    if request.query_params.get('view') == 'true':
        content = log_file.read_text(encoding='utf-8')
        return PlainTextResponse(content)
    
    return FileResponse(log_file, filename=filename, media_type='text/plain')


@router.get('/admin/email-settings/preview-inbox')
async def web_admin_preview_inbox(request: Request, db: AsyncSession = Depends(get_session)):
    """Preview inbox emails without processing them"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    try:
        from app.models.email_settings import EmailSettings
        import imaplib
        import email
        from email.header import decode_header
        from datetime import datetime
        
        # Get email settings
        settings = (await db.execute(
            select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
        )).scalar_one_or_none()
        
        if not settings or not settings.incoming_mail_host:
            return JSONResponse({
                'success': False, 
                'error': 'Incoming mail not configured',
                'details': 'Please configure IMAP settings first'
            })
        
        # Check if mail type is IMAP
        if settings.incoming_mail_type and settings.incoming_mail_type.upper() != 'IMAP':
            return JSONResponse({
                'success': False,
                'error': f'Mail type is set to {settings.incoming_mail_type}',
                'details': 'Inbox preview only works with IMAP. Change mail type to IMAP in settings.'
            })
        
        # Connect to IMAP
        mail = None
        try:
            if settings.incoming_mail_use_ssl:
                logger.debug(f"Connecting to {settings.incoming_mail_host}:{settings.incoming_mail_port or 993} (SSL)")
                mail = imaplib.IMAP4_SSL(
                    settings.incoming_mail_host,
                    settings.incoming_mail_port or 993,
                    timeout=10
                )
            else:
                logger.debug(f"Connecting to {settings.incoming_mail_host}:{settings.incoming_mail_port or 143} (no SSL)")
                mail = imaplib.IMAP4(
                    settings.incoming_mail_host,
                    settings.incoming_mail_port or 143,
                    timeout=10
                )
            
            logger.debug(f"Logging in as {settings.incoming_mail_username}")
            mail.login(settings.incoming_mail_username, settings.incoming_mail_password)
        except (imaplib.IMAP4.error, OSError, TimeoutError) as e:
            error_msg = str(e) if str(e) else 'Connection refused or timeout'
            return JSONResponse({
                'success': False,
                'error': f'IMAP connection failed: {error_msg}',
                'details': f'Could not connect to {settings.incoming_mail_host}:{settings.incoming_mail_port or (993 if settings.incoming_mail_use_ssl else 143)}. Verify server is reachable and port is correct.'
            })
        except Exception as e:
            return JSONResponse({
                'success': False,
                'error': f'Connection error: {str(e)}',
                'details': 'Verify IMAP server settings and credentials'
            })
        mail.select('INBOX')
        
        # Search for last 10 emails (ALL, not just UNSEEN)
        status, messages = mail.search(None, 'ALL')
        email_ids = messages[0].split()
        
        # Get last 10 emails
        email_ids = email_ids[-10:] if len(email_ids) > 10 else email_ids
        email_ids = reversed(email_ids)  # Show newest first
        
        emails = []
        for email_id in email_ids:
            try:
                # Fetch email headers and body
                status, msg_data = mail.fetch(email_id, '(RFC822 FLAGS)')
                flags = msg_data[0]
                msg = email.message_from_bytes(msg_data[0][1])
                
                # Check if unread
                is_unread = b'\\Seen' not in flags
                
                # Decode subject
                subject_header = msg.get('Subject', '')
                if subject_header:
                    decoded_parts = decode_header(subject_header)
                    subject = ''
                    for part, encoding in decoded_parts:
                        if isinstance(part, bytes):
                            subject += part.decode(encoding or 'utf-8', errors='ignore')
                        else:
                            subject += part
                else:
                    subject = '(No Subject)'
                
                # Get from
                from_header = msg.get('From', 'Unknown')
                
                # Get date
                date_header = msg.get('Date', '')
                try:
                    date_obj = email.utils.parsedate_to_datetime(date_header)
                    date_str = date_obj.strftime('%b %d, %H:%M')
                except (ValueError, TypeError):
                    date_str = date_header[:20] if date_header else 'Unknown'
                
                # Get In-Reply-To
                in_reply_to = msg.get('In-Reply-To', '')
                
                # Get body preview
                body = ''
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == 'text/plain':
                            try:
                                body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                                break
                            except (UnicodeDecodeError, AttributeError):
                                pass
                else:
                    try:
                        body = msg.get_payload(decode=True).decode('utf-8', errors='ignore')
                    except (UnicodeDecodeError, AttributeError):
                        body = str(msg.get_payload())
                
                # Clean body for preview
                body = body.replace('\r', '').replace('\n', ' ').strip()
                
                emails.append({
                    'subject': subject,
                    'from': from_header,
                    'date': date_str,
                    'is_unread': is_unread,
                    'in_reply_to': in_reply_to,
                    'preview': body[:200] if body else None
                })
                
            except Exception as e:
                logger.warning(f"Error fetching email {email_id}: {e}")
                continue
        
        mail.close()
        mail.logout()
        
        return JSONResponse({
            'success': True,
            'emails': emails,
            'total': len(emails)
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"Inbox preview error: {error_details}")
        return JSONResponse({
            'success': False, 
            'error': str(e),
            'details': error_details
        })


# --------------------------
# Email Accounts Management (Admin Only)
# --------------------------
@router.get('/admin/email-accounts', response_class=HTMLResponse)
async def web_admin_email_accounts(request: Request, db: AsyncSession = Depends(get_session)):
    """List all email accounts"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.incoming_email_account import IncomingEmailAccount
    
    # Get all email accounts for this workspace
    accounts = (await db.execute(
        select(IncomingEmailAccount).where(IncomingEmailAccount.workspace_id == user.workspace_id)
        .order_by(IncomingEmailAccount.name)
    )).scalars().all()
    
    # Get users for auto-assign dropdown
    users = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id, User.is_active == True)
    )).scalars().all()
    
    # Get projects for linking emails to projects
    projects = (await db.execute(
        select(Project).where(Project.workspace_id == user.workspace_id)
        .order_by(Project.name)
    )).scalars().all()
    
    # Debug: log number of projects found
    import logging
    logging.info(f"Email accounts page: Found {len(projects)} projects for workspace {user.workspace_id}")
    
    return templates.TemplateResponse('admin/email_accounts.html', {
        'request': request,
        'user': user,
        'accounts': accounts,
        'users': users,
        'projects': projects
    })


@router.post('/admin/email-accounts/add')
async def web_admin_email_accounts_add(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Add a new email account"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.incoming_email_account import IncomingEmailAccount
    import imaplib
    import asyncio
    
    form = await request.form()
    
    # Extract form data
    imap_host = form.get('imap_host')
    imap_port = int(form.get('imap_port', 993))
    imap_username = form.get('imap_username')
    imap_password = form.get('imap_password')
    imap_use_ssl = form.get('imap_use_ssl') == 'on'
    skip_test = form.get('skip_test') == 'on'
    protocol = form.get('protocol', 'imap')  # 'imap' or 'pop3'
    
    test_error = None
    
    # Test mail connection before saving (unless skip_test is checked)
    def test_mail_connection():
        """Test IMAP or POP3 connection (runs in thread pool)"""
        import socket
        import ssl
        import poplib
        mail = None
        pop3_conn = None
        step = "initializing"
        try:
            # Set socket timeout to 15 seconds
            socket.setdefaulttimeout(15)
            
            if protocol == 'pop3':
                # POP3 connection test
                step = f"connecting to {imap_host}:{imap_port} via POP3"
                if imap_use_ssl:
                    step = f"SSL connecting to {imap_host}:{imap_port} via POP3"
                    pop3_conn = poplib.POP3_SSL(imap_host, imap_port)
                else:
                    step = f"connecting to {imap_host}:{imap_port} via POP3 (non-SSL)"
                    pop3_conn = poplib.POP3(imap_host, imap_port)
                
                step = f"authenticating as {imap_username}"
                pop3_conn.user(imap_username)
                pop3_conn.pass_(imap_password)
                
                step = "getting mailbox info"
                pop3_conn.stat()
                
                pop3_conn.quit()
                return None  # Success
            else:
                # IMAP connection test (default)
                step = f"connecting to {imap_host}:{imap_port} via IMAP"
                if imap_use_ssl:
                    step = f"SSL connecting to {imap_host}:{imap_port} via IMAP"
                    mail = imaplib.IMAP4_SSL(imap_host, imap_port)
                else:
                    step = f"connecting to {imap_host}:{imap_port} via IMAP (non-SSL)"
                    mail = imaplib.IMAP4(imap_host, imap_port)
                    step = "upgrading to TLS (STARTTLS)"
                    try:
                        mail.starttls()
                    except Exception as e:
                        # Server doesn't support STARTTLS, continue without encryption
                        step = "continuing without STARTTLS"
                
                step = f"authenticating as {imap_username}"
                mail.login(imap_username, imap_password)
                
                step = "selecting INBOX"
                mail.select('INBOX')
                
                mail.close()
                mail.logout()
                return None  # Success
            
        except socket.timeout:
            return f"Timeout while {step} - server not responding within 15 seconds"
        except socket.gaierror as e:
            return f"DNS lookup failed for {imap_host} - check hostname is correct ({e})"
        except ssl.SSLError as e:
            return f"SSL/TLS error while {step}: {e}"
        except ConnectionRefusedError:
            return f"Connection refused while {step} - port {imap_port} may be blocked or wrong"
        except OSError as e:
            if "No route to host" in str(e):
                return f"Cannot reach {imap_host} - server may be down or firewall blocking"
            return f"Network error while {step}: {e}"
        except imaplib.IMAP4.error as e:
            error_msg = str(e)
            if "authentication" in error_msg.lower() or "login" in error_msg.lower():
                return f"Authentication failed - check username/password ({error_msg})"
            return f"IMAP error while {step}: {error_msg}"
        except poplib.error_proto as e:
            error_msg = str(e)
            if "authentication" in error_msg.lower() or "login" in error_msg.lower() or "password" in error_msg.lower():
                return f"POP3 authentication failed - check username/password ({error_msg})"
            return f"POP3 error while {step}: {error_msg}"
        except Exception as e:
            return f"Error while {step}: {type(e).__name__}: {e}"
        finally:
            socket.setdefaulttimeout(None)
    
    # Run test in thread pool with timeout (unless skip_test is checked)
    if not skip_test:
        try:
            test_error = await asyncio.wait_for(
                asyncio.to_thread(test_mail_connection),
                timeout=30.0
            )
        except asyncio.TimeoutError:
            test_error = f"Connection test timed out after 30 seconds - server may be unreachable"
        
        if test_error:
            request.session['flash_message'] = f"✗ {test_error}. Please check your IMAP settings."
            request.session['flash_type'] = 'error'
            return RedirectResponse('/web/admin/email-accounts', status_code=303)
    
    try:
        auto_assign = form.get('auto_assign_to_user_id')
        project_id = form.get('project_id')
        protocol = form.get('protocol', 'imap')  # 'imap' or 'pop3'
        account = IncomingEmailAccount(
            workspace_id=user.workspace_id,
            name=form.get('name'),
            email_address=form.get('email_address'),
            project_id=int(project_id) if project_id else None,
            protocol=protocol,
            imap_host=imap_host,
            imap_port=imap_port,
            imap_username=imap_username,
            imap_password=imap_password,
            imap_use_ssl=imap_use_ssl,
            smtp_host=form.get('smtp_host') or None,
            smtp_port=int(form.get('smtp_port', 587)) if form.get('smtp_port') else 587,
            smtp_username=form.get('smtp_username') or None,
            smtp_password=form.get('smtp_password') or None,
            smtp_use_tls=form.get('smtp_use_tls') == 'on',
            is_active=True,
            auto_assign_to_user_id=int(auto_assign) if auto_assign else None,
            default_priority=form.get('default_priority', 'medium'),
            default_category=form.get('default_category', 'support')
        )
        db.add(account)
        await db.commit()
        
        request.session['flash_message'] = f"✓ Email account '{account.name}' added successfully! Connection verified."
        request.session['flash_type'] = 'success'
        
    except Exception as e:
        request.session['flash_message'] = f"✗ Failed to add account: {str(e)}"
        request.session['flash_type'] = 'error'
    
    return RedirectResponse('/web/admin/email-accounts', status_code=303)


@router.post('/admin/email-accounts/{account_id}/update')
async def web_admin_email_accounts_update(
    account_id: int,
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Update an email account"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.incoming_email_account import IncomingEmailAccount
    from datetime import datetime
    import imaplib
    import asyncio
    
    account = (await db.execute(
        select(IncomingEmailAccount).where(
            IncomingEmailAccount.id == account_id,
            IncomingEmailAccount.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    form = await request.form()
    
    # Extract form data for testing
    imap_host = form.get('imap_host')
    imap_port = int(form.get('imap_port', 993))
    imap_username = form.get('imap_username')
    new_password = form.get('imap_password')
    imap_password = new_password if new_password else account.imap_password
    imap_use_ssl = form.get('imap_use_ssl') == 'on'
    
    # Test IMAP connection before saving
    def test_imap_connection():
        """Test IMAP connection (runs in thread pool)"""
        import socket
        import ssl
        mail = None
        step = "initializing"
        try:
            # Set socket timeout to 15 seconds
            socket.setdefaulttimeout(15)
            
            step = f"connecting to {imap_host}:{imap_port}"
            if imap_use_ssl:
                step = f"SSL connecting to {imap_host}:{imap_port}"
                mail = imaplib.IMAP4_SSL(imap_host, imap_port)
            else:
                step = f"connecting to {imap_host}:{imap_port} (non-SSL)"
                mail = imaplib.IMAP4(imap_host, imap_port)
                step = "upgrading to TLS (STARTTLS)"
                try:
                    mail.starttls()
                except Exception as e:
                    step = "continuing without STARTTLS"
            
            step = f"authenticating as {imap_username}"
            mail.login(imap_username, imap_password)
            
            step = "selecting INBOX"
            mail.select('INBOX')
            
            mail.close()
            mail.logout()
            return None  # Success
            
        except socket.timeout:
            return f"Timeout while {step} - server not responding within 15 seconds"
        except socket.gaierror as e:
            return f"DNS lookup failed for {imap_host} - check hostname ({e})"
        except ssl.SSLError as e:
            return f"SSL/TLS error while {step}: {e}"
        except ConnectionRefusedError:
            return f"Connection refused while {step} - port {imap_port} may be blocked"
        except OSError as e:
            if "No route to host" in str(e):
                return f"Cannot reach {imap_host} - server may be down or firewall blocking"
            return f"Network error while {step}: {e}"
        except imaplib.IMAP4.error as e:
            error_msg = str(e)
            if "authentication" in error_msg.lower() or "login" in error_msg.lower():
                return f"Authentication failed - check username/password ({error_msg})"
            return f"IMAP error while {step}: {error_msg}"
        except Exception as e:
            return f"Error while {step}: {type(e).__name__}: {e}"
        finally:
            socket.setdefaulttimeout(None)
    
    # Run test in thread pool with timeout
    try:
        test_error = await asyncio.wait_for(
            asyncio.to_thread(test_imap_connection),
            timeout=30.0
        )
    except asyncio.TimeoutError:
        test_error = f"Connection test timed out after 30 seconds - server may be unreachable"
    
    if test_error:
        request.session['flash_message'] = f"✗ {test_error}. Settings not saved."
        request.session['flash_type'] = 'error'
        return RedirectResponse('/web/admin/email-accounts', status_code=303)
    
    try:
        auto_assign = form.get('auto_assign_to_user_id')
        project_id = form.get('project_id')
        account.name = form.get('name')
        account.email_address = form.get('email_address')
        account.project_id = int(project_id) if project_id else None
        account.imap_host = imap_host
        account.imap_port = imap_port
        account.imap_username = imap_username
        
        # Only update password if provided
        if new_password:
            account.imap_password = new_password
        
        account.imap_use_ssl = imap_use_ssl
        account.is_active = form.get('is_active') == 'on'
        account.auto_assign_to_user_id = int(auto_assign) if auto_assign else None
        account.default_priority = form.get('default_priority', 'medium')
        account.default_category = form.get('default_category', 'support')
        
        # SMTP settings for outgoing replies
        account.smtp_host = form.get('smtp_host') or None
        account.smtp_port = int(form.get('smtp_port', 587)) if form.get('smtp_port') else 587
        smtp_username_val = form.get('smtp_username') or None
        account.smtp_username = smtp_username_val
        new_smtp_password = form.get('smtp_password')
        if new_smtp_password:
            account.smtp_password = new_smtp_password
        account.smtp_use_tls = form.get('smtp_use_tls') == 'on'
        
        account.updated_at = datetime.utcnow()
        
        await db.commit()
        
        request.session['flash_message'] = f"✓ Email account '{account.name}' updated successfully! Connection verified."
        request.session['flash_type'] = 'success'
        
    except Exception as e:
        request.session['flash_message'] = f"✗ Failed to update account: {str(e)}"
        request.session['flash_type'] = 'error'
    
    return RedirectResponse('/web/admin/email-accounts', status_code=303)


@router.post('/admin/email-accounts/{account_id}/delete')
async def web_admin_email_accounts_delete(
    account_id: int,
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Delete an email account"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.incoming_email_account import IncomingEmailAccount
    
    account = (await db.execute(
        select(IncomingEmailAccount).where(
            IncomingEmailAccount.id == account_id,
            IncomingEmailAccount.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    try:
        account_name = account.name
        await db.delete(account)
        await db.commit()
        
        request.session['flash_message'] = f"✓ Email account '{account_name}' deleted"
        request.session['flash_type'] = 'success'
        
    except Exception as e:
        request.session['flash_message'] = f"✗ Failed to delete account: {str(e)}"
        request.session['flash_type'] = 'error'
    
    return RedirectResponse('/web/admin/email-accounts', status_code=303)


@router.post('/admin/email-accounts/{account_id}/test')
async def web_admin_email_accounts_test(
    account_id: int,
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Test connection to an email account"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'success': False, 'error': 'Not authenticated'})
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return JSONResponse({'success': False, 'error': 'Admin access required'})
    
    from app.models.incoming_email_account import IncomingEmailAccount
    import imaplib
    import socket
    import ssl
    
    account = (await db.execute(
        select(IncomingEmailAccount).where(
            IncomingEmailAccount.id == account_id,
            IncomingEmailAccount.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not account:
        return JSONResponse({'success': False, 'error': 'Account not found'})
    
    def do_test():
        step = "initializing"
        try:
            # Set socket timeout to 15 seconds
            socket.setdefaulttimeout(15)
            
            step = f"connecting to {account.imap_host}:{account.imap_port}"
            # Try to connect
            if account.imap_use_ssl:
                step = f"SSL connecting to {account.imap_host}:{account.imap_port}"
                mail = imaplib.IMAP4_SSL(account.imap_host, account.imap_port)
            else:
                step = f"connecting to {account.imap_host}:{account.imap_port} (non-SSL)"
                mail = imaplib.IMAP4(account.imap_host, account.imap_port)
                step = "upgrading to TLS (STARTTLS)"
                try:
                    mail.starttls()
                except Exception:
                    step = "continuing without STARTTLS"
            
            step = f"authenticating as {account.imap_username}"
            mail.login(account.imap_username, account.imap_password)
            
            step = "selecting INBOX"
            mail.select('INBOX')
            
            # Get unread count
            status, messages = mail.search(None, 'UNSEEN')
            unread_count = len(messages[0].split()) if messages[0] else 0
            
            mail.close()
            mail.logout()
            
            return {'success': True, 'message': f'Connection successful! Found {unread_count} unread email(s)'}
            
        except socket.timeout:
            return {'success': False, 'error': f'Timeout while {step} - server not responding'}
        except socket.gaierror as e:
            return {'success': False, 'error': f'DNS lookup failed for {account.imap_host} ({e})'}
        except ssl.SSLError as e:
            return {'success': False, 'error': f'SSL/TLS error while {step}: {e}'}
        except ConnectionRefusedError:
            return {'success': False, 'error': f'Connection refused while {step} - port may be blocked'}
        except OSError as e:
            if "No route to host" in str(e):
                return {'success': False, 'error': f'Cannot reach {account.imap_host} - firewall blocking?'}
            return {'success': False, 'error': f'Network error while {step}: {e}'}
        except imaplib.IMAP4.error as e:
            error_msg = str(e)
            if "authentication" in error_msg.lower() or "login" in error_msg.lower():
                return {'success': False, 'error': f'Authentication failed ({error_msg})'}
            return {'success': False, 'error': f'IMAP error while {step}: {error_msg}'}
        except Exception as e:
            return {'success': False, 'error': f'Error while {step}: {type(e).__name__}: {e}'}
        finally:
            socket.setdefaulttimeout(None)
    
    try:
        result = await asyncio.wait_for(
            asyncio.to_thread(do_test),
            timeout=30.0
        )
        return JSONResponse(result)
    except asyncio.TimeoutError:
        return JSONResponse({'success': False, 'error': 'Connection test timed out after 30 seconds'})


# --------------------------
# Site Settings (Admin Only)
# --------------------------
@router.get('/admin/site-settings', response_class=HTMLResponse)
async def web_admin_site_settings(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Site branding and customization settings"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return RedirectResponse('/web/dashboard', status_code=303)
    
    # Get workspace settings
    error_message = None
    workspace = None
    try:
        workspace = (await db.execute(
            select(Workspace).where(Workspace.id == user.workspace_id)
        )).scalar_one_or_none()
    except Exception as e:
        # If columns don't exist yet, show migration required message
        if "no such column" in str(e):
            error_message = "Database migration required. Please run: python add_site_settings_columns.py"
        else:
            raise
    
    success_message = request.session.pop('success_message', None)
    if not error_message:
        error_message = request.session.pop('error_message', None)
    
    return templates.TemplateResponse('admin/site_settings.html', {
        'request': request,
        'user': user,
        'workspace': workspace,
        'success_message': success_message,
        'error_message': error_message
    })


@router.post('/admin/site-settings/save')
async def web_admin_site_settings_save(
    request: Request,
    site_title: str = Form(None),
    primary_color: str = Form("#2563eb"),
    timezone: str = Form("UTC"),
    business_hours_start: str = Form("07:30"),
    business_hours_end: str = Form("16:00"),
    business_hours_exclude_weekends: str = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Save site branding settings"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return RedirectResponse('/web/dashboard', status_code=303)
    
    try:
        # Update workspace settings
        workspace = (await db.execute(
            select(Workspace).where(Workspace.id == user.workspace_id)
        )).scalar_one_or_none()
        
        if workspace:
            workspace.site_title = site_title if site_title else None
            workspace.primary_color = primary_color
            workspace.timezone = timezone
            
            # Business hours settings
            workspace.business_hours_start = business_hours_start or "07:30"
            workspace.business_hours_end = business_hours_end or "16:00"
            workspace.business_hours_exclude_weekends = business_hours_exclude_weekends == "1"
            
            await db.commit()
            request.session['success_message'] = 'Site settings saved successfully!'
        else:
            request.session['error_message'] = 'Workspace not found'
            
    except Exception as e:
        if "no such column" in str(e):
            request.session['error_message'] = 'Database migration required. Run: python migrations/add_workspace_timezone.py'
        else:
            request.session['error_message'] = f'Failed to save settings: {str(e)}'
    
    return RedirectResponse('/web/admin/site-settings', status_code=303)


@router.post('/admin/site-settings/upload-logo')
async def web_admin_site_settings_upload_logo(
    request: Request,
    logo: UploadFile = File(...),
    db: AsyncSession = Depends(get_session)
):
    """Upload site logo"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return RedirectResponse('/web/dashboard', status_code=303)
    
    try:
        # Validate file type (SVG excluded - can contain XSS scripts)
        allowed_types = ['image/png', 'image/jpeg', 'image/jpg', 'image/gif']
        if logo.content_type not in allowed_types:
            request.session['error_message'] = 'Invalid file type. Please upload PNG, JPG, or GIF.'
            return RedirectResponse('/web/admin/site-settings', status_code=303)
        
        # Validate file size (max 5MB)
        logo_content = await logo.read()
        if len(logo_content) > 5 * 1024 * 1024:
            request.session['error_message'] = 'Logo file is too large. Maximum size is 5MB.'
            return RedirectResponse('/web/admin/site-settings', status_code=303)
        
        # Create uploads directory if it doesn't exist
        import os
        uploads_dir = os.path.join(os.getcwd(), 'app', 'uploads', 'branding')
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Generate unique filename
        import uuid
        from pathlib import Path
        file_extension = Path(logo.filename).suffix
        filename = f"logo_{uuid.uuid4().hex}{file_extension}"
        file_path = os.path.join(uploads_dir, filename)
        
        # Save file
        with open(file_path, 'wb') as f:
            f.write(logo_content)
        
        # Update workspace
        workspace = (await db.execute(
            select(Workspace).where(Workspace.id == user.workspace_id)
        )).scalar_one_or_none()
        
        if workspace:
            # Delete old logo if exists
            if hasattr(workspace, 'logo_url') and workspace.logo_url:
                old_path = os.path.join(os.getcwd(), 'app', workspace.logo_url.lstrip('/'))
                if os.path.exists(old_path):
                    os.remove(old_path)
            
            workspace.logo_url = f"/uploads/branding/{filename}"
            await db.commit()
            request.session['success_message'] = 'Logo uploaded successfully!'
        
    except Exception as e:
        if "no such column" in str(e):
            request.session['error_message'] = 'Database migration required. Run: python add_site_settings_columns.py'
        else:
            request.session['error_message'] = f'Failed to upload logo: {str(e)}'
    
    return RedirectResponse('/web/admin/site-settings', status_code=303)


# --------------------------
# My tasks view
# --------------------------
@router.get('/my-tasks', response_class=HTMLResponse)
async def web_my_tasks(
    request: Request,
    assignee_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session),
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)

    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)

    # Convert empty strings to None for integer filters
    assignee_id_int = int(assignee_id) if assignee_id and assignee_id.strip() else None
    project_id_int = int(project_id) if project_id and project_id.strip() else None

    # Non-admin: only tasks assigned to me with optional filters
    if not user.is_admin:
        stmt = (
            select(Task)
            .join(Project, Task.project_id == Project.id)
            .join(Assignment, Assignment.task_id == Task.id)
            .where(Assignment.assignee_id == user_id, Project.workspace_id == user.workspace_id)
        )
        if status and status.strip():
            try:
                st = TaskStatus(status)
                stmt = stmt.where(Task.status == st)
            except Exception:
                pass
        if project_id_int:
            stmt = stmt.where(Task.project_id == project_id_int)
        stmt = stmt.order_by(Task.created_at.desc())
        tasks = (await db.execute(stmt)).scalars().all()
        # Non-admin users only see projects they're assigned to
        from app.models.project_member import ProjectMember
        projects = (
            await db.execute(
                select(Project)
                .join(ProjectMember, Project.id == ProjectMember.project_id)
                .where(
                    ProjectMember.user_id == user_id,
                    Project.workspace_id == user.workspace_id
                )
                .order_by(Project.name)
            )
        ).scalars().all()
        return templates.TemplateResponse(
            'tasks/my.html',
            {
                'request': request,
                'tasks': tasks,
                'is_admin': False,
                'projects': projects,
                'selected': {'status': status, 'project_id': project_id},
            },
        )

    # Admin: show all assigned tasks in the workspace, with assignees listed
    tasks_stmt = (
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .join(Assignment, Assignment.task_id == Task.id)
        .where(Project.workspace_id == user.workspace_id)
    )
    if assignee_id_int:
        tasks_stmt = tasks_stmt.where(Assignment.assignee_id == assignee_id_int)
    if status and status.strip():
        try:
            st = TaskStatus(status)
            tasks_stmt = tasks_stmt.where(Task.status == st)
        except Exception:
            pass
    if project_id_int:
        tasks_stmt = tasks_stmt.where(Task.project_id == project_id_int)
    tasks_stmt = tasks_stmt.order_by(Task.created_at.desc())
    tasks = (await db.execute(tasks_stmt)).scalars().all()

    # Build assignees map {task_id: ["Name or Email", ...]}
    assocs = (
        await db.execute(
            select(Assignment.task_id, User.full_name, User.email)
            .join(User, Assignment.assignee_id == User.id)
            .join(Task, Assignment.task_id == Task.id)
            .join(Project, Task.project_id == Project.id)
            .where(Project.workspace_id == user.workspace_id)
        )
    ).all()
    assignees_map: dict[int, list[str]] = {}
    for task_id_val, full_name, email in assocs:
        label = (full_name or '').strip() or email
        assignees_map.setdefault(task_id_val, []).append(label)

    users = (
        await db.execute(select(User).where(User.workspace_id == user.workspace_id, User.is_active == True).order_by(User.full_name, User.email))
    ).scalars().all()
    projects = (
        await db.execute(select(Project).where(Project.workspace_id == user.workspace_id).order_by(Project.name))
    ).scalars().all()
    return templates.TemplateResponse(
        'tasks/my.html',
        {
            'request': request,
            'tasks': tasks,
            'is_admin': True,
            'assignees_map': assignees_map,
            'users': users,
            'projects': projects,
            'selected': {'assignee_id': assignee_id, 'status': status, 'project_id': project_id},
        },
    )

# --------------------------
# Projects (minimal to enable navigation)
# --------------------------
@router.get('/projects', response_class=HTMLResponse)
async def web_projects(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Admin: see all projects in workspace (excluding archived)
    # Regular user: only see projects they're assigned to (excluding archived)
    if user.is_admin:
        result = await db.execute(
            select(Project)
            .where(Project.workspace_id == user.workspace_id, Project.is_archived == False)
            .order_by(Project.created_at.desc())
        )
        projects = result.scalars().all()
    else:
        from app.models.project_member import ProjectMember
        result = await db.execute(
            select(Project)
            .join(ProjectMember, Project.id == ProjectMember.project_id)
            .where(
                ProjectMember.user_id == user_id,
                Project.workspace_id == user.workspace_id,
                Project.is_archived == False
            )
            .order_by(Project.created_at.desc())
        )
        projects = result.scalars().all()
    
    # Get workspace for branding
    workspace = await get_workspace_for_user(user_id, db)
    
    return templates.TemplateResponse('projects/index.html', {
        'request': request, 
        'user': user,
        'projects': projects,
        'workspace': workspace
    })


@router.post('/projects/create')
async def web_projects_create(request: Request, name: str = Form(...), description: Optional[str] = Form(None), db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    project = Project(name=name, description=description, owner_id=user_id, workspace_id=user.workspace_id)
    db.add(project)
    await db.commit()
    await db.refresh(project)
    
    # Auto-assign the creator to the project
    from app.models.project_member import ProjectMember
    member = ProjectMember(project_id=project.id, user_id=user_id, assigned_by=user_id)
    db.add(member)
    await db.commit()
    
    return RedirectResponse('/web/projects', status_code=303)


@router.post('/projects/{project_id}/edit')
async def web_projects_edit(request: Request, project_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can edit projects
    if not user.is_admin:
        return RedirectResponse('/web/projects', status_code=303)
    
    # Get the project
    result = await db.execute(select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id))
    project = result.scalar_one_or_none()
    
    if project:
        form = await request.form()
        project.name = form.get('name', project.name)
        project.description = form.get('description') or None
        project.support_email = form.get('support_email') or None
        # IMAP settings for board email integration
        project.imap_host = form.get('imap_host') or None
        imap_port = form.get('imap_port')
        project.imap_port = int(imap_port) if imap_port else None
        project.imap_username = form.get('imap_username') or None
        # Only update password if provided (don't clear existing)
        imap_password = form.get('imap_password')
        if imap_password:
            project.imap_password = imap_password
        project.imap_use_ssl = form.get('imap_use_ssl') == 'on'
        await db.commit()
    
    return RedirectResponse('/web/projects', status_code=303)


@router.post('/projects/{project_id}/delete')
async def web_projects_delete(request: Request, project_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can delete projects
    if not user.is_admin:
        return RedirectResponse('/web/projects', status_code=303)
    
    # Get the project
    result = await db.execute(select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id))
    project = result.scalar_one_or_none()
    
    if project:
        await db.delete(project)
        await db.commit()
    
    return RedirectResponse('/web/projects', status_code=303)


@router.post('/projects/{project_id}/archive')
async def web_projects_archive(request: Request, project_id: int, db: AsyncSession = Depends(get_session)):
    """Archive a project - preserves all data, comments, and attachments"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can archive projects
    if not user.is_admin:
        return RedirectResponse('/web/projects', status_code=303)
    
    # Get the project
    result = await db.execute(select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id))
    project = result.scalar_one_or_none()
    
    if project:
        from datetime import datetime
        project.is_archived = True
        project.archived_at = datetime.utcnow()
        await db.commit()
    
    return RedirectResponse('/web/projects', status_code=303)


@router.post('/projects/{project_id}/restore')
async def web_projects_restore(request: Request, project_id: int, db: AsyncSession = Depends(get_session)):
    """Restore an archived project"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can restore projects
    if not user.is_admin:
        return RedirectResponse('/web/projects/archived', status_code=303)
    
    # Get the project
    result = await db.execute(select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id))
    project = result.scalar_one_or_none()
    
    if project:
        project.is_archived = False
        project.archived_at = None
        await db.commit()
    
    return RedirectResponse('/web/projects/archived', status_code=303)


@router.get('/projects/archived', response_class=HTMLResponse)
async def web_projects_archived(request: Request, db: AsyncSession = Depends(get_session)):
    """View archived projects"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get archived projects
    if user.is_admin:
        result = await db.execute(
            select(Project)
            .where(Project.workspace_id == user.workspace_id, Project.is_archived == True)
            .order_by(Project.archived_at.desc())
        )
        projects = result.scalars().all()
    else:
        from app.models.project_member import ProjectMember
        result = await db.execute(
            select(Project)
            .join(ProjectMember, Project.id == ProjectMember.project_id)
            .where(
                ProjectMember.user_id == user_id,
                Project.workspace_id == user.workspace_id,
                Project.is_archived == True
            )
            .order_by(Project.archived_at.desc())
        )
        projects = result.scalars().all()
    
    return templates.TemplateResponse('projects/archived.html', {
        'request': request, 
        'user': user,
        'projects': projects
    })


# --------------------------
# Project Members (Admin only)
# --------------------------
@router.get('/projects/{project_id}/members', response_class=HTMLResponse)
async def web_project_members(request: Request, project_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail='Admin access required')
    
    # Get project
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    
    # Get assigned members (exclude deleted users)
    from app.models.project_member import ProjectMember
    assigned_members = (await db.execute(
        select(User, ProjectMember.assigned_at)
        .join(ProjectMember, User.id == ProjectMember.user_id)
        .where(ProjectMember.project_id == project_id)
        .order_by(User.full_name, User.email)
    )).all()
    
    # Get all active workspace users for assignment dropdown
    all_users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id)
        .where(User.is_active == True)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    # Filter out already assigned users
    assigned_user_ids = {m[0].id for m in assigned_members}
    available_users = [u for u in all_users if u.id not in assigned_user_ids]
    
    return templates.TemplateResponse('projects/members.html', {
        'request': request,
        'user': user,
        'project': project,
        'assigned_members': assigned_members,
        'available_users': available_users
    })


@router.post('/projects/{project_id}/members/add')
async def web_project_members_add(
    request: Request, 
    project_id: int, 
    user_identifier: str = Form(...),  # Changed from user_id_to_add to user_identifier 
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail='Admin access required')
    
    # Verify project exists in workspace
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    
    # Find user by username or email in the same workspace
    user_identifier = user_identifier.strip()
    target_user = (await db.execute(
        select(User).where(
            User.workspace_id == user.workspace_id,
            (User.username == user_identifier) | (User.email == user_identifier)
        )
    )).scalar_one_or_none()
    
    if not target_user:
        # User not found - return with error message
        request.session['error_message'] = f'User with username or email "{user_identifier}" not found in your workspace'
        return RedirectResponse(f'/web/projects/{project_id}/members', status_code=303)
    
    # Check if already assigned
    from app.models.project_member import ProjectMember
    existing = (await db.execute(
        select(ProjectMember).where(
            ProjectMember.project_id == project_id,
            ProjectMember.user_id == target_user.id
        )
    )).scalar_one_or_none()
    
    if existing:
        request.session['info_message'] = f'User {target_user.full_name or target_user.email} is already assigned to this project'
    else:
        member = ProjectMember(project_id=project_id, user_id=target_user.id, assigned_by=user_id)
        db.add(member)
        await db.commit()
        request.session['success_message'] = f'Successfully added {target_user.full_name or target_user.email} to the project'
    
    return RedirectResponse(f'/web/projects/{project_id}/members', status_code=303)


@router.post('/projects/{project_id}/members/{member_user_id}/remove')
async def web_project_members_remove(
    request: Request, 
    project_id: int, 
    member_user_id: int, 
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        raise HTTPException(status_code=403, detail='Admin access required')
    
    # Remove the assignment
    from app.models.project_member import ProjectMember
    member = (await db.execute(
        select(ProjectMember).where(
            ProjectMember.project_id == project_id,
            ProjectMember.user_id == member_user_id
        )
    )).scalar_one_or_none()
    
    if member:
        await db.delete(member)
        await db.commit()
    
    return RedirectResponse(f'/web/projects/{project_id}/members', status_code=303)


# Project Report
@router.get('/projects/{project_id}/report', response_class=HTMLResponse)
async def web_project_report(
    request: Request, 
    project_id: int, 
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session)
):
    """Project report showing who did what tasks and time spent"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get project
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    
    # Only admins can access reports
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Admin access required')
    
    # Parse date range (default to last 30 days)
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
        start_date = start_dt.strftime('%Y-%m-%d')
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
        end_date = (end_dt - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Get all tasks for the project
    all_tasks = (await db.execute(
        select(Task).where(Task.project_id == project_id)
    )).scalars().all()
    
    # Get tasks in date range
    tasks_in_range = (await db.execute(
        select(Task).where(
            Task.project_id == project_id,
            Task.created_at >= start_dt,
            Task.created_at < end_dt
        ).order_by(Task.created_at.desc())
    )).scalars().all()
    
    # Get all assignments
    task_ids = [t.id for t in all_tasks]
    assignments = []
    if task_ids:
        assignments = (await db.execute(
            select(Assignment, User)
            .join(User, User.id == Assignment.assignee_id)
            .where(Assignment.task_id.in_(task_ids))
        )).all()
    
    # Build assignees map
    assignees_map = {}  # task_id -> [(name, email), ...]
    for assignment, assignee_user in assignments:
        if assignment.task_id not in assignees_map:
            assignees_map[assignment.task_id] = []
        assignees_map[assignment.task_id].append((
            assignee_user.full_name or assignee_user.username,
            assignee_user.email
        ))
    
    # Get task history for completion dates
    task_completions = {}  # task_id -> completion datetime
    if task_ids:
        completions = (await db.execute(
            select(TaskHistory)
            .where(
                TaskHistory.task_id.in_(task_ids),
                TaskHistory.field == 'status',
                TaskHistory.new_value == 'done'
            )
            .order_by(TaskHistory.created_at.desc())
        )).scalars().all()
        
        for completion in completions:
            if completion.task_id not in task_completions:
                task_completions[completion.task_id] = completion.created_at
    
    # Calculate contributor stats
    contributor_stats = {}  # user_id -> stats
    
    for assignment, assignee_user in assignments:
        uid = assignee_user.id
        if uid not in contributor_stats:
            contributor_stats[uid] = {
                'user': assignee_user,
                'name': assignee_user.full_name or assignee_user.username,
                'email': assignee_user.email,
                'initials': ''.join([n[0].upper() for n in (assignee_user.full_name or assignee_user.email).split()[:2]]),
                'tasks_assigned': 0,
                'tasks_completed': 0,
                'hours_logged': 0.0
            }
        
        # Find the task
        task = next((t for t in all_tasks if t.id == assignment.task_id), None)
        if task:
            contributor_stats[uid]['tasks_assigned'] += 1
            if task.status == TaskStatus.done:
                contributor_stats[uid]['tasks_completed'] += 1
            if task.time_spent_hours:
                contributor_stats[uid]['hours_logged'] += task.time_spent_hours
    
    # Format contributor data
    contributors = []
    for uid, stats in contributor_stats.items():
        completion_rate = round((stats['tasks_completed'] / stats['tasks_assigned'] * 100) if stats['tasks_assigned'] > 0 else 0)
        avg_hours = round(stats['hours_logged'] / stats['tasks_completed'], 1) if stats['tasks_completed'] > 0 else 0
        contributors.append({
            'name': stats['name'],
            'email': stats['email'],
            'initials': stats['initials'],
            'tasks_assigned': stats['tasks_assigned'],
            'tasks_completed': stats['tasks_completed'],
            'completion_rate': completion_rate,
            'hours_logged': round(stats['hours_logged'], 1),
            'avg_hours_per_task': avg_hours
        })
    
    # Sort by tasks completed (descending)
    contributors.sort(key=lambda x: x['tasks_completed'], reverse=True)
    
    # Build task details
    tasks_detail = []
    for task in all_tasks:
        completed_at = task_completions.get(task.id)
        duration_days = None
        if completed_at:
            duration_days = (completed_at - task.created_at).days
        
        tasks_detail.append({
            'id': task.id,
            'title': task.title,
            'description': task.description,
            'status': task.status.value if hasattr(task.status, 'value') else task.status,
            'priority': task.priority.value if hasattr(task.priority, 'value') else task.priority,
            'assignees': [a[0] for a in assignees_map.get(task.id, [])],
            'created_at': task.created_at,
            'completed_at': completed_at,
            'time_spent_hours': task.time_spent_hours,
            'duration_days': duration_days
        })
    
    # Sort by created_at descending
    tasks_detail.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Get activity timeline from TaskHistory
    activities = []
    if task_ids:
        history_entries = (await db.execute(
            select(TaskHistory, User, Task)
            .join(User, User.id == TaskHistory.editor_id)
            .join(Task, Task.id == TaskHistory.task_id)
            .where(
                TaskHistory.task_id.in_(task_ids),
                TaskHistory.created_at >= start_dt,
                TaskHistory.created_at < end_dt
            )
            .order_by(TaskHistory.created_at.desc())
            .limit(50)
        )).all()
        
        for history, editor, task in history_entries:
            action = 'updated'
            description = f'updated {history.field}'
            
            if history.field == 'created':
                action = 'created'
                description = 'created the task'
            elif history.field == 'status':
                action = 'status_changed'
                if history.new_value == 'done':
                    action = 'completed'
                    description = 'completed the task'
                else:
                    description = f'changed status from {history.old_value or "none"} to {history.new_value}'
            elif history.field == 'priority':
                description = f'changed priority from {history.old_value or "none"} to {history.new_value}'
            elif history.field == 'assignee':
                description = f'changed assignee'
            
            activities.append({
                'action': action,
                'description': description,
                'user_name': editor.full_name or editor.username,
                'task_title': task.title,
                'created_at': history.created_at
            })
    
    # Calculate summary
    total_tasks = len(all_tasks)
    completed_tasks = sum(1 for t in all_tasks if t.status == TaskStatus.done)
    completion_rate = round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0)
    total_hours = round(sum(t.time_spent_hours or 0 for t in all_tasks), 1)
    active_contributors = len(contributor_stats)
    
    summary = {
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'completion_rate': completion_rate,
        'total_hours': total_hours,
        'active_contributors': active_contributors
    }
    
    # Get workspace for timezone
    workspace = (await db.execute(
        select(Workspace).where(Workspace.id == user.workspace_id)
    )).scalar_one_or_none()
    
    return templates.TemplateResponse('projects/report.html', {
        'request': request,
        'user': user,
        'project': project,
        'workspace': workspace,
        'start_date': start_date,
        'end_date': end_date,
        'summary': summary,
        'contributors': contributors,
        'tasks_detail': tasks_detail,
        'activities': activities
    })


@router.get('/projects/{project_id}/report/pdf')
async def web_project_report_pdf(
    request: Request, 
    project_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session)
):
    """Generate PDF report for project"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can access reports
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Admin access required')
    
    # Get project
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id)
    )).scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    # Get all tasks for the project
    all_tasks = (await db.execute(
        select(Task).where(Task.project_id == project_id)
    )).scalars().all()
    
    # Get assignments
    task_ids = [t.id for t in all_tasks]
    assignments = []
    if task_ids:
        assignments = (await db.execute(
            select(Assignment, User)
            .join(User, User.id == Assignment.assignee_id)
            .where(Assignment.task_id.in_(task_ids))
        )).all()
    
    # Build assignees map
    assignees_map = {}
    for assignment, assignee_user in assignments:
        if assignment.task_id not in assignees_map:
            assignees_map[assignment.task_id] = []
        assignees_map[assignment.task_id].append(assignee_user.full_name or assignee_user.username)
    
    # Get completion dates
    task_completions = {}
    if task_ids:
        completions = (await db.execute(
            select(TaskHistory)
            .where(
                TaskHistory.task_id.in_(task_ids),
                TaskHistory.field == 'status',
                TaskHistory.new_value == 'done'
            )
            .order_by(TaskHistory.created_at.desc())
        )).scalars().all()
        
        for completion in completions:
            if completion.task_id not in task_completions:
                task_completions[completion.task_id] = completion.created_at
    
    # Calculate contributor stats
    contributor_stats = {}
    for assignment, assignee_user in assignments:
        uid = assignee_user.id
        if uid not in contributor_stats:
            contributor_stats[uid] = {
                'name': assignee_user.full_name or assignee_user.username,
                'tasks_assigned': 0,
                'tasks_completed': 0,
                'hours_logged': 0.0
            }
        
        task = next((t for t in all_tasks if t.id == assignment.task_id), None)
        if task:
            contributor_stats[uid]['tasks_assigned'] += 1
            if task.status == TaskStatus.done:
                contributor_stats[uid]['tasks_completed'] += 1
            if task.time_spent_hours:
                contributor_stats[uid]['hours_logged'] += task.time_spent_hours
    
    # Generate PDF
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    import io
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=20,
        alignment=TA_CENTER,
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20,
    )
    
    # Title
    elements.append(Paragraph(f"Project Report: {project.name}", title_style))
    elements.append(Paragraph(
        f"Period: {start_dt.strftime('%B %d, %Y')} - {(end_dt - timedelta(days=1)).strftime('%B %d, %Y')}",
        styles['Normal']
    ))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    total_tasks = len(all_tasks)
    completed_tasks = sum(1 for t in all_tasks if t.status == TaskStatus.done)
    completion_rate = round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0)
    total_hours = round(sum(t.time_spent_hours or 0 for t in all_tasks), 1)
    
    elements.append(Paragraph("Summary", heading_style))
    summary_data = [
        ['Total Tasks', 'Completed', 'Completion Rate', 'Total Hours'],
        [str(total_tasks), str(completed_tasks), f'{completion_rate}%', f'{total_hours}h']
    ]
    summary_table = Table(summary_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.white),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Contributor Performance
    elements.append(Paragraph("Contributor Performance", heading_style))
    contrib_headers = ['Contributor', 'Assigned', 'Completed', 'Rate', 'Hours']
    contrib_data = [contrib_headers]
    
    for uid, stats in sorted(contributor_stats.items(), key=lambda x: x[1]['tasks_completed'], reverse=True):
        rate = round((stats['tasks_completed'] / stats['tasks_assigned'] * 100) if stats['tasks_assigned'] > 0 else 0)
        contrib_data.append([
            stats['name'],
            str(stats['tasks_assigned']),
            str(stats['tasks_completed']),
            f'{rate}%',
            f"{round(stats['hours_logged'], 1)}h"
        ])
    
    if len(contrib_data) > 1:
        contrib_table = Table(contrib_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        contrib_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(contrib_table)
    else:
        elements.append(Paragraph("No contributors found.", styles['Normal']))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Task Details
    elements.append(Paragraph("Task Details", heading_style))
    task_headers = ['Task', 'Assignees', 'Status', 'Priority', 'Time']
    task_data = [task_headers]
    
    for task in sorted(all_tasks, key=lambda t: t.created_at, reverse=True):
        assignee_names = ', '.join(assignees_map.get(task.id, ['Unassigned']))
        if len(assignee_names) > 25:
            assignee_names = assignee_names[:22] + '...'
        
        title = task.title
        if len(title) > 35:
            title = title[:32] + '...'
        
        status = task.status.value if hasattr(task.status, 'value') else str(task.status)
        priority = task.priority.value if hasattr(task.priority, 'value') else str(task.priority)
        time_str = f"{task.time_spent_hours}h" if task.time_spent_hours else '-'
        
        task_data.append([title, assignee_names, status.title(), priority.title(), time_str])
    
    if len(task_data) > 1:
        task_table = Table(task_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch, 0.7*inch])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(task_table)
    else:
        elements.append(Paragraph("No tasks found.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return PDF
    from fastapi.responses import StreamingResponse
    filename = f"project_report_{project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# Project detail + simple Kanban
@router.get('/projects/{project_id}', response_class=HTMLResponse)
async def web_project_detail(request: Request, project_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    result = await db.execute(select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    
    # Check if user has access to this project (admin or assigned member)
    if not user.is_admin:
        from app.models.project_member import ProjectMember
        member = (await db.execute(
            select(ProjectMember).where(
                ProjectMember.project_id == project_id,
                ProjectMember.user_id == user_id
            )
        )).scalar_one_or_none()
        if not member:
            raise HTTPException(status_code=403, detail='You do not have access to this project')
    
    # Fetch only non-archived tasks for board view (archived tasks go to the Done tab in tasks/list)
    tasks_result = await db.execute(
        select(Task).where(Task.project_id == project_id, Task.is_archived == False)
    )
    tasks = tasks_result.scalars().all()
    
    # Organize tasks by status for kanban columns
    columns = {
        TaskStatus.todo: [],
        TaskStatus.in_progress: [],
        TaskStatus.blocked: [],
        TaskStatus.done: []
    }
    for t in tasks:
        if t.status in columns:
            columns[t.status].append(t)
    
    # Build assignees map: {task_id: [(full_name, email), ...]}
    assignees_map = {}
    if tasks:
        task_ids = [t.id for t in tasks]
        assignments = (await db.execute(
            select(Assignment.task_id, User.full_name, User.email)
            .join(User, User.id == Assignment.assignee_id)
            .where(Assignment.task_id.in_(task_ids))
        )).all()
        for task_id, full_name, email in assignments:
            if task_id not in assignees_map:
                assignees_map[task_id] = []
            assignees_map[task_id].append((full_name or email, email))
    
    # Fetch subtasks for all tasks
    from app.models.subtask import Subtask
    subtasks_map = {}  # {task_id: [subtasks]}
    subtask_stats = {}  # {task_id: {'total': x, 'completed': y}}
    if tasks:
        task_ids = [t.id for t in tasks]
        subtasks = (await db.execute(
            select(Subtask).where(Subtask.task_id.in_(task_ids)).order_by(Subtask.order)
        )).scalars().all()
        
        for subtask in subtasks:
            if subtask.task_id not in subtasks_map:
                subtasks_map[subtask.task_id] = []
                subtask_stats[subtask.task_id] = {'total': 0, 'completed': 0}
            subtasks_map[subtask.task_id].append(subtask)
            subtask_stats[subtask.task_id]['total'] += 1
            if subtask.is_completed:
                subtask_stats[subtask.task_id]['completed'] += 1
    
    # Fetch attachment counts for all tasks
    from app.models.task_extensions import TaskAttachment
    attachment_counts = {}  # {task_id: count}
    if tasks:
        task_ids = [t.id for t in tasks]
        from sqlalchemy import func
        attachment_result = await db.execute(
            select(TaskAttachment.task_id, func.count(TaskAttachment.id).label('count'))
            .where(TaskAttachment.task_id.in_(task_ids))
            .group_by(TaskAttachment.task_id)
        )
        for task_id, count in attachment_result.all():
            attachment_counts[task_id] = count
    
    # Fetch all active users in workspace for assignment dropdown
    users = (await db.execute(select(User).where(User.workspace_id == user.workspace_id, User.is_active == True).order_by(User.full_name, User.email))).scalars().all()
    return templates.TemplateResponse('projects/detail.html', {
        'request': request, 
        'project': project, 
        'tasks': tasks, 
        'TaskStatus': TaskStatus, 
        'columns': columns,
        'assignees_map': assignees_map,
        'subtasks_map': subtasks_map,
        'subtask_stats': subtask_stats,
        'attachment_counts': attachment_counts,
        'users': users,
        'user': user
    })


# Tasks
@router.post('/tasks/create')
async def web_task_create(request: Request, db: AsyncSession = Depends(get_session)):
    # Get all form data
    form_data = await request.form()
    
    # Extract form fields
    project_id = int(form_data.get('project_id'))
    title = form_data.get('title')
    description = form_data.get('description') or None
    subtasks = form_data.get('subtasks') or None
    priority = form_data.get('priority', 'medium')
    start_date_value = form_data.get('start_date_value') or None
    start_time_value = form_data.get('start_time_value') or None
    due_date_value = form_data.get('due_date_value') or None
    due_time_value = form_data.get('due_time_value') or None
    working_days_list = form_data.getlist('working_days')
    # Customer info (optional)
    customer_name = form_data.get('customer_name') or None
    customer_surname = form_data.get('customer_surname') or None
    customer_email = form_data.get('customer_email') or None
    customer_phone = form_data.get('customer_phone') or None
    
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    # Ensure project belongs to user's workspace
    project = (await db.execute(select(Project).where(Project.id == project_id, Project.workspace_id == user.workspace_id))).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    
    # Ensure user has access to this project (admin or assigned member)
    if not user.is_admin:
        from app.models.project_member import ProjectMember
        member = (await db.execute(
            select(ProjectMember).where(
                ProjectMember.project_id == project_id,
                ProjectMember.user_id == user_id
            )
        )).scalar_one_or_none()
        if not member:
            raise HTTPException(status_code=403, detail='You do not have access to this project')
    
    # Parse dates and times
    from datetime import date, time
    start_date = date.fromisoformat(start_date_value) if start_date_value else None
    start_time_obj = time.fromisoformat(start_time_value) if start_time_value else None
    due_date = date.fromisoformat(due_date_value) if due_date_value else None
    due_time_obj = time.fromisoformat(due_time_value) if due_time_value else None
    
    # Parse priority
    from app.models.enums import TaskPriority
    try:
        task_priority = TaskPriority(priority)
    except ValueError:
        task_priority = TaskPriority.medium
    
    # Parse working days (default to Mon-Fri if not provided)
    working_days_str = ','.join(working_days_list) if working_days_list else '0,1,2,3,4'
    
    task = Task(
        title=title, 
        description=description, 
        project_id=project_id,
        creator_id=user_id,
        priority=task_priority,
        start_date=start_date,
        start_time=start_time_obj,
        due_date=due_date,
        due_time=due_time_obj,
        working_days=working_days_str,
        customer_name=customer_name,
        customer_surname=customer_surname,
        customer_email=customer_email,
        customer_phone=customer_phone
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)
    
    # Create subtasks if provided
    if subtasks:
        from app.models.subtask import Subtask
        subtask_titles = [title.strip() for title in subtasks.split('\n') if title.strip()]
        for index, subtask_title in enumerate(subtask_titles):
            new_subtask = Subtask(
                task_id=task.id,
                title=subtask_title,
                is_completed=False,
                order=index
            )
            db.add(new_subtask)
        await db.commit()
    
    # Auto-assign task to creator if they're not an admin
    if not user.is_admin:
        from app.models.assignment import Assignment
        assignment = Assignment(task_id=task.id, assignee_id=user_id)
        db.add(assignment)
        await db.commit()
    
    # Track user behavior for learning
    try:
        from app.core.smart_suggestions import track_user_action
        await track_user_action(
            db=db,
            user_id=user_id,
            workspace_id=user.workspace_id,
            action_type="task_create",
            entity_type="task",
            entity_id=task.id,
            project_id=project_id,
            field_name="priority",
            field_value=priority
        )
    except Exception:
        pass  # Don't fail task creation if tracking fails
    
    return RedirectResponse(f'/web/projects/{project_id}', status_code=303)


@router.get('/tasks/list')
async def web_tasks_list(
    request: Request,
    tab: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    assignee_id: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from datetime import date
    
    # Convert empty strings to None for integer filters
    assignee_id_int = int(assignee_id) if assignee_id and assignee_id.strip() else None
    project_id_int = int(project_id) if project_id and project_id.strip() else None
    
    # Build query with filters
    # Admin: see all tasks in workspace
    # Regular user: only tasks from projects they're assigned to OR tasks they created
    if user.is_admin:
        query = (
            select(Task, Project.name.label('project_name'))
            .join(Project, Task.project_id == Project.id)
            .where(Project.workspace_id == user.workspace_id)
        )
    else:
        from app.models.project_member import ProjectMember
        # Get tasks where user is either assigned OR is the creator
        query = (
            select(Task, Project.name.label('project_name'))
            .join(Project, Task.project_id == Project.id)
            .join(ProjectMember, Project.id == ProjectMember.project_id)
            .where(
                ProjectMember.user_id == user_id,
                Project.workspace_id == user.workspace_id,
                or_(
                    Task.creator_id == user_id,  # Tasks created by the user
                    Task.id.in_(  # Tasks assigned to the user
                        select(Assignment.task_id)
                        .where(Assignment.assignee_id == user_id)
                    )
                )
            )
        )
    
    # Tab filtering - only apply if no specific status filter is set
    if not status or not status.strip():
        if tab == 'done':
            query = query.where(Task.status == 'done')
        else:  # Active tasks (default when tab is None, empty, or 'active')
            query = query.where(Task.status.in_(['todo', 'in_progress', 'blocked']))
    
    if status and status.strip():
        query = query.where(Task.status == status)
    if priority and priority.strip():
        query = query.where(Task.priority == priority)
    if project_id_int:
        query = query.where(Task.project_id == project_id_int)
    if assignee_id_int:
        query = query.join(Assignment, Task.id == Assignment.task_id).where(Assignment.assignee_id == assignee_id_int)
    
    query = query.order_by(Task.due_date.asc().nullslast(), Task.priority.desc())
    
    results = (await db.execute(query)).all()
    tasks = []
    project_names = {}
    for task, project_name in results:
        tasks.append(task)
        project_names[task.id] = project_name
    
    # Get assignees for all tasks
    task_ids = [t.id for t in tasks]
    assocs = (await db.execute(
        select(Assignment.task_id, User.full_name, User.email)
        .join(User, Assignment.assignee_id == User.id)
        .where(Assignment.task_id.in_(task_ids) if task_ids else False)
    )).all()
    
    assignees_map: dict[int, list[str]] = {}
    for task_id_val, full_name, email in assocs:
        label = (full_name or '').strip() or email
        assignees_map.setdefault(task_id_val, []).append(label)
    
    # Get all users and projects for filters
    users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id, User.is_active == True)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    # Get projects user has access to
    if user.is_admin:
        projects = (await db.execute(
            select(Project)
            .where(Project.workspace_id == user.workspace_id)
            .order_by(Project.name)
        )).scalars().all()
    else:
        from app.models.project_member import ProjectMember
        projects = (await db.execute(
            select(Project)
            .join(ProjectMember, Project.id == ProjectMember.project_id)
            .where(
                ProjectMember.user_id == user_id,
                Project.workspace_id == user.workspace_id
            )
            .order_by(Project.name)
        )).scalars().all()
    
    return templates.TemplateResponse('tasks/list.html', {
        'request': request,
        'user': user,
        'tasks': tasks,
        'project_names': project_names,
        'assignees_map': assignees_map,
        'users': users,
        'projects': projects,
        'selected': {
            'tab': tab or 'active',
            'status': status,
            'priority': priority,
            'assignee_id': assignee_id,
            'project_id': project_id,
        },
        'today': date.today(),
    })


@router.get('/tasks/{task_id}')
async def web_task_detail(request: Request, task_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get task with project check
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Get project
    project = (await db.execute(select(Project).where(Project.id == task.project_id))).scalar_one_or_none()
    
    # Get assignments
    assignments = (await db.execute(
        select(User)
        .join(Assignment, User.id == Assignment.assignee_id)
        .where(Assignment.task_id == task_id)
    )).scalars().all()
    
    # Get comments with authors
    comments = (await db.execute(
        select(Comment, User)
        .join(User, Comment.author_id == User.id)
        .where(Comment.task_id == task_id)
        .order_by(Comment.created_at.desc())
    )).all()
    
    # Get attachments for all comments
    comment_ids = [c[0].id for c in comments]
    attachments_by_comment = {}
    if comment_ids:
        attachments = (await db.execute(
            select(CommentAttachment).where(CommentAttachment.comment_id.in_(comment_ids))
        )).scalars().all()
        for attachment in attachments:
            if attachment.comment_id not in attachments_by_comment:
                attachments_by_comment[attachment.comment_id] = []
            attachments_by_comment[attachment.comment_id].append(attachment)
    
    # Check if user can comment (admin or assignee)
    is_assignee = any(a.id == user_id for a in assignments)
    can_comment = user.is_admin or is_assignee
    
    # Get edit history
    history = (await db.execute(
        select(TaskHistory, User)
        .join(User, TaskHistory.editor_id == User.id)
        .where(TaskHistory.task_id == task_id)
        .order_by(TaskHistory.created_at.desc())
    )).all()
    
    # Get all active workspace users
    users = (await db.execute(select(User).where(User.workspace_id == user.workspace_id, User.is_active == True).order_by(User.full_name, User.email))).scalars().all()
    
    # Get subtasks ordered by their order field
    from app.models.subtask import Subtask
    subtasks = (await db.execute(
        select(Subtask).where(Subtask.task_id == task_id).order_by(Subtask.order)
    )).scalars().all()
    
    # Calculate subtask completion stats
    total_subtasks = len(subtasks)
    completed_subtasks = sum(1 for st in subtasks if st.is_completed)
    completion_percentage = int((completed_subtasks / total_subtasks) * 100) if total_subtasks > 0 else 0
    
    # Get task dependencies (blocked by)
    from app.models.task_extensions import TaskDependency
    blocked_by_result = await db.execute(
        select(Task)
        .join(TaskDependency, Task.id == TaskDependency.depends_on_task_id)
        .where(TaskDependency.task_id == task_id)
    )
    blocked_by_tasks = blocked_by_result.scalars().all()
    
    # Get tasks that this task blocks
    blocks_result = await db.execute(
        select(Task)
        .join(TaskDependency, Task.id == TaskDependency.task_id)
        .where(TaskDependency.depends_on_task_id == task_id)
    )
    blocks_tasks = blocks_result.scalars().all()
    
    # Get watchers
    from app.models.task_extensions import TaskWatcher
    watchers_result = await db.execute(
        select(User)
        .join(TaskWatcher, User.id == TaskWatcher.user_id)
        .where(TaskWatcher.task_id == task_id)
    )
    watchers = watchers_result.scalars().all()
    
    # Check if current user is watching
    is_watching = any(w.id == user_id for w in watchers)
    
    # Get available tasks for dependency selection (same project, not this task)
    available_tasks = (await db.execute(
        select(Task)
        .where(Task.project_id == task.project_id, Task.id != task_id, Task.is_archived == False)
        .order_by(Task.title)
    )).scalars().all()
    
    # Get task-level attachments (not comment attachments)
    from app.models.task_extensions import TaskAttachment
    task_attachments = (await db.execute(
        select(TaskAttachment).where(TaskAttachment.task_id == task_id).order_by(TaskAttachment.created_at.desc())
    )).scalars().all()
    
    return templates.TemplateResponse('tasks/detail.html', {
        'request': request,
        'task': task,
        'project': project,
        'assignments': assignments,
        'comments': comments,
        'attachments_by_comment': attachments_by_comment,
        'task_attachments': task_attachments,
        'can_comment': can_comment,
        'history': history,
        'users': users,
        'user': user,
        'subtasks': subtasks,
        'total_subtasks': total_subtasks,
        'completed_subtasks': completed_subtasks,
        'completion_percentage': completion_percentage,
        'blocked_by_tasks': blocked_by_tasks,
        'blocks_tasks': blocks_tasks,
        'watchers': watchers,
        'is_watching': is_watching,
        'available_tasks': available_tasks,
        'TaskStatus': TaskStatus,
        'TaskPriority': TaskPriority
    })


@router.post('/tasks/{task_id}/update')
async def web_task_update(
    request: Request,
    task_id: int,
    title: str = Form(...),
    description: Optional[str] = Form(None),
    status: str = Form(...),
    priority: str = Form(...),
    start_date_value: Optional[str] = Form(None),
    start_time_value: Optional[str] = Form(None),
    due_date_value: Optional[str] = Form(None),
    due_time_value: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get task
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Check if task is archived - only admins can edit archived tasks
    if task.is_archived and not user.is_admin:
        raise HTTPException(status_code=403, detail='This task is archived. Only admins can modify it.')
    
    # Check permission: Admin OR assigned to this task
    if not user.is_admin:
        from app.models.assignment import Assignment
        assignment = (await db.execute(
            select(Assignment).where(
                Assignment.task_id == task_id,
                Assignment.assignee_id == user_id
            )
        )).scalar_one_or_none()
        if not assignment:
            raise HTTPException(status_code=403, detail='You can only edit tasks assigned to you')
    
    # Track changes
    from datetime import date, time
    changes = []
    
    if task.title != title:
        changes.append(('title', task.title, title))
        task.title = title
    
    if task.description != description:
        changes.append(('description', task.description or '', description or ''))
        task.description = description
    
    new_status = TaskStatus(status)
    task_just_completed = False
    if task.status != new_status:
        changes.append(('status', task.status.value, new_status.value))
        old_status_value = task.status.value
        task.status = new_status
        
        # Auto-archive when moved to done
        if new_status.value == 'done':
            task.is_archived = True
            task.archived_at = datetime.utcnow()
            task_just_completed = True  # Flag to send completion notification
        elif old_status_value == 'done' and new_status.value != 'done':
            # Unarchive if moved out of done
            task.is_archived = False
            task.archived_at = None
    
    new_priority = TaskPriority(priority)
    if task.priority != new_priority:
        changes.append(('priority', task.priority.value, new_priority.value))
        task.priority = new_priority
    
    new_start_date = date.fromisoformat(start_date_value) if start_date_value else None
    if task.start_date != new_start_date:
        changes.append(('start_date', str(task.start_date) if task.start_date else '', str(new_start_date) if new_start_date else ''))
        task.start_date = new_start_date
    
    new_start_time = time.fromisoformat(start_time_value) if start_time_value else None
    if task.start_time != new_start_time:
        changes.append(('start_time', str(task.start_time) if task.start_time else '', str(new_start_time) if new_start_time else ''))
        task.start_time = new_start_time
    
    new_due_date = date.fromisoformat(due_date_value) if due_date_value else None
    if task.due_date != new_due_date:
        changes.append(('due_date', str(task.due_date) if task.due_date else '', str(new_due_date) if new_due_date else ''))
        task.due_date = new_due_date
    
    new_due_time = time.fromisoformat(due_time_value) if due_time_value else None
    if task.due_time != new_due_time:
        changes.append(('due_time', str(task.due_time) if task.due_time else '', str(new_due_time) if new_due_time else ''))
        task.due_time = new_due_time
    
    # Save history
    for field, old_value, new_value in changes:
        history_entry = TaskHistory(
            task_id=task_id,
            editor_id=user_id,
            field=field,
            old_value=old_value,
            new_value=new_value
        )
        db.add(history_entry)
    
    await db.commit()
    
    # Send completion notification email if task was just marked as done
    if task_just_completed:
        try:
            # Get project for additional details
            project = (await db.execute(select(Project).where(Project.id == task.project_id))).scalar_one_or_none()
            additional_details = f"Project: {project.name}" if project else ""
            
            await send_completion_notification_email(
                db=db,
                workspace_id=user.workspace_id,
                notification_type='task',
                item_id=str(task_id),
                title=task.title,
                status='Done',
                priority=task.priority.value.title(),
                completed_by_name=user.full_name or user.username,
                created_at=task.created_at,
                completed_at=task.archived_at or datetime.utcnow(),
                additional_details=additional_details
            )
        except Exception as e:
            logger.error(f"Failed to send task completion notification: {e}")
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/complete-with-details')
async def web_task_complete_with_details(
    request: Request,
    task_id: int,
    billable_traveling: Optional[str] = Form(None),
    billable_labour_onsite: Optional[str] = Form(None),
    billable_remote_labour: Optional[str] = Form(None),
    billable_equipment_used: Optional[str] = Form(None),
    non_billable_traveling: Optional[str] = Form(None),
    non_billable_labour_onsite: Optional[str] = Form(None),
    non_billable_remote_labour: Optional[str] = Form(None),
    non_billable_equipment_used: Optional[str] = Form(None),
    completion_notes: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Complete task with optional billing/work details"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return RedirectResponse('/web/login', status_code=303)
    
    # Get task and verify access
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(
        Task.id == task_id,
        Project.workspace_id == user.workspace_id
    )
    task = (await db.execute(stmt)).scalar_one_or_none()
    
    if not task:
        return RedirectResponse('/web/projects', status_code=303)
    
    # Record old status for history
    old_status = task.status.value
    
    # Update task to done
    task.status = TaskStatus.done
    task.updated_at = datetime.utcnow()
    task.is_archived = True
    task.archived_at = datetime.utcnow()
    
    # Save billing details
    task.billable_traveling = billable_traveling.strip() if billable_traveling and billable_traveling.strip() else None
    task.billable_labour_onsite = billable_labour_onsite.strip() if billable_labour_onsite and billable_labour_onsite.strip() else None
    task.billable_remote_labour = billable_remote_labour.strip() if billable_remote_labour and billable_remote_labour.strip() else None
    task.billable_equipment_used = billable_equipment_used.strip() if billable_equipment_used and billable_equipment_used.strip() else None
    task.non_billable_traveling = non_billable_traveling.strip() if non_billable_traveling and non_billable_traveling.strip() else None
    task.non_billable_labour_onsite = non_billable_labour_onsite.strip() if non_billable_labour_onsite and non_billable_labour_onsite.strip() else None
    task.non_billable_remote_labour = non_billable_remote_labour.strip() if non_billable_remote_labour and non_billable_remote_labour.strip() else None
    task.non_billable_equipment_used = non_billable_equipment_used.strip() if non_billable_equipment_used and non_billable_equipment_used.strip() else None
    task.completion_notes = completion_notes.strip() if completion_notes and completion_notes.strip() else None
    
    # Add history
    history_entry = TaskHistory(
        task_id=task_id,
        editor_id=user_id,
        field='status',
        old_value=old_status,
        new_value='done'
    )
    db.add(history_entry)
    
    await db.commit()
    
    # Build billing details string for email
    billing_details = []
    
    # Billable items
    billable_items = []
    if task.billable_traveling:
        billable_items.append(f"  - Traveling: {task.billable_traveling}")
    if task.billable_labour_onsite:
        billable_items.append(f"  - Labour Onsite: {task.billable_labour_onsite}")
    if task.billable_remote_labour:
        billable_items.append(f"  - Remote Labour: {task.billable_remote_labour}")
    if task.billable_equipment_used:
        billable_items.append(f"  - Equipment Used: {task.billable_equipment_used}")
    
    if billable_items:
        billing_details.append("BILLABLE:")
        billing_details.extend(billable_items)
    
    # Non-billable items
    non_billable_items = []
    if task.non_billable_traveling:
        non_billable_items.append(f"  - Traveling: {task.non_billable_traveling}")
    if task.non_billable_labour_onsite:
        non_billable_items.append(f"  - Labour Onsite: {task.non_billable_labour_onsite}")
    if task.non_billable_remote_labour:
        non_billable_items.append(f"  - Remote Labour: {task.non_billable_remote_labour}")
    if task.non_billable_equipment_used:
        non_billable_items.append(f"  - Equipment Used: {task.non_billable_equipment_used}")
    
    if non_billable_items:
        if billing_details:
            billing_details.append("")  # Empty line separator
        billing_details.append("NON-BILLABLE:")
        billing_details.extend(non_billable_items)
    
    # Completion notes
    if task.completion_notes:
        if billing_details:
            billing_details.append("")  # Empty line separator
        billing_details.append(f"COMPLETION NOTES:\n{task.completion_notes}")
    
    # Build additional details for email
    project = (await db.execute(select(Project).where(Project.id == task.project_id))).scalar_one_or_none()
    additional_details_parts = []
    if project:
        additional_details_parts.append(f"Project: {project.name}")
    if billing_details:
        additional_details_parts.append("\n" + "\n".join(billing_details))
    
    additional_details = "\n".join(additional_details_parts)
    
    # Send completion notification email
    try:
        await send_completion_notification_email(
            db=db,
            workspace_id=user.workspace_id,
            notification_type='task',
            item_id=str(task_id),
            title=task.title,
            status='Done',
            priority=task.priority.value.title(),
            completed_by_name=user.full_name or user.username,
            created_at=task.created_at,
            completed_at=task.archived_at,
            additional_details=additional_details
        )
    except Exception as e:
        logger.error(f"Failed to send task completion notification: {e}")
    
    # Check if this is an AJAX request
    if request.headers.get('accept', '').find('application/json') != -1 or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        from fastapi.responses import JSONResponse
        return JSONResponse({
            'success': True, 
            'task_id': task_id,
            'title': task.title,
            'status': 'done'
        })
    
    request.session['success_message'] = f'Task "{task.title}" has been completed.'
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/subtasks')
async def web_task_add_subtasks(
    request: Request,
    task_id: int,
    subtasks: str = Form(...),  # Newline-separated list of subtask titles
    db: AsyncSession = Depends(get_session)
):
    """Add multiple subtasks to a task at once."""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get task and verify access
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(
        Task.id == task_id, 
        Project.workspace_id == user.workspace_id
    )
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Check if task is archived
    if task.is_archived and not user.is_admin:
        raise HTTPException(status_code=403, detail='Cannot add subtasks to archived tasks')
    
    # Check permission: Admin OR assigned to this task
    if not user.is_admin:
        from app.models.assignment import Assignment
        assignment = (await db.execute(
            select(Assignment).where(
                Assignment.task_id == task_id,
                Assignment.assignee_id == user_id
            )
        )).scalar_one_or_none()
        if not assignment:
            raise HTTPException(status_code=403, detail='You can only add subtasks to tasks assigned to you')
    
    # Get current max order
    from app.models.subtask import Subtask
    max_order_result = await db.execute(
        select(Subtask).where(Subtask.task_id == task_id).order_by(Subtask.order.desc()).limit(1)
    )
    max_order_subtask = max_order_result.scalar_one_or_none()
    current_order = max_order_subtask.order if max_order_subtask else -1
    
    # Parse and create subtasks
    subtask_titles = [title.strip() for title in subtasks.split('\n') if title.strip()]
    
    for title in subtask_titles:
        current_order += 1
        new_subtask = Subtask(
            task_id=task_id,
            title=title,
            is_completed=False,
            order=current_order
        )
        db.add(new_subtask)
    
    await db.commit()
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/subtasks/{subtask_id}/toggle')
async def web_task_toggle_subtask(
    request: Request,
    task_id: int,
    subtask_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Toggle a subtask's completion status."""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'error': 'User not found'}, status_code=401)
    
    # Get subtask and verify it belongs to the task
    from app.models.subtask import Subtask
    subtask = (await db.execute(
        select(Subtask).where(Subtask.id == subtask_id, Subtask.task_id == task_id)
    )).scalar_one_or_none()
    
    if not subtask:
        return JSONResponse({'error': 'Subtask not found'}, status_code=404)
    
    # Get task and verify access
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(
        Task.id == task_id,
        Project.workspace_id == user.workspace_id
    )
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        return JSONResponse({'error': 'Task not found'}, status_code=404)
    
    # Check if task is archived
    if task.is_archived and not user.is_admin:
        return JSONResponse({'error': 'Cannot modify subtasks in archived tasks'}, status_code=403)
    
    # Check permission: Admin OR assigned to this task
    if not user.is_admin:
        from app.models.assignment import Assignment
        assignment = (await db.execute(
            select(Assignment).where(
                Assignment.task_id == task_id,
                Assignment.assignee_id == user_id
            )
        )).scalar_one_or_none()
        if not assignment:
            return JSONResponse({'error': 'Permission denied'}, status_code=403)
    
    # Toggle completion
    from datetime import datetime
    subtask.is_completed = not subtask.is_completed
    subtask.completed_at = datetime.utcnow() if subtask.is_completed else None
    
    await db.commit()
    
    # Calculate completion percentage
    all_subtasks = (await db.execute(
        select(Subtask).where(Subtask.task_id == task_id)
    )).scalars().all()
    
    total = len(all_subtasks)
    completed = sum(1 for st in all_subtasks if st.is_completed)
    percentage = int((completed / total) * 100) if total > 0 else 0
    
    return JSONResponse({
        'success': True,
        'is_completed': subtask.is_completed,
        'completed_at': subtask.completed_at.isoformat() if subtask.completed_at else None,
        'total_subtasks': total,
        'completed_subtasks': completed,
        'completion_percentage': percentage
    })


@router.delete('/tasks/{task_id}/subtasks/{subtask_id}')
async def web_task_delete_subtask(
    request: Request,
    task_id: int,
    subtask_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Delete a subtask."""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'error': 'User not found'}, status_code=401)
    
    # Get subtask and verify it belongs to the task
    from app.models.subtask import Subtask
    subtask = (await db.execute(
        select(Subtask).where(Subtask.id == subtask_id, Subtask.task_id == task_id)
    )).scalar_one_or_none()
    
    if not subtask:
        return JSONResponse({'error': 'Subtask not found'}, status_code=404)
    
    # Get task and verify access
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(
        Task.id == task_id,
        Project.workspace_id == user.workspace_id
    )
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        return JSONResponse({'error': 'Task not found'}, status_code=404)
    
    # Check if task is archived
    if task.is_archived and not user.is_admin:
        return JSONResponse({'error': 'Cannot delete subtasks from archived tasks'}, status_code=403)
    
    # Check permission: Admin OR assigned to this task
    if not user.is_admin:
        from app.models.assignment import Assignment
        assignment = (await db.execute(
            select(Assignment).where(
                Assignment.task_id == task_id,
                Assignment.assignee_id == user_id
            )
        )).scalar_one_or_none()
        if not assignment:
            return JSONResponse({'error': 'Permission denied'}, status_code=403)
    
    await db.delete(subtask)
    await db.commit()
    
    return JSONResponse({'success': True})


@router.post('/tasks/{task_id}/comment')
async def web_task_add_comment(
    request: Request,
    task_id: int,
    content: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify task exists and belongs to workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Check if task is archived - only admins can comment on archived tasks
    if task.is_archived and not user.is_admin:
        raise HTTPException(status_code=403, detail='This task is archived. Only admins can add comments.')
    
    # Check permission: only admin or assignees can comment
    is_assignee = (await db.execute(
        select(Assignment).where(Assignment.task_id == task_id, Assignment.assignee_id == user_id)
    )).scalar_one_or_none() is not None
    
    if not user.is_admin and not is_assignee:
        raise HTTPException(status_code=403, detail='Only assigned users and admins can comment on this task')
    
    comment = Comment(task_id=task_id, author_id=user_id, content=content)
    db.add(comment)
    await db.flush()  # Get comment.id for attachments
    
    # Notify all assignees except the commenter
    assignees_stmt = (
        select(User)
        .join(Assignment, User.id == Assignment.assignee_id)
        .where(Assignment.task_id == task_id, User.id != user_id)
    )
    assignees = (await db.execute(assignees_stmt)).scalars().all()
    
    commenter_name = user.full_name or user.username
    for assignee in assignees:
        notification = Notification(
            user_id=assignee.id,
            type='comment',
            message=f'{commenter_name} commented on task: {task.title}',
            url=f'/web/tasks/{task_id}'
        )
        db.add(notification)
    
    # Handle file attachments
    if files:
        # Create uploads directory if it doesn't exist
        upload_dir = BASE_DIR / 'uploads' / 'comments'
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        for file in files:
            if file.filename:  # Only process if file was actually uploaded
                # Read file content
                file_content = await file.read()
                
                # Validate file size (max 10MB)
                if len(file_content) > 10 * 1024 * 1024:
                    raise HTTPException(status_code=400, detail=f'File {file.filename} is too large. Maximum size is 10MB.')
                
                # Generate unique filename
                file_extension = os.path.splitext(file.filename)[1]
                unique_filename = f"{uuid.uuid4()}{file_extension}"
                file_path = upload_dir / unique_filename
                
                # Save file to disk
                with open(file_path, 'wb') as f:
                    f.write(file_content)
                
                # Store relative path from app directory
                relative_path = f"app/uploads/comments/{unique_filename}"
                
                # Create attachment record
                attachment = CommentAttachment(
                    comment_id=comment.id,
                    filename=file.filename,
                    file_path=relative_path,
                    file_size=len(file_content),
                    content_type=file.content_type or 'application/octet-stream',
                    uploaded_by_id=user_id
                )
                db.add(attachment)
    
    await db.commit()
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/watch')
async def web_task_watch(
    request: Request,
    task_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Add current user as a watcher of the task."""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify task exists and belongs to workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Check if already watching
    from app.models.task_extensions import TaskWatcher
    existing = (await db.execute(
        select(TaskWatcher).where(TaskWatcher.task_id == task_id, TaskWatcher.user_id == user_id)
    )).scalar_one_or_none()
    
    if not existing:
        watcher = TaskWatcher(task_id=task_id, user_id=user_id)
        db.add(watcher)
        await db.commit()
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/unwatch')
async def web_task_unwatch(
    request: Request,
    task_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Remove current user as a watcher of the task."""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify task exists and belongs to workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Remove watcher
    from app.models.task_extensions import TaskWatcher
    existing = (await db.execute(
        select(TaskWatcher).where(TaskWatcher.task_id == task_id, TaskWatcher.user_id == user_id)
    )).scalar_one_or_none()
    
    if existing:
        await db.delete(existing)
        await db.commit()
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/dependencies/add')
async def web_task_add_dependency(
    request: Request,
    task_id: int,
    blocker_task_id: int = Form(...),
    db: AsyncSession = Depends(get_session)
):
    """Add a blocking dependency to the task."""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify task exists and belongs to workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Verify blocker task exists and belongs to same project
    blocker = (await db.execute(
        select(Task).where(Task.id == blocker_task_id, Task.project_id == task.project_id)
    )).scalar_one_or_none()
    if not blocker:
        raise HTTPException(status_code=404, detail='Blocker task not found')
    
    # Check for circular dependencies (simple check)
    if blocker_task_id == task_id:
        raise HTTPException(status_code=400, detail='Task cannot block itself')
    
    # Check if dependency already exists
    from app.models.task_extensions import TaskDependency
    existing = (await db.execute(
        select(TaskDependency).where(
            TaskDependency.task_id == task_id, 
            TaskDependency.depends_on_task_id == blocker_task_id
        )
    )).scalar_one_or_none()
    
    if not existing:
        dependency = TaskDependency(
            task_id=task_id, 
            depends_on_task_id=blocker_task_id,
            created_by_id=user_id
        )
        db.add(dependency)
        await db.commit()
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.post('/tasks/{task_id}/dependencies/{dep_task_id}/remove')
async def web_task_remove_dependency(
    request: Request,
    task_id: int,
    dep_task_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Remove a blocking dependency from the task."""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify task exists and belongs to workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Remove dependency
    from app.models.task_extensions import TaskDependency
    existing = (await db.execute(
        select(TaskDependency).where(
            TaskDependency.task_id == task_id,
            TaskDependency.depends_on_task_id == dep_task_id
        )
    )).scalar_one_or_none()
    
    if existing:
        await db.delete(existing)
        await db.commit()
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


@router.get('/attachments/{attachment_id}/preview')
async def preview_comment_attachment(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get attachment with permission check
    attachment = (await db.execute(
        select(CommentAttachment)
        .join(Comment, CommentAttachment.comment_id == Comment.id)
        .join(Task, Comment.task_id == Task.id)
        .join(Project, Task.project_id == Project.id)
        .where(
            CommentAttachment.id == attachment_id,
            Project.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not attachment:
        raise HTTPException(status_code=404, detail='Attachment not found')
    
    # Handle both absolute paths (old) and relative paths (new)
    file_path = Path(attachment.file_path)
    logger.debug(f"Attachment file_path from DB: {attachment.file_path}")
    logger.debug(f"Is absolute: {file_path.is_absolute()}")
    
    if not file_path.is_absolute():
        # Relative path - resolve from current working directory
        file_path = Path.cwd() / file_path
    
    logger.debug(f"Resolved file_path: {file_path}")
    logger.debug(f"File exists: {file_path.exists()}")
    
    if not file_path.exists():
        logger.error(f"File not found: {file_path}")
        raise HTTPException(status_code=404, detail=f'File not found on disk: {file_path}')
    
    # Serve file inline for preview with proper headers for PDF embedding
    return FileResponse(
        path=str(file_path),
        media_type=attachment.content_type,
        filename=attachment.filename,
        headers={
            'Content-Disposition': f'inline; filename="{attachment.filename}"',
            'X-Content-Type-Options': 'nosniff',
            'Content-Security-Policy': "frame-ancestors 'self'"
        }
    )


@router.get('/attachments/{attachment_id}/download')
async def download_comment_attachment(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get attachment with permission check
    attachment = (await db.execute(
        select(CommentAttachment)
        .join(Comment, CommentAttachment.comment_id == Comment.id)
        .join(Task, Comment.task_id == Task.id)
        .join(Project, Task.project_id == Project.id)
        .where(
            CommentAttachment.id == attachment_id,
            Project.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not attachment:
        raise HTTPException(status_code=404, detail='Attachment not found')
    
    # Handle both absolute paths (old) and relative paths (new)
    file_path = Path(attachment.file_path)
    if not file_path.is_absolute():
        # Relative path - resolve from current working directory
        file_path = Path.cwd() / file_path
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail='File not found on disk')
    
    return FileResponse(
        path=str(file_path),
        filename=attachment.filename,
        media_type=attachment.content_type
    )


@router.post('/attachments/{attachment_id}/delete')
async def delete_comment_attachment(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Delete a comment attachment"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse(status_code=401, content={'detail': 'Not authenticated'})
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse(status_code=401, content={'detail': 'User not found'})
    
    # Get attachment with permission check
    attachment = (await db.execute(
        select(CommentAttachment)
        .join(Comment, CommentAttachment.comment_id == Comment.id)
        .join(Task, Comment.task_id == Task.id)
        .join(Project, Task.project_id == Project.id)
        .where(
            CommentAttachment.id == attachment_id,
            Project.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not attachment:
        return JSONResponse(status_code=404, content={'detail': 'Attachment not found'})
    
    # Check permission - only admin or the comment author can delete
    comment = (await db.execute(select(Comment).where(Comment.id == attachment.comment_id))).scalar_one_or_none()
    if not user.is_admin and (not comment or comment.user_id != user.id):
        return JSONResponse(status_code=403, content={'detail': 'Permission denied'})
    
    # Delete file from disk
    file_path = Path(attachment.file_path)
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    
    if file_path.exists():
        try:
            file_path.unlink()
        except Exception as e:
            pass  # File may already be deleted
    
    # Delete from database
    await db.delete(attachment)
    await db.commit()
    
    return JSONResponse(status_code=200, content={'success': True})


@router.post('/tasks/{task_id}/status')
async def web_task_update_status(request: Request, task_id: int, status_value: str = Form(...), db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    # Ensure task belongs to user's workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Check permission: Admin OR project member
    if not user.is_admin:
        from app.models.project_member import ProjectMember
        member = (await db.execute(
            select(ProjectMember).where(
                ProjectMember.project_id == task.project_id,
                ProjectMember.user_id == user_id
            )
        )).scalar_one_or_none()
        if not member:
            raise HTTPException(status_code=403, detail='You must be a project member to update tasks')
    
    if status_value not in [s.value for s in TaskStatus]:
        raise HTTPException(status_code=400, detail='Invalid status')
    
    # Track change
    old_status = task.status.value
    task.status = TaskStatus(status_value)
    
    # Auto-archive when moved to done
    if status_value == 'done':
        task.is_archived = True
        task.archived_at = datetime.utcnow()
    elif old_status == 'done' and status_value != 'done':
        # Unarchive if moved out of done
        task.is_archived = False
        task.archived_at = None
    
    # Save history
    history_entry = TaskHistory(
        task_id=task_id,
        editor_id=user_id,
        field='status',
        old_value=old_status,
        new_value=status_value
    )
    db.add(history_entry)
    
    await db.commit()
    
    # Track user behavior for learning
    try:
        from app.core.smart_suggestions import track_user_action
        await track_user_action(
            db=db,
            user_id=user_id,
            workspace_id=user.workspace_id,
            action_type="task_status_change",
            entity_type="task",
            entity_id=task_id,
            project_id=task.project_id,
            field_name="status",
            field_value=status_value
        )
    except Exception:
        pass
    
    # Check if this is an AJAX request (fetch)
    if request.headers.get('accept', '').find('application/json') != -1 or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        from fastapi.responses import JSONResponse
        return JSONResponse({'success': True, 'status': status_value})
    
    return RedirectResponse(f'/web/projects/{task.project_id}', status_code=303)


@router.post('/tasks/{task_id}/assign')
async def web_task_assign(request: Request, task_id: int, assignee_id: int = Form(...), db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can assign tasks to users
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Only admins can assign tasks')
    
    # Ensure task belongs to user's workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    # Verify assignee is in same workspace and is active
    assignee = (await db.execute(select(User).where(User.id == assignee_id, User.workspace_id == user.workspace_id, User.is_active == True))).scalar_one_or_none()
    if not assignee:
        raise HTTPException(status_code=400, detail='Invalid assignee or user is inactive')
    # Check if already assigned
    existing = (await db.execute(select(Assignment).where(Assignment.task_id == task_id, Assignment.assignee_id == assignee_id))).scalar_one_or_none()
    if not existing:
        assignment = Assignment(task_id=task_id, assignee_id=assignee_id)
        db.add(assignment)
        
        # Create notification for the assignee (don't notify if assigning to self)
        if assignee_id != user_id:
            assigner = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
            assigner_name = assigner.full_name or assigner.username if assigner else "Someone"
            
            notification = Notification(
                user_id=assignee_id,
                type='task_assigned',
                message=f'{assigner_name} assigned you to task: {task.title}',
                url=f'/web/tasks/{task_id}'
            )
            db.add(notification)
        
        await db.commit()
        
        # Track user behavior for learning
        try:
            from app.core.smart_suggestions import track_user_action
            await track_user_action(
                db=db,
                user_id=user_id,
                workspace_id=user.workspace_id,
                action_type="task_assign",
                entity_type="task",
                entity_id=task_id,
                project_id=task.project_id,
                field_name="assignee",
                field_value=str(assignee_id)
            )
        except Exception:
            pass
    return RedirectResponse(f'/web/projects/{task.project_id}', status_code=303)


@router.post('/tasks/{task_id}/unassign')
async def web_task_unassign(request: Request, task_id: int, assignee_id: int = Form(...), db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can unassign tasks
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Only admins can unassign tasks')
    
    # Ensure task belongs to user's workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    # Delete assignment
    assignment = (await db.execute(select(Assignment).where(Assignment.task_id == task_id, Assignment.assignee_id == assignee_id))).scalar_one_or_none()
    if assignment:
        await db.delete(assignment)
        await db.commit()
    return RedirectResponse(f'/web/projects/{task.project_id}', status_code=303)


@router.post('/tasks/{task_id}/delete')
async def web_task_delete(request: Request, task_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can delete tasks
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Only admins can delete tasks')
    
    # Ensure task belongs to user's workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    project_id = task.project_id
    
    # Delete the task (cascade will handle assignments, comments, etc.)
    await db.delete(task)
    await db.commit()
    
    return RedirectResponse(f'/web/projects/{project_id}', status_code=303)


@router.post('/tasks/{task_id}/reopen')
async def web_task_reopen(request: Request, task_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can reopen tasks
    if not user.is_admin:
        raise HTTPException(status_code=403, detail='Only admins can reopen tasks')
    
    # Ensure task belongs to user's workspace
    stmt = select(Task).join(Project, Task.project_id == Project.id).where(Task.id == task_id, Project.workspace_id == user.workspace_id)
    result = await db.execute(stmt)
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail='Task not found')
    
    # Reopen the task
    task.is_archived = False
    task.archived_at = None
    # If status is done, set it back to in_progress
    if task.status == TaskStatus.done:
        task.status = TaskStatus.in_progress
    
    await db.commit()
    
    return RedirectResponse(f'/web/tasks/{task_id}', status_code=303)


# Meetings
@router.get('/meetings')
async def web_meetings_list(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get all meetings where user is an attendee or organizer
    stmt = (
        select(Meeting)
        .join(MeetingAttendee, Meeting.id == MeetingAttendee.meeting_id)
        .where(MeetingAttendee.user_id == user_id)
        .order_by(Meeting.start_time.desc())
    )
    result = await db.execute(stmt)
    meetings = result.scalars().all()
    
    # Get all active workspace users for the create form
    users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id, User.is_active == True)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    return templates.TemplateResponse('meetings/list.html', {
        'request': request,
        'meetings': meetings,
        'users': users,
        'user': user
    })


@router.post('/meetings/create')
async def web_meeting_create(
    request: Request,
    title: str = Form(...),
    start_time_str: str = Form(..., alias='start_time'),
    end_time_str: str = Form(..., alias='end_time'),
    platform: str = Form(...),
    meeting_url: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    auto_generate_meet: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Parse datetime strings
    from datetime import datetime
    try:
        start_datetime = datetime.fromisoformat(start_time_str)
        end_datetime = datetime.fromisoformat(end_time_str)
        
        # Extract date and time components
        meeting_date = start_datetime.date()
        meeting_time = start_datetime.time()
        
        # Calculate duration in minutes
        duration = int((end_datetime - start_datetime).total_seconds() / 60)
        if duration <= 0:
            # Return user to form with error message instead of HTTP exception
            from fastapi.responses import HTMLResponse
            from fastapi.templating import Jinja2Templates
            templates = Jinja2Templates(directory="app/templates")
            
            # Get users for the form
            users_result = await db.execute(
                select(User).where(User.workspace_id == user.workspace_id).order_by(User.full_name)
            )
            users = users_result.scalars().all()
            
            return templates.TemplateResponse("meetings/list.html", {
                "request": request,
                "user": user,
                "meetings": [],
                "users": users,
                "error": "End time must be after start time. Please select a later end time."
            })
            
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f'Invalid datetime format: {str(e)}')
    
    # Auto-generate Google Meet link if requested
    generated_meet_url = None
    if auto_generate_meet == 'on' and platform == 'google_meet' and user.google_access_token:
        try:
            from app.core.google_oauth import create_calendar_event
            
            # Get attendee emails
            form_data = await request.form()
            attendee_ids = form_data.getlist('attendee_ids')
            attendee_emails = []
            
            if attendee_ids:
                attendees_result = await db.execute(
                    select(User).where(User.id.in_([int(aid) for aid in attendee_ids]))
                )
                attendees = attendees_result.scalars().all()
                attendee_emails = [a.email for a in attendees if a.email]
            
            # Create Google Calendar event with Meet link
            calendar_event = create_calendar_event(
                access_token=user.google_access_token,
                refresh_token=user.google_refresh_token,
                token_expiry=user.google_token_expiry,
                summary=title,
                description=description or '',
                start_time=start_datetime,
                end_time=end_datetime,
                attendees=attendee_emails,
                add_google_meet=True
            )
            
            if calendar_event and 'hangoutLink' in calendar_event:
                generated_meet_url = calendar_event['hangoutLink']
                meeting_url = generated_meet_url
                logger.info(f"Auto-generated Google Meet link: {generated_meet_url}")
            else:
                logger.warning("Failed to generate Google Meet link - no hangoutLink in response")
                
        except Exception as e:
            logger.error(f"Error auto-generating Google Meet link: {e}")
            # Continue with meeting creation without the auto-generated link
    
    # Create meeting
    meeting = Meeting(
        title=title,
        description=description,
        date=meeting_date,
        start_time=meeting_time,
        duration_minutes=duration,
        platform=MeetingPlatform(platform),
        url=meeting_url,
        organizer_id=user_id,
        workspace_id=user.workspace_id
    )
    db.add(meeting)
    await db.flush()  # Get the meeting ID
    
    # Add organizer as attendee
    attendee = MeetingAttendee(meeting_id=meeting.id, user_id=user_id)
    db.add(attendee)
    
    # Get attendee IDs from form (if provided)
    form_data = await request.form()
    attendee_ids = form_data.getlist('attendee_ids')
    for attendee_id in attendee_ids:
        if int(attendee_id) != user_id:  # Don't duplicate organizer
            attendee = MeetingAttendee(meeting_id=meeting.id, user_id=int(attendee_id))
            db.add(attendee)
    
    await db.commit()
    return RedirectResponse('/web/meetings', status_code=303)


@router.post('/meetings/{meeting_id}/cancel')
async def web_meeting_cancel(
    request: Request,
    meeting_id: int,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get the meeting
    meeting = (await db.execute(select(Meeting).where(Meeting.id == meeting_id))).scalar_one_or_none()
    if not meeting:
        raise HTTPException(status_code=404, detail='Meeting not found')
    
    # Check if user has permission (organizer or admin)
    if meeting.organizer_id != user_id and not user.is_admin:
        raise HTTPException(status_code=403, detail='Not authorized to cancel this meeting')
    
    # Check if already cancelled
    if meeting.is_cancelled:
        return RedirectResponse('/web/meetings', status_code=303)
    
    # Cancel the meeting
    from datetime import datetime
    meeting.is_cancelled = True
    meeting.cancelled_at = datetime.utcnow()
    meeting.cancelled_by = user_id
    
    await db.commit()
    return RedirectResponse('/web/meetings', status_code=303)


@router.get('/meetings/{meeting_id}/details')
async def web_meeting_details(
    request: Request,
    meeting_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Get meeting details for display in modal"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse({'error': 'Not authenticated'}, status_code=401)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse({'error': 'User not found'}, status_code=401)
    
    # Get the meeting
    meeting = (await db.execute(select(Meeting).where(Meeting.id == meeting_id))).scalar_one_or_none()
    if not meeting:
        return JSONResponse({'error': 'Meeting not found'}, status_code=404)
    
    # Check if user has access to this meeting (is attendee, organizer, or admin)
    is_organizer = meeting.organizer_id == user_id
    attendee = (await db.execute(
        select(MeetingAttendee).where(
            MeetingAttendee.meeting_id == meeting_id,
            MeetingAttendee.user_id == user_id
        )
    )).scalar_one_or_none()
    
    if not (is_organizer or attendee or user.is_admin):
        return JSONResponse({'error': 'Not authorized to view this meeting'}, status_code=403)
    
    # Get organizer details
    organizer = (await db.execute(select(User).where(User.id == meeting.organizer_id))).scalar_one_or_none()
    
    # Get all attendees
    attendees_data = (await db.execute(
        select(MeetingAttendee, User)
        .join(User, MeetingAttendee.user_id == User.id)
        .where(MeetingAttendee.meeting_id == meeting_id)
    )).all()
    
    # Get cancelled by user if meeting is cancelled
    cancelled_by_user = None
    if meeting.is_cancelled and meeting.cancelled_by:
        cancelled_by_user = (await db.execute(
            select(User).where(User.id == meeting.cancelled_by)
        )).scalar_one_or_none()
    
    return JSONResponse({
        'id': meeting.id,
        'title': meeting.title,
        'description': meeting.description,
        'date': meeting.date.strftime('%d/%m/%Y'),
        'date_formatted': meeting.date.strftime('%d/%m/%Y'),
        'start_time': meeting.start_time.strftime('%I:%M %p'),
        'duration_minutes': meeting.duration_minutes,
        'platform': meeting.platform.value,
        'platform_display': meeting.platform.value.replace('_', ' ').title(),
        'url': meeting.url,
        'organizer': {
            'id': organizer.id if organizer else None,
            'name': organizer.full_name if organizer else 'Unknown',
            'email': organizer.email if organizer else ''
        },
        'attendees': [
            {
                'id': user_obj.id,
                'name': user_obj.full_name,
                'email': user_obj.email,
                'status': attendee_obj.status or 'invited'
            }
            for attendee_obj, user_obj in attendees_data
        ],
        'is_cancelled': meeting.is_cancelled,
        'cancelled_at': meeting.cancelled_at.strftime('%d/%m/%Y at %I:%M %p') if meeting.cancelled_at else None,
        'cancelled_by': cancelled_by_user.full_name if cancelled_by_user else None,
        'is_organizer': is_organizer,
        'is_admin': user.is_admin
    })


# Calendar
@router.get('/calendar')
async def web_calendar(
    request: Request,
    year: Optional[int] = None,
    month: Optional[int] = None,
    day: Optional[int] = None,
    view: str = 'month',
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Default to current date if not specified
    today = date.today()
    year = year or today.year
    month = month or today.month
    day = day or today.day
    
    # Calculate prev/next month for navigation
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year
    
    # Determine date range based on view
    if view == 'day':
        current_date = date(year, month, day)
        first_day = current_date
        last_day = current_date
        weeks = []
    elif view == 'week':
        current_date = date(year, month, day)
        # Get Monday of current week
        start_of_week = current_date - timedelta(days=current_date.weekday())
        first_day = start_of_week
        last_day = start_of_week + timedelta(days=6)
        weeks = [[start_of_week + timedelta(days=i) for i in range(7)]]
    else:  # month view
        current_date = date(year, month, day)
        # Build calendar weeks with date objects
        cal = pycalendar.Calendar(firstweekday=0)  # Monday first
        weeks = []
        for week in cal.monthdatescalendar(year, month):
            weeks.append(week)
        # Get all tasks with due dates for display (include adjacent months shown in calendar)
        first_day = weeks[0][0]
        last_day = weeks[-1][-1]
    
    # Admin sees all tasks, regular users see only their assigned tasks
    # Fetch tasks that overlap with the calendar view period (either by start_date or due_date)
    if user.is_admin:
        stmt = (
            select(Task)
            .join(Project, Task.project_id == Project.id)
            .where(
                Project.workspace_id == user.workspace_id,
                # Task has at least a due_date or start_date
                (Task.due_date.isnot(None)) | (Task.start_date.isnot(None)),
                # Task overlaps with calendar period
                (
                    # Tasks with both start and due dates - check if they overlap calendar period
                    ((Task.start_date.isnot(None)) & (Task.due_date.isnot(None)) & 
                     (Task.start_date <= last_day) & (Task.due_date >= first_day)) |
                    # Tasks with only due_date - check if in period
                    ((Task.start_date.is_(None)) & (Task.due_date.isnot(None)) & 
                     (Task.due_date >= first_day) & (Task.due_date <= last_day)) |
                    # Tasks with only start_date - check if in period
                    ((Task.start_date.isnot(None)) & (Task.due_date.is_(None)) & 
                     (Task.start_date >= first_day) & (Task.start_date <= last_day))
                )
            )
            .order_by(Task.start_date, Task.due_date, Task.due_time)
        )
    else:
        stmt = (
            select(Task)
            .join(Project, Task.project_id == Project.id)
            .join(Assignment, Task.id == Assignment.task_id)
            .where(
                Project.workspace_id == user.workspace_id,
                Assignment.assignee_id == user.id,
                # Task has at least a due_date or start_date
                (Task.due_date.isnot(None)) | (Task.start_date.isnot(None)),
                # Task overlaps with calendar period
                (
                    # Tasks with both start and due dates - check if they overlap calendar period
                    ((Task.start_date.isnot(None)) & (Task.due_date.isnot(None)) & 
                     (Task.start_date <= last_day) & (Task.due_date >= first_day)) |
                    # Tasks with only due_date - check if in period
                    ((Task.start_date.is_(None)) & (Task.due_date.isnot(None)) & 
                     (Task.due_date >= first_day) & (Task.due_date <= last_day)) |
                    # Tasks with only start_date - check if in period
                    ((Task.start_date.isnot(None)) & (Task.due_date.is_(None)) & 
                     (Task.start_date >= first_day) & (Task.start_date <= last_day))
                )
            )
            .order_by(Task.start_date, Task.due_date, Task.due_time)
        )
    tasks = (await db.execute(stmt)).scalars().all()
    
    # Fetch projects with date ranges for calendar display
    # Admin sees all projects, regular users see projects they're assigned to (via tasks or project_member)
    from app.models.project_member import ProjectMember
    
    if user.is_admin:
        projects_stmt = (
            select(Project)
            .where(
                Project.workspace_id == user.workspace_id,
                Project.start_date.isnot(None),
                Project.due_date.isnot(None),
                Project.is_archived == False,
                # Project overlaps with calendar view period
                Project.start_date <= last_day,
                Project.due_date >= first_day
            )
            .order_by(Project.start_date)
        )
    else:
        # Get projects where user is a member or has assigned tasks
        projects_stmt = (
            select(Project)
            .join(ProjectMember, Project.id == ProjectMember.project_id)
            .where(
                Project.workspace_id == user.workspace_id,
                ProjectMember.user_id == user.id,
                Project.start_date.isnot(None),
                Project.due_date.isnot(None),
                Project.is_archived == False,
                Project.start_date <= last_day,
                Project.due_date >= first_day
            )
            .distinct()
            .order_by(Project.start_date)
        )
    projects = (await db.execute(projects_stmt)).scalars().all()
    
    # Sort tasks by priority (critical, high, medium, low)
    # Use date.max for None due_dates to put them at the end
    priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
    
    def task_sort_key(t):
        due = t.due_date if t.due_date else date.max
        priority_val = t.priority.value if t.priority else 'low'
        priority = priority_order.get(priority_val, 4)
        due_time = t.due_time if t.due_time else time.max
        return (due, priority, due_time)
    
    tasks = sorted(tasks, key=task_sort_key)
    
    # Get meetings for the calendar period
    # Admin sees all meetings, regular users see only meetings they're attending
    if user.is_admin:
        meetings_stmt = (
            select(Meeting)
            .where(
                Meeting.workspace_id == user.workspace_id,
                Meeting.date >= first_day,
                Meeting.date <= last_day
            )
            .order_by(Meeting.date, Meeting.start_time)
        )
    else:
        meetings_stmt = (
            select(Meeting)
            .join(MeetingAttendee, Meeting.id == MeetingAttendee.meeting_id)
            .where(
                Meeting.workspace_id == user.workspace_id,
                MeetingAttendee.user_id == user.id,
                Meeting.date >= first_day,
                Meeting.date <= last_day
            )
            .order_by(Meeting.date, Meeting.start_time)
        )
    meetings = (await db.execute(meetings_stmt)).scalars().all()
    
    # Get all workspace users for color legend (admin view)
    workspace_users = []
    if user.is_admin:
        workspace_users_stmt = (
            select(User)
            .where(
                User.workspace_id == user.workspace_id,
                User.is_active == True
            )
            .order_by(User.full_name)
        )
        workspace_users = (await db.execute(workspace_users_stmt)).scalars().all()
    
    # Build a map of task/project IDs to their assigned users for color coding
    task_users = {}
    for task in tasks:
        # Get all users assigned to this task
        task_assignments = (await db.execute(
            select(User)
            .join(Assignment, User.id == Assignment.assignee_id)
            .where(Assignment.task_id == task.id)
        )).scalars().all()
        if task_assignments:
            task_users[task.id] = list(task_assignments)  # Store all assigned users
    
    project_users = {}
    for project in projects:
        # Get project owner for color coding
        project_owner = (await db.execute(
            select(User).where(User.id == project.owner_id)
        )).scalar_one_or_none()
        if project_owner:
            project_users[project.id] = project_owner
    
    # Calculate navigation dates based on view
    if view == 'day':
        prev_date = current_date - timedelta(days=1)
        next_date = current_date + timedelta(days=1)
        prev_year, prev_month, prev_day = prev_date.year, prev_date.month, prev_date.day
        next_year, next_month, next_day = next_date.year, next_date.month, next_date.day
    elif view == 'week':
        prev_week_start = first_day - timedelta(days=7)
        next_week_start = first_day + timedelta(days=7)
        prev_year, prev_month, prev_day = prev_week_start.year, prev_week_start.month, prev_week_start.day
        next_year, next_month, next_day = next_week_start.year, next_week_start.month, next_week_start.day
    else:  # month
        prev_day = 1
        next_day = 1
    
    return templates.TemplateResponse('calendar/index.html', {
        'request': request,
        'user': user,
        'view': view,
        'year': year,
        'month': month,
        'day': day,
        'current_date': current_date,
        'first_day': first_day,
        'last_day': last_day,
        'weeks': weeks,
        'tasks': tasks,
        'meetings': meetings,
        'projects': projects,
        'task_users': task_users,
        'project_users': project_users,
        'workspace_users': workspace_users,
        'today': today,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'prev_day': prev_day,
        'next_month': next_month,
        'next_year': next_year,
        'next_day': next_day,
        'TaskStatus': TaskStatus
    })


# ====================================
# Tickets System
# ====================================

@router.get('/tickets/report', response_class=HTMLResponse)
async def web_tickets_report(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),  # Accept as string to handle empty values
    project_id: Optional[str] = Query(None),  # Accept as string to handle empty values
    db: AsyncSession = Depends(get_session),
):
    """Detailed ticket report - admin only"""
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    from sqlalchemy import text, func
    
    # Parse user_id and project_id (handle empty strings)
    user_id_int = None
    if user_id and user_id.strip():
        try:
            user_id_int = int(user_id)
        except ValueError:
            pass
    
    project_id_int = None
    if project_id and project_id.strip():
        try:
            project_id_int = int(project_id)
        except ValueError:
            pass
    
    # Get workspace
    workspace = (await db.execute(
        select(Workspace).where(Workspace.id == user.workspace_id)
    )).scalar_one_or_none()
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    # Get all users for dropdown and lookups
    users_result = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id).order_by(User.full_name, User.username)
    )).scalars().all()
    users_dict = {u.id: u for u in users_result}
    
    # Get projects that have tickets in the date range (projects with actual ticket data)
    # First get all project IDs that have tickets
    from sqlalchemy import distinct
    projects_with_tickets_query = select(distinct(Ticket.related_project_id)).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.related_project_id != None,
        Ticket.created_at >= start_dt,
        Ticket.created_at < end_dt
    )
    project_ids_with_tickets = (await db.execute(projects_with_tickets_query)).scalars().all()
    
    # Get those projects with their details
    all_projects = []
    if project_ids_with_tickets:
        all_projects = (await db.execute(
            select(Project).where(
                Project.id.in_(project_ids_with_tickets),
                Project.is_archived == False
            ).order_by(Project.name)
        )).scalars().all()
    projects_dict = {p.id: p for p in all_projects}
    
    # Count tickets per project in a single GROUP BY query (avoid N+1)
    project_ticket_counts = {}
    if project_ids_with_tickets:
        counts_result = await db.execute(
            select(Ticket.related_project_id, func.count(Ticket.id))
            .where(
                Ticket.workspace_id == user.workspace_id,
                Ticket.related_project_id.in_([pid for pid in project_ids_with_tickets if pid]),
                Ticket.created_at >= start_dt,
                Ticket.created_at < end_dt
            )
            .group_by(Ticket.related_project_id)
        )
        for pid, count in counts_result.all():
            project_ticket_counts[pid] = count
    
    # Get selected user for filtering
    selected_user = None
    selected_user_id = None
    if user_id_int:
        selected_user = users_dict.get(user_id_int)
        if selected_user:
            selected_user_id = user_id_int
    
    # Get selected project for filtering
    selected_project = None
    selected_project_id = None
    if project_id_int:
        selected_project = projects_dict.get(project_id_int)
        if selected_project:
            selected_project_id = project_id_int
    
    # Get all tickets in period - optionally filtered by user and/or project
    ticket_query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.created_at >= start_dt,
        Ticket.created_at < end_dt
    )
    
    if selected_user_id:
        # Filter tickets where user is assigned OR closed by user
        ticket_query = ticket_query.where(
            (Ticket.assigned_to_id == selected_user_id) | (Ticket.closed_by_id == selected_user_id)
        )
    
    if selected_project_id:
        # Filter tickets by project
        ticket_query = ticket_query.where(Ticket.related_project_id == selected_project_id)
    
    all_tickets = (await db.execute(ticket_query.order_by(Ticket.created_at.desc()))).scalars().all()
    
    ticket_ids = [t.id for t in all_tickets]
    
    # Get ALL currently open tickets (regardless of date range) for the "Tickets Open" section
    current_open_query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.status.in_(['open', 'in_progress', 'waiting'])
    )
    if selected_user_id:
        current_open_query = current_open_query.where(Ticket.assigned_to_id == selected_user_id)
    if selected_project_id:
        current_open_query = current_open_query.where(Ticket.related_project_id == selected_project_id)
    all_current_open_tickets = (await db.execute(current_open_query.order_by(Ticket.created_at.desc()))).scalars().all()
    
    # Get all comments for these tickets
    all_comments = []
    if ticket_ids:
        comment_query = select(TicketComment).where(TicketComment.ticket_id.in_(ticket_ids))
        if selected_user_id:
            # Only show comments by this user
            comment_query = comment_query.where(TicketComment.user_id == selected_user_id)
        all_comments = (await db.execute(comment_query.order_by(TicketComment.created_at.desc()))).scalars().all()
    
    # Get ticket history
    all_history = []
    all_history_for_closures = []  # Unfiltered history to determine who closed tickets
    if ticket_ids:
        # Get history for display (filtered by user if selected)
        history_query = select(TicketHistory).where(TicketHistory.ticket_id.in_(ticket_ids))
        if selected_user_id:
            history_query = history_query.where(TicketHistory.user_id == selected_user_id)
        all_history = (await db.execute(history_query.order_by(TicketHistory.created_at.desc()))).scalars().all()
        
        # Get ALL history to determine who closed each ticket (not filtered by user)
        closure_query = select(TicketHistory).where(
            TicketHistory.ticket_id.in_(ticket_ids),
            or_(
                TicketHistory.action == 'closed',
                and_(TicketHistory.action == 'status_changed', TicketHistory.new_value == 'closed')
            )
        )
        all_history_for_closures = (await db.execute(closure_query)).scalars().all()
    
    # Create ticket lookup
    tickets_dict = {t.id: t for t in all_tickets}
    
    # Calculate agent performance
    from app.core.business_hours import calculate_business_hours
    
    # Get business hours settings from workspace
    biz_start = workspace.business_hours_start if workspace and hasattr(workspace, 'business_hours_start') else "07:30"
    biz_end = workspace.business_hours_end if workspace and hasattr(workspace, 'business_hours_end') else "16:00"
    biz_exclude_weekends = workspace.business_hours_exclude_weekends if workspace and hasattr(workspace, 'business_hours_exclude_weekends') else True
    
    agent_stats = {}
    for usr in users_result:
        agent_stats[usr.id] = {
            'user': usr,
            'name': usr.full_name or usr.username,
            'email': usr.email,
            'initials': ''.join([n[0].upper() for n in (usr.full_name or usr.email).split()[:2]]),
            'tickets_assigned': 0,
            'tickets_closed': 0,
            'comments_made': 0,
            'resolution_times': [],
            'first_response_times': [],
        }
    
    # Build a map of who closed each ticket from history (as fallback for closed_by_id)
    ticket_closed_by = {}  # ticket_id -> user_id who closed it
    for history in all_history_for_closures:
        if history.user_id:
            ticket_closed_by[history.ticket_id] = history.user_id
    
    # Process tickets for agent stats
    for ticket in all_tickets:
        if ticket.assigned_to_id and ticket.assigned_to_id in agent_stats:
            agent_stats[ticket.assigned_to_id]['tickets_assigned'] += 1
        
        # Determine who closed the ticket: use closed_by_id if set, otherwise check history
        closer_id = ticket.closed_by_id
        if not closer_id and ticket.status == 'closed':
            closer_id = ticket_closed_by.get(ticket.id)
        
        if closer_id and closer_id in agent_stats:
            agent_stats[closer_id]['tickets_closed'] += 1
            # Calculate resolution time using business hours
            if ticket.closed_at:
                resolution_hours = calculate_business_hours(
                    ticket.created_at, ticket.closed_at,
                    biz_start, biz_end, biz_exclude_weekends
                )
                agent_stats[closer_id]['resolution_times'].append(resolution_hours)
    
    # Process comments
    comment_by_ticket = {}  # ticket_id -> [comments]
    for comment in all_comments:
        if comment.ticket_id not in comment_by_ticket:
            comment_by_ticket[comment.ticket_id] = []
        comment_by_ticket[comment.ticket_id].append(comment)
        
        # Count comments per user
        if comment.user_id and comment.user_id in agent_stats:
            agent_stats[comment.user_id]['comments_made'] += 1
    
    # Calculate first response times
    for ticket in all_tickets:
        ticket_comments = comment_by_ticket.get(ticket.id, [])
        # Find first non-guest comment (staff response)
        staff_comments = [c for c in ticket_comments if c.user_id and c.user_id in agent_stats]
        if staff_comments:
            first_response = min(staff_comments, key=lambda c: c.created_at)
            response_hours = (first_response.created_at - ticket.created_at).total_seconds() / 3600
            if first_response.user_id in agent_stats:
                agent_stats[first_response.user_id]['first_response_times'].append(response_hours)
    
    # Format agent data
    agents = []
    total_tickets_in_range = len(all_tickets)  # Total tickets in the selected date range
    
    for uid, stats in agent_stats.items():
        if stats['tickets_assigned'] > 0 or stats['tickets_closed'] > 0 or stats['comments_made'] > 0:
            # Close rate = (tickets this user closed / total tickets in date range) * 100
            # This shows what percentage of all tickets this user has closed
            close_rate = round((stats['tickets_closed'] / total_tickets_in_range * 100) if total_tickets_in_range > 0 else 0)
            
            avg_resolution = round(sum(stats['resolution_times']) / len(stats['resolution_times']), 1) if stats['resolution_times'] else 0
            avg_first_response = round(sum(stats['first_response_times']) / len(stats['first_response_times']), 1) if stats['first_response_times'] else 0
            
            agents.append({
                'name': stats['name'],
                'email': stats['email'],
                'initials': stats['initials'],
                'tickets_assigned': stats['tickets_assigned'],
                'tickets_closed': stats['tickets_closed'],
                'total_tickets': total_tickets_in_range,  # Total tickets in the date range
                'close_rate': close_rate,
                'avg_resolution_hours': avg_resolution,
                'comments_made': stats['comments_made'],
                'avg_first_response_hours': avg_first_response,
            })
    
    # Sort by tickets closed
    agents.sort(key=lambda x: x['tickets_closed'], reverse=True)
    
    # Status distribution
    status_distribution = {}
    for ticket in all_tickets:
        status = ticket.status
        if status not in status_distribution:
            status_distribution[status] = 0
        status_distribution[status] += 1
    
    # Category distribution
    category_distribution = {}
    for ticket in all_tickets:
        category = ticket.category
        if category not in category_distribution:
            category_distribution[category] = 0
        category_distribution[category] += 1
    
    # Priority distribution with closed counts
    priority_distribution = {}
    for ticket in all_tickets:
        priority = ticket.priority
        if priority not in priority_distribution:
            priority_distribution[priority] = {'count': 0, 'closed': 0}
        priority_distribution[priority]['count'] += 1
        if ticket.status == 'closed':
            priority_distribution[priority]['closed'] += 1
    
    # Closed tickets with details
    closed_tickets = []
    for ticket in all_tickets:
        if ticket.status == 'closed' and ticket.closed_at:
            closed_by = users_dict.get(ticket.closed_by_id)
            # Use business hours for resolution time
            resolution_hours = calculate_business_hours(
                ticket.created_at, ticket.closed_at,
                biz_start, biz_end, biz_exclude_weekends
            )
            closed_tickets.append({
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'subject': ticket.subject,
                'priority': ticket.priority,
                'category': ticket.category,
                'created_at': ticket.created_at,
                'closed_at': ticket.closed_at,
                'closed_by_name': (closed_by.full_name or closed_by.username) if closed_by else 'Unknown',
                'resolution_hours': resolution_hours,
                'billable_traveling': ticket.billable_traveling or '',
                'billable_labour_onsite': ticket.billable_labour_onsite or '',
                'billable_remote_labour': ticket.billable_remote_labour or '',
                'billable_equipment_used': ticket.billable_equipment_used or '',
                'non_billable_traveling': ticket.non_billable_traveling or '',
                'non_billable_labour_onsite': ticket.non_billable_labour_onsite or '',
                'non_billable_remote_labour': ticket.non_billable_remote_labour or '',
                'non_billable_equipment_used': ticket.non_billable_equipment_used or '',
                'closing_notes': ticket.closing_notes or '',
            })
    
    # Sort by closed date
    closed_tickets.sort(key=lambda x: x['closed_at'], reverse=True)
    
    # Comment stats by user
    comment_stats_dict = {}
    for comment in all_comments:
        if comment.user_id:
            if comment.user_id not in comment_stats_dict:
                usr = users_dict.get(comment.user_id)
                if usr:
                    comment_stats_dict[comment.user_id] = {
                        'name': usr.full_name or usr.username,
                        'initials': ''.join([n[0].upper() for n in (usr.full_name or usr.email).split()[:2]]),
                        'total_comments': 0,
                        'public_comments': 0,
                        'internal_comments': 0,
                        'unique_tickets': set(),
                    }
            if comment.user_id in comment_stats_dict:
                comment_stats_dict[comment.user_id]['total_comments'] += 1
                if comment.is_internal:
                    comment_stats_dict[comment.user_id]['internal_comments'] += 1
                else:
                    comment_stats_dict[comment.user_id]['public_comments'] += 1
                comment_stats_dict[comment.user_id]['unique_tickets'].add(comment.ticket_id)
    
    comment_stats = []
    for uid, stats in comment_stats_dict.items():
        comment_stats.append({
            'name': stats['name'],
            'initials': stats['initials'],
            'total_comments': stats['total_comments'],
            'public_comments': stats['public_comments'],
            'internal_comments': stats['internal_comments'],
            'unique_tickets': len(stats['unique_tickets']),
        })
    comment_stats.sort(key=lambda x: x['total_comments'], reverse=True)
    
    # Recent comments with details
    recent_comments = []
    for comment in all_comments[:50]:
        author = users_dict.get(comment.user_id)
        ticket = tickets_dict.get(comment.ticket_id)
        if ticket:
            recent_comments.append({
                'ticket_id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'author_name': (author.full_name or author.username) if author else 'Guest',
                'author_initials': ''.join([n[0].upper() for n in ((author.full_name or author.email) if author else 'GU').split()[:2]]),
                'content': comment.content,
                'is_internal': comment.is_internal,
                'created_at': comment.created_at,
            })
    
    # Activity timeline from history
    activities = []
    for history in all_history[:50]:
        usr = users_dict.get(history.user_id)
        ticket = tickets_dict.get(history.ticket_id)
        if ticket:
            action = history.action
            description = f"{action.replace('_', ' ')}"
            
            if action == 'created':
                description = 'created ticket'
            elif action == 'closed':
                description = 'closed ticket'
            elif action == 'status_changed':
                description = f'changed status to {history.new_value}'
            elif action == 'priority_changed':
                description = f'changed priority to {history.new_value}'
            elif action == 'assigned':
                assignee = users_dict.get(int(history.new_value)) if history.new_value else None
                description = f'assigned to {assignee.full_name or assignee.username}' if assignee else 'assigned'
            elif action == 'commented':
                description = 'added a comment'
            
            activities.append({
                'action': action,
                'description': description,
                'user_name': (usr.full_name or usr.username) if usr else 'System',
                'ticket_id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'ticket_subject': ticket.subject,
                'created_at': history.created_at,
            })
    
    # Summary stats
    total_tickets = len(all_tickets)
    closed_count = sum(1 for t in all_tickets if t.status == 'closed')
    # Use the count from all_current_open_tickets for accurate open ticket count
    current_open_count = len(all_current_open_tickets)
    resolution_rate = round((closed_count / total_tickets * 100) if total_tickets > 0 else 0)
    
    # Average resolution time using business hours
    resolution_times = []
    for ticket in all_tickets:
        if ticket.status == 'closed' and ticket.closed_at:
            hours = calculate_business_hours(
                ticket.created_at, ticket.closed_at,
                biz_start, biz_end, biz_exclude_weekends
            )
            resolution_times.append(hours)
    avg_resolution = round(sum(resolution_times) / len(resolution_times), 1) if resolution_times else 0
    
    summary = {
        'total_tickets': total_tickets,
        'closed_tickets': closed_count,
        'open_tickets': current_open_count,  # All currently open tickets
        'resolution_rate': resolution_rate,
        'avg_resolution_hours': avg_resolution,
        'total_comments': len(all_comments),
        'business_hours': f"{biz_start} - {biz_end}",
        'exclude_weekends': biz_exclude_weekends,
    }
    
    # Count tickets for selected project
    project_ticket_count = len(all_tickets) if selected_project_id else 0
    
    # Build assigned tickets list (tickets that have been assigned to someone)
    assigned_tickets = []
    for ticket in all_tickets:
        if ticket.assigned_to_id:
            assigned_user = users_dict.get(ticket.assigned_to_id)
            project = projects_dict.get(ticket.related_project_id) if ticket.related_project_id else None
            assigned_tickets.append({
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'subject': ticket.subject,
                'priority': ticket.priority,
                'category': ticket.category,
                'status': ticket.status,
                'created_at': ticket.created_at,
                'assigned_to_name': (assigned_user.full_name or assigned_user.username) if assigned_user else 'Unknown',
                'project_name': project.name if project else None,
            })
    assigned_tickets.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Build open tickets list - ALL currently open tickets (not limited to date range)
    open_tickets_list = []
    for ticket in all_current_open_tickets:
        assigned_user = users_dict.get(ticket.assigned_to_id) if ticket.assigned_to_id else None
        project = projects_dict.get(ticket.related_project_id) if ticket.related_project_id else None
        # Calculate time open in hours using business hours
        time_open = calculate_business_hours(
            ticket.created_at, datetime.now(),
            biz_start, biz_end, biz_exclude_weekends
        )
        open_tickets_list.append({
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'subject': ticket.subject,
            'priority': ticket.priority,
            'category': ticket.category,
            'status': ticket.status,
            'created_at': ticket.created_at,
            'assigned_to_name': (assigned_user.full_name or assigned_user.username) if assigned_user else 'Unassigned',
            'project_name': project.name if project else None,
            'time_open_hours': time_open,
        })
    open_tickets_list.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Build created tickets list (all tickets created in the period)
    created_tickets = []
    for ticket in all_tickets:
        created_by_user = users_dict.get(ticket.created_by_id) if ticket.created_by_id else None
        assigned_user = users_dict.get(ticket.assigned_to_id) if ticket.assigned_to_id else None
        project = projects_dict.get(ticket.related_project_id) if ticket.related_project_id else None
        created_tickets.append({
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'subject': ticket.subject,
            'priority': ticket.priority,
            'category': ticket.category,
            'status': ticket.status,
            'created_at': ticket.created_at,
            'created_by_name': (created_by_user.full_name or created_by_user.username) if created_by_user else (ticket.guest_name or 'Guest'),
            'assigned_to_name': (assigned_user.full_name or assigned_user.username) if assigned_user else 'Unassigned',
            'project_name': project.name if project else None,
        })
    created_tickets.sort(key=lambda x: x['created_at'], reverse=True)
    
    return templates.TemplateResponse(
        'tickets/report.html',
        {
            'request': request,
            'user': user,
            'workspace': workspace,
            'start_date': start_dt.strftime('%Y-%m-%d'),
            'end_date': (end_dt - timedelta(days=1)).strftime('%Y-%m-%d'),
            'now': datetime.utcnow(),
            'summary': summary,
            'agents': agents,
            'status_distribution': status_distribution,
            'category_distribution': category_distribution,
            'priority_distribution': priority_distribution,
            'closed_tickets': closed_tickets,
            'assigned_tickets': assigned_tickets,
            'open_tickets_list': open_tickets_list,
            'created_tickets': created_tickets,
            'comment_stats': comment_stats,
            'recent_comments': recent_comments,
            'activities': activities,
            'all_users': users_result,
            'selected_user': selected_user,
            'selected_user_id': selected_user_id,
            'all_projects': all_projects,
            'selected_project': selected_project,
            'selected_project_id': selected_project_id,
            'project_ticket_count': project_ticket_count,
            'project_ticket_counts': project_ticket_counts,
        },
    )


@router.get('/tickets/report/pdf')
async def web_tickets_report_pdf(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session),
):
    """Generate PDF ticket report - admin only"""
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    
    # Parse user_id and project_id filters
    user_id_int = None
    if user_id and user_id.strip():
        try:
            user_id_int = int(user_id)
        except ValueError:
            pass
    
    project_id_int = None
    if project_id and project_id.strip():
        try:
            project_id_int = int(project_id)
        except ValueError:
            pass
    
    # Get workspace for business hours settings
    workspace = await get_workspace_for_user(current_user_id, db)
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    # Build ticket query with filters
    ticket_query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.created_at >= start_dt,
        Ticket.created_at < end_dt
    )
    
    if user_id_int:
        ticket_query = ticket_query.where(
            (Ticket.assigned_to_id == user_id_int) | (Ticket.closed_by_id == user_id_int)
        )
    
    if project_id_int:
        ticket_query = ticket_query.where(Ticket.related_project_id == project_id_int)
    
    all_tickets = (await db.execute(ticket_query.order_by(Ticket.created_at.desc()))).scalars().all()
    
    ticket_ids = [t.id for t in all_tickets]
    
    # Query ALL currently open tickets (regardless of date range) - same as HTML report
    current_open_query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.status.in_(['open', 'in_progress', 'waiting'])
    )
    if user_id_int:
        current_open_query = current_open_query.where(
            (Ticket.assigned_to_id == user_id_int) | (Ticket.created_by_id == user_id_int)
        )
    if project_id_int:
        current_open_query = current_open_query.where(Ticket.related_project_id == project_id_int)
    all_current_open_tickets = (await db.execute(current_open_query.order_by(Ticket.created_at.desc()))).scalars().all()
    
    # Get all comments (filtered by user if specified)
    all_comments = []
    if ticket_ids:
        comment_query = select(TicketComment).where(TicketComment.ticket_id.in_(ticket_ids))
        if user_id_int:
            comment_query = comment_query.where(TicketComment.user_id == user_id_int)
        all_comments = (await db.execute(comment_query.order_by(TicketComment.created_at.desc()))).scalars().all()
    
    # Get all users for lookups
    users_result = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id)
    )).scalars().all()
    users_dict = {u.id: u for u in users_result}
    
    # Get business hours settings
    from app.core.business_hours import calculate_business_hours
    biz_start = workspace.business_hours_start if workspace and hasattr(workspace, 'business_hours_start') else "07:30"
    biz_end = workspace.business_hours_end if workspace and hasattr(workspace, 'business_hours_end') else "16:00"
    biz_exclude_weekends = workspace.business_hours_exclude_weekends if workspace and hasattr(workspace, 'business_hours_exclude_weekends') else True
    
    # Calculate agent performance
    agent_stats = {}
    for usr in users_result:
        agent_stats[usr.id] = {
            'name': usr.full_name or usr.username,
            'tickets_assigned': 0,
            'tickets_closed': 0,
            'comments_made': 0,
            'resolution_times': [],
        }
    
    for ticket in all_tickets:
        if ticket.assigned_to_id and ticket.assigned_to_id in agent_stats:
            agent_stats[ticket.assigned_to_id]['tickets_assigned'] += 1
        if ticket.closed_by_id and ticket.closed_by_id in agent_stats:
            agent_stats[ticket.closed_by_id]['tickets_closed'] += 1
            if ticket.closed_at:
                # Use business hours calculation
                resolution_hours = calculate_business_hours(
                    ticket.created_at, ticket.closed_at,
                    biz_start, biz_end, biz_exclude_weekends
                )
                agent_stats[ticket.closed_by_id]['resolution_times'].append(resolution_hours)
    
    for comment in all_comments:
        if comment.user_id and comment.user_id in agent_stats:
            agent_stats[comment.user_id]['comments_made'] += 1
    
    # Generate PDF
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    
    import io
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#1F2937'), spaceAfter=30, alignment=TA_CENTER)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#374151'), spaceAfter=12, spaceBefore=20)
    subheading_style = ParagraphStyle('CustomSubheading', parent=styles['Heading3'], fontSize=12, textColor=colors.HexColor('#6B7280'), spaceAfter=8)
    
    # Title
    elements.append(Paragraph("Ticket Report", title_style))
    elements.append(Paragraph(f"Period: {start_dt.strftime('%B %d, %Y')} - {(end_dt - timedelta(days=1)).strftime('%B %d, %Y')}", subheading_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subheading_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    total_tickets = len(all_tickets)
    closed_count = sum(1 for t in all_tickets if t.status == 'closed')
    # Use all_current_open_tickets count (ALL currently open, regardless of date range)
    current_open_count = len(all_current_open_tickets)
    resolution_rate = round((closed_count / total_tickets * 100) if total_tickets > 0 else 0)
    
    elements.append(Paragraph("Summary", heading_style))
    summary_data = [
        ['Metric', 'Value'],
        ['Total Tickets (in period)', str(total_tickets)],
        ['Closed Tickets', str(closed_count)],
        ['Currently Open Tickets', str(current_open_count)],
        ['Resolution Rate', f'{resolution_rate}%'],
        ['Total Comments', str(len(all_comments))],
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Agent Performance
    elements.append(Paragraph("Agent Performance", heading_style))
    agent_data = [['Agent', 'Assigned', 'Closed', 'Close Rate', 'Comments']]
    for uid, stats in sorted(agent_stats.items(), key=lambda x: x[1]['tickets_closed'], reverse=True):
        if stats['tickets_assigned'] > 0 or stats['tickets_closed'] > 0 or stats['comments_made'] > 0:
            close_rate = round((stats['tickets_closed'] / stats['tickets_assigned'] * 100) if stats['tickets_assigned'] > 0 else 0)
            agent_data.append([
                Paragraph(stats['name'], styles['Normal']),
                str(stats['tickets_assigned']),
                str(stats['tickets_closed']),
                f'{close_rate}%',
                str(stats['comments_made']),
            ])
    
    if len(agent_data) > 1:
        agent_table = Table(agent_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
        agent_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(agent_table)
    else:
        elements.append(Paragraph("No agent activity during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tickets Closed
    elements.append(Paragraph("Tickets Closed", heading_style))
    closed_data = [['Ticket #', 'Subject', 'Closed By', 'Priority', 'Resolution']]
    closed_tickets_with_billing = []
    for ticket in all_tickets:
        if ticket.status == 'closed' and ticket.closed_at:
            closed_by = users_dict.get(ticket.closed_by_id)
            # Use business hours calculation
            resolution_hours = calculate_business_hours(
                ticket.created_at, ticket.closed_at,
                biz_start, biz_end, biz_exclude_weekends
            )
            closed_data.append([
                ticket.ticket_number,
                Paragraph(ticket.subject, styles['Normal']),
                Paragraph((closed_by.full_name or closed_by.username) if closed_by else 'Unknown', styles['Normal']),
                ticket.priority.title(),
                f'{resolution_hours}h',
            ])
            # Collect billing info for a separate section
            has_billing = (ticket.billable_traveling or ticket.billable_labour_onsite or 
                          ticket.billable_remote_labour or ticket.billable_equipment_used or
                          ticket.non_billable_traveling or ticket.non_billable_labour_onsite or
                          ticket.non_billable_remote_labour or ticket.non_billable_equipment_used)
            if has_billing:
                closed_tickets_with_billing.append(ticket)
    
    if len(closed_data) > 1:
        closed_table = Table(closed_data, colWidths=[1.2*inch, 2*inch, 1.3*inch, 0.9*inch, 0.9*inch])
        closed_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(closed_table)
    else:
        elements.append(Paragraph("No tickets closed during this period.", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Billing Details for Closed Tickets
    if closed_tickets_with_billing:
        elements.append(Paragraph("Billing Details", heading_style))
        billing_data = [['Ticket #', 'Bill. Travel', 'Bill. On-site', 'Bill. Remote', 'Bill. Equip', 'NB Travel', 'NB On-site', 'NB Remote', 'NB Equip']]
        for ticket in closed_tickets_with_billing:
            billing_data.append([
                ticket.ticket_number,
                ticket.billable_traveling or '-',
                ticket.billable_labour_onsite or '-',
                ticket.billable_remote_labour or '-',
                ticket.billable_equipment_used or '-',
                ticket.non_billable_traveling or '-',
                ticket.non_billable_labour_onsite or '-',
                ticket.non_billable_remote_labour or '-',
                ticket.non_billable_equipment_used or '-',
            ])
        billing_table = Table(billing_data, colWidths=[0.9*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
        billing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(billing_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Open Tickets (ALL currently open, regardless of date range)
    elements.append(Paragraph("Open Tickets", heading_style))
    elements.append(Paragraph("All tickets currently open (regardless of date range)", subheading_style))
    open_data = [['Ticket #', 'Subject', 'Assigned To', 'Priority', 'Status', 'Days Open']]
    for ticket in all_current_open_tickets:
        assigned_to = users_dict.get(ticket.assigned_to_id)
        days_open = (datetime.now() - ticket.created_at).days
        open_data.append([
            ticket.ticket_number,
            Paragraph(ticket.subject[:50] + ('...' if len(ticket.subject) > 50 else ''), styles['Normal']),
            Paragraph((assigned_to.full_name or assigned_to.username) if assigned_to else 'Unassigned', styles['Normal']),
            ticket.priority.title(),
            ticket.status.replace('_', ' ').title(),
            f'{days_open}d',
        ])
    
    if len(open_data) > 1:
        open_table = Table(open_data, colWidths=[1.2*inch, 2*inch, 1.3*inch, 0.8*inch, 0.9*inch, 0.6*inch])
        open_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(open_table)
    else:
        elements.append(Paragraph("No tickets currently open.", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Recent Comments
    elements.append(Paragraph("Recent Comments", heading_style))
    tickets_dict = {t.id: t for t in all_tickets}
    comment_data = [['Date', 'Ticket #', 'Author', 'Comment']]
    for comment in all_comments[:30]:  # Limit to 30 most recent
        author = users_dict.get(comment.user_id)
        ticket = tickets_dict.get(comment.ticket_id)
        comment_text = comment.content[:80] + ('...' if len(comment.content) > 80 else '')
        comment_data.append([
            comment.created_at.strftime('%m/%d %H:%M'),
            ticket.ticket_number if ticket else 'N/A',
            Paragraph((author.full_name or author.username) if author else 'Guest', styles['Normal']),
            Paragraph(comment_text, styles['Normal']),
        ])
    
    if len(comment_data) > 1:
        comment_table = Table(comment_data, colWidths=[1*inch, 1*inch, 1.3*inch, 3.2*inch])
        comment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(comment_table)
    else:
        elements.append(Paragraph("No comments during this period.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    from fastapi.responses import StreamingResponse
    filename = f"ticket_report_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.get('/tickets/report/excel')
async def web_tickets_report_excel(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_session),
):
    """Generate Excel ticket report - admin only"""
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    
    # Parse user_id and project_id filters
    user_id_int = None
    if user_id and user_id.strip():
        try:
            user_id_int = int(user_id)
        except ValueError:
            pass
    
    project_id_int = None
    if project_id and project_id.strip():
        try:
            project_id_int = int(project_id)
        except ValueError:
            pass
    
    # Get workspace for business hours settings
    workspace = await get_workspace_for_user(current_user_id, db)
    
    # Parse date range
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_dt = datetime.now() - timedelta(days=30)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end_dt = datetime.now() + timedelta(days=1)
    
    # Build ticket query with filters
    ticket_query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.created_at >= start_dt,
        Ticket.created_at < end_dt
    )
    
    if user_id_int:
        ticket_query = ticket_query.where(
            (Ticket.assigned_to_id == user_id_int) | (Ticket.closed_by_id == user_id_int)
        )
    
    if project_id_int:
        ticket_query = ticket_query.where(Ticket.related_project_id == project_id_int)
    
    all_tickets = (await db.execute(ticket_query.order_by(Ticket.created_at.desc()))).scalars().all()
    
    ticket_ids = [t.id for t in all_tickets]
    
    # Query ALL currently open tickets (regardless of date range) - same as HTML and PDF report
    current_open_query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.status.in_(['open', 'in_progress', 'waiting'])
    )
    if user_id_int:
        current_open_query = current_open_query.where(
            (Ticket.assigned_to_id == user_id_int) | (Ticket.created_by_id == user_id_int)
        )
    if project_id_int:
        current_open_query = current_open_query.where(Ticket.related_project_id == project_id_int)
    all_current_open_tickets = (await db.execute(current_open_query.order_by(Ticket.created_at.desc()))).scalars().all()
    
    # Get all comments (filtered by user if specified)
    all_comments = []
    if ticket_ids:
        comment_query = select(TicketComment).where(TicketComment.ticket_id.in_(ticket_ids))
        if user_id_int:
            comment_query = comment_query.where(TicketComment.user_id == user_id_int)
        all_comments = (await db.execute(comment_query.order_by(TicketComment.created_at.desc()))).scalars().all()
    
    # Get ticket history (filtered by user if specified)
    all_history = []
    if ticket_ids:
        history_query = select(TicketHistory).where(TicketHistory.ticket_id.in_(ticket_ids))
        if user_id_int:
            history_query = history_query.where(TicketHistory.user_id == user_id_int)
        all_history = (await db.execute(history_query.order_by(TicketHistory.created_at.desc()))).scalars().all()
    
    # Get all users for lookups
    users_result = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id)
    )).scalars().all()
    users_dict = {u.id: u for u in users_result}
    
    # Get projects for lookups
    projects_result = (await db.execute(
        select(Project).where(Project.workspace_id == user.workspace_id)
    )).scalars().all()
    projects_dict = {p.id: p for p in projects_result}
    
    # Get business hours settings
    from app.core.business_hours import calculate_business_hours
    biz_start = workspace.business_hours_start if workspace and hasattr(workspace, 'business_hours_start') else "07:30"
    biz_end = workspace.business_hours_end if workspace and hasattr(workspace, 'business_hours_end') else "16:00"
    biz_exclude_weekends = workspace.business_hours_exclude_weekends if workspace and hasattr(workspace, 'business_hours_exclude_weekends') else True
    
    # Calculate agent performance
    agent_stats = {}
    for usr in users_result:
        agent_stats[usr.id] = {
            'name': usr.full_name or usr.username,
            'email': usr.email,
            'tickets_assigned': 0,
            'tickets_closed': 0,
            'comments_made': 0,
            'resolution_times': [],
        }
    
    for ticket in all_tickets:
        if ticket.assigned_to_id and ticket.assigned_to_id in agent_stats:
            agent_stats[ticket.assigned_to_id]['tickets_assigned'] += 1
        if ticket.closed_by_id and ticket.closed_by_id in agent_stats:
            agent_stats[ticket.closed_by_id]['tickets_closed'] += 1
            if ticket.closed_at:
                resolution_hours = calculate_business_hours(
                    ticket.created_at, ticket.closed_at,
                    biz_start, biz_end, biz_exclude_weekends
                )
                agent_stats[ticket.closed_by_id]['resolution_times'].append(resolution_hours)
    
    for comment in all_comments:
        if comment.user_id and comment.user_id in agent_stats:
            agent_stats[comment.user_id]['comments_made'] += 1
    
    # Generate Excel file
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    import io
    
    wb = Workbook()
    
    # Style helpers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_width(ws):
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adjusted_width
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Ticket Report"])
    ws_summary.append([f"Period: {start_dt.strftime('%Y-%m-%d')} to {(end_dt - timedelta(days=1)).strftime('%Y-%m-%d')}"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    if user_id_int:
        filter_user = users_dict.get(user_id_int)
        ws_summary.append([f"Filtered by User: {filter_user.full_name or filter_user.username if filter_user else 'Unknown'}"])
    if project_id_int:
        filter_project = projects_dict.get(project_id_int)
        ws_summary.append([f"Filtered by Project: {filter_project.name if filter_project else 'Unknown'}"])
    ws_summary.append([])
    
    total_tickets = len(all_tickets)
    closed_count = sum(1 for t in all_tickets if t.status == 'closed')
    # Use all_current_open_tickets count (ALL currently open, regardless of date range)
    current_open_count = len(all_current_open_tickets)
    resolution_rate = round((closed_count / total_tickets * 100) if total_tickets > 0 else 0)
    
    ws_summary.append(["Metric", "Value"])
    style_header_row(ws_summary, ws_summary.max_row, 2)
    ws_summary.append(["Total Tickets (in period)", total_tickets])
    ws_summary.append(["Closed Tickets", closed_count])
    ws_summary.append(["Currently Open Tickets", current_open_count])
    ws_summary.append(["Resolution Rate", f"{resolution_rate}%"])
    ws_summary.append(["Total Comments", len(all_comments)])
    ws_summary.append(["Business Hours", f"{biz_start} - {biz_end}"])
    auto_width(ws_summary)
    
    # Sheet 2: All Tickets
    ws_tickets = wb.create_sheet("All Tickets")
    ws_tickets.append(["Ticket #", "Subject", "Status", "Priority", "Category", "Assigned To", "Created", "Closed", "Resolution (hrs)", "Project"])
    style_header_row(ws_tickets, 1, 10)
    for ticket in all_tickets:
        assigned_to = users_dict.get(ticket.assigned_to_id)
        project = projects_dict.get(ticket.related_project_id)
        resolution_hrs = ''
        if ticket.status == 'closed' and ticket.closed_at:
            resolution_hrs = calculate_business_hours(
                ticket.created_at, ticket.closed_at,
                biz_start, biz_end, biz_exclude_weekends
            )
        ws_tickets.append([
            ticket.ticket_number,
            ticket.subject,
            ticket.status,
            ticket.priority,
            ticket.category or '',
            (assigned_to.full_name or assigned_to.username) if assigned_to else '',
            ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            ticket.closed_at.strftime('%Y-%m-%d %H:%M') if ticket.closed_at else '',
            resolution_hrs,
            project.name if project else '',
        ])
    auto_width(ws_tickets)
    
    # Sheet 3: Agent Performance
    ws_agents = wb.create_sheet("Agent Performance")
    ws_agents.append(["Agent", "Email", "Tickets Assigned", "Tickets Closed", "Close Rate", "Comments Made", "Avg Resolution (hrs)"])
    style_header_row(ws_agents, 1, 7)
    for uid, stats in sorted(agent_stats.items(), key=lambda x: x[1]['tickets_closed'], reverse=True):
        if stats['tickets_assigned'] > 0 or stats['tickets_closed'] > 0 or stats['comments_made'] > 0:
            close_rate = round((stats['tickets_closed'] / stats['tickets_assigned'] * 100) if stats['tickets_assigned'] > 0 else 0)
            avg_resolution = round(sum(stats['resolution_times']) / len(stats['resolution_times']), 1) if stats['resolution_times'] else 0
            ws_agents.append([
                stats['name'],
                stats['email'],
                stats['tickets_assigned'],
                stats['tickets_closed'],
                f"{close_rate}%",
                stats['comments_made'],
                avg_resolution,
            ])
    auto_width(ws_agents)
    
    # Sheet 4: Closed Tickets
    ws_closed = wb.create_sheet("Closed Tickets")
    ws_closed.append(["Ticket #", "Subject", "Priority", "Category", "Closed By", "Created", "Closed", "Resolution (hrs)",
                      "Billable Traveling", "Billable Labour On-site", "Billable Remote Labour", "Billable Equipment",
                      "Non-Billable Traveling", "Non-Billable Labour On-site", "Non-Billable Remote Labour", "Non-Billable Equipment",
                      "Closing Notes"])
    style_header_row(ws_closed, 1, 17)
    for ticket in all_tickets:
        if ticket.status == 'closed' and ticket.closed_at:
            closed_by = users_dict.get(ticket.closed_by_id)
            resolution_hours = calculate_business_hours(
                ticket.created_at, ticket.closed_at,
                biz_start, biz_end, biz_exclude_weekends
            )
            ws_closed.append([
                ticket.ticket_number,
                ticket.subject,
                ticket.priority,
                ticket.category or '',
                (closed_by.full_name or closed_by.username) if closed_by else 'Unknown',
                ticket.created_at.strftime('%Y-%m-%d %H:%M'),
                ticket.closed_at.strftime('%Y-%m-%d %H:%M'),
                resolution_hours,
                ticket.billable_traveling or '',
                ticket.billable_labour_onsite or '',
                ticket.billable_remote_labour or '',
                ticket.billable_equipment_used or '',
                ticket.non_billable_traveling or '',
                ticket.non_billable_labour_onsite or '',
                ticket.non_billable_remote_labour or '',
                ticket.non_billable_equipment_used or '',
                ticket.closing_notes or '',
            ])
    auto_width(ws_closed)
    
    # Sheet 5: Open Tickets (ALL currently open, regardless of date range)
    ws_open = wb.create_sheet("Open Tickets")
    ws_open.append(["Ticket #", "Subject", "Status", "Priority", "Category", "Assigned To", "Created", "Days Open", "Project"])
    style_header_row(ws_open, 1, 9)
    for ticket in all_current_open_tickets:
        assigned_to = users_dict.get(ticket.assigned_to_id)
        project = projects_dict.get(ticket.related_project_id)
        days_open = (datetime.now() - ticket.created_at).days
        ws_open.append([
            ticket.ticket_number,
            ticket.subject,
            ticket.status.replace('_', ' ').title(),
            ticket.priority,
            ticket.category or '',
            (assigned_to.full_name or assigned_to.username) if assigned_to else 'Unassigned',
            ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            days_open,
            project.name if project else '',
        ])
    auto_width(ws_open)
    
    # Sheet 6: Comments
    ws_comments = wb.create_sheet("Comments")
    ws_comments.append(["Date", "Ticket #", "Author", "Comment", "Internal"])
    style_header_row(ws_comments, 1, 5)
    tickets_dict = {t.id: t for t in all_tickets}
    for comment in all_comments:
        author = users_dict.get(comment.user_id)
        ticket = tickets_dict.get(comment.ticket_id)
        ws_comments.append([
            comment.created_at.strftime('%Y-%m-%d %H:%M'),
            ticket.ticket_number if ticket else '',
            (author.full_name or author.username) if author else 'Guest',
            comment.content,
            'Yes' if comment.is_internal else 'No',
        ])
    auto_width(ws_comments)
    
    # Sheet 7: Activity History
    ws_history = wb.create_sheet("Activity History")
    ws_history.append(["Date", "Ticket #", "User", "Action", "Details"])
    style_header_row(ws_history, 1, 5)
    for history in all_history:
        usr = users_dict.get(history.user_id)
        ticket = tickets_dict.get(history.ticket_id)
        details = ''
        if history.new_value:
            details = history.new_value
        ws_history.append([
            history.created_at.strftime('%Y-%m-%d %H:%M'),
            ticket.ticket_number if ticket else '',
            (usr.full_name or usr.username) if usr else 'System',
            history.action.replace('_', ' ').title(),
            details,
        ])
    auto_width(ws_history)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from fastapi.responses import StreamingResponse
    filename = f"ticket_report_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.get('/tickets', response_class=HTMLResponse)
async def web_tickets_list(request: Request, db: AsyncSession = Depends(get_session)):
    """List all tickets with filters"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket
    
    # Get filter parameters
    status_filter = request.query_params.get('status', 'all')
    priority_filter = request.query_params.get('priority', 'all')
    assigned_filter = request.query_params.get('assigned', 'all')
    project_filter = request.query_params.get('project', 'all')  # New: filter by project scope
    search_query = request.query_params.get('search', '').strip()
    
    # Base query - exclude closed and resolved (those go to archived tab)
    query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        Ticket.status.notin_(['closed', 'resolved'])
    )
    
    # Project-based visibility based on user permissions
    if not user.is_admin and not user.can_see_all_tickets:
        # Regular users: see only their project tickets and assigned tickets
        # Get projects where user is a member
        from app.models.project_member import ProjectMember
        user_projects_query = select(ProjectMember.project_id).where(
            ProjectMember.user_id == user_id
        )
        user_project_ids = (await db.execute(user_projects_query)).scalars().all()
        
        # User can see:
        # 1. Tickets assigned to them
        # 2. Tickets related to their projects
        # 3. Tickets with no project assignment (general tickets) if they can access main system
        query = query.where(
            (Ticket.assigned_to_id == user_id) |
            (Ticket.related_project_id.in_(user_project_ids)) |
            ((Ticket.related_project_id.is_(None)) & (Ticket.assigned_to_id == user_id))
        )
    
    # Project filter (for admins or filtering within allowed scope)
    if project_filter == 'main':
        # Show only tickets not related to any project (main ticket system)
        query = query.where(Ticket.related_project_id.is_(None))
    elif project_filter != 'all':
        try:
            project_id = int(project_filter)
            query = query.where(Ticket.related_project_id == project_id)
        except ValueError:
            pass
    
    # Apply other filters
    if status_filter != 'all':
        query = query.where(Ticket.status == status_filter)
    if priority_filter != 'all':
        query = query.where(Ticket.priority == priority_filter)
    if assigned_filter == 'me':
        query = query.where(Ticket.assigned_to_id == user_id)
    elif assigned_filter == 'unassigned':
        query = query.where(Ticket.assigned_to_id.is_(None))
    
    # Search filter
    if search_query:
        from sqlalchemy import or_, exists, case, literal
        from app.models.ticket import TicketComment
        search_pattern = f"%{search_query}%"
        
        # Subquery to find tickets with matching comments
        comment_match = exists().where(
            TicketComment.ticket_id == Ticket.id,
            TicketComment.content.ilike(search_pattern)
        )
        
        # Subquery to find tickets assigned to a user matching the search
        assigned_user_match = exists().where(
            User.id == Ticket.assigned_to_id,
            or_(
                User.full_name.ilike(search_pattern),
                User.email.ilike(search_pattern),
            )
        )
        
        # Subquery to find tickets created by a user matching the search
        created_user_match = exists().where(
            User.id == Ticket.created_by_id,
            or_(
                User.full_name.ilike(search_pattern),
                User.email.ilike(search_pattern),
            )
        )
        
        query = query.where(
            or_(
                Ticket.ticket_number.ilike(search_pattern),
                Ticket.subject.ilike(search_pattern),
                Ticket.description.ilike(search_pattern),
                Ticket.guest_email.ilike(search_pattern),
                Ticket.guest_name.ilike(search_pattern),
                Ticket.guest_surname.ilike(search_pattern),
                Ticket.guest_company.ilike(search_pattern),
                Ticket.guest_phone.ilike(search_pattern),
                Ticket.guest_office_number.ilike(search_pattern),
                Ticket.guest_branch.ilike(search_pattern),
                Ticket.closing_notes.ilike(search_pattern),
                comment_match,
                assigned_user_match,
                created_user_match,
            )
        )
        
        # Relevance ordering: name/email/ticket# matches first, then subject, then description/comments
        relevance = case(
            (Ticket.guest_name.ilike(search_pattern), literal(1)),
            (Ticket.guest_surname.ilike(search_pattern), literal(1)),
            (Ticket.guest_email.ilike(search_pattern), literal(1)),
            (Ticket.guest_company.ilike(search_pattern), literal(1)),
            (Ticket.guest_phone.ilike(search_pattern), literal(1)),
            (Ticket.ticket_number.ilike(search_pattern), literal(1)),
            (assigned_user_match, literal(1)),
            (created_user_match, literal(1)),
            (Ticket.subject.ilike(search_pattern), literal(2)),
            else_=literal(3)
        )
        query = query.order_by(relevance, Ticket.created_at.desc())
    else:
        query = query.order_by(Ticket.created_at.desc())
    tickets = (await db.execute(query)).scalars().all()
    
    # Get user's projects for filter dropdown
    user_projects = []
    if not user.is_admin and not user.can_see_all_tickets:
        # Limited users: show only their assigned projects
        from app.models.project_member import ProjectMember
        from app.models.project import Project
        user_projects_query = select(Project).join(
            ProjectMember, Project.id == ProjectMember.project_id
        ).where(ProjectMember.user_id == user_id)
        user_projects = (await db.execute(user_projects_query)).scalars().all()
    else:
        # Admins and users with full ticket access see all projects
        from app.models.project import Project
        user_projects = (await db.execute(
            select(Project).where(Project.workspace_id == user.workspace_id)
        )).scalars().all()
    
    # Get all users for assignment dropdown
    users = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id)
    )).scalars().all()
    
    # Build project lookup dict for showing project names on tickets
    projects_dict = {p.id: p for p in user_projects}
    
    return templates.TemplateResponse('tickets/list.html', {
        'request': request,
        'user': user,
        'tickets': tickets,
        'users': users,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'assigned_filter': assigned_filter,
        'project_filter': project_filter,
        'user_projects': user_projects,
        'projects_dict': projects_dict,
        'search_query': search_query
    })


@router.post('/tickets/create')
async def web_tickets_create(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Create a new ticket"""
    # Get all form data
    form_data = await request.form()
    
    # Extract form fields
    subject = form_data.get('subject')
    description = form_data.get('description') or None
    priority = form_data.get('priority', 'medium')
    category = form_data.get('category', 'general')
    assigned_to_id = form_data.get('assigned_to_id') or None
    scheduled_date = form_data.get('scheduled_date') or None
    ticket_working_days_list = form_data.getlist('ticket_working_days')
    # Customer information (optional)
    customer_name = form_data.get('customer_name') or None
    customer_surname = form_data.get('customer_surname') or None
    customer_email = form_data.get('customer_email') or None
    customer_phone = form_data.get('customer_phone') or None
    customer_company = form_data.get('customer_company') or None
    customer_branch = form_data.get('customer_branch') or None
    
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket, TicketHistory
    from app.models.meeting import Meeting
    from datetime import datetime
    
    # Convert assigned_to_id from string to int or None
    assigned_to_user_id = None
    if assigned_to_id and assigned_to_id.strip():
        try:
            assigned_to_user_id = int(assigned_to_id)
        except ValueError:
            pass
    
    # If non-admin and no assignee specified, auto-assign to creator
    if not user.is_admin and not assigned_to_user_id:
        assigned_to_user_id = user_id
    
    # Parse scheduled date if provided
    scheduled_datetime = None
    if scheduled_date:
        try:
            scheduled_datetime = datetime.fromisoformat(scheduled_date)
        except ValueError:
            pass
    
    # Parse working days (default to Mon-Fri if not provided)
    working_days_str = ','.join(ticket_working_days_list) if ticket_working_days_list else '0,1,2,3,4'
    
    # Generate ticket number using MAX to avoid duplicates after deletions
    year = datetime.utcnow().year
    from sqlalchemy import func, text
    # Get the highest existing ticket number for this year in this workspace
    prefix = f"TKT-{year}-"
    result = await db.execute(
        select(func.max(Ticket.ticket_number)).where(
            Ticket.workspace_id == user.workspace_id,
            Ticket.ticket_number.like(f"{prefix}%")
        )
    )
    max_ticket = result.scalar()
    if max_ticket:
        try:
            last_num = int(max_ticket.split('-')[-1])
        except (ValueError, IndexError):
            last_num = 0
    else:
        last_num = 0
    ticket_number = f"{prefix}{last_num + 1:05d}"
    
    # Create ticket
    ticket = Ticket(
        ticket_number=ticket_number,
        subject=subject,
        description=description,
        priority=priority,
        category=category,
        assigned_to_id=assigned_to_user_id,
        created_by_id=user_id,
        workspace_id=user.workspace_id,
        scheduled_date=scheduled_datetime,
        working_days=working_days_str,
        # Customer information - use guest_* fields
        guest_name=customer_name,
        guest_surname=customer_surname,
        guest_email=customer_email,
        guest_phone=customer_phone,
        guest_company=customer_company,
        guest_branch=customer_branch
    )
    db.add(ticket)
    await db.flush()
    
    # Create history entry
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=user_id,
        action='created',
        new_value=f'Ticket created with priority: {priority}'
    )
    db.add(history)
    
    # Log manual ticket creation for diagnostics
    try:
        from app.core.system_logger import log_fire_and_forget
        log_fire_and_forget('INFO', 'ticket', 'User Action',
            f'Manually created ticket #{ticket_number}',
            f'User={user.full_name or user.username} | Subject={subject[:80]} | Guest={customer_email or "N/A"} | Priority={priority}',
            user.workspace_id)
    except Exception:
        pass
    
    # Create calendar event if scheduled date is set and ticket is assigned
    if scheduled_datetime and assigned_to_user_id:
        # Create a meeting/calendar event for the ticket
        # Split datetime into date and time components for Meeting model
        meeting = Meeting(
            title=f"Ticket: {subject}",
            description=f"Ticket #{ticket_number}\n\n{description or 'No description provided'}",
            date=scheduled_datetime.date(),
            start_time=scheduled_datetime.time(),
            duration_minutes=30,  # Default 30 minutes
            platform='other',
            organizer_id=user_id,
            workspace_id=user.workspace_id
        )
        db.add(meeting)
        await db.flush()
        
        # Add assigned user as attendee
        from app.models.meeting import MeetingAttendee
        attendee = MeetingAttendee(
            meeting_id=meeting.id,
            user_id=assigned_to_user_id,
            status='pending'
        )
        db.add(attendee)
    
    # Create notification if assigned
    if assigned_to_user_id and assigned_to_user_id != user_id:
        # Check if assignee has muted ticket notifications
        assignee_user = (await db.execute(select(User).where(User.id == assigned_to_user_id))).scalar_one_or_none()
        if assignee_user and not getattr(assignee_user, 'mute_ticket_notifications', False):
            calendar_info = ""
            if scheduled_datetime:
                calendar_info = f" scheduled for {scheduled_datetime.strftime('%Y-%m-%d %H:%M')}"
            
            notification = Notification(
                user_id=assigned_to_user_id,
                type='ticket',
                message=f'{user.full_name or user.username} assigned you ticket #{ticket_number}: {subject}{calendar_info}',
                url=f'/web/tickets/{ticket.id}',
                related_id=ticket.id
            )
            db.add(notification)
    
    # Notify admins about the ticket creation (only those who haven't muted ticket notifications)
    admin_users = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id).where(User.is_admin == True)
    )).scalars().all()
    
    for admin in admin_users:
        if admin.id != user_id:  # Don't notify the creator if they're admin
            # Check if this admin has muted ticket notifications
            if getattr(admin, 'mute_ticket_notifications', False):
                continue  # Skip this admin
            
            calendar_info = ""
            if scheduled_datetime and assigned_to_user_id:
                assigned_user = (await db.execute(select(User).where(User.id == assigned_to_user_id))).scalar_one_or_none()
                assigned_name = assigned_user.full_name or assigned_user.username if assigned_user else "Unknown"
                calendar_info = f" and added to calendar for {assigned_name} on {scheduled_datetime.strftime('%Y-%m-%d %H:%M')}"
            
            notification = Notification(
                user_id=admin.id,
                type='ticket',
                message=f'{user.full_name or user.username} created ticket #{ticket_number}: {subject}{calendar_info}',
                url=f'/web/tickets/{ticket.id}',
                related_id=ticket.id
            )
            db.add(notification)
    
    await db.commit()
    
    # Track user behavior for learning
    try:
        from app.core.smart_suggestions import track_user_action
        await track_user_action(
            db=db,
            user_id=user_id,
            workspace_id=user.workspace_id,
            action_type="ticket_create",
            entity_type="ticket",
            entity_id=ticket.id,
            field_name="priority",
            field_value=priority
        )
        if assigned_to_user_id:
            await track_user_action(
                db=db,
                user_id=user_id,
                workspace_id=user.workspace_id,
                action_type="ticket_assign",
                entity_type="ticket",
                entity_id=ticket.id,
                field_name="assigned_to",
                field_value=str(assigned_to_user_id)
            )
    except Exception:
        pass  # Don't fail ticket creation if tracking fails
    
    return RedirectResponse(f'/web/tickets/{ticket.id}', status_code=303)


# Guest ticket routes (must be before /tickets/{ticket_id} to avoid route conflict)
@router.get('/tickets/guest', response_class=HTMLResponse)
async def web_tickets_guest_form(request: Request):
    """Public guest ticket submission form (no login required)"""
    return templates.TemplateResponse('tickets/guest.html', {
        'request': request,
        'success': False
    })


@router.get('/tickets/bubbles', response_class=HTMLResponse)
async def web_tickets_bubbles_chat(request: Request):
    """Bubbles AI Assistant chat page (no login required)"""
    return templates.TemplateResponse('tickets/bubbles.html', {
        'request': request
    })


@router.post('/tickets/guest')
async def web_tickets_guest_submit(
    request: Request,
    guest_name: str = Form(...),
    guest_surname: str = Form(...),
    guest_email: str = Form(...),
    guest_phone: str = Form(...),
    guest_company: str = Form(...),
    guest_office_number: Optional[str] = Form(None),
    guest_branch: Optional[str] = Form(None),
    subject: str = Form(...),
    description: str = Form(...),
    files: list[UploadFile] = File(default=[]),
    camera_files: list[UploadFile] = File(default=[]),
    db: AsyncSession = Depends(get_session)
):
    """Handle guest ticket submission with optional file attachments and camera photos"""
    from app.models.ticket import Ticket, TicketHistory, TicketAttachment
    from app.models.email_settings import EmailSettings
    from datetime import datetime
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import uuid
    import os
    
    try:
        # Combine regular files and camera files
        all_files = list(files) + list(camera_files)
        
        # Use workspace 1 as default for guest tickets
        workspace_id = 1
        
        # Auto-set priority to medium for guest tickets
        priority = 'medium'
        
        # Generate ticket number using MAX to avoid duplicates after deletions
        year = datetime.utcnow().year
        from sqlalchemy import func
        prefix = f"TKT-{year}-"
        result = await db.execute(
            select(func.max(Ticket.ticket_number)).where(
                Ticket.workspace_id == workspace_id,
                Ticket.ticket_number.like(f"{prefix}%")
            )
        )
        max_ticket = result.scalar()
        if max_ticket:
            try:
                last_num = int(max_ticket.split('-')[-1])
            except (ValueError, IndexError):
                last_num = 0
        else:
            last_num = 0
        ticket_number = f"{prefix}{last_num + 1:05d}"
        
        # Create ticket
        ticket = Ticket(
            ticket_number=ticket_number,
            subject=subject,
            description=description,
            priority=priority,
            category='support',
            status='open',
            workspace_id=workspace_id,
            is_guest=True,
            guest_name=guest_name,
            guest_surname=guest_surname,
            guest_email=guest_email,
            guest_phone=guest_phone,
            guest_company=guest_company,
            guest_office_number=guest_office_number,
            guest_branch=guest_branch,
            created_by_id=None  # No user account
        )
        db.add(ticket)
        await db.flush()
        
        # Handle file attachments (including camera photos)
        attachment_count = 0
        if all_files:
            # Create uploads directory if it doesn't exist
            upload_dir = BASE_DIR / 'uploads' / 'tickets'
            upload_dir.mkdir(parents=True, exist_ok=True)
            
            for file in all_files:
                if file.filename:  # Only process if file was actually uploaded
                    # Read file content
                    file_content = await file.read()
                    
                    # Validate file size (max 10MB)
                    if len(file_content) > 10 * 1024 * 1024:
                        continue  # Skip files that are too large
                    
                    # Generate unique filename
                    file_extension = os.path.splitext(file.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"
                    file_path = upload_dir / unique_filename
                    
                    # Save file to disk
                    with open(file_path, 'wb') as f:
                        f.write(file_content)
                    
                    # Store relative path from app directory
                    relative_path = f"app/uploads/tickets/{unique_filename}"
                    
                    # Create attachment record
                    attachment = TicketAttachment(
                        ticket_id=ticket.id,
                        filename=file.filename,
                        file_path=relative_path,
                        file_size=len(file_content),
                        mime_type=file.content_type or 'application/octet-stream',
                        uploaded_by_id=None  # Guest upload
                    )
                    db.add(attachment)
                    attachment_count += 1
        
        # Create history entry
        history = TicketHistory(
            ticket_id=ticket.id,
            user_id=None,
            action='created',
            new_value=f'Guest ticket created from {guest_email}' + (f' with {attachment_count} attachment(s)' if attachment_count > 0 else '')
        )
        db.add(history)
        
        # Notify all admins about new ticket (only those who haven't muted ticket notifications)
        from app.models.notification import Notification
        admin_users = (await db.execute(
            select(User).where(User.workspace_id == workspace_id).where(User.is_admin == True)
        )).scalars().all()
        
        for admin in admin_users:
            # Check if this admin has muted ticket notifications
            if getattr(admin, 'mute_ticket_notifications', False):
                continue  # Skip this admin
            
            notification = Notification(
                user_id=admin.id,
                type='ticket',
                message=f'New guest ticket #{ticket_number}: {subject}',
                url=f'/web/tickets/{ticket.id}',
                related_id=ticket.id
            )
            db.add(notification)
        
        await db.commit()
        
        # Try to send confirmation email
        email_sent = False
        try:
            # Get email settings
            email_settings = (await db.execute(
                select(EmailSettings).where(EmailSettings.workspace_id == workspace_id)
            )).scalar_one_or_none()
            
            if email_settings and email_settings.auto_reply_enabled:
                # Prepare email
                subject_template = email_settings.confirmation_subject
                body_template = email_settings.confirmation_body
                
                # Replace variables
                email_subject = subject_template.format(
                    ticket_number=ticket_number,
                    subject=subject,
                    priority=priority
                )
                
                email_body = body_template.format(
                    guest_name=guest_name,
                    guest_surname=guest_surname,
                    ticket_number=ticket_number,
                    subject=subject,
                    priority=priority,
                    company_name=email_settings.company_name
                )
                
                # Send email
                import uuid
                message_id = f"<{ticket_number}.{uuid.uuid4()}@{email_settings.smtp_host}>"
                
                msg = MIMEMultipart()
                msg['From'] = f"{email_settings.smtp_from_name} <{email_settings.smtp_from_email}>"
                msg['To'] = guest_email
                msg['Subject'] = email_subject
                msg['Message-ID'] = message_id
                msg.attach(MIMEText(email_body, 'plain'))
                
                server = smtplib.SMTP(email_settings.smtp_host, email_settings.smtp_port)
                if email_settings.smtp_use_tls:
                    server.starttls()
                server.login(email_settings.smtp_username, email_settings.smtp_password)
                server.send_message(msg)
                server.quit()
                
                email_sent = True
                
                # Store the message ID so replies can be tracked
                from app.models.processed_mail import ProcessedMail
                processed = ProcessedMail(
                    message_id=message_id,
                    email_from=email_settings.smtp_from_email,
                    subject=email_subject,
                    ticket_id=ticket.id,
                    workspace_id=workspace_id,
                    processed_at=datetime.utcnow()
                )
                db.add(processed)
                await db.commit()
        except Exception as e:
            logger.warning(f"Failed to send confirmation email: {e}")
            # Continue even if email fails
        
        return templates.TemplateResponse('tickets/guest.html', {
            'request': request,
            'success': True,
            'ticket_number': ticket_number,
            'guest_email': guest_email,
            'email_sent': email_sent
        })
        
    except Exception as e:
        logger.error(f"Guest ticket creation failed: {e}")
        return templates.TemplateResponse('tickets/guest.html', {
            'request': request,
            'success': False,
            'error': 'Failed to create ticket. Please try again or contact support.'
        })


# =====================================================
# SUPPORT ASSISTANT - Self-Learning Knowledge Base
# =====================================================

from app.models.support_kb import SupportArticle, SupportConversation, SupportCategory

async def search_duckduckgo(query: str, max_results: int = 5) -> list:
    """Search DuckDuckGo for troubleshooting information"""
    import urllib.parse
    import re
    
    try:
        import httpx
        
        # Use DuckDuckGo HTML search (no API key needed)
        search_url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query + ' troubleshooting solution')}"
        
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(search_url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            
            if response.status_code != 200:
                return []
            
            html = response.text
            results = []
            
            # Extract search results using regex
            # DuckDuckGo HTML results have class="result__a" for links and class="result__snippet" for descriptions
            link_pattern = r'class="result__a"[^>]*href="([^"]+)"[^>]*>([^<]+)</a>'
            snippet_pattern = r'class="result__snippet"[^>]*>([^<]+)</span>'
            
            links = re.findall(link_pattern, html)
            snippets = re.findall(snippet_pattern, html)
            
            for i, (url, title) in enumerate(links[:max_results]):
                snippet = snippets[i] if i < len(snippets) else ""
                # Clean up the snippet
                snippet = snippet.strip().replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
                results.append({
                    'title': title.strip(),
                    'url': url,
                    'snippet': snippet
                })
            
            return results
            
    except Exception as e:
        logger.error(f"DuckDuckGo search error: {e}")
        return []


async def extract_solution_steps(search_results: list) -> list:
    """Extract actionable solution steps from search results"""
    steps = []
    
    for result in search_results:
        snippet = result.get('snippet', '')
        title = result.get('title', '')
        
        # Create a suggested step from the snippet
        if snippet:
            # Clean up and format the step
            step = {
                'source': result.get('url', ''),
                'title': title,
                'suggestion': snippet,
                'confidence': 'medium'
            }
            steps.append(step)
    
    return steps


async def handle_conversational_message(message_lower: str, db: AsyncSession) -> str:
    """Handle greetings and conversational messages for Bubbles the AI Assistant"""
    
    # Use the comprehensive personality module
    response = get_conversational_response(message_lower)
    
    if response == "HUMAN_REQUEST":
        # Get support email from workspace settings
        support_email = "support@company.com"  # Default
        try:
            from app.models.email_settings import EmailSettings
            settings_result = await db.execute(
                select(EmailSettings).where(EmailSettings.workspace_id == 1)
            )
            email_settings = settings_result.scalar_one_or_none()
            if email_settings and email_settings.smtp_from_email:
                support_email = email_settings.smtp_from_email
        except Exception:
            pass
        
        return f"""I understand you'd like to speak with a human support agent! 👨‍💻 No problem at all!

Here are your options:

📧 **Email Support:** Send an email to **{support_email}** and our team will respond as soon as possible.

🎫 **Submit a Ticket:** Click the "Submit a Ticket" button above to create a detailed support request. Our tech team will review it and get back to you.

I'm Bubbles, and while I try my best to help, I know sometimes you need that human touch! 💫 Our support team is awesome and they'll take great care of you!"""
    
    return response


@router.post('/tickets/support/chat', response_class=JSONResponse)
async def support_assistant_chat(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Main support assistant chat endpoint - Bubbles the AI Assistant with advanced features"""
    try:
        from app.core.bubbles_ai import (
            detect_frustration, get_troubleshooting_flow, get_flow_step,
            CLARIFYING_QUESTIONS, get_video_tutorial, get_welcome_back_message,
            generate_ticket_prefill, get_status_message, BubblesLearning, SmartResponder
        )
        
        data = await request.json()
        message = data.get('message', '').strip()
        raw_session_id = data.get('session_id')
        session_id = raw_session_id if raw_session_id and raw_session_id != 'null' else str(uuid.uuid4())
        
        # Get additional context from frontend
        guest_email = data.get('guest_email', '').strip()
        flow_name = data.get('flow_name')  # Active troubleshooting flow
        flow_step = data.get('flow_step')  # Current step in flow
        conversation_history = data.get('conversation_history', [])
        
        if not message:
            return JSONResponse({'error': 'Message is required'}, status_code=400)
        
        message_lower = message.lower()
        
        # Initialize response data
        response_data = {
            'session_id': session_id,
            'is_conversational': False,
            'found_in_kb': False,
            'kb_articles': [],
            'web_results': [],
            'suggestions': [],
            'quick_replies': [],
            'flow_name': None,
            'flow_step': None,
            'empathy_prefix': None,
            'video_tutorial': None,
            'ticket_prefill': None,
            'learned_solution': None  # New: from learning system
        }
        
        # ===== FRUSTRATION DETECTION =====
        frustration_data = detect_frustration(message)
        if frustration_data['level'] >= 3:
            response_data['empathy_prefix'] = frustration_data['empathy_response']
            response_data['frustration_level'] = frustration_data['level']
        
        # ===== SMART CONTEXT DETECTION =====
        context_type, context_response = SmartResponder.detect_context(message)
        if context_type and context_response:
            response_data['is_conversational'] = True
            response_data['bubbles_response'] = context_response
            return JSONResponse(response_data)
        
        # ===== CHECK FOR RETURNING USER =====
        existing_conversation = None
        if guest_email:
            # Check for previous conversations
            result = await db.execute(
                select(SupportConversation).where(
                    SupportConversation.guest_email == guest_email
                ).order_by(SupportConversation.created_at.desc()).limit(1)
            )
            previous_convo = result.scalar_one_or_none()
            if previous_convo and previous_convo.session_id != session_id:
                # This is a returning user!
                welcome_back = get_welcome_back_message({
                    'last_issue': previous_convo.issue_category or previous_convo.initial_problem[:50],
                    'was_resolved': previous_convo.resolved,
                    'last_visit': previous_convo.created_at
                })
                if welcome_back and not flow_name:  # Don't interrupt active flows
                    response_data['welcome_back_message'] = welcome_back
        
        # ===== HANDLE ACTIVE TROUBLESHOOTING FLOW =====
        if flow_name and flow_step:
            flow = get_troubleshooting_flow(flow_name)
            if flow:
                step_data = get_flow_step(flow_name, flow_step)
                if step_data:
                    response_data['is_conversational'] = True
                    response_data['bubbles_response'] = step_data.get('message', '')
                    response_data['quick_replies'] = step_data.get('quick_replies', [])
                    response_data['flow_name'] = flow_name
                    
                    if step_data.get('is_success'):
                        response_data['flow_step'] = None  # End flow
                        response_data['show_success'] = True
                        # Learn from successful resolution
                        BubblesLearning.learn_from_resolution(
                            problem=f"[Flow: {flow_name}] User resolved issue",
                            solution=f"Completed troubleshooting flow: {flow_name}",
                            was_helpful=True,
                            category=flow_name
                        )
                    elif step_data.get('is_escalation'):
                        response_data['flow_step'] = None  # End flow
                        response_data['show_ticket_option'] = True
                        # Generate ticket prefill
                        response_data['ticket_prefill'] = generate_ticket_prefill({
                            'initial_problem': message,
                            'conversation_history': conversation_history,
                            'frustration_level': frustration_data['level']
                        })
                    else:
                        response_data['flow_step'] = flow_step
                    
                    return JSONResponse(response_data)
        
        # ===== CHECK LEARNED SOLUTIONS FIRST =====
        # Try to find solutions from past successful resolutions
        learned_matches = BubblesLearning.find_similar_solutions(message)
        if learned_matches and learned_matches[0]['confidence'] >= 0.6:
            best_match = learned_matches[0]
            response_data['is_conversational'] = True
            prefix = response_data['empathy_prefix'] + "\n\n" if response_data['empathy_prefix'] else ""
            response_data['bubbles_response'] = f"{prefix}🧠 I remember a similar issue! Here's what worked before:\n\n{best_match['solution']}"
            response_data['learned_solution'] = {
                'confidence': best_match['confidence'],
                'similarity': best_match['similarity'],
                'category': best_match.get('category')
            }
            # Don't return yet - still show KB results for comparison
        
        # ===== START NEW TROUBLESHOOTING FLOW =====
        # Check if message matches a flow trigger
        flow_triggers = {
            'printer': ['printer', 'print', 'printing', 'paper jam', 'not printing'],
            'wifi': ['wifi', 'wi-fi', 'internet', 'network', 'not connecting', 'no connection'],
            'email': ['email', 'outlook', 'mail', 'cant send', 'not receiving'],
            'slow_computer': ['slow', 'running slow', 'computer slow', 'laptop slow', 'freezing']
        }
        
        for flow_key, triggers in flow_triggers.items():
            if any(trigger in message_lower for trigger in triggers):
                flow = get_troubleshooting_flow(flow_key)
                if flow:
                    # Check for known issues first
                    status_msg = get_status_message(flow_key)
                    if status_msg:
                        response_data['system_status'] = status_msg
                    
                    response_data['is_conversational'] = True
                    prefix = response_data['empathy_prefix'] + "\n\n" if response_data['empathy_prefix'] else ""
                    response_data['bubbles_response'] = prefix + flow['start_question']
                    response_data['quick_replies'] = flow['quick_replies']
                    response_data['flow_name'] = flow_key
                    response_data['flow_step'] = 'start'
                    
                    # Add video tutorial if available
                    if flow_key == 'printer' and 'jam' in message_lower:
                        response_data['video_tutorial'] = get_video_tutorial('printer_jam')
                    elif flow_key == 'wifi':
                        response_data['video_tutorial'] = get_video_tutorial('wifi_reset')
                    
                    # Save conversation
                    result = await db.execute(
                        select(SupportConversation).where(SupportConversation.session_id == session_id)
                    )
                    conversation = result.scalar_one_or_none()
                    
                    if not conversation:
                        conversation = SupportConversation(
                            workspace_id=1,
                            session_id=session_id,
                            initial_problem=message,
                            guest_email=guest_email if guest_email else None,
                            issue_category=flow_key,
                            frustration_level=frustration_data['level']
                        )
                        db.add(conversation)
                    else:
                        conversation.issue_category = flow_key
                        conversation.frustration_level = frustration_data['level']
                        conversation.total_messages += 1
                        conversation.last_message_at = datetime.utcnow()
                    
                    await db.commit()
                    return JSONResponse(response_data)
        
        # ===== CHECK FOR CONVERSATIONAL MESSAGES =====
        conversational_response = await handle_conversational_message(message_lower, db)
        if conversational_response:
            response_data['is_conversational'] = True
            prefix = response_data['empathy_prefix'] + "\n\n" if response_data['empathy_prefix'] else ""
            response_data['bubbles_response'] = prefix + conversational_response
            return JSONResponse(response_data)
        
        # ===== CHECK FOR PRE-TRAINED BASIC SUPPORT =====
        pretrained = get_pretrained_response(message_lower)
        if pretrained:
            response_data['is_conversational'] = True
            prefix = response_data['empathy_prefix'] + "\n\n" if response_data['empathy_prefix'] else ""
            response_data['bubbles_response'] = prefix + pretrained['response']
            
            # Add clarifying questions if it's a vague issue
            if len(message.split()) < 5:  # Short message, might need clarification
                response_data['follow_up'] = CLARIFYING_QUESTIONS.get('device_type')
            
            return JSONResponse(response_data)
        
        # ===== SEARCH KNOWLEDGE BASE =====
        result = await db.execute(
            select(SupportConversation).where(SupportConversation.session_id == session_id)
        )
        conversation = result.scalar_one_or_none()
        
        if not conversation:
            conversation = SupportConversation(
                workspace_id=1,
                session_id=session_id,
                initial_problem=message,
                guest_email=guest_email if guest_email else None,
                frustration_level=frustration_data['level']
            )
            db.add(conversation)
            await db.commit()
            await db.refresh(conversation)
        
        keywords = message_lower.split()
        kb_results = []
        
        for keyword in keywords[:5]:
            if len(keyword) > 3:
                result = await db.execute(
                    select(SupportArticle).where(
                        and_(
                            SupportArticle.is_active == True,
                            or_(
                                SupportArticle.problem_keywords.ilike(f'%{keyword}%'),
                                SupportArticle.problem_description.ilike(f'%{keyword}%'),
                                SupportArticle.problem_title.ilike(f'%{keyword}%')
                            )
                        )
                    ).order_by(SupportArticle.success_rate.desc()).limit(3)
                )
                kb_results.extend(result.scalars().all())
        
        seen_ids = set()
        unique_kb_results = []
        for article in kb_results:
            if article.id not in seen_ids:
                seen_ids.add(article.id)
                unique_kb_results.append(article)
        
        response_data['found_in_kb'] = len(unique_kb_results) > 0
        
        if unique_kb_results:
            for article in unique_kb_results[:3]:
                response_data['kb_articles'].append({
                    'id': article.id,
                    'title': article.problem_title,
                    'problem': article.problem_description,
                    'solution': article.solution_steps,
                    'success_rate': article.success_rate,
                    'times_used': article.times_shown
                })
                article.times_shown += 1
            
            await db.commit()
            response_data['suggestions'].append({
                'type': 'kb_match',
                'message': f"I found {len(unique_kb_results)} solution(s) that might help with your issue."
            })
        else:
            web_results = await search_duckduckgo(message)
            solution_steps = await extract_solution_steps(web_results)
            response_data['web_results'] = solution_steps
            response_data['suggestions'].append({
                'type': 'web_search',
                'message': "I searched the web for solutions. Here's what I found:"
            })
        
        # Add empathy prefix to bubble response if frustrated
        if response_data['empathy_prefix']:
            response_data['bubbles_response'] = response_data['empathy_prefix']
        
        await db.commit()
        return JSONResponse(response_data)
        
    except Exception as e:
        logger.error(f"Support assistant error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse({'error': str(e)}, status_code=500)


@router.post('/tickets/support/feedback', response_class=JSONResponse)
async def support_assistant_feedback(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Handle feedback on support suggestions - with learning integration"""
    try:
        from app.core.bubbles_ai import BubblesLearning
        
        data = await request.json()
        session_id = data.get('session_id')
        helpful = data.get('helpful', False)
        article_id = data.get('article_id')
        problem = data.get('problem', '')
        solution = data.get('solution', '')
        category = data.get('category', '')
        
        # Get conversation
        result = await db.execute(
            select(SupportConversation).where(SupportConversation.session_id == session_id)
        )
        conversation = result.scalar_one_or_none()
        
        # ===== LEARNING INTEGRATION =====
        # Learn from this interaction
        if problem and solution:
            BubblesLearning.learn_from_resolution(
                problem=problem,
                solution=solution,
                was_helpful=helpful,
                category=category or (conversation.issue_category if conversation else None),
                tags=[]
            )
        
        if article_id:
            # Feedback for existing KB article
            result = await db.execute(
                select(SupportArticle).where(SupportArticle.id == article_id)
            )
            article = result.scalar_one_or_none()
            
            if article:
                if helpful:
                    article.times_helpful += 1
                    article.success_rate = (article.times_helpful / article.times_shown) * 100 if article.times_shown > 0 else 100.0
                    if conversation:
                        conversation.was_helpful = True
                        conversation.resolved = True
                        conversation.resolution_type = 'kb_article'
                else:
                    # Not helpful, decrease success rate
                    article.times_not_helpful += 1
                    article.success_rate = (article.times_helpful / (article.times_helpful + article.times_not_helpful)) * 100 if (article.times_helpful + article.times_not_helpful) > 0 else 0
                
                await db.commit()
        
        elif helpful and problem and solution:
            # Create new KB article from web search that was helpful
            # Extract keywords from the problem
            keywords = ','.join([word for word in problem.lower().split() if len(word) > 3][:10])
            
            new_article = SupportArticle(
                workspace_id=1,
                problem_title=problem[:100] + ('...' if len(problem) > 100 else ''),
                problem_description=problem,
                problem_keywords=keywords,
                solution_steps=solution,
                solution_source='web_search',
                times_shown=1,
                times_helpful=1,
                success_rate=100.0
            )
            db.add(new_article)
            
            if conversation:
                conversation.was_helpful = True
                conversation.resolved = True
                conversation.resolution_type = 'web_search'
            
            await db.commit()
            
            return JSONResponse({
                'success': True,
                'message': 'Thank you! This solution has been saved to our knowledge base. 🧠 I\'m learning!',
                'article_created': True
            })
        
        elif not helpful and conversation:
            # Mark that we need to escalate
            conversation.escalated_to_ticket = True
            await db.commit()
            
            return JSONResponse({
                'success': True,
                'message': 'Sorry the suggestions didn\'t help. You can create a ticket below.',
                'escalate': True
            })
        
        # Return learning stats (for fun!)
        learning_stats = BubblesLearning.get_learning_stats()
        
        return JSONResponse({
            'success': True,
            'learning_stats': {
                'solutions_learned': learning_stats['total_solutions_learned'],
                'patterns_recognized': learning_stats['unique_patterns']
            }
        })
        
    except Exception as e:
        logger.error(f"Support feedback error: {e}")
        return JSONResponse({'error': str(e)}, status_code=500)


@router.get('/tickets/support/search', response_class=JSONResponse)
async def support_search_web(
    q: str = Query(..., description="Search query"),
    db: AsyncSession = Depends(get_session)
):
    """Direct web search endpoint"""
    try:
        results = await search_duckduckgo(q, max_results=5)
        return JSONResponse({
            'query': q,
            'results': results
        })
    except Exception as e:
        logger.error(f"Web search error: {e}")
        return JSONResponse({'error': str(e)}, status_code=500)


# --------------------------
# Bubbles Analytics Dashboard
# --------------------------
@router.get('/support/analytics', response_class=HTMLResponse)
async def bubbles_analytics_dashboard(
    request: Request,
    days: int = Query(30, description="Number of days to analyze"),
    db: AsyncSession = Depends(get_session)
):
    """Bubbles Analytics Dashboard - shows what customers are asking"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Check permissions: must have can_see_all_tickets AND show_bubbles_analytics enabled
    if not (user.can_see_all_tickets and user.show_bubbles_analytics):
        raise HTTPException(status_code=403, detail="You don't have permission to view Bubbles analytics")
    
    from datetime import timedelta
    from collections import Counter
    
    # Date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    prev_start = start_date - timedelta(days=days)
    
    # Get conversations in date range
    convs_result = await db.execute(
        select(SupportConversation)
        .where(
            SupportConversation.workspace_id == user.workspace_id,
            SupportConversation.created_at >= start_date
        )
        .order_by(SupportConversation.created_at.desc())
    )
    conversations = convs_result.scalars().all()
    
    # Previous period for comparison
    prev_convs_result = await db.execute(
        select(SupportConversation)
        .where(
            SupportConversation.workspace_id == user.workspace_id,
            SupportConversation.created_at >= prev_start,
            SupportConversation.created_at < start_date
        )
    )
    prev_conversations = prev_convs_result.scalars().all()
    
    # Calculate stats
    total_conversations = len(conversations)
    prev_total = len(prev_conversations) or 1
    conversations_change = round(((total_conversations - prev_total) / prev_total) * 100, 1)
    
    resolved_by_bubbles = len([c for c in conversations if c.was_helpful])
    escalated_to_tickets = len([c for c in conversations if c.escalated_to_ticket])
    resolution_rate = round((resolved_by_bubbles / total_conversations * 100) if total_conversations > 0 else 0, 1)
    escalation_rate = round((escalated_to_tickets / total_conversations * 100) if total_conversations > 0 else 0, 1)
    
    # Get KB articles stats
    articles_result = await db.execute(
        select(SupportArticle)
        .where(SupportArticle.workspace_id == user.workspace_id)
        .order_by(SupportArticle.success_rate.desc())
    )
    all_articles = articles_result.scalars().all()
    
    kb_articles_used = len([a for a in all_articles if a.times_shown > 0])
    total_helpful = sum(a.times_helpful for a in all_articles)
    total_shown = sum(a.times_shown for a in all_articles) or 1
    avg_helpful_rate = round((total_helpful / total_shown) * 100, 1)
    
    # Top performing articles
    top_articles = sorted(all_articles, key=lambda a: (a.success_rate, a.times_shown), reverse=True)[:10]
    
    # Most common questions (analyze initial_problem)
    question_counter = Counter()
    failed_counter = Counter()
    
    for conv in conversations:
        if conv.initial_problem:
            # Normalize the question
            normalized = conv.initial_problem.lower().strip()[:100]
            question_counter[normalized] += 1
            
            # Track failed questions (not resolved)
            if not conv.was_helpful:
                failed_counter[normalized] += 1
    
    common_questions = []
    for text, count in question_counter.most_common(10):
        resolved_count = count - failed_counter.get(text, 0)
        common_questions.append({
            'text': text.capitalize(),
            'count': count,
            'resolved_rate': round((resolved_count / count) * 100) if count > 0 else 0
        })
    
    # Failed questions (needs improvement)
    failed_questions = []
    for text, count in failed_counter.most_common(10):
        if count >= 2:  # Only show if asked multiple times
            failed_questions.append({
                'text': text.capitalize(),
                'count': question_counter.get(text, count),
                'escalated': count
            })
    
    # Categories breakdown
    categories = [
        {'name': 'Email', 'icon': '📧', 'count': len([c for c in conversations if 'email' in (c.initial_problem or '').lower()])},
        {'name': 'Password', 'icon': '🔑', 'count': len([c for c in conversations if 'password' in (c.initial_problem or '').lower()])},
        {'name': 'Printer', 'icon': '🖨️', 'count': len([c for c in conversations if 'printer' in (c.initial_problem or '').lower() or 'print' in (c.initial_problem or '').lower()])},
        {'name': 'Network', 'icon': '📶', 'count': len([c for c in conversations if 'wifi' in (c.initial_problem or '').lower() or 'network' in (c.initial_problem or '').lower() or 'internet' in (c.initial_problem or '').lower()])},
        {'name': 'Software', 'icon': '💿', 'count': len([c for c in conversations if 'install' in (c.initial_problem or '').lower() or 'software' in (c.initial_problem or '').lower()])},
        {'name': 'Other', 'icon': '❓', 'count': 0}  # Calculate remaining
    ]
    categorized = sum(c['count'] for c in categories[:-1])
    categories[-1]['count'] = total_conversations - categorized
    
    # Recent conversations
    recent_conversations = conversations[:20]
    
    workspace = await get_workspace_for_user(user_id, db)
    
    return templates.TemplateResponse('support/analytics.html', {
        'request': request,
        'user': user,
        'workspace': workspace,
        'stats': {
            'total_conversations': total_conversations,
            'conversations_change': conversations_change,
            'resolved_by_bubbles': resolved_by_bubbles,
            'resolution_rate': resolution_rate,
            'escalated_to_tickets': escalated_to_tickets,
            'escalation_rate': escalation_rate,
            'kb_articles_used': kb_articles_used,
            'avg_helpful_rate': avg_helpful_rate
        },
        'common_questions': common_questions,
        'failed_questions': failed_questions,
        'top_articles': top_articles,
        'recent_conversations': recent_conversations,
        'categories': categories,
        'days': days
    })


# Pre-trained basic support knowledge (Bubbles handles simple issues ONLY)
BUBBLES_BASIC_SUPPORT = {
    # =========================================================================
    # POWER & STARTUP ISSUES
    # =========================================================================
    'not turning on': {
        'keywords': ['not turning on', 'wont turn on', "won't turn on", 'no power', 'dead', 'not starting', "doesn't turn on", 'doesnt turn on', 'wont start', "won't start", 'not powering on', 'not booting', 'wont boot', "won't boot"],
        'response': """I can help with that! 🔌 Let's get it working again:

**Step 1 - Check Power Supply:**
1. ✅ Is the power cable **firmly plugged in** at BOTH ends?
2. ✅ Is the **power outlet working**? (Test with another device like a phone charger)
3. ✅ If using a power strip, is it **switched ON**?
4. ✅ Look for any **LED lights** - even a small flicker

**Step 2 - Power Reset:**
1. ✅ **Unplug** the power cable completely
2. ✅ **Hold the power button** for 15-30 seconds (drains residual power)
3. ✅ **Plug back in** and try again

**Step 3 - For Laptops:**
1. ✅ Try **without the battery** (plugged in only) if removable
2. ✅ Try a **different charger** if available
3. ✅ Check if the **charging light** comes on

**Still no luck?** This likely needs hardware inspection. Would you like to create a ticket? 🎫"""
    },
    
    'blue screen': {
        'keywords': ['blue screen', 'bsod', 'blue screen of death', 'crashed', 'crash', 'system crash', 'keeps crashing', 'random restart', 'randomly restarts', 'sudden shutdown'],
        'response': """Blue screen or crashes? Let's figure this out! 💙

**Immediate Steps:**
1. ✅ **Write down the error code** if visible (like "CRITICAL_PROCESS_DIED")
2. ✅ **Restart** your computer normally
3. ✅ If it happens again, note **what you were doing** when it crashed

**Common Causes & Fixes:**
🔹 **Overheating:** Check vents for dust, make sure fans are running
🔹 **Recent changes:** Did you install new software or hardware recently?
🔹 **Windows Update:** Sometimes an update causes issues

**Quick Fixes to Try:**
1. ✅ **Restart** and let Windows repair itself
2. ✅ **Unplug** any new USB devices
3. ✅ Make sure your computer has **good ventilation**
4. ✅ Run **Windows Update** (sometimes fixes are released)

📸 **Tip:** If you see the blue screen, take a photo of the error code!

If crashes continue, please submit a ticket with the error details! 🎫"""
    },
    
    'startup slow': {
        'keywords': ['slow startup', 'slow boot', 'takes forever to start', 'takes long to boot', 'slow to turn on', 'boots slowly', 'starting slow'],
        'response': """Slow startup can be frustrating! 🐢➡️🚀 Let's speed it up:

**Quick Wins:**
1. ✅ **Restart** your computer (not shut down, actually restart)
2. ✅ Make sure you have at least **20% free disk space**
3. ✅ Close programs you don't need - they might be set to start automatically

**Check What's Running:**
1. ✅ Press **Ctrl+Shift+Esc** to open Task Manager
2. ✅ Click **"Startup"** tab
3. ✅ Look for programs with "High" impact - those slow you down

**Other Tips:**
- ✅ Check for **Windows Updates** (they can fix slow issues)
- ✅ Make sure your **antivirus isn't scanning** at startup
- ✅ Consider if your device is more than 5-7 years old (hardware aging)

Want detailed help optimizing your startup? Submit a ticket! 🎫"""
    },

    # =========================================================================
    # CONNECTION & NETWORK ISSUES  
    # =========================================================================
    'not connecting': {
        'keywords': ['not connecting', 'wont connect', "won't connect", 'no connection', 'cant connect', "can't connect", 'connection failed', 'connection lost', 'keeps disconnecting', 'drops connection'],
        'response': """Let me help you with connection issues! 🔗

**The Magic Reset (works 80% of the time!):**
1. ✅ **Turn OFF** the device completely
2. ✅ **Unplug** your router/modem for 30 seconds
3. ✅ **Plug back in** and wait 2 minutes
4. ✅ **Turn ON** your device

**Check The Basics:**
1. ✅ Are all cables **firmly connected**?
2. ✅ Do other devices work on the same network?
3. ✅ Are you within **range** for wireless?
4. ✅ Is the correct **network selected**?

**If Keeps Disconnecting:**
1. ✅ Move **closer to the router**
2. ✅ Check for **interference** (microwaves, other electronics)
3. ✅ Try connecting with a **cable** instead of WiFi

💡 **Pro Tip:** "Turn it off and on again" is IT advice for a reason - it actually works!

Need more help? Create a ticket for our tech team! 🎫"""
    },
    
    'wifi issues': {
        'keywords': ['wifi', 'wi-fi', 'wireless', 'no internet', 'internet not working', 'internet down', 'wifi slow', 'weak signal', 'wifi keeps dropping', 'cant find wifi', "can't find wifi", 'wifi not showing'],
        'response': """WiFi troubles? Let's fix that! 📶

**Step 1 - Restart Everything:**
1. ✅ Turn off WiFi on your device, wait 10 seconds, turn back on
2. ✅ **Restart your router** - unplug for 30 seconds, plug back in
3. ✅ Wait 2 minutes for it to fully boot up

**Step 2 - Reconnect Properly:**
1. ✅ "Forget" the network on your device
2. ✅ Search for networks again
3. ✅ Reconnect with the **correct password**

**Step 3 - Signal Strength:**
1. ✅ Move **closer to the router** to test
2. ✅ Check if walls/floors are blocking signal
3. ✅ Try different **locations** in the room

**If WiFi is Slow:**
- ✅ Too many devices? Try disconnecting some
- ✅ Check if someone's **streaming or downloading**
- ✅ Test with speedtest.net - is it your connection or device?

**Network Not Showing?**
- ✅ Router might be on **5GHz only** (some devices only see 2.4GHz)
- ✅ Check if router's **WiFi light** is on
- ✅ Try a different device to see if network appears

Still having issues? Submit a ticket and we'll dig deeper! 🎫"""
    },
    
    'ethernet': {
        'keywords': ['ethernet', 'lan cable', 'wired connection', 'network cable', 'cat5', 'cat6', 'rj45', 'cable internet', 'plugged in but no internet'],
        'response': """Wired connection not working? Let's check it! 🔌

**Basic Cable Checks:**
1. ✅ Is the cable **clicked in firmly** at both ends? (You should hear a click)
2. ✅ Check for **bent pins** or damaged cable
3. ✅ Try a **different cable** if you have one
4. ✅ Try a **different port** on the router

**Check The Lights:**
1. ✅ Look at the **port lights** on your router - is your port lit?
2. ✅ Check your computer's **ethernet port** - is the light blinking?
3. ✅ No lights = cable or port issue

**Computer Settings:**
1. ✅ Make sure ethernet is **enabled** in network settings
2. ✅ Check if WiFi is being **prioritized** over ethernet
3. ✅ Try **disabling WiFi** temporarily

**Quick Fix:**
1. ✅ Unplug cable from both ends
2. ✅ Wait 30 seconds
3. ✅ Plug back in firmly

Need hands-on help? Create a ticket! 🎫"""
    },
    
    'vpn': {
        'keywords': ['vpn', 'vpn not working', 'vpn wont connect', 'vpn slow', 'cant connect to vpn', 'remote access', 'work from home', 'company network'],
        'response': """VPN troubles? Let's sort this out! 🔐

**Basic VPN Fixes:**
1. ✅ **Close and reopen** the VPN application
2. ✅ Try **disconnecting and reconnecting**
3. ✅ Make sure your **internet is working** first (test without VPN)
4. ✅ **Restart** your computer

**Check Your Connection:**
1. ✅ Is your regular internet working?
2. ✅ Are you using correct **credentials**?
3. ✅ Is the VPN **server name** correct?

**Common VPN Issues:**
🔹 **"Connection timed out"** - Try a different server location
🔹 **Very slow** - Server might be overloaded, try another server
🔹 **Keeps disconnecting** - Check your internet stability

**For Work VPN:**
1. ✅ Check with IT if there's a **VPN outage**
2. ✅ Make sure your **password hasn't expired**
3. ✅ Some VPNs require you to be on specific networks

⚠️ For company VPN issues, it's best to contact your IT department or submit a ticket! 🎫"""
    },

    # =========================================================================
    # PRINTER ISSUES
    # =========================================================================
    'printer': {
        'keywords': ['printer', 'print', 'printing', 'not printing', 'printer offline', 'printer error', 'print queue', 'print job'],
        'response': """Printer problems? Let's troubleshoot! 🖨️

**Step 1 - Basic Checks:**
1. ✅ Is the printer **turned ON** with lights showing?
2. ✅ Is there **paper loaded** correctly?
3. ✅ Check **ink/toner levels** - is it running low?
4. ✅ Look for any **error lights** or messages on the printer

**Step 2 - The Magic Reset:**
1. ✅ Turn **OFF** the printer
2. ✅ Wait **30 seconds**
3. ✅ Turn it back **ON**
4. ✅ Try printing again

**Step 3 - Connection Check:**
🔹 **USB Printer:** Unplug and replug the USB cable
🔹 **WiFi Printer:** Make sure it's on the **same network** as your computer
🔹 **Network Printer:** Try pinging the printer's IP address

**Step 4 - Print Queue:**
1. ✅ Search "Printers" in Windows
2. ✅ Right-click your printer → **See what's printing**
3. ✅ **Cancel all documents** and try again

**Printer Showing "Offline"?**
1. ✅ Right-click printer → **See what's printing**
2. ✅ Click **Printer** menu → Uncheck **"Use Printer Offline"**

Need more help? Include your printer model in a ticket! 🎫"""
    },
    
    'paper jam': {
        'keywords': ['paper jam', 'jammed paper', 'paper stuck', 'jam', 'jammed', 'paper feed', 'paper not feeding', 'multiple pages', 'grabbing multiple'],
        'response': """Paper jam? ⚠️ **SAFETY FIRST!** 📄

**🛑 IMPORTANT: TURN OFF THE PRINTER FIRST!**
Always turn off the printer before removing jammed paper to avoid injury or damage!

**Step-by-Step Paper Jam Removal:**
1. ✅ **Turn OFF** the printer and unplug it
2. ✅ Wait 30 seconds for it to cool down
3. ✅ **Open all accessible doors and covers**
4. ✅ Look for paper - check **front, back, and sides**
5. ✅ **Pull paper SLOWLY and STRAIGHT** - don't yank!
6. ✅ Remove **ALL torn pieces** - even small bits cause problems
7. ✅ Close all covers and turn printer back on

**After Clearing:**
1. ✅ Let the printer do its startup routine
2. ✅ Print a **test page** to confirm it's clear
3. ✅ If error persists, there may be paper stuck deeper

**Prevent Future Jams:**
- ✅ Don't overfill the paper tray
- ✅ Use the **correct paper size/type**
- ✅ Fan paper before loading to prevent sticking
- ✅ Store paper flat in a dry place

**Still jammed?** Create a ticket with your printer model! 🎫"""
    },
    
    'printer quality': {
        'keywords': ['print quality', 'streaks', 'lines on paper', 'faded print', 'blurry print', 'smudged', 'light printing', 'dark spots', 'bands', 'banding'],
        'response': """Print quality issues? Let's fix that! 🎨

**For Faded/Light Prints:**
1. ✅ Check **ink/toner levels** - probably running low
2. ✅ Try **shaking the toner cartridge** gently (laser printers)
3. ✅ Run a **cleaning cycle** from printer settings

**For Streaks or Lines:**
1. ✅ Run **Print Head Cleaning** (found in printer settings)
2. ✅ Clean the **print heads** with a damp cloth if accessible
3. ✅ For laser printers, the **drum** might need replacing

**For Smudged Prints:**
1. ✅ Let prints **dry completely** before touching
2. ✅ Check if ink cartridge is **leaking**
3. ✅ Clean **inside the printer** for toner buildup

**How to Clean:**
1. ✅ Open printer settings on your computer
2. ✅ Find **Maintenance** or **Tools** section
3. ✅ Run **"Clean Print Heads"** or **"Cleaning Cycle"**
4. ✅ Print a test page

**Quick Tip:** Print a test page weekly if you don't print often - prevents clogs!

Need help identifying the issue? Take a photo and submit a ticket! 📸🎫"""
    },

    # =========================================================================
    # COMPUTER PERFORMANCE
    # =========================================================================
    'slow': {
        'keywords': ['slow', 'running slow', 'too slow', 'sluggish', 'lagging', 'laggy', 'takes forever', 'very slow', 'super slow', 'extremely slow'],
        'response': """Let's speed things up! 🚀

**Immediate Fixes:**
1. ✅ **Restart** your computer (seriously, this helps!)
2. ✅ Close **unused programs** and browser tabs
3. ✅ Check if something's **updating** in the background

**Check What's Using Resources:**
1. ✅ Press **Ctrl+Shift+Esc** to open Task Manager
2. ✅ Click **"More details"** if needed
3. ✅ Sort by **CPU** or **Memory** - what's using the most?
4. ✅ Close any programs using over 50% that you don't need

**Free Up Space:**
1. ✅ Check your **disk space** - keep at least 15-20% free
2. ✅ Empty the **Recycle Bin**
3. ✅ Delete old **downloads** you don't need
4. ✅ Clear **browser cache** (Ctrl+Shift+Delete in browser)

**Check for Issues:**
1. ✅ Is the computer **hot**? Check vents for dust
2. ✅ Run **Windows Update** - fixes often help
3. ✅ Make sure **antivirus** isn't doing a full scan

**Browser Slow?**
- ✅ Too many **extensions**? Disable some
- ✅ Too many **tabs**? (Each tab uses memory!)
- ✅ Try a different browser to test

Still slow? Submit a ticket for deeper troubleshooting! 🎫"""
    },
    
    'freezing': {
        'keywords': ['freezing', 'frozen', 'froze', 'not responding', 'hung', 'hangs', 'unresponsive', 'stuck', 'locked up', 'wont respond'],
        'response': """Computer frozen? Let's unfreeze it! 🥶

**If It's Frozen RIGHT NOW:**
1. ✅ **Wait 2-3 minutes** - it might be processing something heavy
2. ✅ Press **Ctrl+Alt+Delete** - does anything happen?
3. ✅ Try opening **Task Manager** (Ctrl+Shift+Esc)
4. ✅ If Task Manager opens, end the **unresponsive program**

**If Nothing Works:**
1. ✅ **Hold the power button** for 10 seconds to force shutdown
2. ✅ Wait 30 seconds
3. ✅ Turn it back on

**Prevent Future Freezes:**
1. ✅ Don't run too many **programs at once**
2. ✅ Keep **20% disk space free** minimum
3. ✅ Check **temperatures** - overheating causes freezes
4. ✅ Keep **Windows updated**

**If It Freezes Often:**
- ✅ Note **what you're doing** when it freezes
- ✅ Is it always the **same program**?
- ✅ Check if your **hard drive** is making clicking sounds (bad sign!)
- ✅ Run a **virus scan**

Freezing frequently? That needs investigation - submit a ticket! 🎫"""
    },
    
    'overheating': {
        'keywords': ['overheating', 'too hot', 'hot', 'fan loud', 'fans spinning', 'heat', 'thermal', 'getting hot', 'burning hot', 'fan noise', 'loud fan'],
        'response': """Computer running hot? Let's cool it down! 🌡️

**Immediate Actions:**
1. ✅ **Save your work** and close unnecessary programs
2. ✅ Make sure **vents aren't blocked** by blankets, papers, etc.
3. ✅ If laptop, make sure it's on a **hard flat surface** (not bed/couch)
4. ✅ Give it a **15-minute break** if possible

**Check for Dust:**
1. ✅ Look at the **vents** - are they dusty or blocked?
2. ✅ **Gently vacuum** or use compressed air on vents
3. ✅ Desktop? Check if fans are **spinning** inside

**Reduce Heat Generation:**
1. ✅ Close **heavy programs** (games, video editing, etc.)
2. ✅ Reduce **browser tabs**
3. ✅ Lower **screen brightness**
4. ✅ Turn off **Bluetooth and WiFi** if not needed

**Long-term Solutions:**
- ✅ Use a **laptop cooling pad**
- ✅ Keep in a **cool, ventilated area**
- ✅ Clean vents **every few months**
- ✅ Replace **thermal paste** (advanced - submit ticket!)

**Warning Signs to Watch:**
⚠️ Frequent shutdowns = overheating protection
⚠️ Fan ALWAYS loud = possible failing fan
⚠️ Very hot keyboard/bottom = needs cleaning or repair

If it's constantly hot, submit a ticket for deeper cleaning! 🎫"""
    },

    # =========================================================================
    # EMAIL ISSUES
    # =========================================================================
    'email': {
        'keywords': ['email', 'outlook', 'mail', 'cant send email', "can't send email", 'email not working', 'inbox', 'cant receive email', 'email stuck', 'sending failed', 'email error'],
        'response': """Email troubles? Let's fix that! 📧

**Can't Send or Receive:**
1. ✅ Check your **internet connection** first
2. ✅ Try **refreshing** or pressing Send/Receive
3. ✅ Check the **Outbox** - is email stuck there?
4. ✅ Close and **reopen** your email application

**Common Email Issues:**

🔹 **"Message stuck in Outbox":**
- Open Outbox, delete the stuck message
- Check attachment size (usually max 25MB)
- Recreate the email and try again

🔹 **"Can't connect to server":**
- Check internet connection
- Verify email settings haven't changed
- Try webmail (login via browser) to test

🔹 **"Password incorrect":**
- Caps Lock on?
- Password recently changed?
- Try "Forgot Password" option

🔹 **"Mailbox full":**
- Delete old emails
- Empty Deleted Items/Trash folder
- Archive old messages

**Try Web Version:**
1. ✅ Go to your email's website (gmail.com, outlook.com, etc.)
2. ✅ If web works but app doesn't, it's an app issue
3. ✅ Try removing and re-adding your account

Need help with email settings? Submit a ticket! 🎫"""
    },
    
    'outlook': {
        'keywords': ['outlook not opening', 'outlook crash', 'outlook slow', 'pst file', 'outlook stuck', 'outlook error', 'outlook offline', 'send receive error'],
        'response': """Outlook issues? Let's get you back on track! 📬

**Outlook Won't Open:**
1. ✅ **Force close** Outlook (Task Manager → End Task)
2. ✅ Try starting in **Safe Mode**: Hold Ctrl while clicking Outlook
3. ✅ Check for **Outlook updates**

**Outlook is Slow:**
1. ✅ Archive **old emails** (move to separate folder)
2. ✅ Compact your mailbox (File → Account Settings → Data Files → Compact)
3. ✅ Disable **unnecessary add-ins**
4. ✅ Keep mailbox under 2GB if possible

**Outlook Shows "Offline":**
1. ✅ Look at bottom of Outlook - does it say "Working Offline"?
2. ✅ Click **Send/Receive** tab → **Work Offline** to toggle
3. ✅ Check your **internet connection**

**Sync Issues:**
1. ✅ Click **Send/Receive All Folders**
2. ✅ Check **Send/Receive Progress** for errors
3. ✅ Try removing and re-adding the email account

**Profile Issues (Advanced):**
1. ✅ Control Panel → Mail → Show Profiles
2. ✅ Create a **new profile** and add your account
3. ✅ Set new profile as default

Outlook still acting up? Submit a ticket with the error message! 🎫"""
    },

    # =========================================================================
    # PASSWORD & LOGIN ISSUES
    # =========================================================================
    'password': {
        'keywords': ['password', 'forgot password', 'reset password', 'cant login', "can't login", 'login problem', 'locked out', 'account locked', 'wrong password', 'password expired', 'change password'],
        'response': """Account access issues? Let me help! 🔐

**"Forgot My Password":**
1. ✅ Click **"Forgot Password"** on the login page
2. ✅ Check your **email** (and spam folder!) for reset link
3. ✅ Follow the link to create a **new password**

**"Password Not Working":**
1. ✅ Check **Caps Lock** - passwords are case sensitive!
2. ✅ Check **Num Lock** if using numbers
3. ✅ Type password in **Notepad** first to see what you're typing
4. ✅ Has your password **expired**? (Some systems require changes every 90 days)

**"Account Locked":**
1. ✅ Usually unlocks after **15-30 minutes** of waiting
2. ✅ Or use **"Forgot Password"** to reset
3. ✅ Contact IT if it stays locked

**Creating Strong Passwords:**
- ✅ At least **12 characters**
- ✅ Mix of **uppercase, lowercase, numbers, symbols**
- ✅ Use a **passphrase**: "Coffee@Morning#2024!"
- ✅ Don't reuse passwords across sites!

**Password Tips:**
- ✅ Consider a **password manager** (like Bitwarden, LastPass)
- ✅ Write down hints, not actual passwords
- ✅ Set up **recovery options** (phone, backup email)

⚠️ I can't reset passwords directly for security. Use official reset options or submit a ticket! 🎫"""
    },
    
    'two factor': {
        'keywords': ['2fa', 'two factor', 'authenticator', 'verification code', 'mfa', 'multi factor', 'otp', 'one time password', 'cant get code', "can't get code", 'code not working'],
        'response': """Two-factor authentication issues? Let's solve this! 🔐

**Code Not Working:**
1. ✅ Check the **time** on your phone - must be accurate!
2. ✅ Make sure you're using code for the **right account**
3. ✅ Codes change every **30 seconds** - enter quickly!
4. ✅ Try the **next code** that appears

**Not Receiving SMS Codes:**
1. ✅ Check **phone signal** strength
2. ✅ Make sure phone number is **correct** in settings
3. ✅ Check if SMS is **blocked** by your carrier
4. ✅ Wait a few minutes and try **"Resend code"**

**Lost Access to Authenticator App:**
1. ✅ Use **backup codes** if you saved them
2. ✅ Check if you set up a **backup phone**
3. ✅ Contact support for the service to verify identity

**Authenticator App on New Phone:**
1. ✅ Set up new phone **BEFORE** wiping old one
2. ✅ Export accounts from old app if possible
3. ✅ Re-scan QR codes from account security settings

**Prevent Future Lockouts:**
- ✅ **Save backup codes** in a safe place!
- ✅ Set up **multiple methods** (app + phone)
- ✅ Write down **recovery options**

Completely locked out? Submit a ticket - we'll verify your identity! 🎫"""
    },

    # =========================================================================
    # DISPLAY & MONITOR ISSUES
    # =========================================================================
    'display': {
        'keywords': ['display', 'screen', 'monitor', 'black screen', 'blank screen', 'no display', 'flickering', 'resolution', 'screen not working', 'second monitor', 'external monitor', 'monitor not detected'],
        'response': """Display issues? Let's troubleshoot! 🖥️

**No Display at All:**
1. ✅ Is the monitor **powered on**? Check for power light
2. ✅ Is the **correct input** selected? (HDMI, DisplayPort, etc.)
3. ✅ Check **cable connections** at both ends
4. ✅ Try a **different cable** if available
5. ✅ Press a key to wake from **sleep mode**

**Monitor Not Detected:**
1. ✅ Unplug monitor, wait 30 seconds, plug back in
2. ✅ Press **Win+P** and select display mode
3. ✅ Right-click desktop → **Display Settings** → **Detect**
4. ✅ Try a different **port** (HDMI, USB-C, etc.)

**Flickering Screen:**
1. ✅ Check if cable is **loose**
2. ✅ Adjust **refresh rate**: Settings → Display → Advanced → 60Hz
3. ✅ Update **graphics drivers**
4. ✅ Check for **electrical interference**

**Wrong Resolution:**
1. ✅ Right-click desktop → **Display Settings**
2. ✅ Choose **recommended resolution**
3. ✅ Adjust **scaling** if things are too small/big

**Multiple Monitors:**
1. ✅ **Win+P** to switch between display modes
2. ✅ Extend, Duplicate, or Second screen only
3. ✅ Drag monitors in Display Settings to match physical layout

Still having issues? Describe what you see and submit a ticket! 🎫"""
    },
    
    'screen broken': {
        'keywords': ['cracked screen', 'broken screen', 'screen damage', 'lines on screen', 'dead pixels', 'bright spot', 'dark spot', 'bleeding', 'screen bleed'],
        'response': """Screen damage? Let me help assess this! 📱💔

**Types of Screen Issues:**

🔹 **Physical Crack:**
- Don't press on it - can spread
- Use as-is if functional, or get repaired
- Backup important data ASAP!

🔹 **Lines/Bars on Screen:**
- Horizontal/vertical lines = often cable or hardware issue
- Try connecting to external monitor to test
- If external works fine, it's the laptop screen

🔹 **Dead Pixels:**
- Small black dots that never light up
- Try "pixel fixing" videos on YouTube (gentle flashing colors)
- Usually permanent if software fix doesn't work

🔹 **Bright/Dark Spots:**
- Pressure damage or backlight issue
- Usually requires screen replacement

🔹 **Screen Bleeding (Light edges):**
- Light leaking around edges, often on dark screens
- Common on new screens, might reduce over time
- Only fixable by replacement if severe

**Immediate Steps:**
1. ✅ **Backup your data** - screens can fail suddenly
2. ✅ Connect to **external monitor** to keep working
3. ✅ **Don't open the device** yourself

⚠️ Screen repairs need professional help. Submit a ticket with a photo! 📸🎫"""
    },

    # =========================================================================
    # AUDIO ISSUES
    # =========================================================================
    'audio': {
        'keywords': ['audio', 'sound', 'no sound', 'speakers', 'volume', 'microphone', 'mic', 'cant hear', "can't hear", 'muted', 'headphones', 'earphones', 'sound not working', 'speaker not working'],
        'response': """No sound? Let's fix that! 🔊

**Check the Basics:**
1. ✅ Is **volume up** and not muted? (Check speaker icon)
2. ✅ Check **physical volume** on speakers/headphones
3. ✅ Are headphones **plugged in correctly**?
4. ✅ Is the correct **output device** selected?

**Check Output Device:**
1. ✅ Click the **speaker icon** in taskbar
2. ✅ Click the **arrow** to see all devices
3. ✅ Select the **correct speakers/headphones**

**Sound Settings:**
1. ✅ Right-click speaker icon → **Sound settings**
2. ✅ Make sure correct **output device** is selected
3. ✅ Click on the device → **Test** button

**Audio Not Working:**
1. ✅ **Restart** your computer
2. ✅ Unplug headphones/speakers, plug back in
3. ✅ Try a **different port** (front vs back, different USB)
4. ✅ Test with **different headphones/speakers**

**Microphone Issues:**
1. ✅ Check mic isn't **muted** in the app
2. ✅ Give the app **permission** to use microphone
3. ✅ Settings → Privacy → Microphone → Allow apps
4. ✅ Test in Sound settings

**Still No Sound:**
- ✅ Update **audio drivers**
- ✅ Try Windows **audio troubleshooter**
- ✅ Check Device Manager for issues

Need help? Submit a ticket! 🎫"""
    },
    
    'bluetooth audio': {
        'keywords': ['bluetooth', 'bluetooth speaker', 'bluetooth headphones', 'airpods', 'wireless headphones', 'bluetooth not connecting', 'bluetooth audio', 'pair', 'pairing'],
        'response': """Bluetooth audio troubles? Let's pair this up! 🎧

**Pairing New Device:**
1. ✅ Put your Bluetooth device in **pairing mode** (usually hold power button until light flashes)
2. ✅ On computer: Settings → **Bluetooth & devices** → **Add device**
3. ✅ Select your device from the list
4. ✅ Wait for "Connected" message

**Bluetooth Not Working:**
1. ✅ Toggle **Bluetooth off and on** in settings
2. ✅ **Restart** both computer and Bluetooth device
3. ✅ Make sure device is **charged**
4. ✅ Stay within **10 meters/30 feet**

**Audio Not Coming Through:**
1. ✅ Click speaker icon → select your **Bluetooth device**
2. ✅ Make sure it shows as **"Connected - Audio"**
3. ✅ Some devices need to be set as **default audio device**

**Audio Cutting Out:**
- ✅ Move **closer** to computer
- ✅ Remove other Bluetooth devices
- ✅ Check for **interference** (WiFi routers, microwaves)
- ✅ Update Bluetooth drivers

**Remove and Re-pair:**
1. ✅ Settings → Bluetooth → Find device → **Remove**
2. ✅ Restart computer
3. ✅ Put device in pairing mode and add again

Still having issues? Include your device model in a ticket! 🎫"""
    },

    # =========================================================================
    # SOFTWARE & APPLICATIONS
    # =========================================================================
    'error message': {
        'keywords': ['error', 'error message', 'error code', 'getting an error', 'shows error', 'error popup', 'error window', 'application error', 'program error'],
        'response': """I see you're getting an error! 🔍

**First Steps:**
1. ✅ **Write down** the exact error message/code
2. ✅ **Take a screenshot** (use camera button here!)
3. ✅ Note what you were **doing** when it appeared

**Quick Fixes:**
1. ✅ Click **OK/Close** and try the action again
2. ✅ **Restart** the application
3. ✅ **Restart** your computer
4. ✅ Check for **updates** for the software

**Common Error Types:**

🔹 **"Not enough memory/disk space":**
- Close other programs
- Free up disk space

🔹 **"Access denied":**
- Try running as Administrator
- Check if file is open elsewhere

🔹 **"File not found":**
- Check if file was moved/deleted
- Restore from backup

🔹 **"Connection error":**
- Check internet connection
- Try again later

🔹 **"Application has stopped working":**
- Restart the application
- Check for updates
- Reinstall if it keeps happening

**Pro Tip:** Search the exact error text online - often others have solved it!

Share your error with me or submit a ticket with a screenshot! 📸🎫"""
    },
    
    'program not opening': {
        'keywords': ['program not opening', 'app wont open', "app won't open", 'software not starting', 'application not responding', 'click but nothing happens', 'program wont start', "program won't start", 'not launching'],
        'response': """Program won't open? Let's fix that! 🚀

**Quick Fixes:**
1. ✅ **Wait a moment** - sometimes it takes time to load
2. ✅ Check the **taskbar** - it might be minimized
3. ✅ Look in **Task Manager** (Ctrl+Shift+Esc) - is it running?
4. ✅ **End the process** in Task Manager and try again

**If Still Won't Open:**
1. ✅ **Restart** your computer
2. ✅ Try running as **Administrator** (right-click → Run as Administrator)
3. ✅ Check for **updates** for the program
4. ✅ Check if your **computer meets requirements**

**Check for Conflicts:**
1. ✅ Temporarily disable **antivirus** and try
2. ✅ Close **other programs** that might conflict
3. ✅ Check if the program works in **Safe Mode**

**Repair or Reinstall:**
1. ✅ Settings → Apps → Find the program → **Repair** (if available)
2. ✅ If repair fails, try **uninstalling and reinstalling**
3. ✅ Download fresh installer from official website

**Still Not Working?**
- ✅ Note any error messages
- ✅ Check Windows **Event Viewer** for errors
- ✅ The program might be incompatible with your Windows version

Submit a ticket with the program name and what happens! 🎫"""
    },
    
    'install': {
        'keywords': ['install', 'installation', 'how to install', 'installing', 'setup', 'cant install', "can't install", 'installation failed', 'installer', 'download and install'],
        'response': """Need help installing software? Let's do it right! 📦

**Safe Installation Steps:**
1. ✅ **Download from official website** only
2. ✅ Check for the **correct version** (Windows/Mac, 32/64-bit)
3. ✅ **Temporarily disable antivirus** if it blocks installation
4. ✅ **Run as Administrator** if needed

**Installation Won't Start:**
1. ✅ Right-click installer → **Run as Administrator**
2. ✅ Check if download completed fully
3. ✅ Try downloading again from official source
4. ✅ Check **Windows Defender** didn't block it

**Installation Failed:**
1. ✅ Read the **error message** carefully
2. ✅ Make sure you have **enough disk space**
3. ✅ Close **all other programs**
4. ✅ **Restart** computer and try again
5. ✅ Check if program is **compatible** with your Windows version

**"Missing Dependencies":**
- ✅ Install **Visual C++ Redistributable** (search Microsoft website)
- ✅ Install **.NET Framework** if required
- ✅ Check program's requirements

**After Installation:**
1. ✅ **Restart** computer if prompted
2. ✅ Check for **updates** immediately
3. ✅ Enable antivirus again

⚠️ Only install software from **trusted sources**!

Having specific installation issues? Submit a ticket with the software name! 🎫"""
    },
    
    'uninstall': {
        'keywords': ['uninstall', 'remove program', 'delete program', 'remove software', 'uninstaller', 'how to remove', 'get rid of', 'remove app'],
        'response': """Need to uninstall something? Here's how! 🗑️

**Standard Uninstall (Windows):**
1. ✅ Press **Win + I** to open Settings
2. ✅ Click **Apps** → **Installed apps**
3. ✅ Find the program, click **⋮** → **Uninstall**
4. ✅ Follow the prompts

**Alternative Method:**
1. ✅ Type **"Add or remove programs"** in search
2. ✅ Find the program in the list
3. ✅ Click **Uninstall**

**Program Not Listed:**
1. ✅ Check program's folder for **Uninstall.exe**
2. ✅ Usually in C:\\Program Files\\ or C:\\Program Files (x86)\\
3. ✅ Run the uninstaller from there

**Uninstall Won't Complete:**
1. ✅ **Restart** computer and try again
2. ✅ Boot into **Safe Mode** and try
3. ✅ Use the program's **repair** option first, then uninstall

**Leftover Files:**
- ✅ Most programs leave some settings behind
- ✅ Check **%appdata%** folder for leftover folders
- ✅ Use **Disk Cleanup** to clear temp files

**Stubborn Programs:**
- ✅ Try the manufacturer's removal tool
- ✅ Some security software needs special uninstallers

Need help removing something specific? Submit a ticket! 🎫"""
    },
    
    'update': {
        'keywords': ['update', 'updating', 'install update', 'software update', 'system update', 'upgrade', 'windows update', 'update failed', 'update stuck', 'update error'],
        'response': """Updates can be tricky! Let me help! 🔄

**Windows Updates:**
1. ✅ Settings → **Windows Update** → **Check for updates**
2. ✅ Let it download and install
3. ✅ **Restart** when prompted
4. ✅ Be patient - some updates take 30+ minutes!

**Update Stuck:**
1. ✅ **Wait** - sometimes it seems stuck but is working
2. ✅ Leave it for at least **1 hour** before acting
3. ✅ If truly stuck, hold power to force shutdown
4. ✅ Restart and let Windows repair itself

**Update Failed:**
1. ✅ Try **Check for updates** again
2. ✅ Run **Windows Update Troubleshooter** (Settings → Troubleshoot)
3. ✅ Restart and try again
4. ✅ Free up **disk space** (need at least 20GB free)

**Before Big Updates:**
- ✅ **Backup important files** first!
- ✅ Make sure laptop is **plugged in**
- ✅ Have **stable internet**
- ✅ Clear some **disk space**

**Software Updates:**
1. ✅ Check within the program's **settings/help menu**
2. ✅ Or download from **official website**
3. ✅ Install over existing version (usually keeps settings)

**Don't:**
- ❌ Turn off computer during updates
- ❌ Unplug laptop during updates
- ❌ Interrupt the process

Having specific update errors? Screenshot the error and submit a ticket! 📸🎫"""
    },

    # =========================================================================
    # STORAGE & FILES
    # =========================================================================
    'storage': {
        'keywords': ['storage', 'disk space', 'no space', 'low storage', 'hard drive full', 'storage full', 'out of space', 'need space', 'free up space', 'disk full', 'c drive full'],
        'response': """Running out of space? Let's free some up! 💾

**Quick Wins:**
1. ✅ **Empty Recycle Bin** - right-click → Empty
2. ✅ Clear **Downloads folder** - delete old files
3. ✅ Run **Disk Cleanup**: Search "Disk Cleanup" → Select drive → Clean

**Disk Cleanup Options:**
1. ✅ Temporary files
2. ✅ Windows Update Cleanup (can be several GB!)
3. ✅ Delivery Optimization Files
4. ✅ Thumbnails

**Find Large Files:**
1. ✅ Settings → System → **Storage**
2. ✅ Click on drive to see what's using space
3. ✅ **Temporary files** - review and delete
4. ✅ Check **Large or unused files**

**More Space Savers:**
- ✅ Uninstall **programs you don't use**
- ✅ Move photos/videos to **external drive or cloud**
- ✅ Clear **browser cache** (Ctrl+Shift+Delete)
- ✅ Delete old **Windows.old** folder (if exists)

**Browser Cache (often 1-5GB!):**
1. ✅ Chrome: Settings → Privacy → Clear browsing data
2. ✅ Select **Cached images and files**
3. ✅ Time range: All time

**Cloud Options:**
- ✅ Use OneDrive/Google Drive/Dropbox
- ✅ Enable "Files On-Demand" to save space

Need help identifying what's taking space? Submit a ticket! 🎫"""
    },
    
    'files missing': {
        'keywords': ['file missing', 'files gone', 'lost file', 'file disappeared', 'cant find file', "can't find file", 'file deleted', 'accidentally deleted', 'recover file', 'where is my file'],
        'response': """Lost a file? Let's find it! 🔍

**Check These First:**
1. ✅ **Recycle Bin** - open and look for your file
2. ✅ **Downloads folder** - files often go here
3. ✅ **Desktop** - might be hidden behind windows
4. ✅ Check if it's in a **different folder**

**Search for It:**
1. ✅ Press **Win + S** to open search
2. ✅ Type the **filename** (or part of it)
3. ✅ Check results in Documents, Pictures, etc.
4. ✅ Use **wildcards**: *.docx finds all Word docs

**Recent Files:**
1. ✅ File Explorer → **Quick access** → **Recent files**
2. ✅ In the program (Word, Excel, etc.) → **File → Recent**
3. ✅ Check the program's **recent documents**

**Cloud Backup:**
- ✅ **OneDrive:** Check onedrive.live.com → Recycle Bin
- ✅ **Google Drive:** Check drive.google.com → Trash
- ✅ **Dropbox:** Check Deleted files section

**If Deleted:**
1. ✅ Restore from **Recycle Bin**
2. ✅ Check **File History** (if enabled)
3. ✅ Restore from **backup** if you have one

**Important Tips:**
- ⚠️ Don't save new files to the same location
- ⚠️ Stop using the drive if you need data recovery
- ⚠️ Professional recovery might be needed for important files

Can't find it anywhere? Submit a ticket with details! 🎫"""
    },
    
    'usb': {
        'keywords': ['usb', 'flash drive', 'thumb drive', 'usb drive', 'usb not recognized', 'usb not working', 'usb not showing', 'external drive', 'external hard drive', 'usb device'],
        'response': """USB drive not working? Let's troubleshoot! 🔌

**Quick Fixes:**
1. ✅ Try a **different USB port**
2. ✅ Try the **back ports** on desktop (more reliable)
3. ✅ Unplug, wait 10 seconds, plug back in
4. ✅ Try on a **different computer** to test the drive

**Check if Detected:**
1. ✅ Listen for the "device connected" **sound**
2. ✅ Check **File Explorer** for new drive letter
3. ✅ Search **"Disk Management"** - is it listed there?

**If Detected but Can't Access:**
1. ✅ The drive might need a **drive letter**:
   - Disk Management → Right-click drive → Change Drive Letter
2. ✅ Drive might need to be **formatted** (erases data!)
3. ✅ File system might be incompatible

**"USB Device Not Recognized":**
1. ✅ Try different port
2. ✅ Restart computer with USB unplugged
3. ✅ Plug in after restart
4. ✅ Update USB drivers (Device Manager)

**Drive Shows But Won't Open:**
- ✅ Might be **corrupted** - try chkdsk
- ✅ Could be **encrypted** - need password
- ✅ Try on different computer

**Safely Remove:**
- ✅ Always **eject** before unplugging!
- ✅ Click the USB icon in system tray → Eject

Important data on it? Submit a ticket before trying anything risky! 🎫"""
    },

    # =========================================================================
    # KEYBOARD & MOUSE
    # =========================================================================
    'keyboard': {
        'keywords': ['keyboard', 'keys not working', 'keyboard not working', 'key stuck', 'typing wrong', 'wrong characters', 'keyboard unresponsive', 'cant type', "can't type", 'keyboard layout'],
        'response': """Keyboard issues? Let's fix that! ⌨️

**No Response at All:**
1. ✅ Check if keyboard is **plugged in** (for wired)
2. ✅ Check **batteries** (for wireless)
3. ✅ Try a **different USB port**
4. ✅ **Restart** computer

**Typing Wrong Characters:**
1. ✅ Check **Num Lock** isn't making letters into numbers
2. ✅ Check keyboard **language/layout**:
   - Press **Win + Space** to switch layouts
   - Settings → Time & Language → Language
3. ✅ Look for "ENG" in taskbar - click to change

**Keys Stuck or Sticky:**
1. ✅ Gently **clean under the key** with compressed air
2. ✅ Check for **debris** under the key
3. ✅ For mechanical keyboards, the switch might be faulty

**Wireless Keyboard:**
1. ✅ Check/replace **batteries**
2. ✅ Re-pair with the receiver
3. ✅ Move **closer** to the receiver
4. ✅ Try removing and reinserting USB receiver

**On-Screen Keyboard (Emergency):**
1. ✅ Search **"On-Screen Keyboard"**
2. ✅ Or press **Win + Ctrl + O**
3. ✅ Use mouse to type temporarily

**Laptop Keyboard:**
- ✅ Check if you accidentally disabled it
- ✅ Check for **Fn key combinations**
- ✅ Connect external keyboard to test

Keys physically broken? That might need repair - submit a ticket! 🎫"""
    },
    
    'mouse': {
        'keywords': ['mouse', 'mouse not working', 'cursor not moving', 'mouse stuck', 'click not working', 'double click', 'mouse pointer', 'scroll not working', 'trackpad', 'touchpad'],
        'response': """Mouse troubles? Let's get it moving! 🖱️

**Mouse Not Working:**
1. ✅ Check if it's **plugged in** (wired)
2. ✅ Check/replace **batteries** (wireless)
3. ✅ Try a **different USB port**
4. ✅ Try a **different surface** (some surfaces confuse optical mice)

**Wireless Mouse:**
1. ✅ Check **batteries** - most common issue!
2. ✅ Check USB receiver is **plugged in**
3. ✅ Toggle **power switch** on mouse
4. ✅ Try **re-pairing** with receiver

**Cursor Moving Erratically:**
1. ✅ Clean the **sensor** on the bottom
2. ✅ Try a **different surface** or mouse pad
3. ✅ Move **wireless receiver closer**
4. ✅ Check for **interference**

**Clicks Not Working:**
1. ✅ Try **different applications** to test
2. ✅ Clean around the **mouse buttons**
3. ✅ Check mouse button settings in Windows
4. ✅ Mouse might be wearing out

**Touchpad (Laptop):**
1. ✅ Check if **disabled** - look for touchpad key (often F5 or F7)
2. ✅ Settings → Bluetooth & devices → **Touchpad** → Enable
3. ✅ Check if a mouse is disabling touchpad
4. ✅ **Restart** laptop

**Scroll Wheel:**
1. ✅ Clean around the wheel
2. ✅ Try scrolling in different apps
3. ✅ Check scroll settings in Windows

Need a replacement? Submit a ticket! 🎫"""
    },

    # =========================================================================
    # CAMERA & VIDEO
    # =========================================================================
    'camera': {
        'keywords': ['camera', 'webcam', 'video call', 'camera not working', 'cant see camera', "can't see camera", 'black camera', 'camera black', 'zoom camera', 'teams camera', 'video not working'],
        'response': """Camera not working? Let's fix that! 📷

**Quick Checks:**
1. ✅ Is the camera **physically covered**? (Check for privacy cover/slider!)
2. ✅ Is another app **already using** the camera?
3. ✅ Close other apps that might use camera (Zoom, Teams, etc.)

**Enable Camera:**
1. ✅ Settings → **Privacy & Security** → **Camera**
2. ✅ Make sure **Camera access** is ON
3. ✅ Check that your **app has permission** (list below)

**Test the Camera:**
1. ✅ Search for **"Camera"** app in Windows
2. ✅ Open it to test if camera works at all
3. ✅ If it works there, problem is app-specific

**In Video Apps (Zoom, Teams, etc.):**
1. ✅ Check app's **settings/preferences**
2. ✅ Select the **correct camera** from dropdown
3. ✅ Look for camera icon - is it muted/disabled?

**External Webcam:**
1. ✅ Unplug and **replug** the USB cable
2. ✅ Try a **different USB port**
3. ✅ Check if it needs **drivers** installed

**Laptop Camera:**
1. ✅ Check for **function key** to enable/disable (often F8 or F10 with camera icon)
2. ✅ Check Device Manager for camera

**Camera Shows Black:**
- ✅ Check privacy **cover/slider**!
- ✅ Restart the application
- ✅ Restart computer
- ✅ Check lighting - might be working but dark

Still not working? Include your laptop/webcam model in a ticket! 🎫"""
    },

    # =========================================================================
    # BROWSER ISSUES
    # =========================================================================
    'browser': {
        'keywords': ['browser', 'chrome', 'firefox', 'edge', 'browser slow', 'page not loading', 'website not working', 'browser crash', 'browser not responding', 'cant access website', "can't access website"],
        'response': """Browser troubles? Let's get you browsing again! 🌐

**Page Won't Load:**
1. ✅ Check your **internet connection** first
2. ✅ Try a **different website** - is it just one site?
3. ✅ Try a **different browser** to test
4. ✅ **Refresh** with Ctrl+F5 (forces full reload)

**Browser Running Slow:**
1. ✅ **Close extra tabs** - each tab uses memory!
2. ✅ Clear **cache and cookies** (Ctrl+Shift+Delete)
3. ✅ Disable **extensions** you don't use
4. ✅ **Restart** the browser

**Clear Browser Data:**
1. ✅ Press **Ctrl+Shift+Delete**
2. ✅ Select **Cached images and files**
3. ✅ Select **Cookies** (will log you out of sites)
4. ✅ Time range: **All time**
5. ✅ Click **Clear data**

**Browser Crashing:**
1. ✅ **Update** to latest version
2. ✅ Disable **all extensions** and test
3. ✅ Try **resetting** browser settings
4. ✅ Reinstall the browser

**"Can't reach this page":**
- ✅ Check if site is down: downforeveryoneorjustme.com
- ✅ Try clearing DNS: Open Command Prompt → type `ipconfig /flushdns`
- ✅ Try a different browser

**Extensions Causing Issues:**
1. ✅ Menu → Extensions
2. ✅ Disable all, then enable one by one
3. ✅ Find and remove the problematic one

Specific website not working? Submit a ticket with the URL! 🎫"""
    },

    # =========================================================================
    # MOBILE DEVICE ISSUES
    # =========================================================================
    'phone': {
        'keywords': ['phone', 'smartphone', 'mobile', 'iphone', 'android', 'phone slow', 'phone not charging', 'phone battery', 'phone storage', 'phone wont connect', 'phone wifi', 'company phone'],
        'response': """Phone issues? Let me help! 📱

**Phone Running Slow:**
1. ✅ **Restart** your phone (fixes most issues!)
2. ✅ Close **background apps**
3. ✅ Check **storage** - keep 10-15% free
4. ✅ Delete old **photos/videos** or move to cloud

**Phone Won't Charge:**
1. ✅ Try a **different cable**
2. ✅ Try a **different charger**
3. ✅ Clean the **charging port** gently (lint builds up!)
4. ✅ Try a different **outlet**

**WiFi Not Connecting:**
1. ✅ Toggle **WiFi off and on**
2. ✅ Forget network and **reconnect**
3. ✅ **Restart** phone and router
4. ✅ Check if other devices can connect

**Company Email on Phone:**
1. ✅ Check settings haven't changed
2. ✅ Make sure **password is current**
3. ✅ Try removing and re-adding account
4. ✅ Check with IT for server settings

**Battery Draining Fast:**
- ✅ Check **battery usage** in settings
- ✅ Reduce **screen brightness**
- ✅ Turn off **location** when not needed
- ✅ Close unused **background apps**
- ✅ Check for app using too much power

**Phone Frozen:**
1. ✅ **Force restart** - varies by phone:
   - iPhone: Volume up, volume down, then hold power
   - Android: Hold power + volume down for 10 seconds

For company phones, submit a ticket for support! 🎫"""
    },

    # =========================================================================
    # BACKUP & RECOVERY
    # =========================================================================
    'backup': {
        'keywords': ['backup', 'back up', 'save files', 'protect files', 'cloud backup', 'onedrive', 'google drive', 'file history', 'how to backup'],
        'response': """Let's protect your files with backups! 💾

**Quick Backup Options:**

🔹 **OneDrive (Built into Windows):**
1. ✅ Settings → Accounts → **Windows backup**
2. ✅ Turn on **folder backup** for Desktop, Documents, Pictures
3. ✅ Files auto-sync to cloud

🔹 **Google Drive:**
1. ✅ Download Google Drive desktop app
2. ✅ Sign in with Google account
3. ✅ Choose folders to sync

🔹 **External Drive:**
1. ✅ Plug in external USB drive
2. ✅ Copy important folders to it
3. ✅ Keep drive somewhere safe!

**What to Backup:**
- ✅ Documents
- ✅ Photos and Videos
- ✅ Desktop files
- ✅ Email data (if local)
- ✅ Browser bookmarks

**Windows File History:**
1. ✅ Settings → System → **Storage** → Advanced → **Backup options**
2. ✅ Add an external drive
3. ✅ Turn on automatic backup

**3-2-1 Backup Rule:**
- **3** copies of your data
- **2** different types of storage
- **1** copy offsite (cloud)

**Regular Backup Schedule:**
- ✅ Daily: Cloud sync (automatic)
- ✅ Weekly: External drive
- ✅ Monthly: Full system backup

Need help setting up backups? Submit a ticket! 🎫"""
    }
}

# Responses that Bubbles should REFUSE to help with (advanced/dangerous)
BUBBLES_REFUSE_TOPICS = [
    'registry', 'regedit', 'command prompt', 'cmd', 'terminal command', 'powershell',
    'bios', 'firmware', 'boot order', 'format', 'wipe', 'partition',
    'system32', 'delete files', 'admin password', 'hack', 'bypass',
    'root', 'jailbreak', 'unlock bootloader', 'flash', 'overclock',
    'drivers manually', 'dll', 'system restore', 'safe mode'
]


def get_pretrained_response(message_lower: str) -> Optional[dict]:
    """Check if Bubbles has a pre-trained response for this basic issue"""
    # Check for topics Bubbles should refuse
    for topic in BUBBLES_REFUSE_TOPICS:
        if topic in message_lower:
            return {
                'type': 'refuse',
                'response': f"""I appreciate you asking, but that's a bit **advanced** for me! 🙈

Topics like **{topic}** require careful handling by a trained technician. Making changes there without expertise could cause more problems.

🎫 **I recommend submitting a ticket** so our tech team can help you safely!

Is there something simpler I can help with instead? Like connection issues or basic troubleshooting?"""
            }
    
    # Check for pre-trained basic support
    for topic_key, topic_data in BUBBLES_BASIC_SUPPORT.items():
        for keyword in topic_data['keywords']:
            if keyword in message_lower:
                return {
                    'type': 'basic_support',
                    'response': topic_data['response']
                }
    
    return None


@router.post('/tickets/support/analyze-image', response_class=JSONResponse)
async def support_analyze_image(
    request: Request,
    image: UploadFile = File(...),
    session_id: str = Form(''),
    db: AsyncSession = Depends(get_session)
):
    """Analyze uploaded image for error messages using OCR"""
    import uuid
    
    try:
        # Generate session ID if not provided
        if not session_id:
            session_id = str(uuid.uuid4())
        
        # Read image
        image_data = await image.read()
        
        if len(image_data) > 5 * 1024 * 1024:  # 5MB limit
            return JSONResponse({
                'session_id': session_id,
                'error': 'Image is too large. Please use an image under 5MB.'
            })
        
        # Perform OCR using pytesseract if available, otherwise basic analysis
        extracted_text = ""
        ocr_available = False
        
        try:
            import pytesseract
            from PIL import Image
            import io
            
            # Convert to PIL Image
            img = Image.open(io.BytesIO(image_data))
            
            # Perform OCR
            extracted_text = pytesseract.image_to_string(img)
            ocr_available = True
            logger.info(f"OCR extracted text: {extracted_text[:200]}...")
            
        except ImportError:
            logger.warning("pytesseract not available, using basic image analysis")
        except Exception as ocr_error:
            logger.warning(f"OCR failed: {ocr_error}")
        
        # Prepare response
        response_data = {
            'session_id': session_id,
            'is_conversational': True,
            'found_in_kb': False,
            'kb_articles': [],
            'web_results': []
        }
        
        if extracted_text and len(extracted_text.strip()) > 10:
            # We got text from OCR - analyze it
            text_lower = extracted_text.lower()
            
            # Check for pre-trained responses first
            pretrained = get_pretrained_response(text_lower)
            if pretrained:
                response_data['bubbles_response'] = f"""📸 I analyzed your image!

{pretrained['response']}"""
                return JSONResponse(response_data)
            
            # Look for common error patterns
            error_keywords = ['error', 'failed', 'denied', 'unable', 'cannot', 'problem', 'exception', 'crash', 'warning']
            found_errors = [kw for kw in error_keywords if kw in text_lower]
            
            if found_errors:
                # Search KB and web for solutions based on extracted text
                # First try KB
                keywords = [word for word in text_lower.split() if len(word) > 3][:5]
                kb_results = []
                
                for keyword in keywords:
                    result = await db.execute(
                        select(SupportArticle).where(
                            and_(
                                SupportArticle.is_active == True,
                                or_(
                                    SupportArticle.problem_keywords.ilike(f'%{keyword}%'),
                                    SupportArticle.problem_description.ilike(f'%{keyword}%')
                                )
                            )
                        ).order_by(SupportArticle.success_rate.desc()).limit(2)
                    )
                    kb_results.extend(result.scalars().all())
                
                # Deduplicate
                seen = set()
                unique_kb = []
                for article in kb_results:
                    if article.id not in seen:
                        seen.add(article.id)
                        unique_kb.append(article)
                
                if unique_kb:
                    response_data['found_in_kb'] = True
                    response_data['kb_articles'] = [
                        {
                            'id': a.id,
                            'title': a.problem_title,
                            'problem': a.problem_description,
                            'solution': a.solution_steps,
                            'success_rate': a.success_rate
                        }
                        for a in unique_kb[:3]
                    ]
                    response_data['bubbles_response'] = f"""📸 I analyzed your image and found some text!

**I detected:** "{extracted_text[:150]}{'...' if len(extracted_text) > 150 else ''}"

I found some solutions in our knowledge base that might help:"""
                else:
                    # Search web
                    search_query = ' '.join(keywords[:3]) + ' error fix'
                    web_results = await search_duckduckgo(search_query, max_results=3)
                    solution_steps = await extract_solution_steps(web_results)
                    
                    response_data['web_results'] = solution_steps
                    response_data['bubbles_response'] = f"""📸 I analyzed your image!

**I detected:** "{extracted_text[:150]}{'...' if len(extracted_text) > 150 else ''}"

I searched for solutions and found some suggestions:"""
            else:
                # No error keywords found
                response_data['bubbles_response'] = f"""📸 I analyzed your image!

**I found this text:** "{extracted_text[:200]}{'...' if len(extracted_text) > 200 else ''}"

I don't see a specific error message. Could you describe what's happening? For example:
- What were you trying to do?
- What's not working as expected?

This helps me find the right solution for you! 💫"""
        
        else:
            # OCR couldn't extract text or not available
            if ocr_available:
                response_data['bubbles_response'] = """📸 I received your image, but I couldn't read any text from it clearly.

**Tips for better results:**
- Make sure the text in the image is **clear and readable**
- Try to capture just the **error message area**
- Ensure good **lighting** and focus

Could you **describe the error** you're seeing in words? I'd be happy to help! 💫"""
            else:
                response_data['bubbles_response'] = """📸 I received your image! Thanks for sharing.

Unfortunately, I can't analyze images directly right now. But I can still help!

**Please describe what you see:**
- What error message is showing?
- What were you trying to do?
- What device or software is this?

Type your description and I'll find solutions for you! 💫"""
        
        return JSONResponse(response_data)
        
    except Exception as e:
        logger.error(f"Image analysis error: {e}")
        return JSONResponse({
            'session_id': session_id or str(uuid.uuid4()),
            'error': 'Sorry, I had trouble analyzing that image. Please try describing your issue in words.'
        })


@router.get('/tickets/support/articles', response_class=JSONResponse)
async def get_support_articles(
    category: Optional[str] = None,
    limit: int = Query(10, le=50),
    db: AsyncSession = Depends(get_session)
):
    """Get knowledge base articles"""
    try:
        query = select(SupportArticle).where(SupportArticle.is_active == True)
        
        if category:
            query = query.where(SupportArticle.category_id == int(category))
        
        query = query.order_by(SupportArticle.success_rate.desc()).limit(limit)
        
        result = await db.execute(query)
        articles = result.scalars().all()
        
        return JSONResponse({
            'articles': [
                {
                    'id': a.id,
                    'title': a.problem_title,
                    'problem': a.problem_description,
                    'solution': a.solution_steps,
                    'success_rate': a.success_rate,
                    'times_used': a.times_shown
                }
                for a in articles
            ]
        })
    except Exception as e:
        logger.error(f"Get articles error: {e}")
        return JSONResponse({'error': str(e)}, status_code=500)


# Public ticket tracking routes (no login required)
@router.get('/tickets/track', response_class=HTMLResponse)
async def web_tickets_track_form(request: Request):
    """Public ticket tracking form"""
    error_message = request.session.pop('error_message', None)
    return templates.TemplateResponse('tickets/track.html', {
        'request': request,
        'error_message': error_message
    })


@router.post('/tickets/track')
async def web_tickets_track_submit(
    request: Request,
    ticket_number: str = Form(...),
    email: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    """Verify ticket and email, then show tracking details"""
    from app.models.ticket import Ticket
    
    # Find ticket by number
    result = await db.execute(
        select(Ticket).where(Ticket.ticket_number == ticket_number.strip())
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        request.session['error_message'] = 'Ticket not found. Please check the ticket number and try again.'
        return RedirectResponse('/web/tickets/track', status_code=303)
    
    # Verify email matches (check both guest email and requester email if user account)
    email_lower = email.strip().lower()
    ticket_emails = []
    
    if ticket.guest_email:
        ticket_emails.append(ticket.guest_email.lower())
    
    # If ticket has a user, check their email too
    if ticket.created_by_id:
        user = (await db.execute(select(User).where(User.id == ticket.created_by_id))).scalar_one_or_none()
        if user and user.email:
            ticket_emails.append(user.email.lower())
    
    if email_lower not in ticket_emails:
        request.session['error_message'] = 'Email address does not match this ticket. Please use the email you submitted the ticket with.'
        return RedirectResponse('/web/tickets/track', status_code=303)
    
    # Store verified ticket access in session
    verified_tickets = request.session.get('verified_tickets', [])
    if ticket_number not in verified_tickets:
        verified_tickets.append(ticket_number)
        request.session['verified_tickets'] = verified_tickets
    
    # Redirect to tracking detail page
    return RedirectResponse(f'/web/tickets/track/{ticket_number}', status_code=303)


@router.get('/tickets/track/{ticket_number}', response_class=HTMLResponse)
async def web_tickets_track_detail(
    request: Request,
    ticket_number: str,
    db: AsyncSession = Depends(get_session)
):
    """Show ticket tracking details (must have verified via POST first or have session)"""
    from app.models.ticket import Ticket, TicketComment
    
    # Check if user is logged in (staff) or has verified access via email
    user_id = request.session.get('user_id')
    verified_tickets = request.session.get('verified_tickets', [])
    
    if not user_id and ticket_number not in verified_tickets:
        request.session['error_message'] = 'Please verify your email to view this ticket.'
        return RedirectResponse('/web/tickets/track', status_code=303)
    
    # Find ticket
    result = await db.execute(
        select(Ticket).where(Ticket.ticket_number == ticket_number)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        request.session['error_message'] = 'Ticket not found.'
        return RedirectResponse('/web/tickets/track', status_code=303)
    
    # Get comments (only non-internal ones for public view)
    comments_result = await db.execute(
        select(TicketComment)
        .where(TicketComment.ticket_id == ticket.id)
        .where(TicketComment.is_internal == False)
        .order_by(TicketComment.created_at.asc())
    )
    comments = comments_result.scalars().all()
    
    # Load user info for comments in a single batch query (avoid N+1)
    user_ids = list(set(c.user_id for c in comments if c.user_id))
    users_map = {}
    if user_ids:
        users_result = await db.execute(select(User).where(User.id.in_(user_ids)))
        for u in users_result.scalars().all():
            users_map[u.id] = u
    
    comments_with_users = []
    for comment in comments:
        comments_with_users.append({
            'comment': comment,
            'user': users_map.get(comment.user_id) if comment.user_id else None
        })
    
    return templates.TemplateResponse('tickets/track_detail.html', {
        'request': request,
        'ticket': ticket,
        'comments': comments_with_users
    })


@router.post('/tickets/track/{ticket_number}/reply')
async def web_tickets_track_reply(
    request: Request,
    ticket_number: str,
    content: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    """Allow guest/client to reply to their ticket from the tracking page"""
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    
    # Verify access: must be logged in or have verified via email
    user_id = request.session.get('user_id')
    verified_tickets = request.session.get('verified_tickets', [])
    if not user_id and ticket_number not in verified_tickets:
        request.session['error_message'] = 'Please verify your email before replying.'
        return RedirectResponse('/web/tickets/track', status_code=303)
    
    # Find ticket
    result = await db.execute(
        select(Ticket).where(Ticket.ticket_number == ticket_number)
    )
    ticket = result.scalar_one_or_none()
    
    if not ticket:
        request.session['error_message'] = 'Ticket not found.'
        return RedirectResponse('/web/tickets/track', status_code=303)
    
    if ticket.status == 'closed':
        request.session['error_message'] = 'This ticket is closed and cannot accept new replies.'
        return RedirectResponse(f'/web/tickets/track/{ticket_number}', status_code=303)
    
    # Create the comment
    sender_name = f"{ticket.guest_name or ''} {ticket.guest_surname or ''}".strip() or ticket.guest_email or 'Guest'
    comment = TicketComment(
        ticket_id=ticket.id,
        user_id=None,  # Guest comment
        content=f"**Reply from {sender_name}:**\n\n{content}",
        is_internal=False,
        created_at=datetime.utcnow()
    )
    db.add(comment)
    
    # Update ticket timestamp
    ticket.updated_at = datetime.utcnow()
    
    # Add history entry
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=None,
        action='comment_added',
        new_value=f'Guest reply from {sender_name}',
        created_at=datetime.utcnow()
    )
    db.add(history)
    
    # Notify staff about guest reply
    from app.models.notification import Notification
    admin_users = (await db.execute(
        select(User).where(
            User.workspace_id == ticket.workspace_id,
            User.is_active == True,
            (User.is_admin == True) | (User.can_see_all_tickets == True)
        )
    )).scalars().all()
    
    for user in admin_users:
        if getattr(user, 'mute_ticket_notifications', False):
            continue
        notification = Notification(
            user_id=user.id,
            type='ticket',
            message=f'💬 Guest reply on ticket #{ticket.ticket_number} from {sender_name}',
            url=f'/web/tickets/{ticket.id}',
            related_id=ticket.id
        )
        db.add(notification)
    
    await db.commit()
    
    request.session['success_message'] = 'Your reply has been sent successfully.'
    return RedirectResponse(f'/web/tickets/track/{ticket_number}', status_code=303)


@router.get('/tickets/archived', response_class=HTMLResponse)
async def web_tickets_archived(request: Request, db: AsyncSession = Depends(get_session)):
    """View archived tickets"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket
    
    # Get search parameter
    search_query = request.query_params.get('search', '').strip()
    
    # Get archived tickets + closed/resolved tickets
    from sqlalchemy import or_
    query = select(Ticket).where(
        Ticket.workspace_id == user.workspace_id,
        or_(
            Ticket.is_archived == True,
            Ticket.status.in_(['closed', 'resolved'])
        )
    )
    
    # Search filter
    if search_query:
        from sqlalchemy import or_, exists
        from app.models.ticket import TicketComment
        search_pattern = f"%{search_query}%"
        
        # Subquery to find tickets with matching comments
        comment_match = exists().where(
            TicketComment.ticket_id == Ticket.id,
            TicketComment.content.ilike(search_pattern)
        )
        
        query = query.where(
            or_(
                Ticket.ticket_number.ilike(search_pattern),
                Ticket.subject.ilike(search_pattern),
                Ticket.description.ilike(search_pattern),
                Ticket.guest_email.ilike(search_pattern),
                Ticket.guest_name.ilike(search_pattern),
                Ticket.guest_surname.ilike(search_pattern),
                Ticket.guest_company.ilike(search_pattern),
                comment_match
            )
        )
    
    query = query.order_by(Ticket.archived_at.desc())
    tickets = (await db.execute(query)).scalars().all()
    
    return templates.TemplateResponse('tickets/archived.html', {
        'request': request,
        'user': user,
        'tickets': tickets,
        'search_query': search_query
    })


@router.get('/tickets/{ticket_id}', response_class=HTMLResponse)
async def web_tickets_detail(request: Request, ticket_id: int, db: AsyncSession = Depends(get_session)):
    """View ticket details with comments and history"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket, TicketComment, TicketAttachment, TicketHistory
    
    # Get ticket
    ticket = (await db.execute(
        select(Ticket).where(
            Ticket.id == ticket_id,
            Ticket.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not ticket:
        return RedirectResponse('/web/tickets', status_code=303)
    
    # Get creator
    creator = (await db.execute(select(User).where(User.id == ticket.created_by_id))).scalar_one_or_none()
    
    # Get assigned user
    assigned_user = None
    if ticket.assigned_to_id:
        assigned_user = (await db.execute(select(User).where(User.id == ticket.assigned_to_id))).scalar_one_or_none()
    
    # Get comments
    comments_result = await db.execute(
        select(TicketComment)
        .where(TicketComment.ticket_id == ticket_id)
        .order_by(TicketComment.created_at.asc())
    )
    comments = comments_result.scalars().all()
    
    # Get comment authors in a single batch query (avoid N+1)
    comment_user_ids = list(set(c.user_id for c in comments if c.user_id))
    comment_authors = {None: None}  # Pre-set None for guest comments
    if comment_user_ids:
        authors_result = await db.execute(select(User).where(User.id.in_(comment_user_ids)))
        for author in authors_result.scalars().all():
            comment_authors[author.id] = author
    
    # Get attachments
    attachments = (await db.execute(
        select(TicketAttachment).where(TicketAttachment.ticket_id == ticket_id)
    )).scalars().all()
    
    # Get history
    history = (await db.execute(
        select(TicketHistory)
        .where(TicketHistory.ticket_id == ticket_id)
        .order_by(TicketHistory.created_at.desc())
    )).scalars().all()
    
    # Get all users for assignment
    users = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id)
    )).scalars().all()
    
    # Get related project if exists
    related_project = None
    if ticket.related_project_id:
        from app.models.project import Project
        related_project = (await db.execute(
            select(Project).where(Project.id == ticket.related_project_id)
        )).scalar_one_or_none()
    
    # Get closed_by user if exists
    closed_by_user = None
    if ticket.closed_by_id:
        closed_by_user = (await db.execute(
            select(User).where(User.id == ticket.closed_by_id)
        )).scalar_one_or_none()
    
    return templates.TemplateResponse('tickets/detail.html', {
        'request': request,
        'user': user,
        'ticket': ticket,
        'creator': creator,
        'assigned_user': assigned_user,
        'comments': comments,
        'comment_authors': comment_authors,
        'attachments': attachments,
        'history': history,
        'users': users,
        'related_project': related_project,
        'closed_by_user': closed_by_user
    })


@router.post('/tickets/{ticket_id}/comment')
async def web_tickets_add_comment(
    request: Request,
    ticket_id: int,
    content: str = Form(...),
    is_internal: bool = Form(False),
    files: list[UploadFile] = File(default=[]),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: AsyncSession = Depends(get_session)
):
    """Add comment to ticket with optional file attachments"""
    from fastapi import BackgroundTasks
    import logging
    from datetime import datetime
    from pathlib import Path
    import uuid
    import os
    
    logger = logging.getLogger(__name__)
    logger.warning(f"🔔 COMMENT: Received comment for ticket {ticket_id}, is_internal={is_internal}")
    
    # Create detailed log file
    log_dir = Path("logs/comment_emails")
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"comment_{ticket_id}_{timestamp}.log"
    
    def write_log(message: str):
        """Write to both logger and file"""
        logger.warning(message)
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now().isoformat()} - {message}\n")
    
    write_log(f"🔔 COMMENT SUBMITTED: ticket_id={ticket_id}, is_internal={is_internal}")
    
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket, TicketComment, TicketHistory
    
    # Verify ticket exists
    ticket = (await db.execute(select(Ticket).where(Ticket.id == ticket_id))).scalar_one_or_none()
    if not ticket:
        write_log("❌ TICKET NOT FOUND")
        return RedirectResponse('/web/tickets', status_code=303)
    
    write_log(f"✓ Ticket found: #{ticket.ticket_number}, status={ticket.status}, is_guest={ticket.is_guest}, guest_email={ticket.guest_email}")
    
    # Check if ticket is closed
    if ticket.status == 'closed':
        write_log("❌ TICKET IS CLOSED - Cannot add comment")
        # Get user info
        user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
        
        # Send email notification about closed ticket
        if ticket.is_guest and ticket.guest_email:
            try:
                from app.models.email_settings import EmailSettings
                import smtplib
                from email.mime.text import MIMEText
                from email.mime.multipart import MIMEMultipart
                
                # Get email settings
                settings_result = await db.execute(
                    select(EmailSettings).where(EmailSettings.workspace_id == ticket.workspace_id)
                )
                email_settings = settings_result.scalar_one_or_none()
                
                if email_settings and email_settings.smtp_enabled:
                    # Create message
                    msg = MIMEMultipart('alternative')
                    msg['From'] = email_settings.smtp_username
                    msg['To'] = ticket.guest_email
                    msg['Subject'] = f"Ticket #{ticket.ticket_number} is Closed"
                    
                    body = f"""
This ticket has been marked as closed and cannot accept new comments.

Ticket: #{ticket.ticket_number}
Subject: {ticket.subject}

If you need further assistance, please:
- Contact support at {email_settings.smtp_username} to request reopening this ticket
- Or submit a new support ticket

Thank you.
"""
                    msg.attach(MIMEText(body, 'plain'))
                    
                    # Send email
                    if email_settings.smtp_use_ssl:
                        server = smtplib.SMTP_SSL(email_settings.smtp_host, email_settings.smtp_port)
                    else:
                        server = smtplib.SMTP(email_settings.smtp_host, email_settings.smtp_port)
                        if email_settings.smtp_use_tls:
                            server.starttls()
                    
                    server.login(email_settings.smtp_username, email_settings.smtp_password)
                    server.send_message(msg)
                    server.quit()
            except Exception as e:
                logger.warning(f"Error sending closed ticket notification: {e}")
        
        # Redirect with error message
        request.session['error_message'] = 'This ticket is closed and cannot accept new comments. Please contact support to reopen.'
        return RedirectResponse(f'/web/tickets/{ticket_id}', status_code=303)
    
    # Add comment
    comment = TicketComment(
        ticket_id=ticket_id,
        user_id=user_id,
        content=content,
        is_internal=is_internal
    )
    db.add(comment)
    await db.flush()  # Get comment.id for attachments
    
    # Handle file attachments
    from app.models.ticket import TicketAttachment
    attachment_count = 0
    if files:
        # Create uploads directory if it doesn't exist
        upload_dir = BASE_DIR / 'uploads' / 'tickets'
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        for file in files:
            if file.filename:  # Only process if file was actually uploaded
                # Read file content
                file_content = await file.read()
                
                # Validate file size (max 10MB)
                if len(file_content) > 10 * 1024 * 1024:
                    write_log(f"⚠️ File {file.filename} is too large, skipping")
                    continue
                
                # Generate unique filename
                file_extension = os.path.splitext(file.filename)[1]
                unique_filename = f"{uuid.uuid4()}{file_extension}"
                file_path = upload_dir / unique_filename
                
                # Save file to disk
                with open(file_path, 'wb') as f:
                    f.write(file_content)
                
                # Store relative path from app directory
                relative_path = f"app/uploads/tickets/{unique_filename}"
                
                # Create attachment record
                attachment = TicketAttachment(
                    ticket_id=ticket_id,
                    comment_id=comment.id,
                    filename=file.filename,
                    file_path=relative_path,
                    file_size=len(file_content),
                    mime_type=file.content_type or 'application/octet-stream',
                    uploaded_by_id=user_id
                )
                db.add(attachment)
                attachment_count += 1
                write_log(f"✓ Attachment saved: {file.filename} ({len(file_content)} bytes)")
    
    # Update ticket timestamp
    from datetime import datetime
    ticket.updated_at = datetime.utcnow()
    
    # Add history
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        action='commented',
        new_value=f'Added a comment' + (f' with {attachment_count} attachment(s)' if attachment_count > 0 else '')
    )
    db.add(history)
    
    await db.commit()
    await db.refresh(comment)
    
    write_log(f"✓ Comment added successfully, ID={comment.id}, attachments={attachment_count}")
    
    # Send email notification to client in background (if not internal comment)
    write_log(f"📧 EMAIL CHECK: is_internal={is_internal}, guest_email='{ticket.guest_email}', is_guest={ticket.is_guest}")
    if not is_internal and ticket.guest_email:
        write_log(f"✅ WILL SEND EMAIL to {ticket.guest_email} for ticket #{ticket.ticket_number}")
        
        # Send email directly (moved to async task that won't block response)
        try:
            await send_ticket_comment_email(ticket, content, user_id, db, write_log)
            write_log("✅ Email sent successfully")
        except Exception as e:
            write_log(f"❌ EMAIL FAILED: {e}")
            import traceback
            write_log(f"Traceback: {traceback.format_exc()}")
            # Don't fail the request if email fails
    else:
        write_log(f"❌ NOT SENDING EMAIL: is_internal={is_internal}, guest_email='{ticket.guest_email}'")
    
    write_log(f"✓ COMPLETE - Log saved to: {log_file}")
    return RedirectResponse(f'/web/tickets/{ticket_id}', status_code=303)




async def send_ticket_comment_email(ticket, content: str, user_id: int, db: AsyncSession, write_log=None):
    """Send email notification in background (non-blocking)"""
    from app.models.ticket import Ticket
    
    def log(msg):
        if write_log:
            write_log(msg)
        else:
            logger.debug(msg)
    
    try:
        log(f"[EMAIL] send_ticket_comment_email called for ticket #{ticket.ticket_number}")
        
        # Get user info
        user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
        log(f"[EMAIL] User: {user.username if user else 'None'}")
        
        # Get email settings (always needed for SMTP connection)
        from app.models.email_settings import EmailSettings
        settings_result = await db.execute(
            select(EmailSettings).where(EmailSettings.workspace_id == ticket.workspace_id)
        )
        email_settings = settings_result.scalar_one_or_none()
        
        if not email_settings:
            log(f"❌ No email settings configured for workspace {ticket.workspace_id}")
            return
        
        # Check if SMTP settings are properly configured
        if not email_settings.smtp_host or not email_settings.smtp_username or not email_settings.smtp_password:
            log(f"❌ SMTP settings incomplete for workspace {ticket.workspace_id}")
            return
        
        if not email_settings.smtp_from_email:
            log(f"❌ SMTP 'From' email not configured for workspace {ticket.workspace_id}")
            return
        
        log(f"[EMAIL] SMTP settings found: {email_settings.smtp_host}:{email_settings.smtp_port}")
        
        # Always send from the main workspace email
        from_email = email_settings.smtp_from_email
        from_name = email_settings.smtp_from_name or "Support Team"
        
        log(f"[EMAIL] Sending from: {from_name} <{from_email}>")
        
        # Send email if we have a sender address
        if from_email:
            log(f"[EMAIL] Preparing to send email to {ticket.guest_email}")
            
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.utils import make_msgid
            from app.models.processed_mail import ProcessedMail
            
            # Generate unique Message-ID for email threading
            message_id = make_msgid(domain=from_email.split('@')[1])
            
            # Find the original email's Message-ID for threading (In-Reply-To / References)
            original_message_id = None
            all_message_ids = []
            
            # Get all processed emails for this ticket (ordered by date) to build References chain
            processed_emails_result = await db.execute(
                select(ProcessedMail)
                .where(ProcessedMail.ticket_id == ticket.id)
                .order_by(ProcessedMail.processed_at.asc())
            )
            processed_emails = processed_emails_result.scalars().all()
            
            if processed_emails:
                # Original email is the first one
                original_message_id = processed_emails[0].message_id
                # Build References list (all message IDs in the thread)
                all_message_ids = [pe.message_id for pe in processed_emails]
                log(f"[EMAIL] Found {len(processed_emails)} emails in thread, original: {original_message_id[:50]}...")
            
            msg = MIMEMultipart('alternative')
            msg['From'] = f"{from_name} <{from_email}>"
            msg['To'] = ticket.guest_email
            msg['Subject'] = f"Re: Ticket #{ticket.ticket_number} - {ticket.subject}"
            msg['Reply-To'] = from_email
            msg['Message-ID'] = message_id
            
            # Add threading headers if we have the original email
            if original_message_id:
                # In-Reply-To: most recent email in thread (for immediate context)
                last_message_id = all_message_ids[-1] if all_message_ids else original_message_id
                msg['In-Reply-To'] = last_message_id
                # References: chain of all message IDs in the thread (for full context)
                msg['References'] = ' '.join(all_message_ids)
                log(f"[EMAIL] Added In-Reply-To: {last_message_id[:50]}...")
                log(f"[EMAIL] Added References with {len(all_message_ids)} message IDs")
            
            # Build email body
            commenter_name = user.full_name or user.username if user else "Support Team"
            
            email_body = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #2563eb; border-bottom: 2px solid #2563eb; padding-bottom: 10px;">
            New Update on Your Ticket
        </h2>
        
        <div style="background-color: #f3f4f6; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <p><strong>Ticket Number:</strong> #{ticket.ticket_number}</p>
            <p><strong>Subject:</strong> {ticket.subject}</p>
            <p><strong>Status:</strong> {ticket.status.title()}</p>
        </div>
        
        <div style="background-color: #fff; padding: 20px; border-left: 4px solid #2563eb; margin: 20px 0;">
            <p><strong>{commenter_name} commented:</strong></p>
            <div style="margin-top: 10px;">
                {content.replace(chr(10), '<br>')}
            </div>
        </div>
        
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #e5e7eb;">
            <p style="color: #6b7280; font-size: 14px;">
                You can reply directly to this email and your response will be added to the ticket.
            </p>
            <p style="color: #6b7280; font-size: 12px; margin-top: 20px;">
                This is an automated message from {from_name}. Please do not reply to this email if you have no further questions.
            </p>
        </div>
    </div>
</body>
</html>
"""
            
            msg.attach(MIMEText(email_body, 'html'))
            
            log(f"[EMAIL] Email message prepared, attempting to send via SMTP...")
            
            # Extract all SMTP settings BEFORE entering thread pool
            # (ORM objects should not be accessed across threads)
            smtp_host = email_settings.smtp_host
            smtp_port = email_settings.smtp_port
            smtp_username = email_settings.smtp_username
            smtp_password = email_settings.smtp_password
            smtp_use_tls = email_settings.smtp_use_tls
            
            # Send email in thread pool to avoid blocking
            import concurrent.futures
            import asyncio
            loop = asyncio.get_event_loop()
            
            # SMTP timeout in seconds
            SMTP_TIMEOUT = 30
            
            def send_email():
                log(f"[EMAIL] Connecting to SMTP server {smtp_host}:{smtp_port} (timeout: {SMTP_TIMEOUT}s)")
                try:
                    if smtp_use_tls:
                        server = smtplib.SMTP(smtp_host, smtp_port, timeout=SMTP_TIMEOUT)
                        server.starttls()
                    else:
                        server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=SMTP_TIMEOUT)
                    
                    log(f"[EMAIL] Logging in as {smtp_username}")
                    server.login(smtp_username, smtp_password)
                    log(f"[EMAIL] Sending message...")
                    server.send_message(msg)
                    server.quit()
                    log(f"[EMAIL] SMTP connection closed successfully")
                    return True
                except smtplib.SMTPAuthenticationError as e:
                    log(f"[EMAIL] ❌ SMTP Authentication failed: {e}")
                    return False
                except smtplib.SMTPConnectError as e:
                    log(f"[EMAIL] ❌ SMTP Connection failed: {e}")
                    return False
                except smtplib.SMTPServerDisconnected as e:
                    log(f"[EMAIL] ❌ SMTP Server disconnected: {e}")
                    return False
                except TimeoutError as e:
                    log(f"[EMAIL] ❌ SMTP Timeout after {SMTP_TIMEOUT}s: {e}")
                    return False
                except Exception as e:
                    log(f"[EMAIL] ❌ SMTP Error: {e}")
                    return False
            
            with concurrent.futures.ThreadPoolExecutor() as pool:
                email_sent = await loop.run_in_executor(pool, send_email)
            
            if not email_sent:
                log(f"❌ Email was NOT sent to {ticket.guest_email}")
                return
            
            # Store the Message-ID so replies can be threaded
            from app.models.processed_mail import ProcessedMail
            
            # Extract ticket values before database operations
            ticket_workspace_id = ticket.workspace_id
            ticket_id = ticket.id
            msg_subject = msg['Subject']
            guest_email = ticket.guest_email
            
            processed = ProcessedMail(
                workspace_id=ticket_workspace_id,
                message_id=message_id,
                email_from=from_email,
                subject=msg_subject,
                ticket_id=ticket_id,
                processed_at=get_local_time()
            )
            db.add(processed)
            await db.commit()
            
            log(f"✅ Sent email notification to {guest_email} from {from_email} with Message-ID: {message_id}")
    except Exception as e:
        log(f"❌ Error sending email notification: {e}")
        import traceback
        log(f"Traceback: {traceback.format_exc()}")
        # Don't fail if email fails


@router.post('/tickets/{ticket_id}/update-status')
async def web_tickets_update_status(
    request: Request,
    ticket_id: int,
    status: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    """Update ticket status"""
    user_id = request.session.get('user_id')
    if not user_id:
        # Check if AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JSONResponse({'error': 'Not authenticated'}, status_code=401)
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JSONResponse({'error': 'User not found'}, status_code=401)
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket, TicketHistory
    from datetime import datetime
    
    ticket = (await db.execute(select(Ticket).where(Ticket.id == ticket_id, Ticket.workspace_id == user.workspace_id))).scalar_one_or_none()
    if not ticket:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JSONResponse({'error': 'Ticket not found'}, status_code=404)
        return RedirectResponse('/web/tickets', status_code=303)
    
    # Validate status value
    valid_statuses = ['open', 'in_progress', 'waiting', 'resolved', 'closed']
    if status not in valid_statuses:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JSONResponse({'error': f'Invalid status. Must be one of: {valid_statuses}'}, status_code=400)
        return RedirectResponse('/web/tickets', status_code=303)
    
    old_status = ticket.status
    ticket.status = status
    ticket.updated_at = datetime.utcnow()
    
    ticket_just_closed = False
    
    # Set resolved/closed timestamps and track who closed it
    if status == 'resolved' and not ticket.resolved_at:
        ticket.resolved_at = datetime.utcnow()
    elif status == 'closed':
        if not ticket.closed_at:
            ticket.closed_at = datetime.utcnow()
            ticket_just_closed = True  # Flag to send completion notification
        ticket.closed_by_id = user_id  # Track who closed the ticket
        # Auto-archive when closed
        ticket.is_archived = True
        ticket.archived_at = datetime.utcnow()
    
    # Log ticket status change for diagnostics
    try:
        from app.core.system_logger import log_fire_and_forget
        log_fire_and_forget('INFO', 'ticket', 'User Action',
            f'Status changed: {old_status} -> {status} on ticket #{ticket.ticket_number}',
            f'User={user.full_name or user.username} | TicketID={ticket_id} | Subject={ticket.subject[:80]} | Guest={ticket.guest_email or "N/A"}',
            user.workspace_id)
    except Exception:
        pass
    
    # Add history
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        action='status_changed',
        old_value=old_status,
        new_value=status
    )
    db.add(history)
    
    # Notify assigned user (only if they haven't muted ticket notifications)
    if ticket.assigned_to_id and ticket.assigned_to_id != user_id:
        assigned_user = (await db.execute(select(User).where(User.id == ticket.assigned_to_id))).scalar_one_or_none()
        if assigned_user and not getattr(assigned_user, 'mute_ticket_notifications', False):
            notification = Notification(
                user_id=ticket.assigned_to_id,
                type='ticket',
                message=f'{user.full_name or user.username} changed ticket #{ticket.ticket_number} status to {status}',
                url=f'/web/tickets/{ticket_id}',
                related_id=ticket_id
            )
            db.add(notification)
    
    await db.commit()
    
    # Send completion notification email if ticket was just closed
    if ticket_just_closed:
        try:
            additional_details = f"Category: {ticket.category}" if ticket.category else ""
            
            await send_completion_notification_email(
                db=db,
                workspace_id=ticket.workspace_id,
                notification_type='ticket',
                item_id=ticket.ticket_number,
                title=ticket.subject,
                status='Closed',
                priority=ticket.priority.title() if ticket.priority else 'Normal',
                completed_by_name=user.full_name or user.username,
                created_at=ticket.created_at,
                completed_at=ticket.closed_at or datetime.utcnow(),
                additional_details=additional_details
            )
        except Exception as e:
            logger.error(f"Failed to send ticket completion notification: {e}")
    
    # Return JSON for AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JSONResponse({
            'success': True,
            'ticket_id': ticket_id,
            'old_status': old_status,
            'new_status': status,
            'ticket_number': ticket.ticket_number
        })
    
    return RedirectResponse(f'/web/tickets/{ticket_id}', status_code=303)


@router.post('/tickets/{ticket_id}/close-with-details')
async def web_tickets_close_with_details(
    request: Request,
    ticket_id: int,
    billable_traveling: Optional[str] = Form(None),
    billable_labour_onsite: Optional[str] = Form(None),
    billable_remote_labour: Optional[str] = Form(None),
    billable_equipment_used: Optional[str] = Form(None),
    non_billable_traveling: Optional[str] = Form(None),
    non_billable_labour_onsite: Optional[str] = Form(None),
    non_billable_remote_labour: Optional[str] = Form(None),
    non_billable_equipment_used: Optional[str] = Form(None),
    closing_notes: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Close ticket with optional billing/work details"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket, TicketHistory
    from datetime import datetime
    
    ticket = (await db.execute(select(Ticket).where(Ticket.id == ticket_id))).scalar_one_or_none()
    if not ticket:
        return RedirectResponse('/web/tickets', status_code=303)
    
    old_status = ticket.status
    
    # Update status to closed
    ticket.status = 'closed'
    ticket.updated_at = datetime.utcnow()
    ticket.closed_at = datetime.utcnow()
    ticket.closed_by_id = user_id
    ticket.is_archived = True
    ticket.archived_at = datetime.utcnow()
    
    # Log close-with-details for diagnostics
    try:
        from app.core.system_logger import log_fire_and_forget
        log_fire_and_forget('INFO', 'ticket', 'User Action',
            f'Closed with details: ticket #{ticket.ticket_number} ({old_status} -> closed)',
            f'User={user.full_name or user.username} | TicketID={ticket_id} | Subject={ticket.subject[:80]} | Guest={ticket.guest_email or "N/A"}',
            user.workspace_id)
    except Exception:
        pass
    
    # Save billing details (strip whitespace, store None if empty)
    ticket.billable_traveling = billable_traveling.strip() if billable_traveling and billable_traveling.strip() else None
    ticket.billable_labour_onsite = billable_labour_onsite.strip() if billable_labour_onsite and billable_labour_onsite.strip() else None
    ticket.billable_remote_labour = billable_remote_labour.strip() if billable_remote_labour and billable_remote_labour.strip() else None
    ticket.billable_equipment_used = billable_equipment_used.strip() if billable_equipment_used and billable_equipment_used.strip() else None
    ticket.non_billable_traveling = non_billable_traveling.strip() if non_billable_traveling and non_billable_traveling.strip() else None
    ticket.non_billable_labour_onsite = non_billable_labour_onsite.strip() if non_billable_labour_onsite and non_billable_labour_onsite.strip() else None
    ticket.non_billable_remote_labour = non_billable_remote_labour.strip() if non_billable_remote_labour and non_billable_remote_labour.strip() else None
    ticket.non_billable_equipment_used = non_billable_equipment_used.strip() if non_billable_equipment_used and non_billable_equipment_used.strip() else None
    ticket.closing_notes = closing_notes.strip() if closing_notes and closing_notes.strip() else None
    
    # Add history
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        action='status_changed',
        old_value=old_status,
        new_value='closed'
    )
    db.add(history)
    
    # Notify assigned user
    if ticket.assigned_to_id and ticket.assigned_to_id != user_id:
        assigned_user = (await db.execute(select(User).where(User.id == ticket.assigned_to_id))).scalar_one_or_none()
        if assigned_user and not getattr(assigned_user, 'mute_ticket_notifications', False):
            notification = Notification(
                user_id=ticket.assigned_to_id,
                type='ticket',
                message=f'{user.full_name or user.username} closed ticket #{ticket.ticket_number}',
                url=f'/web/tickets/{ticket_id}',
                related_id=ticket_id
            )
            db.add(notification)
    
    await db.commit()
    
    # Build billing details string for email
    billing_details = []
    
    # Billable items
    billable_items = []
    if ticket.billable_traveling:
        billable_items.append(f"  - Traveling: {ticket.billable_traveling}")
    if ticket.billable_labour_onsite:
        billable_items.append(f"  - Labour Onsite: {ticket.billable_labour_onsite}")
    if ticket.billable_remote_labour:
        billable_items.append(f"  - Remote Labour: {ticket.billable_remote_labour}")
    if ticket.billable_equipment_used:
        billable_items.append(f"  - Equipment Used: {ticket.billable_equipment_used}")
    
    if billable_items:
        billing_details.append("BILLABLE:")
        billing_details.extend(billable_items)
    
    # Non-billable items
    non_billable_items = []
    if ticket.non_billable_traveling:
        non_billable_items.append(f"  - Traveling: {ticket.non_billable_traveling}")
    if ticket.non_billable_labour_onsite:
        non_billable_items.append(f"  - Labour Onsite: {ticket.non_billable_labour_onsite}")
    if ticket.non_billable_remote_labour:
        non_billable_items.append(f"  - Remote Labour: {ticket.non_billable_remote_labour}")
    if ticket.non_billable_equipment_used:
        non_billable_items.append(f"  - Equipment Used: {ticket.non_billable_equipment_used}")
    
    if non_billable_items:
        if billing_details:
            billing_details.append("")  # Empty line separator
        billing_details.append("NON-BILLABLE:")
        billing_details.extend(non_billable_items)
    
    # Closing notes
    if ticket.closing_notes:
        if billing_details:
            billing_details.append("")  # Empty line separator
        billing_details.append(f"CLOSING NOTES:\n{ticket.closing_notes}")
    
    # Build additional details for email
    additional_details_parts = []
    if ticket.category:
        additional_details_parts.append(f"Category: {ticket.category}")
    if billing_details:
        additional_details_parts.append("\n" + "\n".join(billing_details))
    
    additional_details = "\n".join(additional_details_parts)
    
    # Send completion notification email
    try:
        await send_completion_notification_email(
            db=db,
            workspace_id=ticket.workspace_id,
            notification_type='ticket',
            item_id=ticket.ticket_number,
            title=ticket.subject,
            status='Closed',
            priority=ticket.priority.title() if ticket.priority else 'Normal',
            completed_by_name=user.full_name or user.username,
            created_at=ticket.created_at,
            completed_at=ticket.closed_at,
            additional_details=additional_details
        )
    except Exception as e:
        logger.error(f"Failed to send ticket completion notification: {e}")
    
    request.session['success_message'] = f'Ticket #{ticket.ticket_number} has been closed.'
    return RedirectResponse(f'/web/tickets/{ticket_id}', status_code=303)


@router.post('/tickets/{ticket_id}/archive')
async def web_tickets_archive(
    request: Request,
    ticket_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Archive a ticket - Admin only"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    # Check if user is admin
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        request.session['error_message'] = 'Only administrators can archive tickets.'
        return RedirectResponse(f'/web/tickets/{ticket_id}', status_code=303)
    
    from app.models.ticket import Ticket, TicketHistory
    from datetime import datetime
    
    ticket = (await db.execute(select(Ticket).where(Ticket.id == ticket_id))).scalar_one_or_none()
    if not ticket:
        return RedirectResponse('/web/tickets', status_code=303)
    
    # Archive the ticket
    ticket.is_archived = True
    ticket.archived_at = datetime.utcnow()
    ticket.updated_at = datetime.utcnow()
    
    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        action='archived',
        new_value='Ticket archived by admin'
    )
    db.add(history)
    
    await db.commit()
    
    try:
        from app.core.system_logger import log_fire_and_forget
        log_fire_and_forget('INFO', 'ticket', 'User Action',
            f'Archived ticket #{ticket.ticket_number}',
            f'User={user.full_name or user.username} | TicketID={ticket_id}',
            user.workspace_id)
    except Exception:
        pass
    
    request.session['success_message'] = f'Ticket #{ticket.ticket_number} has been archived.'
    return RedirectResponse('/web/tickets/archived', status_code=303)


@router.post('/tickets/{ticket_id}/restore')
async def web_tickets_restore(
    request: Request,
    ticket_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Restore archived ticket - Admin only"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    # Check if user is admin
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        request.session['error_message'] = 'Only administrators can restore archived tickets.'
        return RedirectResponse('/web/tickets/archived', status_code=303)
    
    from app.models.ticket import Ticket
    from datetime import datetime
    
    ticket = (await db.execute(select(Ticket).where(Ticket.id == ticket_id))).scalar_one_or_none()
    if not ticket:
        return RedirectResponse('/web/tickets/archived', status_code=303)
    
    # Restore ticket
    ticket.is_archived = False
    ticket.archived_at = None
    ticket.updated_at = datetime.utcnow()
    
    await db.commit()
    
    try:
        from app.core.system_logger import log_fire_and_forget
        log_fire_and_forget('INFO', 'ticket', 'User Action',
            f'Restored ticket #{ticket.ticket_number}',
            f'User={user.full_name or user.username} | TicketID={ticket_id}',
            user.workspace_id)
    except Exception:
        pass
    
    request.session['success_message'] = f'Ticket #{ticket.ticket_number} has been restored.'
    return RedirectResponse('/web/tickets/archived', status_code=303)


@router.post('/tickets/{ticket_id}/assign')
async def web_tickets_assign(
    request: Request,
    ticket_id: int,
    assigned_to_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_session)
):
    """Assign ticket to user"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import Ticket, TicketHistory
    from datetime import datetime
    
    ticket = (await db.execute(select(Ticket).where(Ticket.id == ticket_id))).scalar_one_or_none()
    if not ticket:
        return RedirectResponse('/web/tickets', status_code=303)
    
    old_assigned = ticket.assigned_to_id
    ticket.assigned_to_id = assigned_to_id
    ticket.updated_at = datetime.utcnow()
    
    # Log assignment for diagnostics
    try:
        from app.core.system_logger import log_fire_and_forget
        log_fire_and_forget('INFO', 'ticket', 'User Action',
            f'Assigned ticket #{ticket.ticket_number}',
            f'User={user.full_name or user.username} | OldAssignee={old_assigned} | NewAssignee={assigned_to_id}',
            user.workspace_id)
    except Exception:
        pass
    
    # Add history
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=user_id,
        action='assigned',
        old_value=str(old_assigned) if old_assigned else 'Unassigned',
        new_value=str(assigned_to_id) if assigned_to_id else 'Unassigned'
    )
    db.add(history)
    
    # Notify assigned user (only if they haven't muted ticket notifications)
    if assigned_to_id and assigned_to_id != user_id:
        assigned_user = (await db.execute(select(User).where(User.id == assigned_to_id))).scalar_one_or_none()
        if assigned_user and not getattr(assigned_user, 'mute_ticket_notifications', False):
            notification = Notification(
                user_id=assigned_to_id,
                type='ticket',
                message=f'{user.full_name or user.username} assigned you ticket #{ticket.ticket_number}: {ticket.subject}',
                url=f'/web/tickets/{ticket_id}',
                related_id=ticket_id
            )
            db.add(notification)
    
    await db.commit()
    return RedirectResponse(f'/web/tickets/{ticket_id}', status_code=303)


@router.get('/tickets/attachments/{attachment_id}/download')
async def web_ticket_attachment_download(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Download a ticket attachment"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import TicketAttachment, Ticket
    
    # Get attachment
    attachment = (await db.execute(
        select(TicketAttachment).where(TicketAttachment.id == attachment_id)
    )).scalar_one_or_none()
    
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Verify user has access to this ticket's workspace
    ticket = (await db.execute(
        select(Ticket).where(
            Ticket.id == attachment.ticket_id,
            Ticket.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not ticket:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Build file path
    file_path = BASE_DIR.parent / attachment.file_path if attachment.file_path.startswith('app/') else BASE_DIR / attachment.file_path
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        filename=attachment.filename,
        media_type=attachment.mime_type or 'application/octet-stream'
    )


@router.get('/tickets/attachments/{attachment_id}/preview')
async def web_ticket_attachment_preview(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Preview a ticket attachment (images only)"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.ticket import TicketAttachment, Ticket
    
    # Get attachment
    attachment = (await db.execute(
        select(TicketAttachment).where(TicketAttachment.id == attachment_id)
    )).scalar_one_or_none()
    
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Verify user has access to this ticket's workspace
    ticket = (await db.execute(
        select(Ticket).where(
            Ticket.id == attachment.ticket_id,
            Ticket.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not ticket:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Only allow preview for images
    if not attachment.mime_type or not attachment.mime_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Preview only available for images")
    
    # Build file path
    file_path = BASE_DIR.parent / attachment.file_path if attachment.file_path.startswith('app/') else BASE_DIR / attachment.file_path
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        media_type=attachment.mime_type
    )


@router.post('/tickets/attachments/{attachment_id}/delete')
async def web_ticket_attachment_delete(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Delete a ticket attachment"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse(
            status_code=401,
            content={"detail": "Not authenticated"}
        )
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse(
            status_code=401,
            content={"detail": "User not found"}
        )
    
    from app.models.ticket import TicketAttachment, Ticket
    
    # Get attachment
    attachment = (await db.execute(
        select(TicketAttachment).where(TicketAttachment.id == attachment_id)
    )).scalar_one_or_none()
    
    if not attachment:
        return JSONResponse(
            status_code=404,
            content={"detail": "Attachment not found"}
        )
    
    # Verify user has access to this ticket's workspace
    ticket = (await db.execute(
        select(Ticket).where(
            Ticket.id == attachment.ticket_id,
            Ticket.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not ticket:
        return JSONResponse(
            status_code=403,
            content={"detail": "Access denied"}
        )
    
    # Check permissions: admin or the uploader
    if not user.is_admin and attachment.uploaded_by_id != user.id:
        return JSONResponse(
            status_code=403,
            content={"detail": "You don't have permission to delete this attachment"}
        )
    
    # Delete file from disk
    file_path = BASE_DIR.parent / attachment.file_path if attachment.file_path.startswith('app/') else BASE_DIR / attachment.file_path
    try:
        if file_path.exists():
            file_path.unlink()
    except Exception as e:
        logger.error(f"Failed to delete attachment file {file_path}: {e}")
    
    # Delete from database
    await db.delete(attachment)
    await db.commit()
    
    return JSONResponse(
        status_code=200,
        content={"success": True, "message": "Attachment deleted successfully"}
    )


# ============ TASK ATTACHMENT ROUTES ============

@router.get('/task-attachments/{attachment_id}/download')
async def web_task_attachment_download(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Download a task attachment"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.task_extensions import TaskAttachment
    from app.models.task import Task
    from app.models.project import Project
    
    # Get attachment
    attachment = (await db.execute(
        select(TaskAttachment).where(TaskAttachment.id == attachment_id)
    )).scalar_one_or_none()
    
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Verify user has access to this task's project/workspace
    task = (await db.execute(
        select(Task).where(Task.id == attachment.task_id)
    )).scalar_one_or_none()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    project = (await db.execute(
        select(Project).where(
            Project.id == task.project_id,
            Project.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Build file path
    file_path = BASE_DIR.parent / attachment.file_path if attachment.file_path.startswith('app/') else BASE_DIR / attachment.file_path
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        filename=attachment.filename,
        media_type=attachment.file_type or 'application/octet-stream'
    )


@router.post('/task-attachments/{attachment_id}/delete')
async def web_task_attachment_delete(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Delete a task attachment"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JSONResponse(
            status_code=401,
            content={"detail": "Not authenticated"}
        )
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return JSONResponse(
            status_code=401,
            content={"detail": "User not found"}
        )
    
    from app.models.task_extensions import TaskAttachment
    from app.models.task import Task
    from app.models.project import Project
    
    # Get attachment
    attachment = (await db.execute(
        select(TaskAttachment).where(TaskAttachment.id == attachment_id)
    )).scalar_one_or_none()
    
    if not attachment:
        return JSONResponse(
            status_code=404,
            content={"detail": "Attachment not found"}
        )
    
    # Verify user has access to this task's project/workspace
    task = (await db.execute(
        select(Task).where(Task.id == attachment.task_id)
    )).scalar_one_or_none()
    
    if not task:
        return JSONResponse(
            status_code=404,
            content={"detail": "Task not found"}
        )
    
    project = (await db.execute(
        select(Project).where(
            Project.id == task.project_id,
            Project.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not project:
        return JSONResponse(
            status_code=403,
            content={"detail": "Access denied"}
        )
    
    # Check permissions: admin or the uploader
    if not user.is_admin and attachment.uploaded_by != user.id:
        return JSONResponse(
            status_code=403,
            content={"detail": "You don't have permission to delete this attachment"}
        )
    
    # Delete file from disk
    file_path = BASE_DIR.parent / attachment.file_path if attachment.file_path.startswith('app/') else BASE_DIR / attachment.file_path
    try:
        if file_path.exists():
            file_path.unlink()
    except Exception as e:
        logger.error(f"Failed to delete task attachment file {file_path}: {e}")
    
    # Delete from database
    await db.delete(attachment)
    await db.commit()
    
    return JSONResponse(
        status_code=200,
        content={"success": True, "message": "Attachment deleted successfully"}
    )


@router.post('/tickets/process-emails')
async def web_tickets_process_emails(request: Request, db: AsyncSession = Depends(get_session)):
    """Manually trigger email-to-ticket processing (admin only) - Processes emails immediately"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_admin:
        return RedirectResponse('/web/tickets', status_code=303)
    
    try:
        from app.models.incoming_email_account import IncomingEmailAccount
        from app.core.email_to_ticket_v2 import process_email_account, process_workspace_emails
        
        all_tickets = []
        errors = []
        
        # 1. Process legacy workspace emails (if configured)
        try:
            tickets = await process_workspace_emails(db, user.workspace_id)
            all_tickets.extend(tickets)
        except Exception as e:
            logger.warning(f"[Email] Error processing workspace emails: {e}")
        
        # 2. Get all active email accounts for this workspace
        accounts_result = await db.execute(
            select(IncomingEmailAccount).where(
                IncomingEmailAccount.workspace_id == user.workspace_id,
                IncomingEmailAccount.is_active == True
            )
        )
        accounts = accounts_result.scalars().all()
        
        if not accounts:
            # Check if legacy settings exist
            from app.models.email_settings import EmailSettings
            legacy_settings = (await db.execute(
                select(EmailSettings).where(EmailSettings.workspace_id == user.workspace_id)
            )).scalar_one_or_none()
            
            if not legacy_settings or not legacy_settings.incoming_mail_host:
                request.session['flash_message'] = "✗ No email accounts configured. Add one in Admin → Email Accounts"
                request.session['flash_type'] = 'error'
                return RedirectResponse('/web/tickets', status_code=303)
        
        # 3. Process each email account immediately
        for account in accounts:
            try:
                tickets = await process_email_account(db, account)
                all_tickets.extend(tickets)
            except Exception as e:
                errors.append(f"{account.name}: {str(e)}")
                logger.error(f"[Email] Error processing account {account.name}: {e}")
        
        # Build result message
        if all_tickets:
            ticket_nums = [t.ticket_number for t in all_tickets]
            request.session['flash_message'] = f"✓ Created {len(all_tickets)} ticket(s): {', '.join(ticket_nums)}"
            request.session['flash_type'] = 'success'
        elif errors:
            request.session['flash_message'] = f"✗ Errors: {'; '.join(errors)}"
            request.session['flash_type'] = 'error'
        else:
            request.session['flash_message'] = f"✓ Checked {len(accounts)} email account(s) - no new unprocessed emails found"
            request.session['flash_type'] = 'info'
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        request.session['flash_message'] = f"✗ Error: {str(e)}"
        request.session['flash_type'] = 'error'
    
    return RedirectResponse('/web/tickets', status_code=303)


# Chats
@router.get('/chats')
async def web_chats_list(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get all chats where user is a member
    stmt = (
        select(Chat)
        .join(ChatMember, Chat.id == ChatMember.chat_id)
        .where(ChatMember.user_id == user_id)
        .order_by(Chat.created_at.desc())
    )
    chats = (await db.execute(stmt)).scalars().all()
    
    # Get all active workspace users for creating new chats
    users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id, User.id != user_id, User.is_active == True)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    return templates.TemplateResponse('chats/list.html', {
        'request': request,
        'chats': chats,
        'users': users,
        'user': user
    })


@router.post('/chats/create')
async def web_chat_create(
    request: Request,
    name: Optional[str] = Form(None),
    is_group: bool = Form(False),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Create chat
    chat = Chat(
        name=name,
        is_group=is_group,
        workspace_id=user.workspace_id,
        created_by_id=user_id  # Set the creator
    )
    db.add(chat)
    await db.flush()
    
    # Add creator as member
    member = ChatMember(chat_id=chat.id, user_id=user_id)
    db.add(member)
    
    # Add selected members
    form_data = await request.form()
    member_ids = form_data.getlist('member_ids')
    for member_id in member_ids:
        if int(member_id) != user_id:
            member = ChatMember(chat_id=chat.id, user_id=int(member_id))
            db.add(member)
    
    await db.commit()
    return RedirectResponse(f'/web/chats/{chat.id}', status_code=303)


@router.get('/chats/{chat_id}')
async def web_chat_detail(request: Request, chat_id: int, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify user is member of this chat
    membership = (await db.execute(
        select(ChatMember)
        .where(ChatMember.chat_id == chat_id, ChatMember.user_id == user_id)
    )).scalar_one_or_none()
    
    if not membership:
        raise HTTPException(status_code=403, detail='Not a member of this chat')
    
    # Get chat details
    chat = (await db.execute(select(Chat).where(Chat.id == chat_id))).scalar_one_or_none()
    if not chat:
        raise HTTPException(status_code=404, detail='Chat not found')
    
    # Get all messages
    messages_stmt = (
        select(Message, User.full_name, User.email)
        .join(User, Message.author_id == User.id)
        .where(Message.chat_id == chat_id)
        .order_by(Message.created_at.asc())
    )
    results = (await db.execute(messages_stmt)).all()
    
    # Get attachments for all messages
    from app.models.chat import MessageAttachment
    message_ids = [msg.id for msg, _, _ in results]
    attachments_stmt = (
        select(MessageAttachment)
        .where(MessageAttachment.message_id.in_(message_ids) if message_ids else False)
        .order_by(MessageAttachment.uploaded_at.asc())
    )
    all_attachments = (await db.execute(attachments_stmt)).scalars().all()
    
    # Group attachments by message_id
    attachments_by_message = {}
    for att in all_attachments:
        attachments_by_message.setdefault(att.message_id, []).append(att)
    
    # Combine messages with sender names and attachments
    messages_with_sender = [
        (msg, full_name or email, attachments_by_message.get(msg.id, []))
        for msg, full_name, email in results
    ]
    
    # Get chat members
    members_stmt = (
        select(User)
        .join(ChatMember, User.id == ChatMember.user_id)
        .where(ChatMember.chat_id == chat_id)
        .order_by(User.full_name, User.email)
    )
    members = (await db.execute(members_stmt)).scalars().all()
    
    return templates.TemplateResponse('chats/detail.html', {
        'request': request,
        'chat': chat,
        'messages': messages_with_sender,
        'members': members,
        'user': user
    })


@router.get('/chats/attachments/{attachment_id}/download')
async def download_chat_attachment(
    request: Request,
    attachment_id: int,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.chat import MessageAttachment
    from fastapi.responses import FileResponse
    import os
    
    # Get attachment
    attachment = (await db.execute(
        select(MessageAttachment).where(MessageAttachment.id == attachment_id)
    )).scalar_one_or_none()
    
    if not attachment:
        raise HTTPException(status_code=404, detail='Attachment not found')
    
    # Verify user has access to this chat
    message = (await db.execute(select(Message).where(Message.id == attachment.message_id))).scalar_one_or_none()
    if not message:
        raise HTTPException(status_code=404, detail='Message not found')
    
    membership = (await db.execute(
        select(ChatMember).where(
            ChatMember.chat_id == message.chat_id,
            ChatMember.user_id == user_id
        )
    )).scalar_one_or_none()
    
    if not membership:
        raise HTTPException(status_code=403, detail='Access denied')
    
    # Handle both absolute paths (old) and relative paths (new)
    file_path = Path(attachment.file_path)
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail='File not found on disk')
    
    return FileResponse(
        path=str(file_path),
        filename=attachment.filename,
        media_type=attachment.mime_type or 'application/octet-stream'
    )


@router.post('/chats/{chat_id}/messages')
async def web_chat_send_message(
    request: Request,
    chat_id: int,
    content: Optional[str] = Form(None),
    attachments: Optional[list[UploadFile]] = File(None),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Verify membership
    membership = (await db.execute(
        select(ChatMember)
        .where(ChatMember.chat_id == chat_id, ChatMember.user_id == user_id)
    )).scalar_one_or_none()
    
    if not membership:
        raise HTTPException(status_code=403, detail='Not a member of this chat')
    
    # Require either content or attachments
    if not content and not attachments:
        return RedirectResponse(f'/web/chats/{chat_id}', status_code=303)
    
    # Create message
    message = Message(
        chat_id=chat_id,
        author_id=user_id,
        content=content or ""
    )
    db.add(message)
    await db.flush()  # Get message ID
    
    # Handle file attachments
    if attachments:
        import uuid
        from pathlib import Path
        
        upload_dir = Path('app/uploads/chat_messages')
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        for file in attachments:
            if file.filename:
                # Read file content first for validation
                content_bytes = await file.read()
                
                # Validate file size (max 10MB)
                if len(content_bytes) > 10 * 1024 * 1024:
                    continue  # Skip files that are too large
                
                # Block dangerous file extensions
                file_ext = Path(file.filename).suffix.lower()
                blocked_extensions = {'.exe', '.bat', '.cmd', '.sh', '.php', '.py', '.rb', '.pl', '.cgi', '.js', '.msi', '.ps1', '.vbs', '.wsf'}
                if file_ext in blocked_extensions:
                    continue  # Skip dangerous file types
                
                # Generate unique filename
                unique_filename = f"{uuid.uuid4()}{file_ext}"
                file_path = upload_dir / unique_filename
                
                # Save file
                with open(file_path, 'wb') as f:
                    f.write(content_bytes)
                
                # Create attachment record
                from app.models.chat import MessageAttachment
                attachment = MessageAttachment(
                    message_id=message.id,
                    filename=file.filename,
                    file_path=str(file_path),
                    file_size=len(content_bytes),
                    mime_type=file.content_type
                )
                db.add(attachment)
    
    # Create notifications for other chat members
    chat_members = (await db.execute(
        select(ChatMember)
        .where(ChatMember.chat_id == chat_id, ChatMember.user_id != user_id)
    )).scalars().all()
    
    # Get chat info for notification message
    chat = (await db.execute(select(Chat).where(Chat.id == chat_id))).scalar_one_or_none()
    
    # Build intelligent notification message
    chat_name = chat.name if chat and chat.name else 'chat'
    sender_name = user.full_name or user.username
    
    # Create message preview (keep it short)
    message_preview = ""
    if content:
        # Truncate long messages
        preview_text = content.strip()
        if len(preview_text) > 50:
            preview_text = preview_text[:47] + "..."
        message_preview = f": {preview_text}"
    
    # Check for attachments and add to summary
    attachment_summary = ""
    if attachments and len(attachments) > 0:
        attachment_count = len(attachments)
        
        # Analyze attachment types
        image_count = sum(1 for f in attachments if f.content_type and f.content_type.startswith('image/'))
        video_count = sum(1 for f in attachments if f.content_type and f.content_type.startswith('video/'))
        doc_count = sum(1 for f in attachments if f.content_type and (
            'pdf' in (f.content_type or '') or 
            'document' in (f.content_type or '') or
            'word' in (f.content_type or '') or
            'sheet' in (f.content_type or '') or
            'text' in (f.content_type or '')
        ))
        
        # Build attachment description
        attachment_parts = []
        if image_count > 0:
            attachment_parts.append(f"{image_count} image{'s' if image_count > 1 else ''}")
        if video_count > 0:
            attachment_parts.append(f"{video_count} video{'s' if video_count > 1 else ''}")
        if doc_count > 0:
            attachment_parts.append(f"{doc_count} document{'s' if doc_count > 1 else ''}")
        
        # If there are other files not categorized
        other_count = attachment_count - (image_count + video_count + doc_count)
        if other_count > 0:
            attachment_parts.append(f"{other_count} file{'s' if other_count > 1 else ''}")
        
        if attachment_parts:
            attachment_summary = f" [{', '.join(attachment_parts)}]"
        else:
            attachment_summary = f" [{attachment_count} attachment{'s' if attachment_count > 1 else ''}]"
    
    # Combine everything into a concise notification
    if message_preview and attachment_summary:
        notification_text = f"{sender_name}{message_preview} {attachment_summary}"
    elif attachment_summary:
        notification_text = f"{sender_name} sent {attachment_summary.strip('[]')} in {chat_name}"
    elif message_preview:
        notification_text = f"{sender_name} in {chat_name}{message_preview}"
    else:
        notification_text = f"{sender_name} sent a message in {chat_name}"
    
    for member in chat_members:
        # Create notification for each member
        notification = Notification(
            user_id=member.user_id,
            type='message',
            message=notification_text,
            url=f'/web/chats/{chat_id}',
            related_id=message.id
        )
        db.add(notification)
    
    await db.commit()
    
    return RedirectResponse(f'/web/chats/{chat_id}', status_code=303)


# Invite Users
@router.get('/users/new')
async def web_invite_user_form(request: Request, db: AsyncSession = Depends(get_session)):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Get current workspace users
    users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    return templates.TemplateResponse('users/invite.html', {
        'request': request,
        'users': users,
        'user': user,
        'error': None,
        'success': None
    })


@router.post('/users/invite')
async def web_invite_user(
    request: Request,
    email: str = Form(...),
    full_name: Optional[str] = Form(None),
    is_admin: bool = Form(False),
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    # Only admins can invite users
    if not user.is_admin:
        return RedirectResponse('/web/dashboard', status_code=303)
    
    # Check if user with email already exists
    existing = (await db.execute(select(User).where(User.email == email))).scalar_one_or_none()
    
    users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    if existing:
        return templates.TemplateResponse('users/invite.html', {
            'request': request,
            'users': users,
            'user': user,
            'error': 'User with this email already exists',
            'success': None
        }, status_code=400)
    
    # Create new user in the same workspace with a temporary password
    import secrets
    temp_password = secrets.token_urlsafe(16)
    
    # Generate username from email (before @ symbol)
    username_base = email.split('@')[0].lower()
    # Check if username exists, add number if needed
    username = username_base
    counter = 1
    while True:
        existing_username = (await db.execute(
            select(User).where(User.username == username)
        )).scalar_one_or_none()
        if not existing_username:
            break
        username = f"{username_base}{counter}"
        counter += 1
    
    new_user = User(
        username=username,
        email=email,
        full_name=full_name or '',
        hashed_password=get_password_hash(temp_password),
        workspace_id=user.workspace_id,
        is_admin=is_admin,
        email_verified=True,  # OTP disabled
        profile_completed=True  # Skip profile completion for invited users
    )
    db.add(new_user)
    await db.commit()
    
    # In a real app, you'd send an email with the temp password or invitation link
    # For now, we'll just show a success message
    users = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id)
        .order_by(User.full_name, User.email)
    )).scalars().all()
    
    return templates.TemplateResponse('users/invite.html', {
        'request': request,
        'users': users,
        'user': user,
        'error': None,
        'success': f'User invited successfully! Username: {username}, Temporary password: {temp_password} (share these securely)'
    })


@router.post('/users/{user_id}/delete')
async def web_delete_user(
    request: Request,
    user_id: int,
    db: AsyncSession = Depends(get_session)
):
    current_user_id = request.session.get('user_id')
    if not current_user_id:
        return RedirectResponse('/web/login', status_code=303)
    current_user = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one_or_none()
    if not current_user or not current_user.is_admin:
        return RedirectResponse('/web/projects', status_code=303)
    
    # Get user to delete
    user_to_delete = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user_to_delete or user_to_delete.workspace_id != current_user.workspace_id:
        return RedirectResponse('/web/users/new', status_code=303)
    
    # Prevent deleting yourself
    if user_to_delete.id == current_user.id:
        return RedirectResponse('/web/users/new', status_code=303)
    
    # Deactivate user instead of deleting (preserves audit trail)
    user_to_delete.is_active = False
    await db.commit()
    
    return RedirectResponse('/web/users/new', status_code=303)


@router.get('/set-password')
async def web_set_password_form(request: Request, token: Optional[str] = None):
    return templates.TemplateResponse('auth/set_password.html', {
        'request': request,
        'token': token,
        'error': None
    })


@router.post('/set-password')
async def web_set_password(
    request: Request,
    username: str = Form(...),
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: AsyncSession = Depends(get_session)
):
    # Validate passwords match
    if new_password != confirm_password:
        return templates.TemplateResponse('auth/set_password.html', {
            'request': request,
            'token': None,
            'error': 'New passwords do not match'
        }, status_code=400)
    
    # Find user by username
    user = (await db.execute(select(User).where(User.username == username))).scalar_one_or_none()
    if not user:
        return templates.TemplateResponse('auth/set_password.html', {
            'request': request,
            'token': None,
            'error': 'Invalid username or password'
        }, status_code=400)
    
    # Verify current password
    if not verify_password(current_password, user.hashed_password):
        return templates.TemplateResponse('auth/set_password.html', {
            'request': request,
            'token': None,
            'error': 'Invalid username or password'
        }, status_code=400)
    
    # Update password
    user.hashed_password = get_password_hash(new_password)
    await db.commit()
    
    # Auto-login after password change
    request.session['user_id'] = user.id
    
    # Redirect to profile completion if needed
    if not user.profile_completed:
        return RedirectResponse('/web/profile/complete', status_code=303)
    
    return RedirectResponse('/web/projects', status_code=303)


@router.get('/activity')
async def web_activity_feed(
    request: Request,
    activity_type: Optional[str] = None,
    db: AsyncSession = Depends(get_session)
):
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user or not user.is_active:
        request.session.clear()
        return RedirectResponse('/web/login', status_code=303)
    
    from datetime import datetime, timedelta
    from app.models.task_extensions import ActivityLog
    
    # Get recent activity logs
    query = select(ActivityLog).where(ActivityLog.workspace_id == user.workspace_id)
    if activity_type:
        query = query.where(ActivityLog.action_type == activity_type)
    query = query.order_by(ActivityLog.created_at.desc()).limit(100)
    
    logs = (await db.execute(query)).scalars().all()
    
    # Enhance activity logs with user names and entity titles
    activities = []
    for log in logs:
        # Get user who performed action
        actor = (await db.execute(select(User).where(User.id == log.user_id))).scalar_one_or_none()
        
        # Get entity title based on type
        entity_title = None
        if log.entity_type == 'task':
            task = (await db.execute(select(Task).where(Task.id == log.entity_id))).scalar_one_or_none()
            entity_title = task.title if task else None
        elif log.entity_type == 'project':
            project = (await db.execute(select(Project).where(Project.id == log.entity_id))).scalar_one_or_none()
            entity_title = project.name if project else None
        
        # Calculate time ago
        time_diff = datetime.utcnow() - log.created_at
        if time_diff.total_seconds() < 60:
            time_ago = "just now"
        elif time_diff.total_seconds() < 3600:
            time_ago = f"{int(time_diff.total_seconds() / 60)}m ago"
        elif time_diff.total_seconds() < 86400:
            time_ago = f"{int(time_diff.total_seconds() / 3600)}h ago"
        else:
            time_ago = f"{int(time_diff.total_seconds() / 86400)}d ago"
        
        activities.append({
            'user_name': actor.full_name or actor.email if actor else 'Unknown',
            'action_type': log.action_type,
            'action_text': log.action_type.replace('_', ' '),
            'entity_type': log.entity_type,
            'entity_id': log.entity_id,
            'entity_title': entity_title,
            'details': log.details,
            'time_ago': time_ago,
            'created_at': log.created_at
        })
    
    # Get workspace stats
    today = datetime.utcnow().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    active_tasks = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == user.workspace_id)
        .where(Task.status != TaskStatus.done)
    )).scalars().all()
    
    completed_week = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == user.workspace_id)
        .where(Task.status == TaskStatus.done)
        .where(Task.updated_at >= datetime.combine(week_ago, datetime.min.time()))
    )).scalars().all()
    
    overdue = (await db.execute(
        select(Task)
        .join(Project, Task.project_id == Project.id)
        .where(Project.workspace_id == user.workspace_id)
        .where(Task.status != TaskStatus.done)
        .where(Task.due_date < today)
    )).scalars().all()
    
    team_members = (await db.execute(
        select(User)
        .where(User.workspace_id == user.workspace_id)
        .where(User.is_active == True)
    )).scalars().all()
    
    # Calculate per-user statistics
    user_stats = []
    
    if user.is_admin:
        # Admin sees all users' statistics
        for member in team_members:
            # Get tasks assigned to this user
            member_assignments = (await db.execute(
                select(Assignment)
                .where(Assignment.assignee_id == member.id)
            )).scalars().all()
            
            task_ids = [a.task_id for a in member_assignments]
            
            if task_ids:
                # Get all tasks for this user
                member_tasks = (await db.execute(
                    select(Task)
                    .where(Task.id.in_(task_ids))
                )).scalars().all()
                
                # Tasks completed in the last month
                completed_month = [t for t in member_tasks 
                                  if t.status == TaskStatus.done 
                                  and t.updated_at 
                                  and t.updated_at.date() >= month_ago]
                
                # Tasks completed late (had due date, completed after due date)
                completed_late = [t for t in completed_month
                                 if t.due_date and t.updated_at 
                                 and t.updated_at.date() > t.due_date]
                
                # Currently overdue tasks
                overdue_tasks = [t for t in member_tasks
                                if t.status != TaskStatus.done
                                and t.due_date
                                and t.due_date < today]
                
                # Currently active (in progress) tasks
                active_member_tasks = [t for t in member_tasks
                                      if t.status == TaskStatus.in_progress]
                
                # Get current task (most recently updated in-progress task)
                current_task = None
                if active_member_tasks:
                    current_task = sorted(active_member_tasks, 
                                        key=lambda x: x.updated_at if x.updated_at else x.created_at,
                                        reverse=True)[0]
                
                user_stats.append({
                    'user_id': member.id,
                    'user_name': member.full_name or member.email,
                    'user_email': member.email,
                    'completed_month': len(completed_month),
                    'completed_late': len(completed_late),
                    'overdue_count': len(overdue_tasks),
                    'active_count': len(active_member_tasks),
                    'current_task': current_task.title if current_task else None,
                    'current_task_id': current_task.id if current_task else None
                })
    else:
        # Regular user sees only their own statistics
        member_assignments = (await db.execute(
            select(Assignment)
            .where(Assignment.assignee_id == user.id)
        )).scalars().all()
        
        task_ids = [a.task_id for a in member_assignments]
        
        if task_ids:
            member_tasks = (await db.execute(
                select(Task)
                .where(Task.id.in_(task_ids))
            )).scalars().all()
            
            # Tasks completed in the last month
            completed_month = [t for t in member_tasks 
                              if t.status == TaskStatus.done 
                              and t.updated_at 
                              and t.updated_at.date() >= month_ago]
            
            # Separate completed on time vs late
            completed_on_time = []
            completed_late = []
            
            for t in completed_month:
                if t.due_date and t.updated_at:
                    if t.updated_at.date() > t.due_date:
                        completed_late.append(t)
                    else:
                        completed_on_time.append(t)
                else:
                    # No due date set, just mark as completed
                    completed_on_time.append(t)
            
            # Currently overdue tasks
            overdue_tasks = [t for t in member_tasks
                            if t.status != TaskStatus.done
                            and t.due_date
                            and t.due_date < today]
            
            # Currently active tasks
            active_member_tasks = [t for t in member_tasks
                                  if t.status == TaskStatus.in_progress]
            
            user_stats.append({
                'user_id': user.id,
                'user_name': user.full_name or user.email,
                'user_email': user.email,
                'completed_month': len(completed_month),
                'completed_on_time': len(completed_on_time),
                'completed_late': len(completed_late),
                'completed_late_tasks': completed_late,
                'overdue_count': len(overdue_tasks),
                'overdue_tasks': overdue_tasks,
                'active_count': len(active_member_tasks)
            })
    
    stats = {
        'active_tasks': len(active_tasks),
        'completed_week': len(completed_week),
        'overdue': len(overdue),
        'team_members': len(team_members)
    }
    
    return templates.TemplateResponse('activity/feed.html', {
        'request': request,
        'user': user,
        'activities': activities,
        'stats': stats,
        'user_stats': user_stats
    })


# --------------------------
# Workload View (Asana-inspired)
# --------------------------
@router.get('/workload', response_class=HTMLResponse)
async def web_workload(
    request: Request,
    project_id: int = None,
    time_range: str = 'week',
    db: AsyncSession = Depends(get_session)
):
    """Team workload visualization"""
    user_id = request.session.get('user_id')
    if not user_id:
        return RedirectResponse('/web/login', status_code=303)
    
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        return RedirectResponse('/web/login', status_code=303)
    
    from app.models.task import Task
    from app.models.assignment import Assignment
    from app.models.project import Project
    from app.models.project_member import ProjectMember
    from datetime import timedelta
    
    today = date.today()
    
    # Determine date range
    if time_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif time_range == 'month':
        start_date = today.replace(day=1)
        next_month = today.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)
    else:  # quarter
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=quarter_start_month, day=1)
        end_date = (start_date + timedelta(days=92)).replace(day=1) - timedelta(days=1)
    
    # Get accessible projects
    if user.is_admin:
        projects_query = select(Project).where(
            Project.workspace_id == user.workspace_id,
            Project.is_archived == False
        )
    else:
        projects_query = (
            select(Project)
            .join(ProjectMember, ProjectMember.project_id == Project.id)
            .where(
                Project.workspace_id == user.workspace_id,
                ProjectMember.user_id == user_id,
                Project.is_archived == False
            )
        )
    projects = (await db.execute(projects_query)).scalars().all()
    project_ids = [p.id for p in projects]
    
    if project_id and project_id in project_ids:
        project_ids = [project_id]
    
    # Base task query for active tasks
    task_query = select(Task).where(
        Task.project_id.in_(project_ids),
        Task.is_archived == False,
        Task.status != 'done'
    )
    all_tasks = (await db.execute(task_query)).scalars().all()
    
    # Get assignments
    task_ids = [t.id for t in all_tasks]
    assignments = (await db.execute(
        select(Assignment.task_id, Assignment.user_id)
        .where(Assignment.task_id.in_(task_ids))
    )).all() if task_ids else []
    
    assignments_map = {}
    for a in assignments:
        if a.task_id not in assignments_map:
            assignments_map[a.task_id] = []
        assignments_map[a.task_id].append(a.user_id)
    
    # Get team members
    team_members = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id, User.is_active == True)
    )).scalars().all()
    
    # Calculate workload per member
    team_workload = []
    for member in team_members:
        member_tasks = [t for t in all_tasks if member.id in assignments_map.get(t.id, [])]
        todo_count = len([t for t in member_tasks if t.status.value == 'todo'])
        in_progress_count = len([t for t in member_tasks if t.status.value == 'in_progress'])
        blocked_count = len([t for t in member_tasks if t.status.value == 'blocked'])
        overdue_count = len([t for t in member_tasks if t.due_date and t.due_date < today])
        
        team_workload.append({
            'id': member.id,
            'username': member.username,
            'full_name': member.full_name,
            'task_count': len(member_tasks),
            'todo_count': todo_count,
            'in_progress_count': in_progress_count,
            'blocked_count': blocked_count,
            'overdue_count': overdue_count,
        })
    
    # Sort by task count (busiest first)
    team_workload.sort(key=lambda x: x['task_count'], reverse=True)
    
    # Calculate summary stats
    total_tasks = len(all_tasks)
    in_progress_tasks = len([t for t in all_tasks if t.status.value == 'in_progress'])
    overdue_tasks = len([t for t in all_tasks if t.due_date and t.due_date < today])
    unassigned_tasks = len([t for t in all_tasks if t.id not in assignments_map or not assignments_map[t.id]])
    
    # Priority breakdown
    priority_breakdown = {
        'critical': len([t for t in all_tasks if t.priority.value == 'critical']),
        'high': len([t for t in all_tasks if t.priority.value == 'high']),
        'medium': len([t for t in all_tasks if t.priority.value == 'medium']),
        'low': len([t for t in all_tasks if t.priority.value == 'low']),
    }
    
    # Due date breakdown
    week_end = today + timedelta(days=(6 - today.weekday()))
    due_breakdown = {
        'overdue': len([t for t in all_tasks if t.due_date and t.due_date < today]),
        'today': len([t for t in all_tasks if t.due_date and t.due_date == today]),
        'this_week': len([t for t in all_tasks if t.due_date and today < t.due_date <= week_end]),
        'later': len([t for t in all_tasks if t.due_date and t.due_date > week_end]),
        'no_date': len([t for t in all_tasks if not t.due_date]),
    }
    
    # Unassigned tasks list with project names
    project_names = {p.id: p.name for p in projects}
    unassigned_task_list = []
    for t in all_tasks:
        if t.id not in assignments_map or not assignments_map[t.id]:
            unassigned_task_list.append({
                'id': t.id,
                'title': t.title,
                'priority': t.priority.value,
                'due_date': t.due_date,
                'project_name': project_names.get(t.project_id, 'Unknown')
            })
    
    return templates.TemplateResponse('workload/index.html', {
        'request': request,
        'user': user,
        'projects': projects,
        'selected_project_id': project_id,
        'time_range': time_range,
        'team_workload': team_workload,
        'total_tasks': total_tasks,
        'in_progress_tasks': in_progress_tasks,
        'overdue_tasks': overdue_tasks,
        'unassigned_tasks': unassigned_tasks,
        'priority_breakdown': priority_breakdown,
        'due_breakdown': due_breakdown,
        'unassigned_task_list': unassigned_task_list,
    })


# Goals feature removed - routes deleted


# ============ TASK TEMPLATES ============

@router.get('/templates', response_class=HTMLResponse)
async def templates_page(
    request: Request,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """Task templates management page"""
    from app.models.task_extensions import TaskTemplate
    if not user or not user.workspace_id:
        return RedirectResponse(url='/web/login', status_code=303)
    
    # Get templates for this workspace
    templates_list = (await db.execute(
        select(TaskTemplate).where(
            TaskTemplate.workspace_id == user.workspace_id,
            or_(TaskTemplate.is_shared == True, TaskTemplate.created_by_id == user.id)
        ).order_by(TaskTemplate.use_count.desc())
    )).scalars().all()
    
    # Get projects for the "use template" modal
    projects = (await db.execute(
        select(Project).where(Project.workspace_id == user.workspace_id, Project.is_archived == False)
        .order_by(Project.name)
    )).scalars().all()
    
    # Get workspace members
    members = (await db.execute(
        select(User).where(User.workspace_id == user.workspace_id, User.is_active == True)
        .order_by(User.full_name)
    )).scalars().all()
    
    return templates.TemplateResponse("templates/index.html", {
        "request": request,
        "user": user,
        "templates": templates_list,
        "projects": projects,
        "members": members
    })


@router.post('/templates/create')
async def create_template(
    request: Request,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """Create a new task template"""
    from app.models.task_extensions import TaskTemplate
    if not user or not user.workspace_id:
        return RedirectResponse(url='/web/login', status_code=303)
    
    form = await request.form()
    
    template = TaskTemplate(
        workspace_id=user.workspace_id,
        name=form.get('name'),
        title_template=form.get('title_template'),
        description_template=form.get('description_template') or None,
        priority=form.get('priority', 'medium'),
        estimated_hours=float(form.get('estimated_hours')) if form.get('estimated_hours') else None,
        default_tags=form.get('default_tags') or None,
        is_shared='is_shared' in form,
        created_by_id=user.id
    )
    
    db.add(template)
    await db.commit()
    
    return RedirectResponse(url='/web/templates', status_code=303)


@router.post('/templates/{template_id}/update')
async def update_template(
    request: Request,
    template_id: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """Update a task template"""
    from app.models.task_extensions import TaskTemplate
    if not user or not user.workspace_id:
        return RedirectResponse(url='/web/login', status_code=303)
    
    template = (await db.execute(
        select(TaskTemplate).where(
            TaskTemplate.id == template_id,
            TaskTemplate.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not template:
        return RedirectResponse(url='/web/templates', status_code=303)
    
    form = await request.form()
    
    template.name = form.get('name')
    template.title_template = form.get('title_template')
    template.description_template = form.get('description_template') or None
    template.priority = form.get('priority', 'medium')
    template.estimated_hours = float(form.get('estimated_hours')) if form.get('estimated_hours') else None
    template.default_tags = form.get('default_tags') or None
    template.is_shared = 'is_shared' in form
    
    await db.commit()
    
    return RedirectResponse(url='/web/templates', status_code=303)


@router.post('/templates/{template_id}/delete')
async def delete_template(
    template_id: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """Delete a task template"""
    from app.models.task_extensions import TaskTemplate
    if not user or not user.workspace_id:
        return JSONResponse(status_code=401, content={'error': 'Unauthorized'})
    
    template = (await db.execute(
        select(TaskTemplate).where(
            TaskTemplate.id == template_id,
            TaskTemplate.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not template:
        return JSONResponse(status_code=404, content={'error': 'Template not found'})
    
    await db.delete(template)
    await db.commit()
    
    return JSONResponse(status_code=200, content={'success': True})


@router.post('/templates/use')
async def use_template(
    request: Request,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(get_current_user)
):
    """Create a task from a template"""
    from app.models.task_extensions import TaskTemplate
    if not user or not user.workspace_id:
        return RedirectResponse(url='/web/login', status_code=303)
    
    form = await request.form()
    template_id = int(form.get('template_id'))
    
    template = (await db.execute(
        select(TaskTemplate).where(
            TaskTemplate.id == template_id,
            TaskTemplate.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    
    if not template:
        return RedirectResponse(url='/web/templates', status_code=303)
    
    # Create the task from template
    from datetime import datetime
    task = Task(
        title=form.get('title') or template.title_template,
        description=template.description_template,
        project_id=int(form.get('project_id')),
        priority=template.priority,
        status='pending',
        creator_id=user.id,
        due_date=datetime.strptime(form.get('due_date'), '%Y-%m-%d') if form.get('due_date') else None
    )
    
    db.add(task)
    await db.flush()
    
    # Auto-assign to creator
    from app.models.assignment import Assignment
    assignment = Assignment(task_id=task.id, assignee_id=user.id)
    db.add(assignment)
    
    # If specific user was assigned, add them too
    assigned_to_id = form.get('assigned_to_id')
    if assigned_to_id and int(assigned_to_id) != user.id:
        extra_assignment = Assignment(task_id=task.id, assignee_id=int(assigned_to_id))
        db.add(extra_assignment)
    
    # Increment template use count
    template.use_count += 1
    
    await db.commit()
    await db.refresh(task)
    
    return RedirectResponse(url=f'/web/tasks/{task.id}', status_code=303)


# Calls feature removed - routes deleted


# =====================================================
# IT KNOWLEDGE BASE - Solutions, Diagnostics, Resolved Cases
# =====================================================

from app.models.knowledge_base import KBDiagnosticTree, KBResolvedCase


@router.get('/knowledge-base', response_class=HTMLResponse)
async def web_knowledge_base(
    request: Request,
    category: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_session)
):
    """Main Knowledge Base / Solutions page"""
    user = await get_current_user(request, db)

    # Recent articles from SupportArticle
    article_query = select(SupportArticle).where(
        SupportArticle.workspace_id == user.workspace_id,
        SupportArticle.is_active == True
    )
    if category:
        article_query = article_query.where(SupportArticle.category == category)
    article_query = article_query.order_by(SupportArticle.created_at.desc()).limit(12)
    recent_articles = (await db.execute(article_query)).scalars().all()

    # Recent resolved cases
    case_query = select(KBResolvedCase).where(
        KBResolvedCase.workspace_id == user.workspace_id
    )
    if category:
        case_query = case_query.where(KBResolvedCase.category == category)
    case_query = case_query.order_by(KBResolvedCase.created_at.desc()).limit(12)
    recent_cases = (await db.execute(case_query)).scalars().all()

    # Categories
    categories = (await db.execute(
        select(SupportCategory).where(
            SupportCategory.workspace_id == user.workspace_id,
            SupportCategory.is_active == True
        ).order_by(SupportCategory.name)
    )).scalars().all()

    return templates.TemplateResponse('knowledge_base/index.html', {
        'request': request,
        'user': user,
        'recent_articles': recent_articles,
        'recent_cases': recent_cases,
        'categories': categories,
        'all_categories': categories,
        'search_query': search or '',
        'active_tab': 'search',
    })


@router.get('/knowledge-base/search', response_class=JSONResponse)
async def web_kb_search(
    request: Request,
    q: str = '',
    db: AsyncSession = Depends(get_session)
):
    """Fuzzy search across articles and resolved cases"""
    user = await get_current_user(request, db)
    if not q.strip():
        return {'results': []}

    search_term = q.strip()
    keywords = [w for w in search_term.lower().split() if len(w) >= 2]
    results = []

    # Search SupportArticle
    for kw in keywords:
        pattern = f'%{kw}%'
        articles = (await db.execute(
            select(SupportArticle).where(
                SupportArticle.workspace_id == user.workspace_id,
                SupportArticle.is_active == True,
                or_(
                    SupportArticle.problem_title.ilike(pattern),
                    SupportArticle.problem_description.ilike(pattern),
                    SupportArticle.problem_keywords.ilike(pattern),
                    SupportArticle.solution_steps.ilike(pattern),
                )
            ).order_by(SupportArticle.times_helpful.desc()).limit(10)
        )).scalars().all()
        for a in articles:
            if not any(r['id'] == f'article-{a.id}' for r in results):
                results.append({
                    'id': f'article-{a.id}',
                    'type': 'article',
                    'title': a.problem_title,
                    'description': a.problem_description[:200],
                    'category': a.category,
                    'tags': a.problem_keywords or '',
                    'is_verified': a.is_verified,
                    'helpful_votes': a.times_helpful,
                    'url': f'/web/knowledge-base/article/{a.id}'
                })

    # Search KBResolvedCase
    for kw in keywords:
        pattern = f'%{kw}%'
        cases = (await db.execute(
            select(KBResolvedCase).where(
                KBResolvedCase.workspace_id == user.workspace_id,
                or_(
                    KBResolvedCase.problem_title.ilike(pattern),
                    KBResolvedCase.problem_description.ilike(pattern),
                    KBResolvedCase.solution_steps.ilike(pattern),
                    KBResolvedCase.error_message.ilike(pattern),
                    KBResolvedCase.tags.ilike(pattern),
                    KBResolvedCase.device_brand.ilike(pattern),
                    KBResolvedCase.root_cause.ilike(pattern),
                )
            ).order_by(KBResolvedCase.helpful_votes.desc()).limit(10)
        )).scalars().all()
        for c in cases:
            if not any(r['id'] == f'case-{c.id}' for r in results):
                results.append({
                    'id': f'case-{c.id}',
                    'type': 'resolved_case',
                    'title': c.problem_title,
                    'description': c.problem_description[:200],
                    'category': c.category,
                    'tags': c.tags or '',
                    'is_verified': c.is_verified,
                    'helpful_votes': c.helpful_votes,
                    'url': f'/web/knowledge-base/resolved/{c.id}'
                })

    return {'results': results[:20]}


@router.get('/knowledge-base/article/{article_id}', response_class=HTMLResponse)
async def web_kb_article_detail(
    request: Request,
    article_id: int,
    db: AsyncSession = Depends(get_session)
):
    """View a single KB article"""
    user = await get_current_user(request, db)
    article = (await db.execute(
        select(SupportArticle).where(
            SupportArticle.id == article_id,
            SupportArticle.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")

    # Increment view count
    article.times_shown += 1
    await db.commit()

    # Related articles (same category)
    related_articles = (await db.execute(
        select(SupportArticle).where(
            SupportArticle.workspace_id == user.workspace_id,
            SupportArticle.category == article.category,
            SupportArticle.id != article.id,
            SupportArticle.is_active == True
        ).order_by(SupportArticle.times_helpful.desc()).limit(5)
    )).scalars().all()

    return templates.TemplateResponse('knowledge_base/article.html', {
        'request': request,
        'user': user,
        'article': article,
        'related_articles': related_articles,
    })


@router.post('/knowledge-base/article/{article_id}/rate')
async def web_kb_article_rate(
    request: Request,
    article_id: int,
    helpful: str = Form('true'),
    db: AsyncSession = Depends(get_session)
):
    """Rate an article helpful/not helpful"""
    user = await get_current_user(request, db)
    article = (await db.execute(
        select(SupportArticle).where(
            SupportArticle.id == article_id,
            SupportArticle.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    if article:
        if helpful == 'true':
            article.times_helpful += 1
        else:
            article.times_not_helpful += 1
        # Recalculate success rate
        total = article.times_helpful + article.times_not_helpful
        article.success_rate = (article.times_helpful / total * 100) if total > 0 else 0
        await db.commit()
    return RedirectResponse(url=f'/web/knowledge-base/article/{article_id}', status_code=303)


@router.post('/knowledge-base/articles', response_class=JSONResponse)
async def web_kb_create_article(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Create a new KB article (from the Add Solution modal)"""
    user = await get_current_user(request, db)
    data = await request.json()

    problem_title = (data.get('problem_title') or '').strip()
    problem_description = (data.get('problem_description') or '').strip()
    category = (data.get('category') or '').strip()
    solution_steps = (data.get('solution_steps') or '').strip()

    if not problem_title or not problem_description or not category or not solution_steps:
        return JSONResponse({'error': 'Missing required fields'}, status_code=400)

    article = SupportArticle(
        workspace_id=user.workspace_id,
        problem_title=problem_title,
        problem_description=problem_description,
        problem_keywords=(data.get('keywords') or '').strip(),
        category=category,
        solution_steps=solution_steps,
        solution_source='manual',
        created_by_id=user.id,
        is_active=True,
    )
    db.add(article)
    await db.commit()
    return {'ok': True, 'id': article.id}


# --- Diagnostic Engine routes ---

@router.get('/knowledge-base/diagnostic/start', response_class=JSONResponse)
async def web_kb_diagnostic_start(
    request: Request,
    category: str = '',
    db: AsyncSession = Depends(get_session)
):
    """Get root diagnostic node(s) for a category"""
    user = await get_current_user(request, db)

    # Find root nodes (parent_id is None) for this category
    roots = (await db.execute(
        select(KBDiagnosticTree).where(
            KBDiagnosticTree.workspace_id == user.workspace_id,
            KBDiagnosticTree.parent_id == None,
            KBDiagnosticTree.is_active == True,
            KBDiagnosticTree.category == category
        ).order_by(KBDiagnosticTree.sort_order)
    )).scalars().all()

    if not roots:
        return {'node': None, 'children': []}

    # If single root, return it with its children
    root = roots[0]
    children = (await db.execute(
        select(KBDiagnosticTree).where(
            KBDiagnosticTree.parent_id == root.id,
            KBDiagnosticTree.is_active == True
        ).order_by(KBDiagnosticTree.sort_order)
    )).scalars().all()

    return {
        'node': {
            'id': root.id,
            'title': root.title,
            'node_type': root.node_type,
            'question_text': root.question_text,
            'solution_text': root.solution_text,
        },
        'children': [
            {
                'id': c.id,
                'title': c.title,
                'node_type': c.node_type,
                'question_text': c.question_text,
                'solution_text': c.solution_text,
            } for c in children
        ]
    }


@router.get('/knowledge-base/diagnostic/step/{node_id}', response_class=JSONResponse)
async def web_kb_diagnostic_step(
    request: Request,
    node_id: int,
    db: AsyncSession = Depends(get_session)
):
    """Get a diagnostic tree node and its children"""
    user = await get_current_user(request, db)

    node = (await db.execute(
        select(KBDiagnosticTree).where(
            KBDiagnosticTree.id == node_id,
            KBDiagnosticTree.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()

    if not node:
        return {'node': None, 'children': []}

    children = (await db.execute(
        select(KBDiagnosticTree).where(
            KBDiagnosticTree.parent_id == node.id,
            KBDiagnosticTree.is_active == True
        ).order_by(KBDiagnosticTree.sort_order)
    )).scalars().all()

    return {
        'node': {
            'id': node.id,
            'title': node.title,
            'node_type': node.node_type,
            'question_text': node.question_text,
            'solution_text': node.solution_text,
        },
        'children': [
            {
                'id': c.id,
                'title': c.title,
                'node_type': c.node_type,
                'question_text': c.question_text,
                'solution_text': c.solution_text,
            } for c in children
        ]
    }


@router.post('/knowledge-base/diagnostic/rate', response_class=JSONResponse)
async def web_kb_diagnostic_rate(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Rate a diagnostic solution node"""
    user = await get_current_user(request, db)
    data = await request.json()
    # Placeholder acknowledgment - could track per-node feedback in the future
    return {'ok': True}


# --- Resolved Cases routes ---

@router.get('/knowledge-base/resolved-cases', response_class=HTMLResponse)
async def web_kb_resolved_cases(
    request: Request,
    search: Optional[str] = None,
    category: Optional[str] = None,
    device_type: Optional[str] = None,
    verified: Optional[str] = None,
    db: AsyncSession = Depends(get_session)
):
    """List resolved cases"""
    user = await get_current_user(request, db)

    query = select(KBResolvedCase).where(
        KBResolvedCase.workspace_id == user.workspace_id
    )

    if search:
        pattern = f'%{search.strip()}%'
        query = query.where(or_(
            KBResolvedCase.problem_title.ilike(pattern),
            KBResolvedCase.problem_description.ilike(pattern),
            KBResolvedCase.solution_steps.ilike(pattern),
            KBResolvedCase.error_message.ilike(pattern),
            KBResolvedCase.tags.ilike(pattern),
            KBResolvedCase.device_brand.ilike(pattern),
            KBResolvedCase.ticket_number.ilike(pattern),
        ))
    if category:
        query = query.where(KBResolvedCase.category == category)
    if device_type:
        query = query.where(KBResolvedCase.device_type == device_type)
    if verified == 'yes':
        query = query.where(KBResolvedCase.is_verified == True)

    query = query.order_by(KBResolvedCase.created_at.desc()).limit(50)
    cases = (await db.execute(query)).scalars().all()

    return templates.TemplateResponse('knowledge_base/resolved_cases.html', {
        'request': request,
        'user': user,
        'cases': cases,
        'search_query': search or '',
        'category_filter': category or '',
        'device_filter': device_type or '',
        'verified_filter': verified or '',
    })


@router.get('/knowledge-base/resolved-cases/new', response_class=HTMLResponse)
async def web_kb_resolved_case_new(
    request: Request,
    db: AsyncSession = Depends(get_session)
):
    """Form to log a new resolved case"""
    user = await get_current_user(request, db)
    return templates.TemplateResponse('knowledge_base/resolved_new.html', {
        'request': request,
        'user': user,
        'error': None,
        'form_data': None,
    })


@router.post('/knowledge-base/resolved-cases/new')
async def web_kb_resolved_case_create(
    request: Request,
    problem_title: str = Form(...),
    problem_description: str = Form(...),
    category: str = Form(...),
    solution_steps: str = Form(...),
    error_message: str = Form(''),
    device_type: str = Form(''),
    device_brand: str = Form(''),
    device_model: str = Form(''),
    connection_type: str = Form(''),
    tags: str = Form(''),
    root_cause: str = Form(''),
    time_to_resolve: Optional[int] = Form(None),
    ticket_number: str = Form(''),
    db: AsyncSession = Depends(get_session)
):
    """Save a new resolved case"""
    user = await get_current_user(request, db)

    if not problem_title.strip() or not problem_description.strip() or not solution_steps.strip():
        return templates.TemplateResponse('knowledge_base/resolved_new.html', {
            'request': request,
            'user': user,
            'error': 'Problem title, description, and solution steps are required.',
            'form_data': {
                'problem_title': problem_title,
                'problem_description': problem_description,
                'category': category,
                'solution_steps': solution_steps,
                'error_message': error_message,
                'device_type': device_type,
                'device_brand': device_brand,
                'device_model': device_model,
                'connection_type': connection_type,
                'tags': tags,
                'root_cause': root_cause,
                'time_to_resolve': time_to_resolve,
                'ticket_number': ticket_number,
            }
        }, status_code=400)

    case = KBResolvedCase(
        workspace_id=user.workspace_id,
        problem_title=problem_title.strip(),
        problem_description=problem_description.strip(),
        error_message=error_message.strip() or None,
        category=category.strip(),
        device_type=device_type.strip() or None,
        device_brand=device_brand.strip() or None,
        device_model=device_model.strip() or None,
        connection_type=connection_type.strip() or None,
        tags=tags.strip() or None,
        solution_steps=solution_steps.strip(),
        root_cause=root_cause.strip() or None,
        time_to_resolve=time_to_resolve,
        ticket_number=ticket_number.strip() or None,
        resolved_by_id=user.id,
        resolved_by_name=user.full_name or user.username,
    )
    db.add(case)
    await db.commit()

    request.session['success_message'] = 'Resolved case logged successfully! Thank you for contributing to the knowledge base.'
    return RedirectResponse(url='/web/knowledge-base/resolved-cases', status_code=303)


@router.get('/knowledge-base/resolved/{case_id}', response_class=HTMLResponse)
async def web_kb_resolved_detail(
    request: Request,
    case_id: int,
    db: AsyncSession = Depends(get_session)
):
    """View a resolved case in detail"""
    user = await get_current_user(request, db)
    case = (await db.execute(
        select(KBResolvedCase).where(
            KBResolvedCase.id == case_id,
            KBResolvedCase.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Resolved case not found")

    # Increment view count
    case.times_viewed += 1
    await db.commit()

    # Related cases (same category)
    related_cases = (await db.execute(
        select(KBResolvedCase).where(
            KBResolvedCase.workspace_id == user.workspace_id,
            KBResolvedCase.category == case.category,
            KBResolvedCase.id != case.id
        ).order_by(KBResolvedCase.helpful_votes.desc()).limit(5)
    )).scalars().all()

    return templates.TemplateResponse('knowledge_base/resolved_detail.html', {
        'request': request,
        'user': user,
        'case': case,
        'related_cases': related_cases,
    })


@router.post('/knowledge-base/resolved/{case_id}/rate')
async def web_kb_resolved_rate(
    request: Request,
    case_id: int,
    helpful: str = Form('true'),
    db: AsyncSession = Depends(get_session)
):
    """Rate a resolved case"""
    user = await get_current_user(request, db)
    case = (await db.execute(
        select(KBResolvedCase).where(
            KBResolvedCase.id == case_id,
            KBResolvedCase.workspace_id == user.workspace_id
        )
    )).scalar_one_or_none()
    if case:
        if helpful == 'true':
            case.helpful_votes += 1
        else:
            case.not_helpful_votes += 1
        await db.commit()
    return RedirectResponse(url=f'/web/knowledge-base/resolved/{case_id}', status_code=303)
